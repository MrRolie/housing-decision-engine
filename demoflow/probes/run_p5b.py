"""P5b — pick the temporary-resident STOCK source (spec §11 item 5b, codex F5).

WHO USES THIS (spec §7 / spec:473): the TRIPWIRE BASELINE registry carries a
`temporary-resident stock` indicator declared `wired`, i.e. (indicator, current value,
source, as_of, threshold band) with a freshness limit. Until that indicator is wired the
tripwire reports **UNKNOWN**, never a stale within-band. P5b's job is therefore to PICK
ONE source and RECORD its schema + cadence from live responses — never to invent one.

THE DISCRIMINATOR, stated before the hunt so the verdict cannot drift to fit the answer:
a candidate must carry temporary-resident **STOCK** (a count of persons present at a
reference date — spec:473 consumes a level, not a movement), at a stated **GEOGRAPHY**,
at a stated **CADENCE**, and be **CURRENTLY MAINTAINED** (a tripwire needs a fresh
current value). All four are computed below from live responses AND all four GATE the
verdict: geography, cadence and currency are enforced by the pick rule and the floor
guard, and the STOCK discriminator decides between the `LOCATED` and `LOCATED-NOT-STOCK`
tokens. (It did not always: `measure` was once computed AFTER the pick and could not
affect the verdict, leaving this very sentence an untied claim in the file that lectures
about them.)

WHAT THE HUNT ESTABLISHED (recorded live, so a re-run re-derives it):
  * The pick is resolved by a SWEEP over the full StatCan cube catalogue
    (`getAllCubesListLite`), not from a guessed product id. Every absence claim in this
    note is SCOPED to that sweep ("not among the N cubes swept") — never "does not exist".
  * The pick carries temporary-resident stock at PROVINCE/TERRITORY level only. The two
    CMAs demoflow models are ABSENT from its geography members, searched BY NAME. That
    limit is RECORDED as evidence, not filled in from elsewhere (Ruling D): the swept
    cube that DOES carry the modeled CMAs is a one-time Census snapshot whose
    `cubeStartDate == cubeEndDate`, so it cannot supply a fresh current value. This note
    proposes NO combination of the two — they are different programs, and the pick's own
    live footnotes caution against mixing programs (quoted verbatim below).

ANTI-FABRICATION (the cardinal discipline). Every VALUE reported as observed is emitted
by THIS run from a live response: the resolved product id, the title, the cadence label
(resolved from the live `getCodeSets`, never typed beside a raw frequency code), the
geography member COUNT and the member list itself, the CMA name-search results, the
measure-type marker hits, and the archive status. Interpretive glosses are FUNCTIONS of
those measured values, so a gloss cannot contradict the number it sits beside. The three
verbatim footnote quotes are CITED, resolved from the live footnote list by predicate.

FLOOR GUARD (NS #1 — a verification gate must REFUSE when it cannot verify). The WDS
answers 200 with vacuous bodies: an empty catalogue, a sweep matching zero candidates, a
`getCubeMetadata` object carrying no `dimension` list, a geography dimension with zero
members, or a `frequencyCode` the live code set does not define WITH A NON-EMPTY LABEL
(a code present but described by a blank string published `DECISION-CADENCE: (frequencyCode
9, ...)` with an empty label under a LOCATED — a proven false-green on this guard's own
declared axis, closed by dropping label-less codes from the map). None may launder into
a LOCATED. The guard
RAISES `VacuousProbeError` -> UNKNOWN-PROBE-FAILED with the failing boundary recorded.

TWO COMPARISONS ARE RECORDED BUT DO NOT GATE THE VERDICT (§4, §5): the spec-named IRCC
temporary-resident tables (open.canada.ca) and the in-repo ISQ compo column. The pick is
EARNED from the StatCan sweep; these two are alternatives it is measured against. Each
emits its measured result OR an explicit `NOT MEASURED THIS RUN`, and the pick's scope
clause is a function of which — so an unmeasured comparison can never read as a
comparison that was made.

Run:  cd demoflow && uv run python probes/run_p5b.py
"""

import json
import urllib.request
from pathlib import Path

# Flat, NOT `probes._wds`: probes/ is deliberately not a package, so in script mode
# sys.path[0] IS probes/ and this resolves natively. See probes/_wds.py.
from _wds import (WDS_LIST, WDS_META, WDS_TIMEOUT, Fact, new_run, post as _post,
                  provenance_header, table_number, table_url)

# --- verdict-gating boundaries (all www150; the reachability probe is one POST) ------
# `getCodeSets` resolves frequencyCode -> label LIVE, so the cadence in this note is
# never a label hand-typed beside a raw integer.
WDS_CODESETS = "https://www150.statcan.gc.ca/t1/wds/rest/getCodeSets"

# --- recorded-but-NON-GATING comparison boundaries ----------------------------------
CKAN_SEARCH = ("https://open.canada.ca/data/api/3/action/package_search"
               "?q=temporary+residents&rows=100")
IRCC_ORG_MARK = "immigration, refugees and citizenship canada"
# A package titled for ONE permit class publishes a subset of the temporary-resident
# population; spec:473 consumes the total. Applied as a scope predicate in §4.
PERMIT_CLASS_TERMS = ("study permit", "work permit", "visa", "application")

# The ISQ compo workbook committed with the grounding research. Read-only, and OUTSIDE
# demoflow/ — hence non-gating: its absence says nothing about the StatCan pick, and
# gating on it would turn a relocated research file into a red on a located source.
_REPO_ROOT = Path(__file__).resolve().parents[2]
ISQ_COMPO = (_REPO_ROOT / "docs" / "research"
             / "2026-07-21-demographic-housing-flow-grounding" / "data"
             / "compo-rmr-base.xlsx")
# 0-indexed, per the plan's pinned ISQ compo layout. The header rows are SEARCHED for the
# NPR column rather than indexed at the plan's column 18 — the found index is emitted and
# cross-checked against 18, so a layout drift surfaces instead of reading a wrong cell.
ISQ_HEADER_ROWS = range(5, 10)
ISQ_FIRST_DATA_ROW = 10
ISQ_PLAN_COL = 18

# --- sweep predicate ----------------------------------------------------------------
# POPULATION terms name the temporary-resident population itself — only these cubes are
# eligible to be picked. PERIPHERAL terms name permit/worker programmes; they are swept
# (so the absence claim is scoped to a WIDER population than the pick pool) and each is
# rejected with the measured reason. Multi-word PHRASES, never the bare "NPR"
# abbreviation: §1 measures what a bare-abbreviation predicate would have matched.
POPULATION_TERMS = ("non-permanent resident", "non permanent resident", "temporary resident")
PERIPHERAL_TERMS = ("temporary foreign worker", "study permit", "work permit")
ABBREV_TRAP_TERM = "npr"

# The two CMAs demoflow models, mapped to the province that CONTAINS each. The map is the
# single datum: MODELED_CMAS is derived from its keys, so the two cannot drift apart.
#
# The containment is DECLARED here rather than asserted in prose downstream. The note used
# to call a member "the coarsest geography containing them that IS present" — a containment
# relation nothing computed or cited: swap these keys for Toronto/Vancouver and the note
# would still hunt "Quebec" and still claim it contained them. Declaring it means the claim
# moves with the constant it depends on.
MODELED_CMA_PROVINCE = {"Montréal": "Quebec", "Québec": "Quebec"}
# A cube TITLE saying "census metropolitan areas" is NOT evidence its members contain these
# — only the live member search is.
MODELED_CMAS = tuple(MODELED_CMA_PROVINCE)
CMA_MARKERS = ("(CMA)", "census metropolitan", "RMR")
CA_MARKER = "(CA)"

# Measure-type markers, matched over title + dimension names + member names. Reported as
# the MARKER HITS themselves, never as a bare verdict token per candidate: a one-token
# match is a weak classifier and must not launder into a stated fact.
STOCK_MARKERS = ("number of", "as of", "on july 1", "on january 1", "on december 31",
                 "count", "population", "stock", "living in", "with a valid permit")
FLOW_MARKERS = ("solde", "components of", "growth", "arrival", "admission", "landing",
                "entries", "flow", "emigration", "became effective", "balance",
                "change in", "net migration")

# Footnote predicates — each resolves a VERBATIM quote out of the pick's live footnotes.
POINT_IN_TIME_MARKERS = ("q1 =", "q2 =", "q3 =", "q4 =", "as of", "on july 1",
                         "on january 1", "on december 31")
DEFINITION_MARKER = "non-permanent resident refers to"
CROSS_PROGRAMME_MARKERS = ("caution", "compar")

OUT = Path(__file__).resolve().parent / "P5b-temp-resident-stock.md"
CKAN_TIMEOUT = 120


# --- this note's provenance prose (the shared header skeleton lives in _wds) ---------
# The filename this note must attribute itself to. DERIVED from __file__, never typed:
# `written_by` is the one header field a copy-pasted call block carries forward silently
# — a probe cloned from p5 would publish "Written by run_p5.py" over its own computed
# body, exactly the untied claim this registry exists to stop.
_WRITTEN_BY = Path(__file__).name
_SCOPE = ("SCOPE OF THIS HEADER (it claims only what it can enforce): the catalogue size, the "
          "swept-candidate list, the resolved product id and title, the cadence label and its "
          "reference-period span, the archive status, the geography member COUNT and the member "
          "list, the CMA name-search results and the measure-type marker hits in §1-§3 are all "
          "emitted by this run from the live StatCan WDS responses — the cadence label is "
          "resolved from the live getCodeSets, never typed beside a raw frequency code. The "
          "footnote quotes are verbatim from the pick's live footnote list. §4 (IRCC / CKAN) and "
          "§5 (ISQ compo) are RECORDED COMPARISONS that do NOT gate the verdict: each states "
          "either its measured result or an explicit NOT MEASURED THIS RUN, and the pick's scope "
          "clause is a function of which. Every absence claim is scoped to what was swept.")
_CITED_LABEL = "Quoted verbatim from the live response:"


def _summary(*, total: int, derived: int, cited: int) -> str:
    """The provenance sentence, sized to what this run actually registered."""
    return (
        f"This run registered {total} provenance-tagged figures: {derived} DERIVED "
        f"(computed from the live responses of this run) and {cited} CITED (verbatim from "
        f"the live cube footnotes). Untagged numerals elsewhere are audit metadata (candidate "
        f"counts, resource counts, member counts, column indices) and reference labels "
        f"(product ids, reference dates, the ISQ header/data row positions), each traceable to "
        f"the live response or the named file this run read."
    )


# --- network boundaries (injectable seams so the floor-guard test runs OFFLINE) ------
def _codesets() -> dict:
    """`getCodeSets` -> parsed JSON. Boundary `wds-codesets` (www150).

    Deliberately NOT `pd.read_json` (the repeat trap this probe family keeps hitting on
    both the WDS and the CKAN shapes): navigate the documented envelope
    {"status", "object": {"frequency": [...], ...}} explicitly.
    """
    raw = urllib.request.urlopen(WDS_CODESETS, timeout=WDS_TIMEOUT).read()
    return json.loads(raw)


def _catalogue() -> list[dict]:
    """`getAllCubesListLite` -> the full cube catalogue. Boundary `wds-list` (www150)."""
    raw = urllib.request.urlopen(WDS_LIST, timeout=WDS_TIMEOUT).read()
    return json.loads(raw)


def _meta(pids: list[str]) -> dict[str, dict]:
    """`getCubeMetadata` for many productIds -> {pid: object}. Boundary `wds-meta` (POST).

    One batch: the swept pool is small by construction (a phrase predicate over titles),
    so no chunking constant is introduced for a population that does not need one.
    """
    out: dict[str, dict] = {}
    for r in _post(WDS_META, [{"productId": int(p)} for p in pids]):
        obj = r.get("object") or {}
        pid = str(obj.get("productId") or "")
        if pid:
            out[pid] = obj
    return out


def _ckan_search() -> dict:
    """CKAN `package_search` -> parsed JSON. NON-GATING boundary (open.canada.ca)."""
    raw = urllib.request.urlopen(CKAN_SEARCH, timeout=CKAN_TIMEOUT).read()
    return json.loads(raw)


def _isq_rows() -> list[tuple]:
    """The ISQ compo header+first-data block. NON-GATING boundary (a committed workbook).

    `openpyxl` directly, NOT pandas: the plan's sketch reached for pandas on a JSON
    boundary and mis-parsed it; nothing here needs a DataFrame, and the raw cell grid is
    what the column search reads.
    """
    from openpyxl import load_workbook

    wb = load_workbook(ISQ_COMPO, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        return [row for row in ws.iter_rows(min_row=1, max_row=ISQ_FIRST_DATA_ROW + 1,
                                            values_only=True)]
    finally:
        wb.close()


# --- small computed helpers ---------------------------------------------------------
def _norm(text: str) -> str:
    """The ONE name-normalisation rule in this file.

    `_cma_hits` and the containing-province search previously used two different rules for
    the same names (substring-lower vs strip-lower against an accent-variant tuple). Two
    rules for one job is what a later probe copies wrongly; both now call this.
    """
    return (text or "").strip().lower()


def _lower_terms(text: str, terms) -> list[str]:
    low = _norm(text)
    return [t for t in terms if t.lower() in low]


def _count_str(result: dict) -> str:
    """CKAN's total-match count, or an explicit marker when the key is absent.

    `result.get("count")` printed a bare `None` into the note if CKAN omitted the key — a
    missing measurement rendered as if it were one.
    """
    n = result.get("count")
    return str(n) if isinstance(n, int) else "an UNREPORTED number of"


def _trunc_str(result: dict, results: list) -> str:
    """Says so when `rows=100` truncated the population the sweep actually saw.

    Without this, a sweep over the first 100 of N results reads as a sweep over all N —
    which would silently widen every absence claim scoped to it.
    """
    n = result.get("count")
    if isinstance(n, int) and n > len(results):
        return f" (TRUNCATED: {len(results)} of {n} — absence claims below cover only these)"
    return ""


def _dimensions(obj: dict) -> list[dict]:
    return obj.get("dimension") or []


def _geo_dimension(obj: dict) -> dict | None:
    for d in _dimensions(obj):
        if (d.get("dimensionNameEn") or "").lower().startswith("geo"):
            return d
    return None


def _geo_members(obj: dict) -> list[str]:
    d = _geo_dimension(obj) or {}
    return [(m.get("memberNameEn") or "") for m in (d.get("member") or [])]


def _searchable_text(obj: dict) -> str:
    """Title + every dimension name + every member name, for marker matching."""
    parts = [obj.get("cubeTitleEn") or ""]
    for d in _dimensions(obj):
        parts.append(d.get("dimensionNameEn") or "")
        parts += [(m.get("memberNameEn") or "") for m in (d.get("member") or [])]
    return " | ".join(parts)


def _cma_hits(members: list[str], name: str) -> tuple[list[str], list[str]]:
    """(name-only hits, name+CMA-marker hits) for one modeled CMA.

    Two tiers on purpose. The marker tier is what earns CMA-AVAILABLE, because a bare
    name substring also matches a province member. The name-only tier exists so a cube
    whose CMA members are formatted WITHOUT a "(CMA)" marker surfaces as a visible
    discrepancy (name hits > 0 while marker hits == 0) instead of a silent NO.
    """
    low = _norm(name)
    named = [m for m in members if low in _norm(m)]
    marked = [m for m in named if any(k.lower() in _norm(m) for k in CMA_MARKERS)]
    return named, marked


def _footnote_quote(obj: dict, *, all_of=(), any_of=()) -> str:
    """The first live English footnote matching the predicate, verbatim. "" if none."""
    for f in obj.get("footnote") or []:
        text = (f.get("footnotesEn") or "").strip()
        low = text.lower()
        if all_of and not all(t.lower() in low for t in all_of):
            continue
        if any_of and not any(t.lower() in low for t in any_of):
            continue
        if text:
            return text
    return ""


def _is_current(obj: dict) -> bool:
    return (obj.get("archiveStatusEn") or "").strip().upper().startswith("CURRENT")


# --- floor guard (standalone so it can be mutation-tested) --------------------------
class VacuousProbeError(RuntimeError):
    """A LOCATED must be EARNED. Raised when a boundary answered but with a body that
    cannot support the verdict: an empty code set, an empty catalogue, a sweep matching
    zero candidates, a pick whose metadata carries no dimensions or a geography dimension
    with zero members, or a frequencyCode the live code set does not define. Routes to
    UNKNOWN-PROBE-FAILED with the failing boundary, never a fabricated LOCATED."""


def _guard_codesets(freq_map: dict) -> None:
    """`wds-codesets` guard: without a live frequency map the cadence label would have to
    be hand-typed beside a raw integer — the exact untied claim house rule #1 forbids."""
    if not freq_map:
        raise VacuousProbeError(
            "getCodeSets answered but defined no frequency codes — the cadence label "
            "cannot be resolved live, and a typed label beside a raw frequencyCode is a "
            "claim nothing computed"
        )


def _guard_sweep(cubes: list, swept: list, pool: list) -> None:
    """`wds-list` guard: an absence claim must be scoped to a population that EXISTS.
    An empty-but-200 catalogue, or a sweep resolving no eligible candidate, cannot earn
    a pick over a population of zero."""
    if not cubes:
        raise VacuousProbeError("getAllCubesListLite answered 200 but with an empty catalogue")
    if not swept:
        raise VacuousProbeError(
            f"the catalogue sweep over {len(cubes)} cubes matched ZERO titles on the "
            f"population/peripheral term predicate — the sweep resolved nothing to pick from"
        )
    if not pool:
        raise VacuousProbeError(
            f"{len(swept)} title(s) matched the sweep but NONE matched a temporary-resident "
            f"POPULATION term — no candidate is eligible to be picked"
        )


def _guard_pick(pid: str, obj: dict, geo_dim: dict | None, members: list,
                freq_map: dict) -> None:
    """`wds-meta` guard — the load-bearing one (see the mutation test).

    A `getCubeMetadata` 200 can carry an object stripped of its `dimension` list, or a
    geography dimension with zero members, or a frequencyCode the code set does not
    define. Each still flows through every downstream computation without raising, and
    would publish a LOCATED whose geography evidence is an empty list and whose cadence
    is an undefined integer. REFUSE instead."""
    if not obj:
        raise VacuousProbeError(f"getCubeMetadata resolved no object for the pick {pid}")
    if not _dimensions(obj):
        raise VacuousProbeError(
            f"the pick {pid} came back with an EMPTY dimension list — its schema, geography "
            f"and measure evidence would all be computed over nothing"
        )
    if geo_dim is None:
        raise VacuousProbeError(
            f"the pick {pid} carries no Geography dimension — its geography level and the "
            f"CMA search would both be claims over an absent axis"
        )
    if not members:
        raise VacuousProbeError(
            f"the pick {pid}'s Geography dimension carries ZERO members — a CMA-AVAILABLE "
            f"answer over an empty member list is unearned"
        )
    code = obj.get("frequencyCode")
    if code not in freq_map:
        raise VacuousProbeError(
            f"the pick {pid}'s frequencyCode {code!r} is not defined in the live code set "
            f"({len(freq_map)} codes) — the cadence label cannot be resolved"
        )


# --- the live hunt ------------------------------------------------------------------
def _hunt(note: list[str], stage: dict) -> tuple[str, dict]:
    """Run the live sweep, append its record to `note`, return (verdict, evidence).

    `stage["at"]` tracks the current boundary so a raise can be attributed. Verdict is
    "LOCATED" on success; any raise is handled by main() -> UNKNOWN-PROBE-FAILED.
    """
    # ============ boundary 1: getCodeSets (cadence labels, resolved live) ============
    stage["at"] = "wds-codesets"
    cs = (_codesets().get("object") or {})
    # A code whose DESCRIPTION is missing/blank is DROPPED, not mapped to "". Coercing it
    # to "" put a code in the map with no label, so `_guard_codesets` (map non-empty) and
    # `_guard_pick` (`code in freq_map`) both passed while the note published
    # "DECISION-CADENCE: (frequencyCode 9, label from the live getCodeSets)" with an EMPTY
    # label — a false-green on the exact axis `_guard_codesets` exists to defend. Dropping
    # the entry makes the existing `code not in freq_map` check cover it for free.
    freq_map = {f.get("frequencyCode"): (f.get("frequencyDescEn") or "").strip()
                for f in (cs.get("frequency") or [])
                if f.get("frequencyCode") is not None
                and (f.get("frequencyDescEn") or "").strip()}
    _guard_codesets(freq_map)  # RAISES (wds-codesets)

    # ============ boundary 2: getAllCubesListLite (the enumerated sweep) =============
    stage["at"] = "wds-list"
    cubes = _catalogue()
    swept: list[tuple[str, str, list[str], list[str]]] = []
    for c in cubes:
        pid = str(c.get("productId") or "")
        title = c.get("cubeTitleEn") or ""
        pop = _lower_terms(title, POPULATION_TERMS)
        per = _lower_terms(title, PERIPHERAL_TERMS)
        if pid and (pop or per):
            swept.append((pid, title, pop, per))
    swept.sort()
    pool = [s for s in swept if s[2]]  # matched >=1 POPULATION term
    _guard_sweep(cubes, swept, pool)  # RAISES (wds-list)

    # What a bare-abbreviation predicate would have matched instead — MEASURED, so the
    # phrase predicate's choice is evidenced rather than asserted.
    abbrev_hits = [(str(c.get("productId") or ""), c.get("cubeTitleEn") or "")
                   for c in cubes if ABBREV_TRAP_TERM in (c.get("cubeTitleEn") or "").lower()]
    abbrev_pop = [h for h in abbrev_hits if _lower_terms(h[1], POPULATION_TERMS)]

    f_cat = Fact.derived(str(len(cubes)), "cubes returned by the live getAllCubesListLite")
    f_swept = Fact.derived(str(len(swept)), "catalogue titles matching the sweep predicate")
    f_pool = Fact.derived(str(len(pool)), "swept titles matching a POPULATION term (pick-eligible)")

    note += [
        "## 1. Catalogue sweep — the enumerated population the pick is earned over",
        "",
        f"- `getAllCubesListLite` -> **{f_cat}** cubes in the live StatCan catalogue.",
        f"- Sweep predicate over `cubeTitleEn` (case-insensitive phrase match). POPULATION "
        f"terms {list(POPULATION_TERMS)} make a cube PICK-ELIGIBLE; PERIPHERAL terms "
        f"{list(PERIPHERAL_TERMS)} are swept too, so every absence claim below is scoped to a "
        f"WIDER population than the pick pool.",
        f"- **{f_swept}** titles matched; **{f_pool}** of those matched a POPULATION term.",
        f"- Abbreviation check (why the predicate uses phrases, MEASURED not assumed): a bare "
        f"`\"{ABBREV_TRAP_TERM}\"` substring predicate matches **{len(abbrev_hits)}** catalogue "
        f"titles, of which **{len(abbrev_pop)}** also match a population term"
        + (" — so it surfaces no population cube the phrase predicate misses. What it DID match "
           "is emitted verbatim below rather than glossed by naming the enclosing word by hand "
           "(that gloss was true only by accident of today's catalogue):"
           if not abbrev_pop
           else " — so it DOES surface population cubes the phrase predicate would miss; those "
                "are folded into the swept list above. What it matched:"),
        "",
    ]
    note += [f"  - `{pid}` {title}" for pid, title in sorted(abbrev_hits)]
    note += [
        "",
        "**Every swept candidate, with the measured attributes the pick is decided on:**",
        "",
        "| productId | title | terms matched | eligible |",
        "|---|---|---|---|",
    ]
    for pid, title, pop, per in swept:
        note.append(f"| `{pid}` | {title} | {', '.join(pop + per)} | "
                    f"{'POPULATION' if pop else 'peripheral only'} |")
    note.append("")

    # ============ boundary 3: getCubeMetadata (schema / geography / cadence) =========
    stage["at"] = "wds-meta"
    meta = _meta([s[0] for s in swept])

    # Per-candidate measured attributes. Every field here is read off the live object.
    cand: dict[str, dict] = {}
    for pid, title, pop, per in swept:
        obj = meta.get(pid) or {}
        members = _geo_members(obj)
        text = _searchable_text(obj)
        named_all, marked_all = {}, {}
        for cma in MODELED_CMAS:
            named_all[cma], marked_all[cma] = _cma_hits(members, cma)
        cand[pid] = {
            "title": obj.get("cubeTitleEn") or title,
            "eligible": bool(pop),
            "obj": obj,
            "resolved": bool(obj),
            "members": members,
            "n_geo": len(members),
            "n_cma_marked": sum(1 for m in members
                                if any(k.lower() in m.lower() for k in CMA_MARKERS)),
            "n_ca_marked": sum(1 for m in members if CA_MARKER.lower() in m.lower()),
            "freq_code": obj.get("frequencyCode"),
            "cadence": freq_map.get(obj.get("frequencyCode"), ""),
            "start": obj.get("cubeStartDate") or "",
            "end": obj.get("cubeEndDate") or "",
            "archive": (obj.get("archiveStatusEn") or "").strip(),
            "current": _is_current(obj),
            "release": obj.get("releaseTime") or "",
            "stock_hits": _lower_terms(text, STOCK_MARKERS),
            "flow_hits": _lower_terms(text, FLOW_MARKERS),
            "cma_named": named_all,
            "cma_marked": marked_all,
        }

    # --- the PICK rule, deterministic and stated before its result ------------------
    # spec:473 wants a CURRENT VALUE with an as_of and a freshness limit, so among the
    # PICK-ELIGIBLE (population-term) candidates that are CURRENT, take the one whose
    # series runs LATEST; ties break on series count, then productId. The eligibility
    # filter runs FIRST so the freshness ranking cannot be hijacked by a cube that is not
    # a temporary-resident population source at all.
    eligible_current = sorted(
        [pid for pid in cand if cand[pid]["eligible"] and cand[pid]["current"]
         and cand[pid]["resolved"]],
        key=lambda p: (cand[p]["end"], cand[p]["obj"].get("nbSeriesCube") or 0, p),
        reverse=True,
    )
    if not eligible_current:
        raise VacuousProbeError(
            f"none of the {len(pool)} pick-eligible candidates resolved as CURRENT — there is "
            f"no maintained temporary-resident population cube among the {len(swept)} swept"
        )
    pick = eligible_current[0]
    p = cand[pick]
    geo_dim = _geo_dimension(p["obj"])
    _guard_pick(pick, p["obj"], geo_dim, p["members"], freq_map)  # RAISES (wds-meta)

    # --- the recorded rejection reason for every candidate that is NOT the pick -----
    # COMPUTED from the measured attributes above, never a per-pid typed string.
    def _reject(pid: str) -> str:
        c = cand[pid]
        why = []
        if not c["resolved"]:
            return "getCubeMetadata resolved no object for it this run"
        if not c["eligible"]:
            why.append(f"title matched only a peripheral permit/worker term, not a "
                       f"temporary-resident population term (its subject is \"{c['title']}\")")
        if not c["current"]:
            why.append(f"archive status is \"{c['archive']}\" (not CURRENT)")
        if c["start"] and c["start"] == c["end"]:
            why.append(f"one-time reference period ({c['start']}) — cubeStartDate equals "
                       f"cubeEndDate, so it cannot supply a fresh current value")
        if c["end"] < p["end"]:
            why.append(f"series ends {c['end']}, earlier than the pick's {p['end']}")
        if c["flow_hits"] and not c["stock_hits"]:
            why.append(f"only FLOW markers matched ({c['flow_hits']}); no stock marker")
        return "; ".join(why) or "eligible and current, but the pick's series runs at least as late"

    note += [
        "## 2. Live metadata for every swept candidate (boundary `wds-meta`)",
        "",
        f"`getCubeMetadata` resolved **{len(meta)}** of the {len(swept)} swept product ids. "
        f"Cadence labels are resolved from the live `getCodeSets` "
        f"({len(freq_map)} frequency codes defined), never typed beside the raw code.",
        "",
        "| productId | cadence (code) | reference span | archive status | geo members "
        "| (CMA)-marked | modeled CMAs found | stock markers | flow markers |",
        "|---|---|---|---|---:|---:|---|---|---|",
    ]
    for pid, _t, _pop, _per in swept:
        c = cand[pid]
        found = ", ".join(
            f"{k}:{len(c['cma_marked'][k])}" for k in MODELED_CMAS
        )
        note.append(
            f"| `{pid}` | {c['cadence'] or '?'} ({c['freq_code']}) | {c['start']}..{c['end']} "
            f"| {c['archive'] or '?'} | {c['n_geo']} | {c['n_cma_marked']} | {found} "
            f"| {', '.join(c['stock_hits']) or 'none'} "
            f"| {', '.join(c['flow_hits']) or 'none'} |"
        )
    note += [
        "",
        "Marker columns are the RAW HITS, not a classification: a one-token match is a weak "
        "classifier and must not read as a stated measure type. The pick's measure type is "
        "decided in §3 on its title plus a verbatim footnote, and is the only measure-type "
        "claim this note makes about a swept cube.",
        "",
        "**Why each non-pick candidate was rejected (computed from the row above, not typed "
        "per product):**",
        "",
    ]
    for pid, _t, _pop, _per in swept:
        if pid != pick:
            note.append(f"- `{pid}` — {_reject(pid)}.")
    note.append("")

    # ============ §3 the pick, in full ==============================================
    f_pid = Fact.derived(pick, "productId of the picked cube, resolved from the live sweep")
    f_cad = Fact.derived(p["cadence"],
                         f"frequencyDescEn for frequencyCode {p['freq_code']}, from the live "
                         f"getCodeSets")
    f_ngeo = Fact.derived(str(p["n_geo"]), "members of the pick's live Geography dimension")
    n_named = {k: len(p["cma_named"][k]) for k in MODELED_CMAS}
    n_marked = {k: len(p["cma_marked"][k]) for k in MODELED_CMAS}
    cma_available = all(n_marked[k] > 0 for k in MODELED_CMAS)
    # A name matched but no CMA marker => the member is almost certainly the province row,
    # not a CMA. Surfaced explicitly rather than collapsing into a silent NO.
    name_only_discrepancy = {k: n_named[k] for k in MODELED_CMAS
                             if n_named[k] > 0 and n_marked[k] == 0}
    f_cmafound = Fact.derived(
        "YES" if cma_available else "NO",
        f"name-search of the pick's {p['n_geo']} geography members for {list(MODELED_CMAS)} "
        f"requiring a {list(CMA_MARKERS)} marker",
    )

    q_point = _footnote_quote(p["obj"], any_of=POINT_IN_TIME_MARKERS)
    q_def = _footnote_quote(p["obj"], all_of=(DEFINITION_MARKER,))
    q_cross = _footnote_quote(p["obj"], all_of=CROSS_PROGRAMME_MARKERS)
    for label, quote in (("point-in-time reference dates (the STOCK evidence)", q_point),
                         ("the definition of a non-permanent resident", q_def),
                         ("the cross-programme comparison caution", q_cross)):
        if quote:
            Fact.cited(label, f"live footnote of {pick}: \"{quote}\"")

    # Measure type — a FUNCTION of the measured evidence, so the token cannot contradict it.
    title_stock = _lower_terms(p["title"], STOCK_MARKERS)
    measure = "STOCK" if (title_stock or q_point) and not p["flow_hits"] else (
        "FLOW" if p["flow_hits"] and not (title_stock or q_point) else "AMBIGUOUS")
    f_measure = Fact.derived(
        measure,
        f"title stock markers {title_stock or 'none'}; flow markers {p['flow_hits'] or 'none'}; "
        f"a point-in-time footnote was {'found' if q_point else 'NOT found'}",
    )

    note += [
        f"## 3. The pick — `{pick}` (StatCan table {table_number(pick)})",
        "",
        f"- Title (live): *\"{p['title']}\"* — {table_url(pick)}",
        f"- **Cadence: {f_cad}** (frequencyCode {p['freq_code']}, label resolved from the live "
        f"code set), reference periods **{p['start']}..{p['end']}**.",
        f"- **Currency: {p['archive']}**; last released {p['release']}.",
        f"- **Geography: {f_ngeo} members**, of which {p['n_cma_marked']} carry a "
        f"{list(CMA_MARKERS)} marker and {p['n_ca_marked']} carry a `{CA_MARKER}` marker. "
        f"The full member list, verbatim from the live response, so the level is "
        f"self-evidencing rather than glossed:",
        "",
    ]
    note += [f"  {i}. {m}" for i, m in enumerate(p["members"], start=1)]
    note += [
        "",
        f"- **Modeled-CMA search: {f_cmafound}.** Searched the {p['n_geo']} members BY NAME for "
        + ", ".join(f"**{k}** (name hits {n_named[k]}, of which {n_marked[k]} carry a CMA "
                    f"marker)" for k in MODELED_CMAS)
        + ". A cube title is never treated as evidence of its members; only this search is."
        + (f" DISCREPANCY: {name_only_discrepancy} matched by name with NO CMA marker — inspect "
           f"those members before reading this as a plain absence."
           if name_only_discrepancy else ""),
        f"- **Measure type: {f_measure}.** "
        + (f"Title stock markers {title_stock}. " if title_stock else "No title stock marker. ")
        + (f"Flow markers {p['flow_hits']}. " if p["flow_hits"] else "No flow markers. ")
        + ("Point-in-time reference dates, quoted verbatim from the live footnotes: "
           f"\"{q_point}\""
           + (" — a value stated AT a reference date is a level, not a movement, which is "
              "what spec:473 consumes."
              if measure == "STOCK" else
              f" — but the measure type computed {measure}, so this note does NOT conclude "
              f"the value is a level; the marker evidence above is mixed and must be "
              f"resolved before the source is wired.")
           if q_point else
           "NO point-in-time footnote was found in the live footnote list, so the measure type "
           "rests on the title markers alone — weaker evidence, recorded as such."),
        "",
        "- Schema (dimensions and member counts, live):",
        "",
        "  | # | dimension | members |",
        "  |---:|---|---:|",
    ]
    note += [f"  | {d.get('dimensionPositionId')} | {d.get('dimensionNameEn')} | "
             f"{len(d.get('member') or [])} |" for d in _dimensions(p["obj"])]
    note += [
        "",
        f"- Definition (verbatim, live footnote): "
        + (f"\"{q_def}\"" if q_def else "NOT STATED in the live footnote list."),
        "",
    ]

    # --- the CMA-bearing candidate, recorded as a SEPARATE source (Ruling D) --------
    cma_bearers = [pid for pid in cand
                   if pid != pick and cand[pid]["eligible"]
                   and all(len(cand[pid]["cma_marked"][k]) > 0 for k in MODELED_CMAS)]
    note += ["## 3b. Does any swept candidate carry the modeled CMAs?", ""]
    if cma_bearers:
        for pid in cma_bearers:
            c = cand[pid]
            note += [
                f"- **`{pid}`** *\"{c['title']}\"* DOES carry both modeled CMAs: "
                + "; ".join(f"{k} -> {c['cma_marked'][k]}" for k in MODELED_CMAS)
                + f" (among {c['n_geo']} geography members, {c['n_cma_marked']} CMA-marked and "
                f"{c['n_ca_marked']} CA-marked).",
                f"  - Measured disqualifier for the tripwire slot: cadence "
                f"**{c['cadence']}** ({c['freq_code']}), reference span "
                f"**{c['start']}..{c['end']}**"
                + (" — cubeStartDate EQUALS cubeEndDate, i.e. a single reference period, so it "
                   "cannot supply the fresh current value spec:473 requires."
                   if c["start"] == c["end"] else
                   f" — its series ends {c['end']}, before the pick's {p['end']}.")
                + f" Its archive status is \"{c['archive']}\".",
            ]
        note += [
            "",
            "**This note proposes NO combination of the two.** They are separate products with "
            "separate reference periods and methods; nothing here licenses disaggregating or "
            "benchmarking the pick's province total to a CMA using the cube above. The pick's "
            "own live footnotes carry the programme-comparison caution: "
            + (f"\"{q_cross}\"" if q_cross
               else "(no comparison-caution footnote was found in the live footnote list)")
            + " The CMA absence in the pick is a RECORDED LIMIT, not a gap to fill.",
            "",
        ]
    else:
        note += [
            f"- NO candidate among the {len(swept)} swept carries both modeled CMAs in its live "
            f"geography members. Scoped to this sweep — not a claim that no such cube exists.",
            "",
        ]

    evidence = {
        "pick": pick,
        "title": p["title"],
        "cadence": p["cadence"],
        "freq_code": p["freq_code"],
        "start": p["start"],
        "end": p["end"],
        "archive": p["archive"],
        "release": p["release"],
        "n_geo": p["n_geo"],
        "n_cma_marked": p["n_cma_marked"],
        "n_ca_marked": p["n_ca_marked"],
        "members": p["members"],
        "cma_available": cma_available,
        "n_named": n_named,
        "n_marked": n_marked,
        "measure": measure,
        "title_stock": title_stock,
        "flow_hits": p["flow_hits"],
        "q_point": q_point,
        "q_cross": q_cross,
        "n_cubes": len(cubes),
        "n_swept": len(swept),
        "n_pool": len(pool),
        "cma_bearers": cma_bearers,
        "cand": cand,
        "dims": [(d.get("dimensionNameEn"), len(d.get("member") or []))
                 for d in _dimensions(p["obj"])],
    }
    # The STOCK discriminator is ENFORCED, not decorative. spec:473 consumes a stock, so a
    # pick that does not compute as one cannot be published under a bare LOCATED — that
    # would leave three of the four declared discriminators gating the verdict and the
    # fourth as prose. It is a DISTINCT verdict token, deliberately NOT a VacuousProbeError:
    # the boundary answered perfectly well, so routing it through the failure path would
    # record a boundary failure THAT DID NOT HAPPEN — a false statement in a generated note.
    # R6's vocabulary exists precisely to express "located, with a limit".
    return ("LOCATED" if measure == "STOCK" else "LOCATED-NOT-STOCK"), evidence


# --- §4 / §5: RECORDED comparisons that do NOT gate the verdict ---------------------
def _record_ircc(note: list[str]) -> dict:
    """The spec-named IRCC alternative, swept live on CKAN. NON-GATING.

    Returns {"measured": bool, ...}. Any failure is recorded as NOT MEASURED THIS RUN and
    the pick's scope clause narrows accordingly — an unmeasured comparison must never
    read as a comparison that was made.
    """
    note += ["## 4. The spec-named alternative — IRCC temporary-resident tables "
             "(RECORDED, non-gating)", ""]
    try:
        payload = _ckan_search()
        result = payload.get("result") or {}
        results = result.get("results") or []
        # An empty-but-200 CKAN body is NOT a measured comparison. Reported as measured, it
        # inflated the pick's scope clause into "chosen over the 0 IRCC packages swept" — a
        # claimed comparison against a set of zero, the vacuous-absence shape in miniature.
        if not results:
            note += [
                "- `IRCC COMPARISON NOT MEASURED THIS RUN: EmptyPopulation: package_search "
                "answered but returned zero results`",
                "",
                "  A zero-result search is not a comparison. It does NOT change the verdict — "
                "the pick is earned from the StatCan sweep in §1-§3 — and the DECISION block "
                "below therefore does not count the IRCC family among what the pick was "
                "compared against.",
                "",
            ]
            return {"measured": False, "n_pkgs": 0, "n_combined": 0}
        pkgs = []
        for pkg in results:
            org = ((pkg.get("organization") or {}).get("title") or "").lower()
            title = pkg.get("title") or ""
            if IRCC_ORG_MARK in org and "temporary resident" in title.lower():
                res = pkg.get("resources") or []
                cma_res = [r for r in res
                           if any(k.lower() in (r.get("name") or "").lower()
                                  for k in ("census metropolitan", "cma"))]
                cma_live = [r for r in cma_res
                            if "[archived]" not in (r.get("name") or "").lower()]
                stock_res = [r for r in res
                             if any(k in (r.get("name") or "").lower()
                                    for k in ("on december 31", "on july 1",
                                              "with a valid permit"))]
                pkgs.append({
                    "id": pkg.get("id") or "", "title": title,
                    "freq": pkg.get("frequency") or "?",
                    "archived_title": title.lower().startswith("[archived]"),
                    "n_res": len(res), "n_cma": len(cma_res), "n_cma_live": len(cma_live),
                    "n_stock": len(stock_res),
                    # A package titled for one PERMIT CLASS publishes a subset of the
                    # temporary-resident population, not the total spec:473 consumes. Same
                    # kind of scope predicate as POPULATION_TERMS on the StatCan side.
                    "class_scoped": _lower_terms(title, PERMIT_CLASS_TERMS),
                    "cma_live_names": [r.get("name") or "" for r in cma_live],
                })
        # The combined predicate: a package usable for the tripwire slot needs a NON-archived
        # CMA-named resource AND a point-in-time-named resource AND a total (not permit-class)
        # scope. All three, because any one alone is satisfied by a table that is not the
        # temporary-resident STOCK the spec asks for.
        combined = [p for p in pkgs
                    if p["n_cma_live"] > 0 and p["n_stock"] > 0 and not p["class_scoped"]]
        note += [
            f"- Live query: `{CKAN_SEARCH}`",
            f"- `package_search` -> {_count_str(result)} total matches, {len(results)} "
            f"returned and swept{_trunc_str(result, results)}; **{len(pkgs)}** are IRCC "
            f"packages whose title names temporary residents.",
            "",
            "| package | frequency | resources | CMA-named | CMA-named & not [ARCHIVED] "
            "| point-in-time-named | permit-class scope |",
            "|---|---|---:|---:|---:|---:|---|",
        ]
        for pk in sorted(pkgs, key=lambda x: x["title"]):
            note.append(f"| {pk['title']} | {pk['freq']} | {pk['n_res']} | {pk['n_cma']} "
                        f"| {pk['n_cma_live']} | {pk['n_stock']} "
                        f"| {', '.join(pk['class_scoped']) or 'total (no class term)'} |")
        n_class = sum(1 for pk in pkgs if pk["class_scoped"])
        note += [
            "",
            f"- **Measured reason the IRCC family is not picked:** **{len(combined)}** of the "
            f"{len(pkgs)} swept IRCC temporary-resident packages satisfy ALL THREE of "
            f"(a) a CMA-named resource that is not [ARCHIVED]-labelled, (b) a "
            f"point-in-time-named resource, (c) a total scope rather than a single permit "
            f"class — {n_class} of the {len(pkgs)} are titled for a permit class "
            f"{list(PERMIT_CLASS_TERMS)}, which publishes a SUBSET of the temporary-resident "
            f"population, not the total spec:473 consumes"
            + (". So none of them offers a CMA-level, point-in-time count of the "
               "temporary-resident total. (Maintenance is NOT among the three tested "
               "conditions — the frequency column above is displayed, not asserted on.)"
               if not combined else
               f": {[c['title'] for c in combined]} — inspect these before treating the "
               f"StatCan pick as the only maintained option.")
            + " The counts above are what this run measured, scoped to this query — not a "
            "claim about IRCC holdings outside it.",
            "",
        ]
        return {"measured": True, "n_pkgs": len(pkgs), "n_combined": len(combined)}
    except Exception as exc:
        note += [
            f"- `IRCC COMPARISON NOT MEASURED THIS RUN: {type(exc).__name__}: {exc}`",
            "",
            "  The CKAN boundary did not answer usefully. This does NOT change the verdict — "
            "the pick is earned from the StatCan sweep in §1-§3 — but it DOES narrow what the "
            "pick was compared against, and the DECISION block below says so rather than "
            "implying a comparison that did not run.",
            "",
        ]
        return {"measured": False, "n_pkgs": 0, "n_combined": 0}


def _record_isq(note: list[str]) -> dict:
    """The in-repo ISQ compo alternative. NON-GATING (the workbook lives outside demoflow/).

    The column is LOCATED by searching the header rows for the NPR label rather than
    indexed at the plan's pinned column, and the found index is cross-checked against it.
    """
    note += ["## 5. The in-repo alternative — ISQ compo NPR column (RECORDED, non-gating)", ""]
    try:
        rows = _isq_rows()
        ncol = max(len(r) for r in rows)

        def stack(col: int) -> str:
            return " ".join(
                str(rows[r][col]).strip() for r in ISQ_HEADER_ROWS
                if col < len(rows[r]) and rows[r][col] is not None
            ).strip()

        found = [c for c in range(ncol)
                 if "solde" in stack(c).lower() and "non permanents" in stack(c).lower()]
        if not found:
            note += [
                f"- Searched header rows {list(ISQ_HEADER_ROWS)} of `{ISQ_COMPO.name}` across "
                f"{ncol} columns: NO column carries a 'Solde ... non permanents' label this "
                f"run. The in-repo NPR column could not be located, so no rejection reason is "
                f"claimed for it.",
                "",
            ]
            return {"measured": False}
        col = found[0]
        label = stack(col)
        value = rows[ISQ_FIRST_DATA_ROW][col] if col < len(rows[ISQ_FIRST_DATA_ROW]) else None
        numeric = isinstance(value, (int, float)) and not isinstance(value, bool)
        negative = numeric and value < 0
        note += [
            f"- File read: `{ISQ_COMPO.relative_to(_REPO_ROOT)}` "
            f"(header rows {list(ISQ_HEADER_ROWS)} searched across {ncol} columns).",
            f"- NPR column LOCATED at 0-indexed column **{col}** "
            + (f"(matches the plan's pinned column {ISQ_PLAN_COL})."
               if col == ISQ_PLAN_COL
               else f"— NOTE: the plan pins column {ISQ_PLAN_COL}; the live workbook puts it at "
                    f"{col}. The searched index is what this run read."),
            f"- Label read from the workbook: **\"{label}\"**.",
            f"- First data row ({ISQ_FIRST_DATA_ROW}, 0-indexed) value: **{value}**.",
            "- **Measured reason it does not fill the STOCK slot:** the label CONTAINS "
            "\"Solde\" (a net balance) — the predicate is a substring test, not a prefix test"
            + (f", and the first data value {value} is NEGATIVE — a population STOCK cannot be "
               f"below zero, so this column is a net FLOW over the year, not a level."
               if negative else
               (f", and the first data value {value} is not negative this run, so the FLOW "
                f"reading rests on the label alone — weaker evidence, recorded as such."
                if numeric else
                f", and the first data cell is non-numeric ({value!r}), so the FLOW reading "
                f"rests on the label alone — weaker evidence, recorded as such."))
            + " spec:473 consumes a STOCK, so this column is recorded as a complement (it is "
            "already the demand model's NPR input), never as the stock indicator.",
            "",
        ]
        return {"measured": True, "col": col, "label": label, "value": value,
                "negative": negative}
    except Exception as exc:
        note += [
            f"- `ISQ COMPARISON NOT MEASURED THIS RUN: {type(exc).__name__}: {exc}`",
            "",
            "  The committed workbook could not be read. This does NOT change the verdict "
            "(the pick is earned from the StatCan sweep) and no rejection reason is claimed "
            "for the ISQ column on the strength of a read that did not happen.",
            "",
        ]
        return {"measured": False}


def main() -> None:
    # Per-run registry (see run_p3.py): `_wds` is one cached module shared by every probe,
    # so a module-global list would let one run's figures inflate another's.
    facts = new_run()
    title = ["# P5b — Temporary-resident STOCK source pick (RECORDED OBSERVATION)", ""]
    body: list[str] = []
    stage = {"at": "wds-codesets"}

    verdict = "UNKNOWN-PROBE-FAILED"
    evidence: dict = {}
    try:
        verdict, evidence = _hunt(body, stage)
    except Exception as exc:  # outage, a vacuous 200, or a wrong-body page — never a false LOCATED
        boundary = stage.get("at", "wds-codesets")
        body += [
            "## LIVE HUNT VERDICT: UNKNOWN-PROBE-FAILED",
            "",
            f"- `LIVE PROBE FAILED-AT: {boundary}` (host: www150.statcan.gc.ca)",
            f"- `LIVE PROBE FAILED: {type(exc).__name__}: {exc}`",
            "",
            "  The source did not answer usefully at the boundary named above — an outage, or a "
            "200 carrying an empty/wrong body (caught by the floor guard). This run therefore "
            "did NOT pick a temporary-resident stock source, and deliberately records UNKNOWN "
            "rather than a fabricated pick. Per spec:473 the tripwire's temporary-resident-stock "
            "indicator reports **UNKNOWN** while this source is unwired (never a stale "
            "within-band). Re-run against a live service to record the pick.",
            "",
        ]

    # NON-GATING comparisons — they run whatever the verdict, and cannot change it.
    ircc = _record_ircc(body)
    isq = _record_isq(body)

    body += ["## DECISION", ""]
    if verdict in ("LOCATED", "LOCATED-NOT-STOCK"):
        e = evidence
        # The comparison scope is a FUNCTION of which comparisons actually ran, so the pick
        # can never claim to have been chosen over an alternative nobody measured.
        scope_bits = [f"the {e['n_swept']} StatCan cubes swept in §1 (the pick is one of them)"]
        if ircc["measured"]:
            scope_bits.append(f"the {ircc['n_pkgs']} IRCC temporary-resident packages swept in §4")
        if isq["measured"]:
            scope_bits.append("the in-repo ISQ compo NPR column measured in §5")
        unmeasured = [n for n, ok in (("IRCC/CKAN (§4)", ircc["measured"]),
                                      ("ISQ compo (§5)", isq["measured"])) if not ok]
        # Every part COUNTED off the live member list. No level NAME is glossed here: the
        # full member list is emitted verbatim in §3, which is what states the levels — a
        # typed "national + province/territory" gloss would be an inference sitting beside
        # a computed count, which is the shape house rule #1 forbids.
        n_canada = sum(1 for m in e["members"] if m.strip().lower() == "canada")
        geo_levels = (f"{e['n_geo']} members: {n_canada} named exactly \"Canada\", "
                      f"{e['n_cma_marked']} carrying a CMA marker, {e['n_ca_marked']} carrying "
                      f"a {CA_MARKER} marker, "
                      f"{e['n_geo'] - n_canada - e['n_cma_marked'] - e['n_ca_marked']} other; "
                      f"the full member list is emitted verbatim in §3")
        # The DECLARED containing provinces (MODELED_CMA_PROVINCE), matched against the live
        # member list. Both halves are now tied: the value half is counted off the response,
        # and the containment half is a declared datum that moves with MODELED_CMAS rather
        # than a relation asserted in prose. (Two earlier glosses died here: a hand-typed
        # "PROVINCE-level (Quebec)" that survived ZERO geography members, and a "coarsest
        # geography containing them" that would have survived swapping in Toronto.)
        want_prov = sorted({MODELED_CMA_PROVINCE[k] for k in MODELED_CMAS})
        prov_members = [m for m in e["members"] if _norm(m) in {_norm(x) for x in want_prov}]
        limit = (
            f"no CMA breakdown — the modeled CMAs "
            + ", ".join(f"{k} (name hits {e['n_named'][k]}, CMA-marked {e['n_marked'][k]})"
                        for k in MODELED_CMAS)
            + f" are ABSENT from the pick's {e['n_geo']} geography members"
            + (f"; their DECLARED containing province(s) {want_prov} — declared beside "
               f"MODELED_CMAS, not inferred here — ARE present as member(s) {prov_members}, "
               f"so that is the finest geography this source offers for them"
               if prov_members else
               f"; no member matches their DECLARED containing province(s) {want_prov} either, "
               f"so this source offers no geography for them at all")
            + (f"; the only swept cube carrying both is {e['cma_bearers']}, disqualified in §3b "
               f"and NOT combinable with the pick"
               if e["cma_bearers"] else
               "; no swept cube carries both, so no CMA-level substitute was found either")
        ) if not e["cma_available"] else "none"

        body += [
            f"- `DECISION-VERDICT: {verdict}`"
            + ("" if verdict == "LOCATED" else
               "  (a source WAS located and every field below is measured, but its measure type "
               "did NOT compute as STOCK — spec:473 consumes a stock, so this pick must not be "
               "wired until the mismatch in §3 is resolved. The boundaries all answered; this "
               "is a limit on the ANSWER, not a probe failure.)"),
            f"- `DECISION-SOURCE-PID: {e['pick']}`  (StatCan table {table_number(e['pick'])}; "
            f"resolved from the live sweep of {e['n_cubes']} catalogue cubes, not typed)",
            f"- `DECISION-SOURCE-TITLE: {e['title']}`  (cubeTitleEn from the live "
            "getCubeMetadata response)",
            f"- `DECISION-CADENCE: {e['cadence']} (frequencyCode {e['freq_code']}, label from "
            f"the live getCodeSets); reference periods {e['start']}..{e['end']}`",
            f"- `DECISION-GEO-LEVEL: {geo_levels}`",
            f"- `DECISION-CMA-AVAILABLE: {'YES' if e['cma_available'] else 'NO'}`  "
            f"(computed by searching the pick's {e['n_geo']} live geography members BY NAME for "
            + ", ".join(f"{k}: {e['n_marked'][k]} CMA-marked hit(s) of {e['n_named'][k]} name "
                        f"hit(s)" for k in MODELED_CMAS)
            + " — never inferred from the cube title)",
            f"- `DECISION-MEASURE-TYPE: {e['measure']}`  "
            + (f"(title stock markers {e['title_stock']}; flow markers "
               f"{e['flow_hits'] or 'none'}; "
               + ("the live footnote states point-in-time reference dates"
                  + (", so the value is a level at a date, not a movement — which is what "
                     "spec:473 consumes"
                     if e["measure"] == "STOCK" else
                     f", but the computed measure type is {e['measure']}, so this note draws NO "
                     f"level-vs-movement conclusion")
                  if e["q_point"] else
                  "NO point-in-time footnote was found, so this rests on title markers alone")
               + ")"),
            f"- `DECISION-CURRENCY: {e['archive']}; last released {e['release']}; series runs to "
            f"{e['end']}`",
            f"- `DECISION-PICK: StatCan {table_number(e['pick'])} (productId {e['pick']}) — "
            f"chosen from {', '.join(scope_bits)}`"
            + (f"  (NOT MEASURED this run, so NOT part of the comparison: "
               f"{', '.join(unmeasured)})" if unmeasured else ""),
            f"- `DECISION-PICK-LIMIT: {limit}`",
            "- `DECISION-TRIPWIRE-STATUS: UNKNOWN — the source is RECORDED here, not yet wired; "
            "per spec:473 the temporary-resident-stock indicator reports UNKNOWN until wired, "
            "never a stale within-band`",
            "",
            "- Standing rule (spec:473): this source feeds the TRIPWIRE BASELINE registry's "
            "temporary-resident-stock indicator (current value + as_of + freshness limit), NOT "
            "the demand model — the demand model's NPR input is the ISQ compo net-flow column"
            + (" measured in §5" if isq["measured"]
               else " (NOT measured this run — see §5)")
            + ". The two are different quantities and this note proposes no substitution "
            "between them.",
            "",
        ]
    else:
        body += [
            f"- `DECISION-VERDICT: {verdict}`",
            "- No temporary-resident stock source was picked this run (see the boundary failure "
            "above). Per spec:473 the tripwire's temporary-resident-stock indicator reports "
            "UNKNOWN until this source is wired. This is a recorded observation, not an "
            "invented pick.",
            "- `DECISION-TRIPWIRE-STATUS: UNKNOWN — no source was picked this run; per spec:473 "
            "the indicator reports UNKNOWN, never a stale within-band`",
            "",
        ]

    header = provenance_header(facts, written_by=_WRITTEN_BY, scope=_SCOPE,
                               summary=_summary, cited_label=_CITED_LABEL)
    text = "\n".join(title + header + body) + "\n"
    for placeholder in ("[FILL:", "[FILL]", "[FILL "):
        if placeholder in text:
            raise AssertionError(f"run_p5b.py emitted an unresolved {placeholder!r} placeholder")
    OUT.write_text(text, encoding="utf-8")
    print(text)


if __name__ == "__main__":
    main()
