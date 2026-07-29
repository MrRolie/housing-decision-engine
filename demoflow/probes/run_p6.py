"""P6 — MRC-level ISQ source hunt for couronne-nord precision (spec §11 item 6).

WHO USES THIS (spec §8, Geography junction): the `Geography` enum carries
LANAUDIERE_RA14_PROXY / LAURENTIDES_RA15_PROXY / MONTEREGIE_RA16_PROXY, flagged `ra_proxy`
— exact administrative-region data standing in for the finer couronne geography demoflow
would rather model. §8 defers couronne-nord precision to v1 and this probe is the hunt that
§11.6 charters. **v0 PROCEEDS REGARDLESS** — a find enables a v1 Geography-enum extension
only, never a v0 change (that sentence is written unconditionally, so no gate may rest on
it).

WHY NO §8 SENTENCE IS QUOTED ANYWHERE IN THIS FILE. Reproducing another artifact's TEXT
creates a dependency nothing checks, and amending the cited text is exactly when it breaks
— the CROSS-ARTIFACT STALENESS class, which bit this probe once already. So §8's premise is
reproduced in exactly ONE place — §4 of the note, where `_record_spec_premise` READS the
spec at run time and quotes the span it actually found, on every verdict branch. That
reproduction is live by construction and cannot go stale. Every other reference in this
file, and every hand-written string in it, names the section instead of restating it.

THE DISCRIMINATOR, stated before the hunt so the verdict cannot drift to fit the answer.
A LOCATED needs THREE pieces of evidence about ONE resource, all computed here:
  (a) a resource URL resolved from a swept population — never a guessed slug;
  (b) the OBSERVED HTTP status of a real request to it;
  (c) a BODY-SHAPE check proving the bytes really are MRC-level — the file is opened, its
      `MRC` header cell located, and its geography labels counted and name-searched.
A bare 200 FAILS: this host answers 200 with an HTML page body for some paths and 404 with
a 45KB HTML body for others, so status alone cannot tell a workbook from an error page.

THE SEARCHED POPULATION (Ruling R7 — a NOT-FOUND is unearned without one). Two
boundaries, each enumerated, so an absence claim is scoped to something real:
  * BOUNDARY A — Données Québec CKAN (`donneesquebec.ca`). The ISQ organization slug is
    RESOLVED LIVE from `organization_list` by a title predicate, never typed: a guessed
    slug (`organization:institut-de-la-statistique-du-quebec`) returns zero, and a zero
    from a wrong slug is not an absence.
  * BOUNDARY B — ISQ's own product pages / full-edition downloads
    (`statistique.quebec.ca`), which is what §11.6 actually asks for. Its `sitemap.xml`
    enumerates the site, so the sweep runs over a real population rather than two guesses.
Every absence claim below is scoped to "not among the N locs and the M packages swept" —
never "no MRC source exists".

METHOD IS LOAD-BEARING HERE, so this run MEASURES it instead of asserting it: §3 issues a
HEAD and a GET against the SAME url and prints both statuses. The plan body's hunt is
HEAD-only, so if the two disagree, a HEAD-based hunt can report a live workbook as absent.
Every request this probe makes for evidence is a GET; the HEAD is issued solely as the
measured comparison, and the note's sentence about it is a function of the two observed
codes — not a claim that survives them agreeing.

FLOOR GUARD (NS #1 — a verification gate must REFUSE when it cannot verify). An empty
sitemap, an empty CKAN catalogue, a sweep over an unswept population, or a candidate that
answers 200 with a non-workbook body must NOT become a NOT-FOUND or a LOCATED. Each raises
`VacuousProbeError` -> UNKNOWN-PROBE-FAILED with the failing boundary recorded. A NOT-FOUND
in particular requires BOTH populations to have been swept and to be NON-EMPTY.

ANTI-FABRICATION (the cardinal discipline). Every value reported as observed is emitted by
THIS run: the resolved org slug, the catalogue and sitemap sizes, the swept and eligible
candidate lists, each candidate's observed status / content-type / declared length / magic
bytes, the opened workbook's sheet names, header position, geography-label count and label
list, its scenario labels, and the per-target couronne search. The verbatim quotes are
CITED, resolved from live responses by predicate.

Run:  cd demoflow && uv run python probes/run_p6.py
"""

import io
import json
import re
import urllib.error
import urllib.request
from pathlib import Path

# Flat, NOT `probes._wds`: probes/ is deliberately not a package, so in script mode
# sys.path[0] IS probes/ and this resolves natively. See probes/_wds.py.
from _wds import Fact, new_run, provenance_header

# --- boundary A: Données Québec CKAN (the enumerable open-data catalogue) ------------
CKAN_ACTION = "https://www.donneesquebec.ca/recherche/api/3/action"
CKAN_ORGS = f"{CKAN_ACTION}/organization_list?all_fields=true&limit=1000"
CKAN_COUNT = f"{CKAN_ACTION}/package_search?rows=0"
CKAN_ORG_SEARCH = CKAN_ACTION + "/package_search?fq=organization:{slug}&rows=100"
CKAN_TERM_SEARCH = (CKAN_ACTION + "/package_search?q=MRC+projection+population"
                    "+perspectives+demographiques&rows=100")
# The ISQ organization is resolved by THIS predicate over the live `title` field. A slug is
# never assumed: `organization:institut-de-la-statistique-du-quebec` (the obvious guess)
# returns 0 packages, and a zero from a wrong slug would be a fabricated absence.
CKAN_ORG_TITLE_MARK = "institut de la statistique"

# --- boundary B: ISQ's own site (product pages / full-edition downloads) -------------
ISQ_SITEMAP = "https://statistique.quebec.ca/sitemap.xml"
ISQ_HOST = "statistique.quebec.ca"
# The sitemap lists `/fr/fichier/<slug>` and `/en/fichier/<slug>` for the SAME workbook.
# When the sweep collapses them, keep the `/fr/` one: the ISQ workbooks demoflow already
# pins are `/fr/` urls, so the url this note publishes stays on the repo's own convention.
PREFERRED_LANG_PATH = "/fr/"

# The plan body's two GUESSED slugs, probed live so the "the 404 was the slug convention,
# not the data" statement in §4 is a COMPUTED comparison against the located resource's own
# observed status — not a retelling of what someone recorded in 2026-07.
PLAN_GUESSED_SLUGS = (
    "https://statistique.quebec.ca/fr/fichier/pop-as-mrc-base.xlsx",
    "https://statistique.quebec.ca/fr/fichier/pop-mrc-base.xlsx",
)

# --- the sitemap sweep predicate (stated before its result) --------------------------
# A candidate must be an .xlsx naming BOTH an MRC geography term and a population term:
# that is the SWEPT population. ELIGIBILITY additionally requires a PROJECTION term,
# because spec §8's junction consumes projected population by scenario — an estimates
# workbook is a different product. Both tiers are emitted, so every absence claim is scoped
# to the WIDER swept set rather than to the narrower eligible one.
MRC_TERMS = ("mrc", "municipalites-regionales-de-comte")
POPULATION_TERMS = ("population", "composantes-demographiques", "menages", "demographiques")
PROJECTION_TERMS = ("projet", "scenario", "perspectives-demographiques")

# The two ISQ workbook families demoflow already consumes at RMR level (the files committed
# with the grounding research: `pop-as-rmr-base.xlsx`, `compo-rmr-base.xlsx`). DECLARED here
# rather than asserted downstream, so the note's family attribution moves with this map
# instead of being a sentence about files nothing in this run inspected.
DEMOFLOW_RMR_FAMILY = {
    "pop-as-* (population by age and sex)": ("population", "age", "sexe"),
    "compo-* (projected demographic components)": ("composantes-demographiques",),
}

# The couronne MRCs, DECLARED per spec §8's three `ra_proxy` rows, each keyed by the RA
# NUMBER the spec proxies. Declared here, not inferred — and deliberately falsifiable: the
# opened workbook carries its OWN administrative-region column, so the declared RA number is
# checked against the code the live response puts beside each MRC. Flip a key to RA99 and
# the corroboration stops, instead of the note republishing a false grouping.
#
# What the declaration does NOT claim, and what this run does not compute: that these MRCs
# EXHAUST their RA, or that they exactly compose the Montréal RMR's couronne. The per-target
# search establishes membership, never partition.
COURONNE_MRC_BY_RA = {
    "RA14 Lanaudière": ("Les Moulins", "L'Assomption"),
    "RA15 Laurentides": ("Thérèse-De Blainville", "Deux-Montagnes", "Mirabel",
                         "La Rivière-du-Nord"),
    "RA16 Montérégie": ("Roussillon", "Marguerite-D'Youville", "La Vallée-du-Richelieu",
                        "Vaudreuil-Soulanges"),
}

# The geography header cell is matched by PREFIX, not by equality and not by substring.
# Measured reason for each rejection: equality misses the 2016-2041 edition, whose header
# cell reads "MRC par région administrative"; a substring test matches the CAPTION row
# ("Population projetée des MRC du Québec, …") and would count zero labels below it. A
# prefix admits both real headers and excludes both captions.
MRC_HEADER_PREFIX = "mrc"
# The administrative-region column, when the edition publishes one (the A2021 components
# edition heads it `RA1`). This is the axis spec §8's RA14/15/16 proxies turn on, so it is
# SEARCHED FOR and its absence is reported as an absence — never assumed either way.
RA_HEADER_PATTERN = re.compile(r"^ra\s*\d*$")
# Both spellings, because these workbooks mix accented and unaccented headers and this file
# does not normalise accents away (doing so would also fold the MRC labels, which are
# compared by exact name).
RA_HEADER_MARKS = ("région administrative", "region administrative")
SCENARIO_HEADER_MARKS = ("scénario", "scenario")
HEADER_SCAN_ROWS = 12
# These geography columns are NOT homogeneous: ISQ interleaves administrative-region
# SUBTOTAL rows, published in a `NN  Name` form ("01  Bas-Saint-Laurent"), among the MRC
# rows. The label count alone would therefore read as an MRC count and be wrong. This
# pattern splits the two so the decomposition is emitted instead of the raw total.
AGGREGATE_LABEL_PATTERN = re.compile(r"^\d{2}\s")
# A geography column carries NAMES. The prefix header search can still land on a numeric
# column if a sheet ever heads its code column with an MRC token, and the resulting LOCATED
# would publish codes as geography labels. MEASURED on the live picked workbook: the correct
# (prefix) column is 0/122 numeric, while the column a substring search reaches is 106/109 —
# so the two are separated by an enormous margin and a half threshold is not a close call.
# Refusing a genuinely numeric geography column is the SAFE direction: digits cannot evidence
# MRC-level, so a gate that cannot verify must refuse rather than pass.
GEO_LABEL_MAX_NUMERIC_FRACTION = 0.5
# An administrative-region CODE is a short integer (Québec publishes 01..17). A column whose
# values are longer or non-numeric is not a per-MRC RA code column whatever its header says.
RA_CODE_MAX_DIGITS = 3
# How many rows past the header block the per-edition probe reads, so the RA VALUE check has
# data to sample. Small: the question is what KIND of value the column holds, not its range.
RA_VALUE_SAMPLE_ROWS = 5
# Residual (ii) asks whether the declared couronne MRCs COMPOSE the Montréal RMR couronne.
# That is a metropolitan-area question, so the note must MEASURE whether this file carries a
# metropolitan axis at all before saying it cannot answer — "no RMR column" is itself a claim.
RMR_MARKERS = ("rmr", "cma", "métropolitaine", "metropolitaine", "metropolitan")
# Candidate caption markers §3b tests for CO-OCCURRENCE with the presence/absence of the RA
# axis. Declared here, tested against the live captions, and reported only when the
# separation is perfect — never as a cause, and never as a recency ranking.
EDITION_CAPTION_MARKERS = ("scénarios de 2025", "scenarios de 2025", "scénario référence",
                           "a2021")
# One gloss per set-algebra verdict, selected by the relations a run ACTUALLY computed. A
# single unconditional paragraph explaining PROPER SUBSET asserts about "this file" under a
# hypothesis that is false whenever the computed relation is PROPER SUPERSET or EMPTY.
RELATION_GLOSS = {
    "EQUAL": "An EQUAL relation says the declared targets are exactly the MRCs this workbook "
             "puts under that RA — the declared set exhausts it.",
    "PROPER SUBSET": "A PROPER SUBSET says the declared targets are SOME of the MRCs this "
                     "workbook puts under that RA, so this file's own RA grouping is WIDER "
                     "than the couronne set declared here and the declared set does NOT "
                     "exhaust the RA.",
    "PROPER SUPERSET": "A PROPER SUPERSET says this file declares MRCs the workbook does not "
                       "place under that RA — the declaration reaches past the workbook's "
                       "own grouping.",
    "OVERLAPPING": "An OVERLAPPING relation says neither set contains the other.",
    "DISJOINT": "A DISJOINT relation says no declared target sits under that RA here.",
    "EMPTY": "An EMPTY relation says this workbook assigns no MRC at all to that RA number.",
    "NOT COMPUTABLE": "NOT COMPUTABLE says this workbook publishes no separate RA column, so "
                      "no exhaustion relation exists to compute from it.",
}

# --- the spec premise this hunt was launched against (READ-ONLY cross-check) ---------
# The note used to QUOTE spec §8's premise as a typed string. Steering amended §8 after this
# probe's first run, which turned that quote into a claim about a locked artifact that the
# artifact no longer made. The premise is therefore READ LIVE from the spec file and the
# DECISION token is a function of what is found — so the note cannot go stale against the
# document it cites. Read-only: this probe never writes to docs/specs.
_REPO_ROOT = Path(__file__).resolve().parents[2]
SPEC_PATH = (_REPO_ROOT / "docs" / "specs"
             / "2026-07-21-demoflow-demographic-scenario-module-design.md")
SPEC_ROW_MARK = "couronne-nord precision"
# Two markers, not one, so the check is three-valued. Absence of the old marker alone would
# not distinguish "amended" from "the row was rewritten in a way this predicate cannot read";
# the second marker makes the amended state POSITIVELY identified and leaves a third,
# honestly-labelled INDETERMINATE branch. The old marker is the FULL phrase "no MRC workbook
# EXISTS": the amended text quotes "'no MRC workbook (404)' finding", so a shorter marker
# would match the amendment itself and report the premise as still standing.
SPEC_OLD_PREMISE_MARK = "no mrc workbook exists"
SPEC_AMENDED_MARK = "mrc-level isq projection workbooks exist"
# ZIP local-file-header magic. An HTML error page served at 200 — the wrong-body case R7
# names — cannot start with these four bytes.
XLSX_MAGIC = b"PK\x03\x04"
PREFIX_BYTES = 8

OUT = Path(__file__).resolve().parent / "P6-mrc-isq-hunt.md"
CKAN_TIMEOUT = 120
ISQ_TIMEOUT = 180  # the sitemap is ~13MB; WDS_TIMEOUT governs a different boundary


# --- this note's provenance prose (the shared header skeleton lives in _wds) ---------
# The filename this note must attribute itself to. DERIVED from __file__, never typed:
# `written_by` is the one header field a copy-pasted call block carries forward silently —
# a probe cloned from run_p5b.py would publish "Written by run_p5b.py" over its own
# computed body, exactly the untied claim this registry exists to stop.
_WRITTEN_BY = Path(__file__).name
# The header must claim only what the note it heads actually CONTAINS. Written as a flat
# inventory it forward-referenced §3/§3b/§3c — sections three of the four verdict branches
# never write — so on a NOT-FOUND or UNKNOWN run it promised evidence the file did not hold.
# run_p5b.py:152-162 already solved this in wording ("each states either its measured result
# or an explicit NOT MEASURED THIS RUN"); that is the pattern adopted here: the header
# describes the CONTRACT every section satisfies, and names no section it cannot guarantee.
_SCOPE = ("SCOPE OF THIS HEADER (it claims only what it can enforce): EVERY figure, list, "
          "status, header position, label and per-target result this note states as observed "
          "is emitted by this run from a live read — the CKAN organization list and package "
          "counts, the sitemap loc and .xlsx counts, the swept and eligible candidate lists, "
          "each candidate's observed HTTP status / content-type / declared length / "
          "magic-byte result, the HEAD-vs-GET comparison, and, for whichever workbooks this "
          "run actually opened, their sheet names, header rows, geography labels, scenario "
          "and administrative-region columns, per-target couronne search and per-RA "
          "membership sets. Quoted strings are verbatim and contiguous. A section whose "
          "boundary did not answer states an explicit NOT MEASURED THIS RUN rather than "
          "falling silent, and a run that reached no workbook writes no body-shape section "
          "at all — so this header promises no section the file does not contain. Every "
          "absence claim is scoped to what was actually swept, never to what exists. What "
          "this run does NOT compute, and therefore does not claim: which of the swept "
          "editions is CURRENT (no live response states it — captions are emitted, none "
          "ranked); that any cross-edition join is VALID (no label-set agreement or vintage "
          "compatibility is tested between two workbooks); and whether the declared MRCs "
          "compose the Montréal RMR couronne (measured only against the header row and "
          "geography column this run read of one file, and recorded as a limit rather than "
          "closed from a second source). AT MOST ONE workbook is opened in full; any others "
          "are opened HEADER ROW ONLY, so their rows carry header-scoped evidence and say "
          "nothing about their data below the header.")
_CITED_LABEL = "Quoted verbatim from the live responses:"


def _summary(*, total: int, derived: int, cited: int) -> str:
    """The provenance sentence, sized to what this run actually registered."""
    return (
        f"This run registered {total} provenance-tagged figures: {derived} DERIVED "
        f"(computed from the live responses of this run) and {cited} CITED (verbatim from "
        f"a live response body). Untagged numerals elsewhere are audit metadata (candidate "
        f"counts, byte lengths, row/column positions, HTTP status codes) and reference "
        f"labels (slugs, urls, sheet names), each traceable to the live response this run "
        f"read."
    )


# --- network boundaries (injectable seams so the floor-guard test runs OFFLINE) ------
def _ckan_get(url: str) -> dict:
    """A CKAN action endpoint -> parsed JSON. Boundary `ckan-*` (donneesquebec.ca).

    Deliberately NOT `pd.read_json` (the repeat trap this probe family keeps hitting on
    the CKAN shape): navigate the documented envelope
    {"success", "result": {"count", "results": [...]}} explicitly.
    """
    raw = urllib.request.urlopen(url, timeout=CKAN_TIMEOUT).read()
    return json.loads(raw)


def _sitemap() -> str:
    """`sitemap.xml` -> decoded text. Boundary `isq-sitemap` (statistique.quebec.ca)."""
    raw = urllib.request.urlopen(ISQ_SITEMAP, timeout=ISQ_TIMEOUT).read()
    return raw.decode("utf-8", errors="replace")


def _probe_url(url: str, *, method: str = "GET", nbytes: int = PREFIX_BYTES) -> dict:
    """Observe one url: status, content-type, declared length, first `nbytes` bytes.

    Only a PREFIX is read, then the connection closes — so a 17MB candidate costs a
    handshake, not a download, while still yielding real magic bytes.

    `urllib.error.HTTPError` is caught and REPORTED AS THE OBSERVED STATUS: a 404 is an
    answer, not a probe failure, and this hunt's whole job is to record which URLs answer
    what. Every other exception propagates — a DNS failure or a timeout is NOT an observed
    status and must not be laundered into one.
    """
    req = urllib.request.Request(url, method=method)
    try:
        with urllib.request.urlopen(req, timeout=ISQ_TIMEOUT) as resp:
            return {
                "status": resp.status,
                "ctype": (resp.headers.get("Content-Type") or "").split(";")[0].strip(),
                "length": resp.headers.get("Content-Length") or "",
                "prefix": resp.read(nbytes) if nbytes else b"",
                "error": "",
            }
    except urllib.error.HTTPError as exc:
        return {"status": exc.code,
                "ctype": (exc.headers.get("Content-Type") or "").split(";")[0].strip()
                if exc.headers else "",
                "length": (exc.headers.get("Content-Length") or "") if exc.headers else "",
                "prefix": b"", "error": f"HTTPError {exc.code}"}


def _spec_text() -> str:
    """The spec file's text. NON-GATING boundary (a committed, READ-ONLY repo artifact)."""
    return SPEC_PATH.read_text(encoding="utf-8")


def _download(url: str) -> bytes:
    """Fetch one candidate in full, for the body-shape check. Boundary `isq-file`."""
    return urllib.request.urlopen(url, timeout=ISQ_TIMEOUT).read()


def _workbook_rows(data: bytes, *, max_rows: int = 0) -> tuple[list[str], list[tuple]]:
    """(sheet names, rows of the FIRST sheet) from xlsx bytes. Boundary `isq-file`.

    `openpyxl` directly, NOT pandas: nothing here needs a DataFrame, and the raw cell grid
    is what the header search and the label count read. Raises on a non-xlsx body — which
    is a REFUSAL, routed to UNKNOWN by the caller, never a silent empty result.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = [r for r in ws.iter_rows(max_row=max_rows or None, values_only=True)]
        return list(wb.sheetnames), rows
    finally:
        wb.close()


# --- small computed helpers ---------------------------------------------------------
def _norm(text: object) -> str:
    """The ONE name-normalisation rule in this file (see run_p5b.py: two rules for one job
    is what a later probe copies wrongly)."""
    return str(text or "").strip().lower()


def _terms(text: str, terms) -> list[str]:
    low = _norm(text)
    return [t for t in terms if t in low]


def _slug(url: str) -> str:
    return url.rsplit("/", 1)[-1]


def _count_str(result: dict) -> str:
    """CKAN's total-match count, or an explicit marker when the key is absent.

    `result.get("count")` would print a bare `None` into the note if CKAN omitted the key —
    a missing measurement rendered as if it were one.
    """
    n = result.get("count")
    return str(n) if isinstance(n, int) else "an UNREPORTED number of"


def _pkg_text(pkg: dict) -> str:
    """Title + notes + every resource name and url, for predicate matching."""
    parts = [pkg.get("title") or "", pkg.get("notes") or ""]
    for r in pkg.get("resources") or []:
        parts += [r.get("name") or "", r.get("url") or ""]
    return " | ".join(parts)


def _locs(xml: str) -> list[str]:
    """Every `<loc>` in the sitemap, IN DOCUMENT ORDER and NOT deduped.

    The caller dedupes and reports the measured effect. This used to return the deduped set
    behind the explanation "the raw document repeats each url once per hreflang alternate" —
    which the document refutes: there is exactly one `<loc>` per `<url>`, and the hreflang
    alternates are `<xhtml:link href=...>` attributes this regex never sees. The repeats
    that DO exist are a small number of genuinely duplicated page urls. Returning the raw
    list forces the note to state what deduping actually removed rather than why.
    """
    return re.findall(r"<loc>(.*?)</loc>", xml)


def _find_header(rows: list[tuple]) -> tuple[int | None, int | None]:
    """(row, col) of the geography header cell, or (None, None) when the sheet has none.

    PREFIX match — see `MRC_HEADER_PREFIX` for the measured reason equality and substring
    both fail across the editions this hunt sweeps.
    """
    for r, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        for c, value in enumerate(row):
            if _norm(value).startswith(MRC_HEADER_PREFIX):
                return r, c
    return None, None


def _find_column(rows: list[tuple], header_row: int, predicate) -> tuple[int, str]:
    """(column index, header text) of the first cell in `header_row` satisfying `predicate`,
    or (-1, "") when the sheet publishes no such axis. -1 rather than None so the caller
    cannot confuse "absent" with column 0."""
    for c, value in enumerate(rows[header_row]):
        text = str(value).strip() if value is not None else ""
        if text and predicate(_norm(text)):
            return c, text
    return -1, ""


def _ra_axis_usable(ra_col: int, geo_col: int) -> bool:
    """Is there a SEPARATE administrative-region column, i.e. a per-MRC RA code?

    ONE rule, three call sites (§3's corroboration, §3b's per-edition probe, §3c's
    membership sets). Some editions head their GEOGRAPHY column "MRC par région
    administrative": that cell NAMES the grouping but carries no code, so treating it as
    an RA column compares every label against itself — which reports NOT CORROBORATED for
    a file that simply does not publish the axis, and manufactures a DISJOINT relation in
    §3c. Three copies of this test is exactly the "two rules for one job" defect a later
    probe copies wrongly (run_p5b.py's `_norm`), so it lives here."""
    return ra_col >= 0 and ra_col != geo_col


def _ra_values_are_codes(rows: list[tuple], header_row: int, col: int) -> bool:
    """Does this column actually carry per-MRC RA CODES, or does its header merely say so?

    `_ra_axis_usable` tests INDICES — that the RA column is not the geography column. That is
    necessary and it survived mutation, but it is a PROXY: a column headed e.g. "Population de
    la région administrative" satisfies the header predicate at a non-geography index and
    would be read as RA codes, silently comparing populations against declared RA numbers.
    So the column is verified by its CONTENT, not by its name: every non-empty value below the
    header must be a short numeric code (the live workbook's are '11'..'17'), and a column
    with no sampled value at all cannot be confirmed and is REFUSED — the fail-safe direction,
    since an unverifiable axis must not read as a verified one.
    """
    if col < 0:
        return False
    values = [str(r[col]).strip() for r in rows[header_row + 1:]
              if col < len(r) and r[col] is not None and str(r[col]).strip()]
    return bool(values) and all(v.isdigit() and len(v) <= RA_CODE_MAX_DIGITS for v in values)


def _relation_head(relation: str) -> str:
    """The set-algebra VERDICT of a relation string, without its explanatory tail.

    `_ra_membership` returns e.g. `"PROPER SUBSET — declared targets do NOT exhaust this RA"`
    and `"NOT COMPUTABLE (no separate administrative-region column in this workbook)"`. The DECISION
    token needs the verdict alone, and it must come out the same way for BOTH tail forms —
    otherwise the token is machine-checkable on one shape and free prose on the other, which
    is how a gate ends up unable to see the case that matters.
    """
    return re.split(r"\s+[—(]", relation, maxsplit=1)[0].strip()


def _declared_ra_number(key: str) -> str:
    """`"RA14 Lanaudière"` -> `"14"`. Derived from the declaration's own key, so the number
    checked against the live response cannot drift from the key it is grouped under."""
    found = re.search(r"ra\s*(\d+)", key, re.IGNORECASE)
    return found.group(1) if found else ""


def _is_numeric_label(label: str) -> bool:
    """Is this geography label just a number? One rule, used by the wrong-column guard and by
    the note's own reporting, so the guard and the sentence describing it cannot disagree."""
    return label.replace(".", "", 1).strip().isdigit()


def _labels(rows: list[tuple], header_row: int, col: int) -> list[str]:
    """Distinct non-empty values below the header in the geography column, sorted."""
    out = set()
    for row in rows[header_row + 1:]:
        if col < len(row) and row[col] is not None and str(row[col]).strip():
            out.add(str(row[col]).strip())
    return sorted(out)


def _first_cell_matching(rows: list[tuple], needle: str, *, max_rows: int) -> str:
    """The first cell in the top `max_rows` whose text contains `needle`, verbatim."""
    for row in rows[:max_rows]:
        for value in row:
            if value is not None and needle in _norm(value):
                return str(value).strip()
    return ""


# --- floor guard (standalone so it can be mutation-tested) --------------------------
class VacuousProbeError(RuntimeError):
    """A verdict must be EARNED. Raised when a boundary answered but with a body that
    cannot support one: an empty sitemap, a candidate whose 200 carries a non-workbook
    body, a workbook with no MRC header or zero geography labels, or — for a NOT-FOUND —
    a searched population that was empty or never swept. Routes to UNKNOWN-PROBE-FAILED
    with the failing boundary, never a fabricated LOCATED and never a hollow NOT-FOUND."""


def _guard_sitemap(xml: str, locs: list[str]) -> None:
    """`isq-sitemap` guard: the ISQ-side population must EXIST before anything is claimed
    about it — a sweep over zero urls can neither locate nor rule out."""
    if not xml.strip():
        raise VacuousProbeError("sitemap.xml answered 200 but with an EMPTY body")
    if not locs:
        raise VacuousProbeError(
            "sitemap.xml answered 200 but parsed to ZERO <loc> entries — the ISQ-side sweep "
            "would run over an empty population"
        )


def _is_workbook_response(probe: dict) -> bool:
    """The CHEAP screen every eligible candidate passes through: a 200 whose first bytes are
    the ZIP magic. A bare 200 is not enough — this host serves HTML at 200 on some paths and
    a 45KB HTML body at 404 on others, so status alone cannot tell a workbook from an error
    page. Standalone so the screen and the deep guard below cannot drift apart."""
    return probe["status"] == 200 and probe["prefix"].startswith(XLSX_MAGIC)


def _guard_body(url: str, probe: dict, sheets: list[str], header_row: int | None,
                col: int | None, labels: list[str]) -> None:
    """`isq-file` guard — the one that keeps a LOCATED honest (see the mutation test).

    Which branches do the work, stated honestly because the mutation test grades them:

      * SAFETY-load-bearing — the three SHAPE branches (opened to at least one sheet, an
        MRC-named header cell found, that column carrying at least one label). Nothing else
        in the run inspects the opened bytes, so neutering this publishes a LOCATED whose
        geography evidence is an empty list. That is the mutation the test performs.
      * BACKSTOP only — the status and magic-byte branches. `_is_workbook_response` already
        screens both when `verified` is built, so on today's call path they cannot fire.
        They are kept because they state this function's PRECONDITION at the point that
        depends on it: a future pick rule that stopped screening would otherwise reach the
        `openpyxl` call with an HTML body and surface as a raw parse error rather than a
        refusal. They are not claimed to be doing work today.
    """
    if probe["status"] != 200:
        raise VacuousProbeError(
            f"the selected candidate {url} answered HTTP {probe['status']} — it cannot carry "
            f"the body-shape evidence a LOCATED requires"
        )
    if not probe["prefix"].startswith(XLSX_MAGIC):
        raise VacuousProbeError(
            f"the selected candidate {url} answered 200 but its first bytes are "
            f"{probe['prefix']!r}, not the {XLSX_MAGIC!r} workbook magic — a 200-but-wrong-body "
            f"page (this host serves HTML at 200), never a workbook"
        )
    if not sheets:
        raise VacuousProbeError(f"{url} opened but carries ZERO sheets")
    if header_row is None or col is None:
        raise VacuousProbeError(
            f"{url} opened but carries no cell beginning {MRC_HEADER_PREFIX!r} in its first "
            f"{HEADER_SCAN_ROWS} rows — nothing in this body identifies an MRC geography "
            f"column, so an MRC-level claim over it would be unearned"
        )
    if not labels:
        raise VacuousProbeError(
            f"{url}'s MRC column (row {header_row}, column {col}) carries ZERO labels — an "
            f"MRC-level answer over an empty column is unearned"
        )
    # THE WRONG-COLUMN CASE, which zero-label checking does NOT cover. A header search that
    # lands on a CODE column yields plenty of labels — they are just digits — so it sails
    # through every emptiness test and publishes a LOCATED whose "geography labels" are
    # numbers. Measured on the live workbook: the correct column is 0/122 numeric, the code
    # column reachable by a looser header search is 106/109. Refuse on digits.
    numeric = sum(1 for lb in labels if _is_numeric_label(lb))
    if numeric / len(labels) > GEO_LABEL_MAX_NUMERIC_FRACTION:
        raise VacuousProbeError(
            f"{url}'s geography column (row {header_row}, column {col}) is "
            f"{numeric}/{len(labels)} numeric labels — that is a CODE column, not a geography "
            f"column, and an MRC-level claim over digits is unearned. Sample: {labels[:5]}"
        )


def _guard_not_found(locs: list[str], eligible: list, verified: list, ckan: dict) -> None:
    """The NOT-FOUND guard — an ABSENCE claim needs a population that was actually swept,
    AND must not be standing in for a boundary that answered badly.

    R7: a zero-result sweep over an EMPTY catalogue, over a boundary that never answered, or
    a 200-but-wrong-body product page must each route to UNKNOWN, never NOT-FOUND. All three
    are checked here, so this single guard is the whole gate on an absence claim.

    Note the distinction the first check draws, because it is the one that decides whether
    NOT-FOUND means anything: ZERO eligible candidates is a real absence — the sweep looked
    and found nothing to look at. Eligible candidates that NONE verify is not an absence at
    all: the sweep found things and the boundary failed to serve them, which is exactly the
    wrong-body case. Collapsing the two would let an outage publish "no MRC source among the
    N swept" while N candidate urls sat in the table above it.
    """
    if eligible and not verified:
        raise VacuousProbeError(
            f"the sweep resolved {len(eligible)} eligible candidate(s) but NONE answered 200 "
            f"with a workbook body — that is a boundary serving wrong bodies (or refusing), "
            f"not an absence; a NOT-FOUND here would report the sweep's own failure as a "
            f"finding about ISQ's holdings"
        )
    if not locs:
        raise VacuousProbeError(
            f"the ISQ sitemap sweep ran over {len(locs)} loc(s) — a NOT-FOUND scoped to an "
            f"empty population is not an earned absence"
        )
    if not ckan["measured"]:
        raise VacuousProbeError(
            f"the CKAN boundary did not answer usefully ({ckan['why']}) — with only one of "
            f"the two searched populations swept, a NOT-FOUND would be scoped to a "
            f"population this run never enumerated"
        )
    if not ckan["n_catalogue"] or not ckan["n_swept"]:
        raise VacuousProbeError(
            f"the CKAN catalogue reported {ckan['n_catalogue']} package(s) and the sweep "
            f"enumerated {ckan['n_swept']} — an absence claim over an EMPTY catalogue is the "
            f"vacuous-absence shape, not a NOT-FOUND"
        )


# --- residual probes (RECORDED OBSERVATIONS — steering ruling G; never verdict-moving) ---
def _probe_editions(verified: list[str], picked: str, picked_data: bytes) -> list[dict]:
    """Open every verified candidate and record whether it carries an RA column.

    Residual (i). Only the first `HEADER_SCAN_ROWS` rows are materialised — the question is
    about the HEADER, so a 17MB candidate costs its download and a dozen rows, not a full
    parse. The picked workbook REUSES the bytes already downloaded rather than fetching them
    twice, so the two sections cannot disagree about the same file.

    A candidate that fails to open is recorded as `opened: False` WITH its exception type and
    message, never silently dropped: a workbook this run could not read is a real observation
    about that candidate, and dropping it would shrink the population the §3b counts are
    scoped to without saying so. This cannot mask a fault in the picked workbook — that one
    is already open and guarded before this runs.
    """
    out: list[dict] = []
    for url in verified:
        row = {"url": url, "opened": False, "error": "", "caption": "", "mrc_head": "",
               "ra_head": "", "ra_col": -1, "geo_col": -1, "ra_separate": False,
               "sheets": []}
        try:
            raw = picked_data if url == picked else _download(url)
            # A few rows PAST the header block, so the RA value check has something to
            # sample; a header-only read could not tell a code column from a population one.
            sheets, rows = _workbook_rows(raw, max_rows=HEADER_SCAN_ROWS + RA_VALUE_SAMPLE_ROWS)
            hr, hc = _find_header(rows)
            row["sheets"] = sheets
            row["caption"] = (str(rows[0][0]).strip()
                              if rows and rows[0] and rows[0][0] else "")
            if hr is not None and hc is not None:
                row["mrc_head"] = str(rows[hr][hc]).strip()
                row["geo_col"] = hc
                col, head = _find_column(
                    rows, hr,
                    lambda t: bool(RA_HEADER_PATTERN.match(t))
                    or any(k in t for k in RA_HEADER_MARKS))
                row["ra_col"], row["ra_head"] = col, head
                # THE DISTINCTION THAT MATTERS, and the one a symmetric gloss would destroy:
                # some editions head their GEOGRAPHY column "MRC par région administrative".
                # That cell NAMES an RA grouping; it is not a separate RA column, and no RA
                # code can be read from it per MRC. Counting those together with the editions
                # that publish a real `RA1` column would report a machine-readable axis where
                # none exists — the v1 constraint turns on exactly this difference.
                # Same two rules as §3, over the SAMPLED rows this header-only read holds —
                # stated in the note as header/sample-scoped, not whole-file-scoped.
                row["ra_separate"] = (_ra_axis_usable(col, hc)
                                      and _ra_values_are_codes(rows, hr, col))
            row["opened"] = True
        except Exception as exc:
            row["error"] = f"{type(exc).__name__}: {exc}"
        out.append(row)
    return out


def _ra_membership(rows: list[tuple], header_row: int, geo_col: int, ra_col: int, *,
                   usable: bool) -> tuple[dict, dict]:
    """For each DECLARED RA, the full MRC set this workbook assigns to it, and the relation
    the declared targets bear to that set.

    Residual (ii). The relation is computed with set algebra over the two sets — never
    described in prose — so the word printed in the note cannot disagree with the members
    printed beside it. `usable` is `_ra_axis_usable`'s verdict, passed in rather than
    re-derived: everything here reports `NOT COMPUTABLE` when it is False, and re-deriving
    the rule would let this section and §3 drift into disagreeing about the same workbook.
    """
    members: dict[str, dict] = {}
    relation: dict[str, str] = {}
    for ra_key, targets in COURONNE_MRC_BY_RA.items():
        declared = set(targets)
        if not usable:
            members[ra_key] = {"members": [], "declared": declared, "excluded": 0,
                               "undeclared": [], "missing": sorted(declared)}
            relation[ra_key] = (
                "NOT COMPUTABLE (no separate administrative-region column in this workbook)"
                if ra_col < 0 else
                "NOT COMPUTABLE (this workbook names its RA grouping in the geography header "
                "only — there is no separate column carrying a per-MRC RA code)")
            continue
        want = _declared_ra_number(ra_key)
        found_set, excluded = set(), 0
        for row in rows[header_row + 1:]:
            if geo_col >= len(row) or ra_col >= len(row):
                continue
            if row[geo_col] is None or row[ra_col] is None:
                continue
            label = str(row[geo_col]).strip()
            if not label or str(row[ra_col]).strip() != want:
                continue
            # A defensive exclusion, and UNEXERCISED on this data — say so rather than imply
            # it is doing work. MEASURED on the live workbook: all 17 RA-subtotal-form rows
            # carry an EMPTY RA cell, so they are already filtered out by the RA-code test
            # above and never reach this branch (the per-RA "excluded" count the note prints
            # is 0 for every RA, which is how a reader can see that). It is kept because an
            # edition that DOES code its subtotal rows would otherwise inflate every member
            # set by one and make an EQUAL relation unreachable — a silent wrong answer,
            # where the cost of the guard is one comparison.
            if AGGREGATE_LABEL_PATTERN.match(label):
                excluded += 1
                continue
            found_set.add(label)
        members[ra_key] = {
            "members": sorted(found_set), "declared": declared, "excluded": excluded,
            "undeclared": sorted(found_set - declared),
            "missing": sorted(declared - found_set),
        }
        if not found_set:
            relation[ra_key] = "EMPTY (this workbook assigns no MRC to that RA number)"
        elif declared == found_set:
            relation[ra_key] = "EQUAL — the declared set exhausts this RA"
        elif declared < found_set:
            relation[ra_key] = "PROPER SUBSET — declared targets do NOT exhaust this RA"
        elif found_set < declared:
            relation[ra_key] = "PROPER SUPERSET — this file declares MRCs the workbook does " \
                               "not put under that RA"
        elif declared & found_set:
            relation[ra_key] = "OVERLAPPING — neither contains the other"
        else:
            relation[ra_key] = "DISJOINT — no declared target is under that RA here"
    return members, relation


def _record_spec_premise(note: list[str]) -> dict:
    """Read spec §8's CURRENT text and compute the state of the premise this hunt tested.

    NON-GATING, and READ-ONLY. This exists because the note previously quoted the premise as
    a typed string; steering then amended §8, and the quote became a claim about a locked
    artifact that the artifact no longer made. Reading it live means the note cannot go stale
    against the document it cites.

    Three-valued on purpose. Absence of the old marker alone cannot distinguish "amended"
    from "rewritten in a way this predicate cannot read", so the amended state is identified
    POSITIVELY and anything else is reported as INDETERMINATE rather than guessed.
    """
    note += ["## 4. Spec-premise cross-check (RECORDED, non-gating, READ-ONLY)", ""]
    out = {"measured": False, "state": "NOT MEASURED THIS RUN", "why": "", "quote": ""}
    try:
        # EVERY matching line is kept and the count is reported. `SPEC_ROW_MARK` also occurs
        # in the spec's §11.6 line, so `rows[0]` is the §8 Geography row only because §8
        # precedes §11 in file order — a positional assumption. Reporting how many lines
        # matched makes that assumption visible instead of silent: a second match is a signal
        # to look, not something the reader is left to discover.
        rows = [ln for ln in _spec_text().splitlines() if SPEC_ROW_MARK in _norm(ln)]
        if not rows:
            out.update(measured=True, state="INDETERMINATE",
                       why=f"no line in the spec contains {SPEC_ROW_MARK!r}")
        else:
            line = rows[0]
            low = _norm(line)
            old = SPEC_OLD_PREMISE_MARK in low
            amended = SPEC_AMENDED_MARK in low
            # The premise sentences of that row — quoted as a CONTIGUOUS SPAN from the
            # first match to the last, never as a filtered re-join. `". ".join(filter(...))`
            # keeps only matching sentences and re-joins them, which is a true substring
            # ONLY while the matches happen to be adjacent; the moment it keeps 1 and 3 and
            # drops 2 it splices non-adjacent text and publishes the splice under "Quoted
            # verbatim", registered via `Fact.cited`. In a family where "verbatim" is
            # load-bearing that is a latent depth-1, so the span is taken by INDEX and any
            # dropped interior sentence is marked with an explicit ellipsis.
            parts = line.split(". ")
            keep = [i for i, sent in enumerate(parts)
                    if "mrc" in _norm(sent) or SPEC_ROW_MARK in _norm(sent)]
            if keep:
                span = parts[keep[0]:keep[-1] + 1]
                dropped = (keep[-1] - keep[0] + 1) - len(keep)
                quote = ". ".join(span)
                if dropped:
                    quote += f"  [contiguous span; {dropped} interior sentence(s) not matched]"
            else:
                quote = ""
            out.update(
                measured=True, quote=quote,
                state=("PREMISE STANDS" if old and not amended else
                       "AMENDED" if amended and not old else
                       "INDETERMINATE"),
                why=(f"old-premise marker {SPEC_OLD_PREMISE_MARK!r} "
                     f"{'PRESENT' if old else 'absent'}; amended marker "
                     f"{SPEC_AMENDED_MARK!r} {'PRESENT' if amended else 'absent'}"),
            )
        if out["quote"]:
            Fact.cited("spec §8's CURRENT text on the MRC premise",
                       f"{SPEC_PATH.name}: \"{out['quote']}\"")
        note += [
            f"- Read live from `{SPEC_PATH.relative_to(_REPO_ROOT)}` (READ-ONLY; this probe "
            f"never writes there). Marker `{SPEC_ROW_MARK!r}` matched **{len(rows)}** line(s); "
            f"the FIRST is read, which is the §8 Geography row only because §8 precedes §11 "
            f"in file order — a positional assumption, stated so a second match is visible "
            f"rather than silently outranked.",
            f"- **State: {out['state']}** — {out['why']}.",
            (f"- Quoted verbatim from the spec as it stands NOW: *\"{out['quote']}\"*"
             if out["quote"] else
             "- No premise sentence could be quoted from that row this run."),
            "",
            "  Why this is read rather than typed: this note's DECISION block cites spec §8. "
            "A typed quote of a locked artifact goes stale the moment the artifact is amended "
            "— which is exactly what happened here — and a stale quote is a claim nothing "
            "computed. Nothing in this section moves the verdict.",
            "",
        ]
        return out
    except Exception as exc:
        out["why"] = f"{type(exc).__name__}: {exc}"
        note += [
            f"- `SPEC CROSS-CHECK NOT MEASURED THIS RUN: {out['why']}`",
            "",
            "  The spec file could not be read. The DECISION block below therefore states "
            "that the premise was NOT CHECKED rather than asserting anything about it.",
            "",
        ]
        return out


# --- boundary A: Données Québec CKAN ------------------------------------------------
def _sweep_ckan(note: list[str]) -> dict:
    """Sweep the CKAN catalogue. Contributes the searched population a NOT-FOUND is scoped
    to; a failure here CANNOT demote a body-verified LOCATED (a located source is located),
    but it DOES block a NOT-FOUND — see `_guard_not_found`.

    Returns {"measured": bool, ...}; every failure is recorded as NOT MEASURED THIS RUN so
    an unswept catalogue can never read as a catalogue that was swept.
    """
    note += ["## 1. Boundary A — Données Québec CKAN (the enumerable open-data catalogue)",
             ""]
    out = {"measured": False, "why": "", "slug": "", "n_orgs": 0, "n_catalogue": 0,
           "n_swept": 0, "n_isq": 0, "n_match": 0, "matches": [], "quote": ""}
    try:
        orgs = (_ckan_get(CKAN_ORGS).get("result") or [])
        # The slug is RESOLVED, never typed. `organization:institut-de-la-statistique-du-quebec`
        # — the obvious guess — returns zero packages, and a zero from a wrong slug is a
        # fabricated absence, not a measurement.
        hits = sorted(
            (o.get("name") or "") for o in orgs
            if CKAN_ORG_TITLE_MARK in _norm(o.get("title"))
        )
        slug = hits[0] if hits else ""
        total = (_ckan_get(CKAN_COUNT).get("result") or {})
        n_catalogue = total.get("count") if isinstance(total.get("count"), int) else 0

        swept: dict[str, dict] = {}
        n_isq = 0
        if slug:
            isq_res = (_ckan_get(CKAN_ORG_SEARCH.format(slug=slug)).get("result") or {})
            for pkg in isq_res.get("results") or []:
                swept[pkg.get("id") or pkg.get("name") or ""] = pkg
            n_isq = len(swept)
        term_res = (_ckan_get(CKAN_TERM_SEARCH).get("result") or {})
        for pkg in term_res.get("results") or []:
            swept.setdefault(pkg.get("id") or pkg.get("name") or "", pkg)

        # A SINGLE three-way conjunction over title + notes + every resource name and url.
        # This is NOT "the same two-tier predicate boundary B uses": B has two tiers (SWEPT =
        # MRC AND population; ELIGIBLE = swept AND projection) and scopes its absence claims
        # to the WIDER swept tier, while this conjunction equals B's NARROWER eligible tier
        # only. The two boundaries' absence claims therefore do NOT mean the same thing, and
        # the note below says so instead of asserting a symmetry that does not hold.
        matches = []
        for pkg in swept.values():
            text = _pkg_text(pkg)
            if _terms(text, MRC_TERMS) and _terms(text, PROJECTION_TERMS) \
                    and _terms(text, POPULATION_TERMS):
                matches.append({"title": pkg.get("title") or "",
                                "org": ((pkg.get("organization") or {}).get("name") or ""),
                                "n_res": len(pkg.get("resources") or [])})

        # A CITED corroboration, resolved BY PREDICATE from a live package's own notes: a
        # statement about ISQ's publication practice. It is NOT evidence about the file
        # opened in §3 — that file is verified by its own bytes — and nothing downstream
        # gates on it.
        quote = ""
        for pkg in swept.values():
            for sentence in (pkg.get("notes") or "").replace("\n", " ").split("."):
                low = _norm(sentence)
                if "mrc" in low and "isq" in low and "diffuse" in low:
                    quote = sentence.strip()
                    break
            if quote:
                break

        out.update(measured=True, slug=slug, n_orgs=len(orgs), n_catalogue=n_catalogue,
                   n_swept=len(swept), n_isq=n_isq, n_match=len(matches),
                   matches=sorted(m["title"] for m in matches), quote=quote)

        f_orgs = Fact.derived(str(len(orgs)), "organizations in the live CKAN organization_list")
        f_cat = Fact.derived(str(n_catalogue), "packages in the live CKAN catalogue "
                                               "(package_search rows=0 count)")
        f_swept = Fact.derived(str(len(swept)), "distinct CKAN packages enumerated and swept "
                                                "by this run")
        if quote:
            Fact.cited("ISQ's own diffusion geographies, per a live CKAN package's notes",
                       f"donneesquebec.ca package notes: \"{quote}\"")

        note += [
            f"- `organization_list` -> **{f_orgs}** organizations. The ISQ slug is RESOLVED "
            f"from that live list by the title predicate `{CKAN_ORG_TITLE_MARK!r}`: "
            + (f"**`{slug}`** (title match)." if slug else
               "**NO organization matched the predicate this run** — so no org-scoped sweep "
               "ran, and the swept population below is the term query alone."),
            f"- `package_search?rows=0` -> **{f_cat}** packages in the catalogue "
            f"(`{_count_str(total)}` reported).",
            f"- Swept: **{f_swept}** distinct packages — {n_isq} from "
            + (f"`organization:{slug}`" if slug else "no org-scoped query")
            + f" and the remainder from the live term query `{CKAN_TERM_SEARCH}`.",
            f"- Candidate predicate — a SINGLE three-way conjunction over title + notes + "
            f"every resource name and url: an MRC term {list(MRC_TERMS)} AND a projection "
            f"term {list(PROJECTION_TERMS)} AND a population term {list(POPULATION_TERMS)}. "
            f"**{len(matches)}** of the {len(swept)} swept packages matched"
            + (f": {sorted(m['title'] for m in matches)}." if matches else "."),
            f"  Note the ASYMMETRY with boundary B rather than assuming the two match: this "
            f"conjunction is equivalent to B's NARROWER *eligible* tier, whereas B scopes its "
            f"absence claims to its WIDER *swept* tier (MRC AND population, projection term "
            f"not required). So the two boundaries' absence claims do not cover the same "
            f"thing, and neither is used to reinforce the other.",
            "",
            # A FUNCTION of `matches`, not an unconditional sentence: with a match present,
            # calling this an absence would contradict the count one line above it — the
            # exact adjective-beside-a-correct-value defect this family keeps reintroducing.
            (f"  This boundary is therefore NOT an absence: {len(matches)} swept package(s) "
             f"matched the candidate predicate. None of them is opened or body-checked here — "
             f"this boundary contributes the second searched population, and the verdict is "
             f"earned on boundary B below, where a candidate's bytes are actually inspected. "
             f"Whether a match is a real MRC-projection dataset or a slug-predicate false "
             f"positive is left to the reader: the titles are printed above, unglossed."
             if matches else
             f"  Scoped exactly: this is an absence **among the {len(swept)} packages this "
             f"run enumerated out of a {n_catalogue}-package catalogue** — not a claim about "
             f"the catalogue as a whole, and not a claim that ISQ publishes no MRC data. The "
             f"verdict is earned on boundary B below; this boundary contributes the second "
             f"searched population a NOT-FOUND would have to be scoped to."),
            "",
        ]
        if quote:
            note += [
                f"- CITED, verbatim from a live package's own `notes` (resolved by predicate, "
                f"not typed): *\"{quote}.\"* This is a statement about ISQ's publication "
                f"practice quoted from CKAN — it is NOT evidence about the file opened in §3, "
                f"which is verified by its own bytes, and nothing in the verdict rests on it.",
                "",
            ]
        return out
    except Exception as exc:
        out["why"] = f"{type(exc).__name__}: {exc}"
        note += [
            f"- `CKAN SWEEP NOT MEASURED THIS RUN: {out['why']}`",
            "",
            "  The CKAN boundary did not answer usefully. This does NOT demote a "
            "body-verified find on boundary B — a located source is located — but it DOES "
            "block a NOT-FOUND: with only one searched population swept, an absence claim "
            "would be scoped to a catalogue this run never enumerated. `_guard_not_found` "
            "enforces exactly that, and the run would record UNKNOWN-PROBE-FAILED instead.",
            "",
        ]
        return out


# --- the live hunt ------------------------------------------------------------------
def _section_3_body_shape(note: list[str], observed: dict, verified: list[str],
                          eligible: list[str]) -> dict:
    """§3 — pick ONE verified candidate, open it, and prove its body is MRC-level.

    Extracted from `_hunt`, which had grown past 600 lines. That length was not cosmetic:
    every gloss defect this file has shipped lived in a long run of nested conditional
    string concatenation, where a conditional clause and an unconditional one are visually
    identical. Sections that can be read whole can be reviewed whole.

    RAISES `VacuousProbeError` (boundary `isq-file`) via `_guard_body`. Returns the opened
    workbook's measured state; the caller passes it to §3b and §3c so all three read the
    SAME numbers rather than recomputing them.
    """
    # ============ §3 the body-shape check on ONE candidate ==========================
    # SELECTION RULE, stated before its result so it cannot drift to fit the answer:
    #   1. restrict to verified candidates matching a DECLARED demoflow family
    #      (`DEMOFLOW_RMR_FAMILY`) — the MRC analogue of a file demoflow already consumes at
    #      RMR level is what a v1 extension would read, so the witness must be one of those
    #      rather than merely the cheapest workbook on the site. This makes the family map
    #      LOAD-BEARING rather than a decorative table in §3;
    #   2. among those, the smallest by declared Content-Length (a shape witness only has to
    #      be sufficient; a 17MB download buys no extra evidence), ties broken by url;
    #   3. if NO verified candidate matches a family, fall back to the smallest verified
    #      overall — and say so in the note, because the witness is then weaker evidence
    #      about what a v1 extension could consume.
    def _declared_len(url: str) -> int:
        raw = observed[url]["length"]
        return int(raw) if str(raw).isdigit() else 1 << 62

    def _family_of(url: str) -> str:
        for family, terms in DEMOFLOW_RMR_FAMILY.items():
            if all(t in _norm(url) for t in terms):
                return family
        return ""

    family_verified = [u for u in verified if _family_of(u)]
    picked = sorted(family_verified or verified, key=lambda u: (_declared_len(u), u))[0]
    picked_family = _family_of(picked)
    # The METHOD comparison, MEASURED on the same url rather than asserted. The plan body's
    # hunt is HEAD-only; whether that is fatal on this host is an empirical question, and
    # the sentence emitted below is a function of these two observed codes.
    head = _probe_url(picked, method="HEAD", nbytes=0)
    head_disagrees = head["status"] != observed[picked]["status"]
    data = _download(picked)
    sheets, rows = _workbook_rows(data)
    header_row, header_col = _find_header(rows)
    labels = _labels(rows, header_row, header_col) if header_row is not None else []
    _guard_body(picked, observed[picked], sheets, header_row, header_col, labels)  # RAISES

    # The scenario axis, located by its OWN header cell — never by assuming column 0. An
    # earlier version read column 0 unconditionally and would have published this edition's
    # `Code` column as a list of "scenario labels": a fabricated axis with real-looking
    # values in it.
    scen_col, scen_head = _find_column(
        rows, header_row, lambda t: any(k in t for k in SCENARIO_HEADER_MARKS))
    scenarios = _labels(rows, header_row, scen_col) if scen_col >= 0 else []
    caption = str(rows[0][0]).strip() if rows and rows[0] and rows[0][0] else ""
    diffusion = _first_cell_matching(rows, "diffusion", max_rows=HEADER_SCAN_ROWS)
    if caption:
        Fact.cited("the opened workbook's own caption",
                   f"cell A1 of {_slug(picked)}: \"{caption}\"")
    if diffusion:
        Fact.cited("the opened workbook's own release line",
                   f"{_slug(picked)}: \"{diffusion}\"")

    # The per-target couronne search — the measurement that BEARS on the conclusion. The
    # count that matters is NOT "how many MRC labels exist" (105 labels would be equally
    # consistent with the couronne being absent); it is whether each DECLARED target is
    # among them, searched by name.
    found: dict[str, list[str]] = {}
    for ra, targets in COURONNE_MRC_BY_RA.items():
        for target in targets:
            found[target] = [m for m in labels if _norm(m) == _norm(target)]
    n_targets = sum(len(t) for t in COURONNE_MRC_BY_RA.values())
    hits = sorted(t for t, m in found.items() if m)
    misses = sorted(t for t, m in found.items() if not m)
    couronne_complete = not misses

    # Does the opened sheet carry an administrative-region axis at all? SEARCHED FOR, because
    # the RA↔MRC correspondence is the one thing spec §8's `ra_proxy` rows turn on: its
    # presence AND its absence both have to be measured to be stated. Editions differ — the
    # A2021 components sheet heads this column `RA1`; the 2025 sheets publish none.
    # WHAT THE REJECTED PREDICATES ACTUALLY DO — run, not asserted. The note used to claim a
    # substring search "counts zero labels below it"; on the live workbook it returns 109
    # labels from the CODE column, which is strictly WORSE than zero: zero trips the
    # empty-column guard and fails safe, while a populated wrong column sails through it. The
    # rejected predicate is therefore EXECUTED here and its real outcome published.
    sub_row, sub_col = None, None
    for _r, _row in enumerate(rows[:HEADER_SCAN_ROWS]):
        for _c, _v in enumerate(_row):
            if _v is not None and MRC_HEADER_PREFIX in _norm(_v):
                sub_row, sub_col = _r, _c
                break
        if sub_row is not None:
            break
    sub_labels = _labels(rows, sub_row, sub_col) if sub_row is not None else []
    sub_numeric = sum(1 for lb in sub_labels if _is_numeric_label(lb))

    header_cells = [str(v).strip() for v in rows[header_row] if v is not None and str(v).strip()]
    ra_col, ra_head = _find_column(
        rows, header_row,
        lambda t: bool(RA_HEADER_PATTERN.match(t)) or any(k in t for k in RA_HEADER_MARKS))
    # Whether the axis is READABLE here — see `_ra_axis_usable`, which is the single rule
    # §3, §3b and §3c all consult, so none of the three can disagree with the others about
    # whether this workbook publishes a per-MRC RA code.
    # Index rule AND value rule. The index rule alone is a proxy — see `_ra_values_are_codes`.
    # Scope here: the FULL sheet, every data row below the header.
    ra_usable = (_ra_axis_usable(ra_col, header_col)
                 and _ra_values_are_codes(rows, header_row, ra_col))

    # The DECLARED RA number of each target, checked against the code the LIVE response puts
    # beside that MRC — an independent witness this file does not control (P5b's
    # declared-province pattern). Per target: the observed codes, and whether they agree.
    # `None` where the edition publishes no RA column: NOT CHECKABLE, never a silent pass.
    ra_observed: dict[str, list[str]] = {}
    ra_agrees: dict[str, bool | None] = {}
    for ra_key, targets in COURONNE_MRC_BY_RA.items():
        want = _declared_ra_number(ra_key)
        for target in targets:
            if not ra_usable or not found[target]:
                ra_observed[target], ra_agrees[target] = [], None
                continue
            codes = sorted({
                str(row[ra_col]).strip()
                for row in rows[header_row + 1:]
                if header_col < len(row) and ra_col < len(row)
                and row[header_col] is not None and row[ra_col] is not None
                and _norm(row[header_col]) == _norm(target) and str(row[ra_col]).strip()
            })
            ra_observed[target] = codes
            ra_agrees[target] = (bool(codes) and all(c == want for c in codes)) if want else None
    ra_checked = [t for t in ra_agrees if ra_agrees[t] is not None]
    ra_disagree = sorted(t for t in ra_checked if ra_agrees[t] is False)

    # The DECOMPOSITION, not the raw total: the column interleaves administrative-region
    # subtotal rows with the MRC rows, so "122 labels" would read as an MRC count and be
    # wrong by 17. What the non-aggregate remainder contains is NOT asserted — the full list
    # is emitted above and the reader judges it.
    aggregate_labels = [m for m in labels if AGGREGATE_LABEL_PATTERN.match(m)]
    fine_labels = [m for m in labels if not AGGREGATE_LABEL_PATTERN.match(m)]

    f_nlabels = Fact.derived(str(len(labels)), "distinct geography labels in the opened "
                                               "workbook's MRC column")
    f_ndecomp = Fact.derived(
        f"{len(aggregate_labels)} + {len(fine_labels)}",
        f"of those labels, the count in the `NN  Name` administrative-region-subtotal form "
        f"and the remainder")
    f_hits = Fact.derived(f"{len(hits)} of {n_targets}",
                          "declared couronne MRC targets found by exact name search of the "
                          "opened workbook's own labels")
    f_scen = Fact.derived(str(len(scenarios)),
                          f"distinct labels in the opened workbook's scenario column"
                          + (f" (header {scen_head!r})" if scen_col >= 0
                             else " — no column in its header row names a scenario axis"))
    f_ra = Fact.derived(
        f"{len(ra_checked) - len(ra_disagree)} of {len(ra_checked)}",
        "declared couronne targets whose DECLARED RA number equals the administrative-region "
        "code the opened workbook puts beside that MRC"
        + (f" (column {ra_col}, header {ra_head!r})" if ra_usable
           else " — this edition publishes no SEPARATE RA column (a geography header that "
                "merely NAMES the grouping carries no per-MRC code), so NONE was checkable"))

    note += [
        f"## 3. Body-shape check — is `{_slug(picked)}` really MRC-level?",
        "",
        f"- Selected DETERMINISTICALLY from the {len(verified)} verified candidates by the "
        f"rule stated in the code before its result: "
        + (f"{len(family_verified)} of them match a DECLARED demoflow family, and this is the "
           f"smallest of those by declared `Content-Length` ({observed[picked]['length']} "
           f"bytes) — family **{picked_family}**."
           if picked_family else
           f"NONE of them matches a declared demoflow family, so this is the fallback — the "
           f"smallest verified candidate overall ({observed[picked]['length']} bytes). The "
           f"witness is therefore weaker evidence about what a v1 extension could consume, "
           f"and this note says so rather than implying a family match.")
        + " A shape witness only has to be sufficient; the note does NOT claim this is the "
          "newest edition — "
        + ("the caption and release line below are read from its own bytes and state which "
           "edition it is." if caption and diffusion else
           "the caption below is read from its own bytes and states which edition it is (this "
           "workbook publishes NO release line — see below)." if caption else
           "this workbook publishes no release line, and cell A1 carries no caption either, "
           "so nothing in its own bytes states which edition it is."),
        f"- Full GET -> {len(data)} bytes; prefix `{observed[picked]['prefix'].hex()}` matches "
        f"the `{XLSX_MAGIC.hex()}` workbook magic; opened with {len(sheets)} sheet(s): "
        f"{sheets}.",
        f"- **Method comparison, measured on this same url:** GET -> "
        f"**{observed[picked]['status']}**, HEAD -> **{head['status']}**"
        + (f". The two DISAGREE, so on this host a HEAD-only hunt (which is what the plan "
           f"body's P6 sketch performs) would record this live workbook as absent. That is a "
           f"measured property of this endpoint, not a general rule."
           if head_disagrees else
           ". The two AGREE on this url, so this run records no HEAD/GET discrepancy here — "
           "the note draws no conclusion about HEAD-based hunts from it."),
        f"- Caption cell A1 (verbatim): *\"{caption}\"*" if caption
        else "- Cell A1 carries no caption this run.",
        f"- Release line (verbatim, resolved by predicate): *\"{diffusion}\"*" if diffusion
        else "- No cell in the header block names a diffusion date this run.",
        f"- Geography column located by a header cell BEGINNING `{MRC_HEADER_PREFIX.upper()}` "
        f"at row {header_row}, column {header_col} (0-indexed); the cell reads "
        f"`{header_cells[header_col] if header_col < len(header_cells) else '?'}`. Prefix, not "
        f"equality and not substring, and BOTH rejected predicates were RUN on this workbook "
        f"so the reason is measured rather than asserted: an equality test would miss the "
        f"editions whose geography header carries an RA qualifier (the §3b table lists every "
        f"opened candidate's header), while a substring test locks onto row {sub_row} column "
        f"{sub_col} and returns **{len(sub_labels)}** labels of which **{sub_numeric}** are "
        f"numeric"
        + (f" — e.g. {sub_labels[:5]}. That is the CODE column, and it is strictly more "
           f"dangerous than an empty one: zero labels trip the empty-column guard and fail "
           f"safe, whereas a populated wrong column would sail through it. `_guard_body` "
           f"therefore also refuses a geography column that is more than "
           f"{GEO_LABEL_MAX_NUMERIC_FRACTION:.0%} numeric."
           if sub_labels and sub_numeric / len(sub_labels) > GEO_LABEL_MAX_NUMERIC_FRACTION
           else f" — this run records no numeric-column hazard from the substring path on "
                f"this workbook.")
        + " No edition or year is named here; the §3b table is what states which files have "
          "which header.",
        f"- Full header row (verbatim): {header_cells}.",
        f"- **{f_nlabels} distinct geography labels** below that header, which decompose as "
        f"**{f_ndecomp}**: labels in the `NN  Name` administrative-region-SUBTOTAL form, plus "
        f"the remainder. That split is emitted because the raw total would read as an MRC "
        f"count and be wrong — this column interleaves RA subtotals with the MRC rows. What "
        f"the remainder contains is NOT asserted here; the full list is emitted verbatim from "
        f"the live response, so the LEVEL is self-evidencing rather than glossed — a count "
        f"alone would leave \"MRC-level\" a word beside a number (P5b's precedent):",
        "",
    ]
    note += [f"  {i}. {label}" for i, label in enumerate(labels, start=1)]
    note += [
        "",
        f"- Scenario axis: "
        + (f"**{f_scen}** distinct labels in column {scen_col} (header `{scen_head}`): "
           f"{scenarios}."
           if scen_col >= 0 else
           f"NONE — no cell in this sheet's header row names a scenario axis (so the "
           f"{f_scen} labels counted for it are counted over nothing). The "
           f"sheet-name list above is emitted verbatim; this note draws no conclusion from it "
           f"about how this edition separates its scenarios."),
        "",
        f"**The declared couronne targets, searched BY NAME in those labels — {f_hits} found.** "
        f"The label COUNT above does not bear on couronne precision (a large MRC set is "
        f"equally consistent with the couronne being absent); this per-target search is the "
        f"measurement that does.",
        "",
        "| declared RA (spec §8 `ra_proxy`) | declared MRC target | found in the opened "
        "workbook? | RA code observed beside it | agrees with the declared RA number? |",
        "|---|---|---|---|---|",
    ]
    for ra_key, targets in COURONNE_MRC_BY_RA.items():
        want = _declared_ra_number(ra_key)
        for target in targets:
            agree = ra_agrees[target]
            note.append(
                f"| {ra_key} | {target} | "
                f"{'YES — ' + repr(found[target][0]) if found[target] else '**NO**'} | "
                f"{ra_observed[target] or '—'} | "
                + ("CORROBORATED" if agree is True
                   else f"**NOT CORROBORATED (declared {want})**" if agree is False
                   else "NOT CHECKABLE") + " |"
            )
    note += [
        "",
        f"- The target list is **DECLARED in this file** from spec §8's three `ra_proxy` rows, "
        f"not derived from the response; what is COMPUTED is the search result per target, and "
        f"a miss is published as a miss "
        + ("(none missed this run)." if couronne_complete
           else f"— MISSING this run: {misses}."),
        f"- **The RA↔MRC correspondence — {f_ra} declared targets corroborated.** "
        + (f"This edition DOES publish a SEPARATE administrative-region column (column {ra_col}, header "
           f"`{ra_head}`), so the RA number this file DECLARES for each target is checked "
           f"against the code the live response puts beside that MRC — an independent witness "
           f"nothing here controls. Flip a declared key to the wrong RA and the check stops "
           f"corroborating."
           + (f" **{len(ra_disagree)} target(s) DISAGREE: {ra_disagree} — treat every RA "
              f"grouping in this note as UNSUPPORTED until reconciled.**" if ra_disagree
              else "")
           if ra_usable else
           "This edition publishes NO SEPARATE administrative-region column — its header row is listed "
           "verbatim above — so NO target was checkable here and this run corroborates no RA "
           "grouping at all.")
        # CONDITIONAL, like the DECISION-side gloss it is the sibling of. Concatenated flat,
        # this sentence followed "NO target was checkable here" with "each declared target is
        # present and sits under the RA number this file declares for it" — self-contradictory
        # in one breath, and false twice on the header-named-only fixture (2 of 10 present,
        # none checkable). Every clause below is now a function of what §3 measured.
        + " What this establishes is MEMBERSHIP, and only as far as it was measured: "
        + (f"all {n_targets} declared targets are present in this workbook's labels"
           if couronne_complete else
           f"{len(hits)} of {n_targets} declared targets are present in this workbook's "
           f"labels ({len(misses)} absent)")
        + (f", and each of the {len(ra_checked)} checkable ones sits under the RA number this "
           f"file declares for it"
           if ra_checked and not ra_disagree else
           f", and of those, {len(ra_checked)} had a checkable RA code with "
           f"{len(ra_disagree)} DISAGREEING" if ra_disagree else
           ", and NONE had a checkable RA code, so this section corroborates no RA grouping")
        + ". It is NOT a partition, and this section makes no partition claim: §3c computes "
          "what the workbook can actually say about exhaustion. Scoped to the ONE workbook "
          "opened here; §3b opens the other verified candidates and reports the RA axis "
          "across them, so the edition scope of this corroboration is measured rather than "
          "assumed.",
        "",
        f"- **Which swept files would feed a v1 extension.** demoflow consumes two ISQ families "
        f"at RMR level; the DECLARED term map {({k: list(v) for k, v in DEMOFLOW_RMR_FAMILY.items()})} "
        f"is matched against the ELIGIBLE slugs above (a slug match, not a schema comparison — "
        f"no equivalence between the RMR and MRC editions is tested here):",
        "",
    ]
    for family, terms in DEMOFLOW_RMR_FAMILY.items():
        fam = sorted(_slug(u) for u in eligible if all(t in _norm(u) for t in terms))
        note.append(f"  - **{family}** -> {len(fam)} eligible slug(s) match: {fam or 'none'}")
    note.append("")


    return {
        "picked": picked, "picked_family": picked_family, "data": data, "sheets": sheets,
        "rows": rows, "header_row": header_row, "header_col": header_col,
        "header_cells": header_cells, "labels": labels, "aggregate_labels": aggregate_labels,
        "fine_labels": fine_labels, "scenarios": scenarios, "scen_col": scen_col,
        "scen_head": scen_head, "caption": caption, "diffusion": diffusion,
        "ra_col": ra_col, "ra_head": ra_head, "ra_usable": ra_usable,
        "ra_observed": ra_observed, "ra_agrees": ra_agrees, "ra_checked": ra_checked,
        "ra_disagree": ra_disagree, "found": found, "hits": hits, "misses": misses,
        "n_targets": n_targets, "couronne_complete": couronne_complete,
        "head": head, "head_disagrees": head_disagrees,
        "family_verified": family_verified,
    }


def _section_3b_editions(note: list[str], verified: list[str], wb: dict) -> dict:
    """§3b — residual (i): is the RA↔MRC axis edition-specific? RECORDED, never verdict-moving."""
    picked, data = wb["picked"], wb["data"]
    # ===== §3b — RESIDUAL (i): is the RA↔MRC axis EDITION-SPECIFIC? =================
    # A RECORDED OBSERVATION (steering ruling G), not a verdict: nothing below can move
    # DECISION-VERDICT. Every verified candidate is opened — the picked one reuses the bytes
    # already downloaded — and each is asked the SAME two questions its own header answers.
    editions = _probe_editions(verified, picked, data)
    opened = [e for e in editions if e["opened"]]
    # THREE states, not two. Collapsing the middle one would report a machine-readable axis
    # where the edition merely NAMES the grouping in its geography header.
    ra_separate = [e for e in opened if e["ra_separate"]]
    ra_named_only = [e for e in opened if e["ra_col"] >= 0 and not e["ra_separate"]]
    ra_absent = [e for e in opened if e["ra_col"] < 0]

    f_editions = Fact.derived(
        f"{len(ra_separate)} / {len(ra_named_only)} / {len(ra_absent)}",
        f"of the {len(opened)} candidate workbooks opened this run: those with a SEPARATE "
        f"administrative-region column, those whose GEOGRAPHY header merely names an RA "
        f"grouping, and those with neither")

    note += [
        "## 3b. Residual (i) — is the RA↔MRC axis EDITION-SPECIFIC? (RECORDED, non-gating)",
        "",
        (f"All {len(verified)} verified candidates were opened and asked the same question "
         f"their own header rows answer."
         if len(opened) == len(editions) else
         f"Of the {len(verified)} verified candidates, **{len(opened)} opened** and "
         f"{len(editions) - len(opened)} did NOT — the failures are named in the table below "
         f"and every count in this section covers the opened ones only.")
        + f" The split is **{f_editions}** — a SEPARATE RA column / an RA grouping NAMED in "
          f"the geography header only / neither.",
        "",
        # A RATIONALE paragraph, carrying NO population count and no header literal. It used
        # to open "Three editions head their GEOGRAPHY column `MRC par région administrative`"
        # — both typed, and both true only by coincidence: forced to a 1/0/0 or 0/1/0 split
        # the paragraph still read "Three editions", one paragraph above the correctly
        # computed token. The class-level fix is that a rationale states the RULE and the
        # table beneath states the population; a count here can only ever duplicate a
        # computed one, and duplication is what goes stale.
        "That middle state is kept distinct on purpose. When an edition heads its GEOGRAPHY "
        "column with an RA grouping rather than publishing a second column, that cell NAMES "
        "the grouping but carries **no per-MRC RA code** — so nothing in such a file assigns "
        "an individual MRC to a region. Counting those together with the editions that do "
        "publish a separate column would report a machine-readable axis where none exists, "
        "and the v1 constraint turns on exactly that difference. Which candidates fall in "
        "which state is the table below, computed per file.",
        "",
        "Each row's edition is its OWN caption cell, verbatim. This run does NOT rank the "
        "editions by recency: no live response states which is current, so that judgment is "
        "left to the reader with the captions in front of them.",
        "",
        "| candidate (slug) | opened? | geography header | RA axis | caption cell A1 "
        "(verbatim) |",
        "|---|---|---|---|---|",
    ]
    for e in editions:
        note.append(
            f"| `{_slug(e['url'])}` | "
            + ("yes" if e["opened"] else f"**NO — {e['error']}**") + " | "
            + (f"`{e['mrc_head']}`" if e["opened"] else "—") + " | "
            + (f"**SEPARATE column `{e['ra_head']}`** (col {e['ra_col']})" if e["ra_separate"]
               else f"NAMED IN THE GEOGRAPHY HEADER ONLY (`{e['ra_head']}`, col {e['ra_col']} "
                    f"= the geography column) — no per-MRC RA code" if e["ra_col"] >= 0
               else ("**none**" if e["opened"] else "—")) + " | "
            + (f"*{e['caption']}*" if e["caption"] else "(no caption cell)") + " |"
        )

    # A computed CO-OCCURRENCE over the caption text, not a recency ranking: which caption
    # markers partition the has-axis group from the no-axis group. Emitted only when the
    # separation is PERFECT in this population, and labelled as co-occurrence — a marker that
    # sorts 15 files is not thereby a cause, and this run tests no causal claim.
    # Counted across ALL THREE groups, not two. An earlier version compared only
    # `ra_separate` against `ra_absent` — 12 of the 15 files — and then described the result
    # as "a perfect separation across the 15 files opened": a two-state measurement glossed
    # over a three-state population, by the very edit that created the third state.
    cooccur = []
    for marker in EDITION_CAPTION_MARKERS:
        n_sep = sum(1 for e in ra_separate if marker in _norm(e["caption"]))
        n_named = sum(1 for e in ra_named_only if marker in _norm(e["caption"]))
        n_abs = sum(1 for e in ra_absent if marker in _norm(e["caption"]))
        if ra_absent and n_abs == len(ra_absent) and not n_sep:
            cooccur.append((marker, n_sep, n_named, n_abs))

    note += [
        "",
        f"- **Measured consequence for a v1 extension.** "
        + (f"The axis IS edition-specific: {len(ra_separate)} opened candidate(s) publish a "
           f"separate RA column, {len(ra_named_only)} name the grouping in the geography "
           f"header only, and {len(ra_absent)} carry neither. So a v1 that pins a workbook "
           f"from either of the latter two groups would get its projection values and its "
           f"machine-readable RA↔MRC axis from DIFFERENT workbooks — a cross-edition join. "
           f"This run does NOT validate such a join: it tests no label-set agreement, no "
           f"vintage compatibility and no reconciliation between any two editions. That is a "
           f"v1 design constraint, recorded here."
           if ra_separate and (ra_named_only or ra_absent) else
           f"All {len(opened)} opened candidate(s) publish a separate RA column, so this run "
           f"finds NO edition-specificity across the population it opened."
           if ra_separate else
           f"NO opened candidate publishes a separate RA column. If the picked workbook is "
           f"among them this CONTRADICTS §3's corroboration — inspect before relying on "
           f"either."),
        (f"- **Caption co-occurrence (computed over ALL THREE groups, NOT a recency "
         f"ranking).** Marker(s) {[m for m, _, _, _ in cooccur]}: carried by "
         + "; ".join(
             f"{n_abs}/{len(ra_absent)} of the no-axis files, {n_sep}/{len(ra_separate)} of "
             f"the separate-RA-column files and {n_named}/{len(ra_named_only)} of the "
             f"header-named-only files (marker `{m}`)"
             for m, n_sep, n_named, n_abs in cooccur)
         + f". The separation is between the no-axis and separate-column groups; the "
           f"header-named-only group's counts are printed rather than folded into that claim, "
           f"because it is a THIRD state and a two-group separation says nothing about it. "
           f"This run tests only CO-OCCURRENCE: it does not claim the marker causes the "
           f"absence, does not rank the editions, and says nothing about files outside the "
           f"swept set."
         if cooccur else
         f"- **Caption co-occurrence:** no caption marker in {list(EDITION_CAPTION_MARKERS)} "
         f"separates the no-axis group from the separate-RA-column group in this population, "
         f"so this run offers no caption-level explanation of the split."),
        "- Scope asymmetry, stated because the two halves differ: the *candidate population* "
        "is sweep-scoped (every verified candidate in §2), while each *edition label* is "
        "file-scoped (that workbook's own caption). And this whole section is "
        f"HEADER-AND-SAMPLE-scoped — only the first "
        f"{HEADER_SCAN_ROWS + RA_VALUE_SAMPLE_ROWS} rows of each non-picked candidate were "
        f"read: enough to locate the header row and to sample whether an RA column holds "
        f"short numeric CODES rather than merely carrying an RA-shaped name, but nothing "
        f"about the rest of their data. The picked workbook alone was read in full.",
        "",
    ]


    return {"editions": editions, "opened": opened, "ra_separate": ra_separate,
            "ra_named_only": ra_named_only, "ra_absent": ra_absent}


def _section_3c_membership(note: list[str], wb: dict) -> dict:
    """§3c — residual (ii): membership vs partition. RECORDED, never verdict-moving."""
    rows, header_row, header_col = wb["rows"], wb["header_row"], wb["header_col"]
    ra_col, ra_head, ra_usable = wb["ra_col"], wb["ra_head"], wb["ra_usable"]
    labels, header_cells = wb["labels"], wb["header_cells"]
    # ===== §3c — RESIDUAL (ii): membership vs partition ============================
    # Also a RECORDED OBSERVATION. §3 established MEMBERSHIP (each declared target is
    # present). The question here is EXHAUSTION: what is the FULL set of MRCs this workbook
    # assigns to each declared RA, and how do the declared targets relate to it?
    ra_members, ra_relation = _ra_membership(rows, header_row, header_col, ra_col,
                                             usable=ra_usable)
    # Can this file answer the OTHER half — whether those MRCs compose the Montréal RMR
    # couronne? That is a metropolitan-area question, so MEASURE whether the file carries a
    # metropolitan axis at all rather than asserting it does not.
    rmr_cells = [h for h in header_cells if any(k in _norm(h) for k in RMR_MARKERS)]
    rmr_labels = [m for m in labels if any(k in _norm(m) for k in RMR_MARKERS)]
    f_rmr = Fact.derived(
        f"{len(rmr_cells)} + {len(rmr_labels)}",
        "header cells and geography labels of the opened workbook matching a metropolitan-area "
        f"marker {list(RMR_MARKERS)}")

    note += [
        "## 3c. Residual (ii) — membership vs partition (RECORDED, non-gating)",
        "",
        f"§3 established MEMBERSHIP: each declared target is present, under the RA number this "
        f"file declares for it. This section asks the harder question — **EXHAUSTION**: what is "
        f"the FULL set of MRCs the opened workbook assigns to each declared RA, and how do the "
        f"declared targets relate to it? Rows in the `NN  Name` administrative-region-SUBTOTAL "
        f"form are EXCLUDED from these sets (they are the RA's own subtotal line, not one of "
        f"its MRCs); the exclusion count is printed per RA so it can be checked.",
        "",
    ]
    if not ra_usable:
        note += [
            "- **NOT COMPUTABLE from this workbook**: "
            + ("it publishes no separate administrative-region column at all"
               if ra_col < 0 else
               f"its RA grouping is NAMED in the geography header (`{ra_head}`, column "
               f"{ra_col}) but there is no SEPARATE column, so no per-MRC RA code exists to "
               f"read")
            + ". No MRC can be assigned to an RA here and no exhaustion relation exists to "
            "compute. Recorded as a limit, not resolved elsewhere.",
            "",
        ]
    else:
        note += [
            "| declared RA | MRCs the workbook assigns to it | declared targets | relation "
            "| RA-subtotal rows excluded |",
            "|---|---:|---:|---|---:|",
        ]
        for ra_key in COURONNE_MRC_BY_RA:
            m = ra_members[ra_key]
            note.append(
                f"| {ra_key} | {len(m['members'])} | {len(m['declared'])} | "
                f"**{ra_relation[ra_key]}** | {m['excluded']} |"
            )
        note += [
            "",
            "The MRCs each declared RA carries, verbatim, with the declared targets marked — "
            "so the relation above is checkable rather than taken:",
            "",
        ]
        for ra_key in COURONNE_MRC_BY_RA:
            m = ra_members[ra_key]
            marked = ", ".join(
                f"**{x}**" if x in m["declared"] else x for x in m["members"]
            ) or "none"
            note.append(f"- **{ra_key}** -> {marked}")
            if m["undeclared"]:
                note.append(f"  - present in the workbook but NOT declared by this file: "
                            f"{m['undeclared']}")
            if m["missing"]:
                note.append(f"  - declared by this file but ABSENT from the workbook's set for "
                            f"this RA: {m['missing']}")
        note += [
            "",
            # Explains the relations this run ACTUALLY computed. Written flat it explained
            # `PROPER SUBSET` unconditionally — asserting about "this file" under a hypothesis
            # that is false whenever the computed relation is PROPER SUPERSET or EMPTY.
            f"- **What the computed relation(s) mean, and only that.** This run computed "
            + ", ".join(sorted({_relation_head(v) for v in ra_relation.values()}))
            + ". " + " ".join(RELATION_GLOSS[r] for r in sorted({
                _relation_head(v) for v in ra_relation.values()}) if r in RELATION_GLOSS)
            + " Each is a statement about THIS FILE's RA assignment; none is a statement "
            "about which MRCs belong to the couronne, which nothing here measures.",
            "",
        ]
    note += [
        f"- **What this workbook CANNOT answer, measured rather than asserted.** Whether the "
        f"declared MRCs exactly COMPOSE the Montréal RMR couronne is a metropolitan-area "
        f"question. Matching the metropolitan-area markers {list(RMR_MARKERS)} against this "
        f"file's own header cells and geography labels yields **{f_rmr}** hits respectively"
        + (f" — with zero of each, nothing **this run read of this file** (its header row and "
           f"its geography column) names a metropolitan area, so nothing it read supplies RMR "
           f"membership for any MRC and the couronne-composition question is NOT ANSWERABLE "
           f"from it. Stated at that scope deliberately: an unnamed column carrying RMR codes, "
           f"or a footnote row below the header block, is OUTSIDE what this run inspected, so "
           f"a flat \"this file cannot answer it\" would claim more than the search covers. "
           f"Per steering ruling G this run does NOT consult a second source to close it: the "
           f"limit is the result."
           if not rmr_cells and not rmr_labels else
           f" — a metropolitan-area marker DOES appear here, so this file may carry an RMR "
           f"axis after all; inspect the hits above before treating the couronne question as "
           f"unanswerable from it. Cells: {rmr_cells}; labels: {rmr_labels[:10]}.")
        + " Either way this changes no verdict: §11.6 stands — a find enables v1, never v0.",
        "",
    ]


    return {"ra_members": ra_members, "ra_relation": ra_relation,
            "rmr_cells": rmr_cells, "rmr_labels": rmr_labels}


def _hunt(note: list[str], stage: dict) -> tuple[str, dict]:
    """Run the live two-boundary hunt, append its record to `note`, return
    (verdict, evidence). `stage["at"]` tracks the current boundary so a raise can be
    attributed. Any raise is handled by main() -> UNKNOWN-PROBE-FAILED.
    """
    # ============ boundary A: the CKAN catalogue (searched population #2) ============
    stage["at"] = "ckan"
    ckan = _sweep_ckan(note)

    # ============ boundary B1: the ISQ sitemap (searched population #1) ==============
    stage["at"] = "isq-sitemap"
    xml = _sitemap()
    raw_locs = _locs(xml)
    locs = sorted(set(raw_locs))
    n_repeated = len(raw_locs) - len(locs)
    _guard_sitemap(xml, locs)  # RAISES (isq-sitemap)

    xlsx = [u for u in locs if u.lower().endswith(".xlsx")]
    # DEDUPED BY FILE, not by url: the sitemap publishes `/fr/fichier/<slug>` and
    # `/en/fichier/<slug>` as separate locs pointing at the SAME workbook. Counting both
    # would double every candidate figure below and print visibly duplicated table rows,
    # inflating the population the verdict is scoped over with files that do not exist.
    # The surviving url prefers PREFERRED_LANG_PATH — a DECLARED preference, not an
    # arbitrary sort artifact: the ISQ workbooks demoflow already pins are `/fr/` urls, so a
    # v1 extension pinning the url this note publishes stays on the repo's own convention.
    by_slug: dict[str, str] = {}
    for url in sorted(xlsx):
        if _terms(url, MRC_TERMS) and _terms(url, POPULATION_TERMS):
            slug = _slug(url)
            if slug not in by_slug or (PREFERRED_LANG_PATH in url
                                       and PREFERRED_LANG_PATH not in by_slug[slug]):
                by_slug[slug] = url
    swept = sorted(by_slug.values(), key=_slug)
    n_swept_urls = sum(1 for u in xlsx if _terms(u, MRC_TERMS) and _terms(u, POPULATION_TERMS))
    eligible = [u for u in swept if _terms(u, PROJECTION_TERMS)]

    f_locs = Fact.derived(str(len(locs)), "distinct <loc> entries in the live ISQ sitemap.xml")
    f_xlsx = Fact.derived(str(len(xlsx)), "of those locs that are .xlsx download urls")
    f_swept = Fact.derived(str(len(swept)), "swept: distinct .xlsx FILES (deduped by slug "
                                            "across language paths) naming an MRC term AND a "
                                            "population term")
    f_elig = Fact.derived(str(len(eligible)), "of the swept that ALSO name a projection term "
                                              "(pick-eligible)")

    note += [
        "## 2. Boundary B — ISQ's own product pages / full-edition downloads",
        "",
        f"- `{ISQ_SITEMAP}` -> **{len(raw_locs)}** `<loc>` entries, **{f_locs}** of them "
        f"distinct ({n_repeated} duplicate entr{'y' if n_repeated == 1 else 'ies'} removed). "
        f"That is the MEASURED effect of deduping, stated instead of a mechanism: this run "
        f"parses `<loc>` elements only, and makes no claim about why any url repeats. Of the "
        f"distinct urls, **{f_xlsx}** are `.xlsx` download urls.",
        f"- Sweep predicate, stated in full so the swept population is exactly what it says: "
        f"a url is SWEPT iff it ends `.xlsx` AND its slug carries (case-insensitive "
        f"substring) an MRC term "
        f"{list(MRC_TERMS)} AND a population term {list(POPULATION_TERMS)}. "
        f"A projection term {list(PROJECTION_TERMS)} additionally makes it ELIGIBLE (spec §8's "
        f"junction consumes projected population by scenario, so an estimates workbook is a "
        f"different product). **{f_swept}** swept, **{f_elig}** eligible — every absence claim "
        f"is therefore scoped to the WIDER swept set.",
        f"- The swept count is DEDUPED BY FILE: {n_swept_urls} matching locs collapse to "
        f"{len(swept)} distinct slugs, because the sitemap lists `/fr/fichier/<slug>` and "
        f"`/en/fichier/<slug>` separately for the same workbook.",
        "",
    ]

    # ============ boundary B2: observe every eligible candidate =====================
    stage["at"] = "isq-file"
    observed: dict[str, dict] = {u: _probe_url(u) for u in eligible}
    verified = sorted(u for u in eligible if _is_workbook_response(observed[u]))

    note += [
        f"**Every eligible candidate, observed live by GET** (status, content-type, declared "
        f"length and the first {PREFIX_BYTES} bytes — only that prefix is read and the "
        f"connection is then closed, so a candidate costs a handshake rather than its full "
        f"body whatever its size; the declared lengths are in the table). A bare 200 is NOT "
        f"treated as evidence: the magic-byte column is what separates a workbook from an "
        f"HTML page served at 200.",
        "",
        "",
        "| candidate (slug) | HTTP | content-type | Content-Length | magic bytes | "
        "workbook prefix? |",
        "|---|---:|---|---:|---|---|",
    ]
    for url in sorted(eligible):
        o = observed[url]
        note.append(
            f"| `{_slug(url)}` | {o['status']} | {o['ctype'] or '?'} | {o['length'] or '?'} "
            f"| `{o['prefix'].hex() or 'none'}` "
            f"| {'YES' if o['prefix'].startswith(XLSX_MAGIC) else 'no'} |"
        )
    note += [
        "",
        f"- **{len(verified)}** of the {len(eligible)} eligible candidates answered 200 with a "
        f"workbook magic-byte prefix. Note the exact scope of that number: it is a "
        f"STATUS-AND-PREFIX result, not a body-shape result"
        # Conditioned on `verified`: §3 exists only when a candidate survives to be opened.
        # Written unconditionally, this sentence promised a section that a NOT-FOUND or an
        # UNKNOWN run never writes — a forward reference to evidence that is not in the file.
        + (" — exactly ONE candidate is opened and shape-checked in §3, and only that one "
           "carries the three evidence pieces a LOCATED requires."
           if verified else
           ". NO candidate survived this screen, so this run opens NOTHING and there is no "
           "body-shape section below: no resource here carries the three evidence pieces a "
           "LOCATED requires."),
        "",
        f"**The plan body's two GUESSED slugs, probed live by this run** (so the comparison in "
        f"§4 is measured here rather than recalled):",
        "",
        "| guessed slug | HTTP (GET) |",
        "|---|---:|",
    ]
    guessed = {u: _probe_url(u, nbytes=0) for u in PLAN_GUESSED_SLUGS}
    for url in PLAN_GUESSED_SLUGS:
        note.append(f"| `{_slug(url)}` | {guessed[url]['status']} |")
    note.append("")

    if not verified:
        # An earned NOT-FOUND: both populations were swept and non-empty, and the ISQ sweep
        # resolved NO eligible candidate at all. The guard refuses every other shape —
        # notably candidates that exist but do not serve a workbook body.
        _guard_not_found(locs, eligible, verified, ckan)  # RAISES (isq-file) if unearned
        note += [
            f"- **The sweep resolved no eligible MRC-projection candidate.** Scoped exactly: "
            f"not among the {len(locs)} sitemap locs ({len(swept)} swept, {len(eligible)} "
            f"eligible) and the {ckan['n_swept']} CKAN packages swept from a "
            f"{ckan['n_catalogue']}-package catalogue. This is NOT a claim that no MRC-level "
            f"ISQ source exists.",
            "",
        ]
        return "NOT-FOUND", {
            "ckan": ckan, "n_locs": len(locs), "n_xlsx": len(xlsx), "n_swept": len(swept),
            "n_eligible": len(eligible), "n_verified": 0, "eligible": sorted(eligible),
            "guessed": {u: guessed[u]["status"] for u in PLAN_GUESSED_SLUGS},
        }

    wb = _section_3_body_shape(note, observed, verified, eligible)
    ed = _section_3b_editions(note, verified, wb)
    mem = _section_3c_membership(note, wb)

    picked = wb["picked"]
    evidence = {
        # §3b / §3c residual state
        "editions": ed["editions"], "n_editions_opened": len(ed["opened"]),
        "n_ra_separate": len(ed["ra_separate"]),
        "n_ra_named_only": len(ed["ra_named_only"]),
        "n_ra_absent": len(ed["ra_absent"]),
        "ra_members": mem["ra_members"], "ra_relation": mem["ra_relation"],
        "n_rmr_cells": len(mem["rmr_cells"]), "n_rmr_labels": len(mem["rmr_labels"]),
        # sweep state
        "ckan": ckan,
        "n_locs": len(locs), "n_xlsx": len(xlsx), "n_swept": len(swept),
        "n_eligible": len(eligible), "n_verified": len(verified),
        "eligible": sorted(eligible), "verified": verified,
        "guessed": {u: guessed[u]["status"] for u in PLAN_GUESSED_SLUGS},
        # the opened workbook (§3)
        "url": picked, "status": observed[picked]["status"],
        "head_status": wb["head"]["status"], "head_disagrees": wb["head_disagrees"],
        "ctype": observed[picked]["ctype"], "length": observed[picked]["length"],
        "nbytes": len(wb["data"]), "sheets": wb["sheets"],
        "header_row": wb["header_row"], "header_col": wb["header_col"],
        "labels": wb["labels"], "n_labels": len(wb["labels"]),
        "n_aggregate": len(wb["aggregate_labels"]), "n_fine": len(wb["fine_labels"]),
        "scenarios": wb["scenarios"], "scen_col": wb["scen_col"],
        "scen_head": wb["scen_head"], "caption": wb["caption"],
        "diffusion": wb["diffusion"], "family": wb["picked_family"],
        "hits": wb["hits"], "misses": wb["misses"], "n_targets": wb["n_targets"],
        "couronne_complete": wb["couronne_complete"],
        "ra_col": wb["ra_col"], "ra_head": wb["ra_head"], "ra_usable": wb["ra_usable"],
        "ra_observed": wb["ra_observed"],
        "n_ra_checked": len(wb["ra_checked"]), "ra_disagree": wb["ra_disagree"],
    }
    return "LOCATED", evidence


def _guessed_str(evidence: dict) -> str:
    """The plan's guessed slugs and the status each ANSWERED, for the §4 comparison.

    Module-level, not inline in `main()`: `test_probe_contracts.py` locates the note
    assembly by SHAPE (the first `<sep>.join(...)` in `main()`), so a second join anywhere
    above it hides the real one from the contract gate. Keeping helper joins out of `main()`
    keeps that gate pointed at the assembly it exists to check.
    """
    return ", ".join(
        f"`{_slug(url)}` -> HTTP {status}"
        for url, status in (evidence.get("guessed") or {}).items()
    ) or "not probed this run"


def _premise_token(state: str) -> str:
    """The DECISION clause for a spec-premise STATE. One table, so the LOCATED branch and any
    other consumer cannot drift apart — and a state outside the four raises here rather than
    printing an unmapped verdict into a generated note."""
    return {
        "PREMISE STANDS": "CONTRADICTED — ESCALATION",
        "AMENDED": "ALREADY AMENDED — no live conflict remains",
        "INDETERMINATE": "INDETERMINATE — the spec row did not match either marker",
        "NOT MEASURED THIS RUN": "NOT CHECKED — the spec file was not read this run",
    }[state]


def _decision_located(e: dict, spec: dict, guessed_str: str) -> list[str]:
    """The DECISION block for a LOCATED verdict.

    Extracted from `main()` for the same reason `_hunt` was split: four of this review's
    findings (#2 #3 #4 #5) lived inside this one ~100-line run of nested conditional string
    concatenation, where a conditional gloss and an unconditional one look identical. As its
    own function it is readable end to end, and each clause is reachable from a fixture.

    Takes `spec` (the LIVE §4 read) and `guessed_str` rather than closing over them, so no
    clause here can quietly depend on state `main()` happens to have in scope.
    """
    body: list[str] = []
    ck = e["ckan"]
    premise_token = _premise_token(spec["state"])
    scope = (f"the {e['n_locs']} ISQ sitemap locs swept in §2 ({e['n_swept']} matched the "
             f"MRC×population predicate, {e['n_eligible']} eligible)")
    if ck["measured"]:
        scope += (f" and the {ck['n_swept']} CKAN packages swept in §1 from a "
                  f"{ck['n_catalogue']}-package catalogue")
    else:
        scope += " (the CKAN boundary was NOT MEASURED this run — see §1)"
    # Every clause below is a FUNCTION of a computed value, so a gloss cannot contradict
    # the number it sits beside.
    couronne = (f"{len(e['hits'])} of {e['n_targets']} declared couronne MRC targets found "
                f"by exact name search")
    couronne += (" — ALL declared targets present" if e["couronne_complete"]
                 else f" — MISSING: {e['misses']}")
    body += [
        "- `DECISION-VERDICT: LOCATED`",
        f"- `DECISION-RESOURCE-URL: {e['url']}`",
        f"- `DECISION-HTTP-STATUS: {e['status']} ({e['ctype'] or 'no content-type'}, "
        f"Content-Length {e['length'] or 'unreported'})`  (observed by GET this run; the "
        f"same url answered HTTP {e['head_status']} to HEAD"
        + (" — the methods DISAGREE here, so a HEAD-only hunt would miss this file)"
           if e["head_disagrees"] else " — the methods agree here)"),
        f"- `DECISION-BODY-SHAPE: {e['nbytes']} bytes downloaded, magic-byte prefix "
        f"matches {XLSX_MAGIC.hex()}, opened to sheets {e['sheets']}, MRC header cell at "
        f"row {e['header_row']} column {e['header_col']}, {e['n_labels']} distinct "
        f"geography labels, "
        + (f"{len(e['scenarios'])} scenario labels {e['scenarios']}"
           if e["scen_col"] >= 0 else "no scenario column in this sheet")
        + "`",
        # NO backticks inside the token value: `_probe_asserts.token` parses the
        # backtick-delimited span non-greedily, so an inner backtick truncates the value
        # and the gate would read a decomposition as a bare number.
        f"- `DECISION-MRC-LABEL-COUNT: {e['n_labels']} ({e['n_aggregate']} of them in the "
        f"NN-Name administrative-region-subtotal form + {e['n_fine']} others)`  "
        f"(distinct labels below the MRC-named header of the opened workbook. Precisely: "
        f"the header NAMES an MRC axis and this count says that column is populated. It "
        f"is NOT an MRC count — the column interleaves RA subtotals, hence the split; the "
        f"full label list is emitted verbatim in §3. And it says NOTHING about the "
        f"couronne — a large label set is equally consistent with the couronne being "
        f"absent — which is why the per-target search below is a separate token)",
        f"- `DECISION-COURONNE-TARGETS: {couronne}`",
        f"- `DECISION-RA-CORRESPONDENCE: "
        # Branches on `ra_usable` — the SHARED predicate — never on `ra_col >= 0`.
        # With the loose test this token read "0 of 0 declared targets corroborated
        # against the opened workbook's own RA column (column 1, header 'MRC par région
        # administrative')" on the header-named-only fixture: claiming an RA column at the
        # GEOGRAPHY column, i.e. exactly the machine-readable axis the three-state split
        # exists to deny, and contradicting §3's own prose in the same note.
        + (f"{e['n_ra_checked'] - len(e['ra_disagree'])} of {e['n_ra_checked']} declared "
           f"targets corroborated against the opened workbook's own SEPARATE RA column "
           f"(column {e['ra_col']}, header {e['ra_head']!r})"
           + (f"; DISAGREEING: {e['ra_disagree']}" if e["ra_disagree"] else "")
           if e["ra_usable"] else
           f"NOT CHECKABLE — the opened workbook publishes no SEPARATE administrative-"
           f"region column"
           + (f" (its geography header {e['ra_head']!r} NAMES an RA grouping but carries "
              f"no per-MRC code)" if e["ra_col"] >= 0 else ""))
        # MEMBERSHIP is stated as a FUNCTION of what §3 measured. This gloss used to
        # assert "each declared target is present and carries the RA code" flat out —
        # and an offline fixture published it beside `2 of 10` and `NOT CHECKABLE`, with
        # the whole suite green. A gloss beside a conditional value must be conditional.
        + "`  (MEMBERSHIP only, and only what was measured: "
        + (f"all {e['n_targets']} declared targets present"
           if e["couronne_complete"] else
           f"{len(e['hits'])} of {e['n_targets']} declared targets present")
        + (f", each carrying the RA code this file declares for it"
           if e["ra_usable"] and not e["ra_disagree"] and e["n_ra_checked"] else
           f", and the RA code was NOT CHECKABLE for {e['n_targets'] - e['n_ra_checked']} "
           f"of them" if e["n_ra_checked"] < e["n_targets"] else
           f", with {len(e['ra_disagree'])} RA disagreement(s)")
        # These clauses sit AFTER the closing backtick, so `token()` never returns them
        # and the residual-II gate cannot see them — which is exactly why they must be
        # conditional at the source rather than merely gated. Flat, they asserted that
        # §3c "measured" the composition "unanswerable" on runs where §3c computed NOT
        # COMPUTABLE (middle-state fixture) or found a metropolitan marker (RMR fixture).
        + ". EXHAUSTION is a different question, computed separately in §3c: "
        + ("; ".join(f"{k} -> {_relation_head(v)}" for k, v in e["ra_relation"].items()))
        + ". The RMR-couronne composition is "
        + ("measured there to be unanswerable from this workbook"
           if not e["n_rmr_cells"] and not e["n_rmr_labels"] else
           f"NOT settled there: {e['n_rmr_cells']} header cell(s) and "
           f"{e['n_rmr_labels']} label(s) DO match a metropolitan-area marker")
        + ". See DECISION-RESIDUAL-II below)",
        f"- `DECISION-SWEPT-POPULATION: {scope}`",
        f"- `DECISION-SPEC-PREMISE: {premise_token}`  (state read LIVE from the spec in "
        f"§4 — {spec['why'] or 'not measured'}. MEASURED THIS RUN: the plan's guessed "
        f"slugs {guessed_str}, while the resource above answers {e['status']} with a body "
        f"this run opened and shape-checked. The two together say the 404 was a property "
        f"of the GUESSED SLUG CONVENTION, not of the data. This note never edits the "
        f"spec.)",
        # --- the two RESIDUALS, recorded per steering ruling G. They are OBSERVATIONS:
        # neither appears in any verdict expression, and §11.6 stands either way.
        f"- `DECISION-RESIDUAL-I-RA-AXIS: of {e['n_editions_opened']} candidate workbooks "
        f"opened, {e['n_ra_separate']} publish a SEPARATE administrative-region column, "
        f"{e['n_ra_named_only']} name the grouping in the geography header ONLY (no "
        f"per-MRC RA code), {e['n_ra_absent']} carry neither"
        + ("; the axis is EDITION-SPECIFIC, so a v1 pinning an edition from the latter "
           "two groups must source it from a different workbook — an unvalidated "
           "cross-edition join"
           if e["n_ra_separate"] and (e["n_ra_named_only"] or e["n_ra_absent"]) else
           "; no edition-specificity found across the candidates opened")
        + "`  (RECORDED OBSERVATION — see §3b; changes no verdict. This run does not rank "
        "the editions by recency: no live response states which is current)",
        # `membership` is COUNTED, never the literal "YES": the same offline fixture that
        # falsified the gloss above published `membership YES` beside `2 of 10`.
        f"- `DECISION-RESIDUAL-II-PARTITION: membership {len(e['hits'])} of "
        f"{e['n_targets']} (§3); exhaustion "
        + "; ".join(f"{k} -> {_relation_head(v)}" for k, v in e["ra_relation"].items())
        # The conclusion FOLLOWS the counts inside its own parenthesis. Written flat, an
        # RMR-marker fixture published "NOT ANSWERABLE ... (1 header cells ... match)" —
        # detectably dishonest (a gate caught it), but the artifact must be written
        # honest, not merely written so that a test can catch the lie.
        + "; RMR-couronne composition "
        + (f"NOT ANSWERABLE from what this run read of this workbook "
           f"({e['n_rmr_cells']} header cells and {e['n_rmr_labels']} geography labels "
           f"match a metropolitan-area marker)"
           if not e["n_rmr_cells"] and not e["n_rmr_labels"] else
           f"UNSETTLED — {e['n_rmr_cells']} header cell(s) and {e['n_rmr_labels']} "
           f"geography label(s) DO match a metropolitan-area marker; inspect them before "
           f"treating the question as unanswerable from this file")
        + "`  (RECORDED OBSERVATION — see §3c; changes no "
        "verdict. No second source was consulted to close it, per steering ruling G)",
        "",
        "- **Standing rule (spec §11.6): v0 PROCEEDS REGARDLESS.** A find enables a **v1** "
        "`Geography` enum extension for couronne-nord precision — never a v0 change. In v0 "
        "the RA14/15/16 rows keep their `ra_proxy` flag (spec §8): they remain ranking "
        "members, never balance participants, never emitted in `ScenarioPrior`. Nothing in "
        "this note licenses a v0 loader change.",
        "",
    ]
    return body


def main() -> None:
    # Per-run registry (see run_p3.py): `_wds` is one cached module shared by every probe,
    # so a module-global list would let one run's figures inflate another's.
    facts = new_run()
    title = ["# P6 — MRC-level ISQ source hunt (RECORDED OBSERVATION)", ""]
    body: list[str] = []
    stage = {"at": "ckan"}

    verdict = "UNKNOWN-PROBE-FAILED"
    evidence: dict = {}
    try:
        verdict, evidence = _hunt(body, stage)
    except Exception as exc:  # outage, a vacuous 200, or a wrong-body page — never a false verdict
        boundary = stage.get("at", "ckan")
        host = "donneesquebec.ca" if boundary == "ckan" else ISQ_HOST
        body += [
            "## LIVE HUNT VERDICT: UNKNOWN-PROBE-FAILED",
            "",
            f"- `LIVE PROBE FAILED-AT: {boundary}` (host: {host})",
            f"- `LIVE PROBE FAILED: {type(exc).__name__}: {exc}`",
            "",
            "  The hunt did not answer usefully at the boundary named above — an outage, or a "
            "200 carrying an empty/wrong body, or a searched population too thin to earn an "
            "absence (all caught by the floor guard). This run therefore records neither a "
            "located source NOR a not-found: an unearned NOT-FOUND would be the cheap all-clear "
            "this gate exists to refuse. Re-run against live sources to record the hunt.",
            "",
        ]

    # NON-GATING and runs on EVERY branch, so the DECISION block's premise statement is a
    # function of the spec as it stands rather than of what this run happened to find.
    spec = _record_spec_premise(body)

    body += ["## 5. DECISION", ""]
    guessed_str = _guessed_str(evidence)
    # The premise clause, a FUNCTION of the live read. A LOCATED contradicts the premise only
    # while the premise is actually in the spec; once amended, "CONTRADICTED" would be the
    # note asserting a conflict with text that no longer exists.
    if verdict == "LOCATED":
        body += _decision_located(evidence, spec, guessed_str)
    elif verdict == "NOT-FOUND":
        e = evidence
        ck = e["ckan"]
        body += [
            "- `DECISION-VERDICT: NOT-FOUND`",
            f"- `DECISION-SWEPT-POPULATION: the {e['n_locs']} ISQ sitemap locs swept in §2 "
            f"({e['n_swept']} matched the MRC×population predicate, {e['n_eligible']} eligible, "
            f"{e['n_verified']} verified) and the {ck['n_swept']} CKAN packages swept in §1 "
            f"from a {ck['n_catalogue']}-package catalogue`",
            "- `DECISION-NOT-FOUND-SCOPE: an absence AMONG THE POPULATIONS NAMED ABOVE — this "
            "run does NOT claim that no MRC-level ISQ source exists`",
            "- `DECISION-RESIDUAL-I-RA-AXIS: NOT MEASURED THIS RUN`  (no candidate workbook "
            "was opened, so the RA axis was not read in any edition)",
            "- `DECISION-RESIDUAL-II-PARTITION: NOT MEASURED THIS RUN`  (same reason)",
            # Two DIFFERENT facts, kept apart: what THIS RUN did (located nothing, so it
            # contradicts nothing) and what the spec CURRENTLY says (read live in §4). Fusing
            # them would let a run report the spec's state as its own finding.
            f"- `DECISION-SPEC-PREMISE: NOT CONTRADICTED BY THIS RUN (spec state read live: "
            f"{spec['state']})`  (this sweep located nothing, so it challenges no premise; the "
            f"plan's guessed slugs answered {guessed_str})",
            "",
            "- **Standing rule (spec §11.6): v0 PROCEEDS REGARDLESS.** A find enables a **v1** "
            "`Geography` enum extension for couronne-nord precision — never a v0 change. In v0 "
            "the RA14/15/16 rows keep their `ra_proxy` flag (spec §8): they remain ranking "
            "members, never balance participants, never emitted in `ScenarioPrior`. Nothing in "
            "this note licenses a v0 loader change.",
            "",
        ]
    else:
        body += [
            f"- `DECISION-VERDICT: {verdict}`",
            "- No MRC-level source was located this run AND no absence was earned (see the "
            "boundary failure above). The searched population was too thin, or a boundary "
            "answered with a body that cannot support a verdict — either way this is a "
            "recorded observation, not an invented find and not a hollow not-found.",
            f"- `DECISION-SPEC-PREMISE: NOT CONTRADICTED BY THIS RUN (spec state read live: "
            f"{spec['state']})`  (a run that recorded no finding challenges no premise)",
            "- `DECISION-RESIDUAL-I-RA-AXIS: NOT MEASURED THIS RUN`  (the hunt failed before "
            "any candidate workbook was opened — see the boundary failure above)",
            "- `DECISION-RESIDUAL-II-PARTITION: NOT MEASURED THIS RUN`  (same reason)",
            "",
            "- **Standing rule (spec §11.6): v0 PROCEEDS REGARDLESS.** A find enables a **v1** "
            "`Geography` enum extension for couronne-nord precision — never a v0 change. In v0 "
            "the RA14/15/16 rows keep their `ra_proxy` flag (spec §8): they remain ranking "
            "members, never balance participants, never emitted in `ScenarioPrior`. Nothing in "
            "this note licenses a v0 loader change.",
            "",
        ]

    header = provenance_header(facts, written_by=_WRITTEN_BY, scope=_SCOPE,
                               summary=_summary, cited_label=_CITED_LABEL)
    text = "\n".join(title + header + body) + "\n"
    for placeholder in ("[FILL:", "[FILL]", "[FILL "):
        if placeholder in text:
            raise AssertionError(f"run_p6.py emitted an unresolved {placeholder!r} placeholder")
    OUT.write_text(text, encoding="utf-8")
    print(text)


if __name__ == "__main__":
    main()
