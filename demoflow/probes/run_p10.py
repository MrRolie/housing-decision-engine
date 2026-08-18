"""P10 — HORS_RMR OPERAND ALIGNMENT: the aligned immigrant inputs (spec §6 amendment #12(A)).

Writes `probes/P10-hors-operand-alignment.md`. This module is that file's only writer.

THE DEFECT. Ruling S measures HORS_RMR's immigrant headship and ownership ratio over a census
residual — the Québec province NET of the six wholly-QC CMAs — and that residual INCLUDES the
Québec side of Ottawa-Gatineau, because 98-10-0621-01 carries exactly one Ottawa-Gatineau
member and parents it to ONTARIO, so it is not among the six geoLevel-503 children of Quebec.
The arrival flows those rates multiply come from ISQ's own hors-RMR row, which EXCLUDES that
territory: ISQ publishes `RMR d'Ottawa-Gatineau` as its own row, footnoted *Partie québécoise
uniquement*. The territory is SEPARABLE on the ISQ side and INSEPARABLE on the census side, so
a rate measured on one was multiplying a flow defined on the other. Amendment #12(A) rules the
principle: **the rate's territory must match the flow's territory.**

WHAT THIS RUN DOES ABOUT IT. It does not bracket the contaminant with whole census divisions —
the Ottawa-Gatineau CMA is defined at CSD level and no whole-CD union equals it. It resolves
the CMA's Québec-part CSD membership from a live StatCan source (98-10-0003-01 publishes the
CMA's constituent census subdivisions as its geography-dimension CHILDREN) and subtracts those
CSDs' cells, read from 98-10-0622-01 — the ruled cube's own census-division/subdivision
sibling, one universe at two grains. The whole-CD bracket is published beside it as the
sensitivity, and it does NOT enclose the exact headship: that is the measured cost of the
construction this run did not use.

    aligned(g) = province − Σ six wholly-QC CMAs − Σ Ottawa-Gatineau Québec-part CSDs
    headship   = maintainers ÷ persons on the ruled `Before 2016` member
    ratio      = settled owner-maintainer propensity ÷ the non-immigrant one

WHAT IT DOES NOT DO. It wires NOTHING. `demand/immigrant_inputs.py`, `probes/run_p8.py` and
`probes/P8-immigrant-inputs.md` are untouched by this run, and must stay so until the spec
carries the aligned values: P8's note is citation-coupled to §6's stated figures, so wiring
first would couple that note to numbers the ruling no longer carries. The ordering is
deliberate — spec ruling, then P8 regeneration, then the join table.

THE OWNERSHIP LEG (#12(B)), measured and NOT corrected. The same contamination sits in the
ownership propensity ρ, and #12(B) rules it must not be "fixed" because a BAND-UNIFORM relative
scaling of ρ cancels exactly in ED. This run measures ρ's contamination in the MODEL'S OWN
lattice — 98-10-0231-01 (the cube `loaders/census.py` reads) and its CSD sibling
98-10-0232-01, on `census._AGE_BAND_SPEC`'s own four bands — rather than on a proxy. The
band-uniform premise is measured FALSE in its strict form and the residual it leaves is
measured second-order; both halves are published, because the premise is what the ruling
rests on and the size is what makes the ruling survive.

THE FLOOR GUARDS (each raises `ProbeRefusal`, which routes the run to UNKNOWN-PROBE-FAILED —
never to a friendlier verdict; the name is deliberately NOT network-class so a refusal cannot
launder itself into `pytest.skip`):

  * `_guard_meta`            — every cube answered SUCCESS with a populated dimension list.
  * `_member`                — every member resolves BY NAME to EXACTLY one candidate.
  * `_guard_pinned_id`       — the resolved id equals the id the mandate pinned.
  * `_guard_response`        — every requested cell came back at refPer 2021, EXCEPT cells in
                               the declared suppressible scope (CSD-grain immigrant members and
                               CSD-grain age bands, where StatCan withholds small counts). A
                               missing cell anywhere else refuses: suppression is a property of
                               tiny geographies, and accepting it cube-wide would let an outage
                               read as a publication rule.
  * `_guard_required_complete` — the two members the suppression BOUND is built from
                               (`Total - Age`, `Non-immigrants`) are published at every QC-part
                               CSD. A bound over an unpublished bound is not a bound.
  * `_guard_membership_closes` — the CMA's children sum EXACTLY to the CMA's own population, so
                               the QC/ON split is a PARTITION and not a selection.
  * `_guard_qc_split`        — SGC prefix and structural province ancestry agree on every child.
  * `_guard_membership_pop`  — the QC-part census population matches the ISQ row it aligns to
                               within a threshold DERIVED from innocent controls (the six
                               wholly-QC CMAs) measured in the same construction.
  * `_guard_universe`        — the province rows of each cube pair are bit-identical on the
                               quantities this run subtracts across them, ±5 elsewhere.
  * `_guard_isq_parts`       — each ISQ workbook's own parts close on its province row (the
                               population workbook exactly; the flow workbook within its
                               published rounding step, measured rather than assumed).
  * `_isq_code`              — every ISQ row is resolved BY LABEL and its pinned code asserted.
  * `_guard_footnote`        — the Gatineau row's footnote marker and the footnote line the
                               note QUOTES are both read live; a quotation this run cannot read
                               is not printed.
  * `_guard_no_direct_source` — none of the four maintainer-cross cubes P9's closure names
                               publishes the Québec part directly. If one did, this
                               construction would be an approximation with a published exact
                               beside it, and the run refuses rather than approximating.
  * `_guard_ratio_band`      — the suppression bound may not straddle 1.0, or the crossing this
                               note reports is not earned.
  * `_guard_citation`        — every figure this run REPRODUCES from §6 and from the Task-26
                               QFE record matches those documents' own digits.

Run:  cd demoflow && uv run python probes/run_p10.py
"""
import re
from dataclasses import dataclass
from pathlib import Path

# Flat, NOT `probes._wds`: probes/ is deliberately not a package, so in script mode
# sys.path[0] IS probes/ and this resolves natively. See probes/_wds.py.
from _wds import (WDS_DATA, WDS_META, Fact, new_run, post, provenance_header, table_number,
                  table_url)

from demoflow.loaders.pins import DATA_DIR, verify_pin

_WRITTEN_BY = Path(__file__).name
OUT = Path(__file__).resolve().parent / "P10-hors-operand-alignment.md"
_TITLE = "# P10 — HORS_RMR operand alignment: the ALIGNED immigrant inputs (amendment #12(A))"

_REPO_ROOT = Path(__file__).resolve().parents[2]
SPEC = (_REPO_ROOT / "docs" / "specs"
        / "2026-07-21-demoflow-demographic-scenario-module-design.md")
QFE = (_REPO_ROOT / "docs" / "audits" / "dispatch" / "2026-08-15-qfe-retriage-task-26.md")
P9_NOTE = Path(__file__).resolve().parent / "P9-catalogue-closure.md"
CONSTANTS = _REPO_ROOT / "demoflow" / "src" / "demoflow" / "loaders" / "constants.py"

# --- the cubes ----------------------------------------------------------------------------
CMA_PID = 98100621          # ruling S's source: the immigrant inputs at CMA grain
CD_PID = 98100622           # its CD/CSD sibling: the QC-part CSDs, one universe at two grains
MEMBER_PID = 98100003       # the CMA -> constituent-CSD membership, and its population
RHO_CMA_PID = 98100231      # the cube loaders/census.py reads for ownership by maintainer age
RHO_CD_PID = 98100232       # its CD/CSD sibling — the aligned ρ curve's other half
SHELTER_CMA_PID = 98100623  # the other two maintainer-cross cubes P9's closure names:
SHELTER_CD_PID = 98100624   #   scanned for a Québec-part member, never read for data
PARTS_PID = 43100060        # the ONE cube publishing the Québec part — with the wrong metric

MAINTAINER_CROSS = (CMA_PID, CD_PID, SHELTER_CMA_PID, SHELTER_CD_PID)

# 98100621/98100622 share this dimension order, name for name (asserted live).
POS_GEO, POS_TENURE, POS_GENDER, POS_MAINT, POS_STAT, POS_POPCHAR, POS_SUIT = range(1, 8)
TENURE_TOTAL = ("Total - Tenure including presence of mortgage payments and subsidized "
                "housing (totals include farm operators)")
TENURE_OWNER = "Owner"
GENDER_TOTAL = "Total - Gender"
MAINT_ALL_PERSONS = "Total - Household maintainer"
MAINT_PRIMARY = "Person is primary household maintainer"
STAT_PEOPLE = "Number of people"
SUIT_TOTAL = "Total - Housing suitability"

PC_TOTAL_AGE = "Total - Age"
PC_NONIMM = "Non-immigrants"
PC_SETTLED = "Before 2016"                        # the RULED member — untouched by #12(A)
POPCHARS = (PC_TOTAL_AGE, PC_NONIMM, PC_SETTLED)
PINNED_POPCHAR_IDS = {PC_TOTAL_AGE: 1, PC_NONIMM: 10, PC_SETTLED: 12}

QUEBEC = "Quebec"
ONTARIO = "Ontario"
PROVINCE_GEO_LEVEL = 2
CMA_GEO_LEVEL = 503
CMA_PART_GEO_LEVEL = 505     # the grain that would publish the Québec part directly
CD_GEO_LEVEL = 3
CSD_GEO_LEVEL = 5
PINNED_GEO_IDS = {(CMA_PID, QUEBEC): 24, (CD_PID, QUEBEC): 884}
PINNED_CMA_IDS = (30, 35, 36, 40, 48, 51)
# The census-side inseparability, pinned as the mandate states it: ONE Ottawa-Gatineau member,
# at CMA grain, parented to Ontario — which is exactly why it sits inside a province residual
# taken net of Quebec's own CMA children.
OG_CMA_NAME = "Ottawa - Gatineau (CMA), Ont./Que."
PINNED_OG_CMA_ID = 81
PINNED_OG_CMA_CODE = "505"
# 98-10-0003-01's own Ottawa-Gatineau CMA member, whose CHILDREN are the membership.
OG_MEMBER_NAME = "Ottawa - Gatineau"
PINNED_OG_MEMBER_ID = 594
MEMBER_COUNT_2021 = "Population, 2021"
# CD Gatineau, for the whole-CD bracket. Its CSD namesake one level down is the trap the
# mandate names: id 1944, geoLevel 5, code 2481017 — a name-only lookup takes the city.
BRACKET_CDS = (("2481", "Gatineau"), ("2482", "Les Collines-de-l'Outaouais"),
               ("2480", "Papineau"))
PINNED_CD_GATINEAU_ID = 1943
PINNED_CSD_GATINEAU_ID = 1944
# The Québec side of the CMA, pinned by SGC code. Recorded so a delineation change REDS; the
# membership is RESOLVED from the live parent-child hierarchy and this set is asserted against
# what came back, never used to select.
PINNED_QC_PART_CODES = ("2480050", "2480055", "2480060", "2480065", "2480085", "2480140",
                        "2480145", "2481017", "2482005", "2482010", "2482015", "2482020",
                        "2482025", "2482030", "2482035", "2483005")
QC_SGC_PREFIX = "24"

# --- the ρ cubes (the model's own ownership curve) ------------------------------------------
# Two DIFFERENT dimension layouts, and that is why each is described separately rather than
# shared: 98100231 carries a `Condominium status` dimension 98100232 does not, so a shared
# coordinate builder would silently shift every member id by one position.
RHO_CMA_POS = {"geo": 1, "struct": 2, "condo": 3, "hh": 4, "stat": 5, "age": 6, "tenure": 7}
RHO_CD_POS = {"geo": 1, "struct": 2, "hh": 3, "stat": 4, "age": 5, "tenure": 6}
RHO_STRUCT_TOTAL = "Total - Structural type of dwelling"
RHO_CONDO_TOTAL = "Total - Condominium status"
RHO_HH_TOTAL_CMA = "Total - Household type including census family structure"
RHO_HH_TOTAL_CD = "Total - Household type including family structure"
RHO_STAT = "Number of private households"
RHO_TENURE_TOTAL = "Total - Tenure"
RHO_TENURE_OWNER = "Owner"
RHO_AGE_TOTAL = "Total - Age of primary household maintainer"
# `census._AGE_BAND_SPEC`'s own bands, expressed in each cube's published age members. The
# band EDGES are the model's; the constituents are whatever each cube publishes, and the two
# cubes publish different granularities of the same partition.
RHO_BANDS_CMA = {
    "25-54": ("25 to 29 years", "30 to 34 years", "35 to 39 years", "40 to 44 years",
              "45 to 49 years", "50 to 54 years"),
    "55-64": ("55 to 59 years", "60 to 64 years"),
    "65-74": ("65 to 69 years", "70 to 74 years"),
    "75+": ("75 to 84 years", "85 years and over"),
}
RHO_BANDS_CD = {
    "25-54": ("25 to 34 years", "35 to 44 years", "45 to 54 years"),
    "55-64": ("55 to 64 years",),
    "65-74": ("65 to 74 years",),
    "75+": ("75 to 84 years", "85 years and over"),
}
RHO_BAND_ORDER = ("25-54", "55-64", "65-74", "75+")

# --- 43-10-0060-01: the one cube that publishes the part, and cannot serve ------------------
# Resolved by SGC code, not by name: this cube's geography labels carry trailing non-breaking
# spaces (P8 measured 32 of 63) and an EN DASH inside `Ottawa–Gatineau`, so a name match here
# would be a typography test wearing a territory's clothes.
PARTS_QC_PART_CODE = "24505"

# --- the ISQ side ---------------------------------------------------------------------------
ISQ_POP_BOOK = "pop-as-rmr-base.xlsx"      # the population row the membership is validated on
ISQ_FLOW_BOOK = "compo-rmr-base.xlsx"      # the OPERAND's own workbook — the flows themselves
ISQ_PROVINCE_LABEL = "Le Québec"
ISQ_GATINEAU_LABEL = "RMR d'Ottawa-Gatineau"
ISQ_HORS_POP_LABEL = "Territoire hors des RMR"
ISQ_HORS_FLOW_LABEL = "Hors RMR"           # the flow workbook's own, SHORTER label
ISQ_PROVINCE_CODE = 0
ISQ_GATINEAU_CODE = 505
ISQ_HORS_CODE = 999
ISQ_YEAR = 2021
ISQ_SEX_TOTAL = 3
ISQ_SCENARIOS = ("Référence (A2026)", "Faible (D2026)", "Fort (E2026)")
ISQ_HEADER_CELL = "Scénario"
ISQ_REFERENCE = "Référence (A2026)"
ISQ_NOTE_MARKER = "découpage"
ISQ_PART_FOOTNOTE_MARKER = "québécoise"
ISQ_FLOW_COLUMN = "Immigrants permanents"
# The compo sheet stacks a column's NAME across three header rows and puts its UNIT (`n`) on a
# fourth. Three is therefore the name's depth, not a guess: reading four would append the unit
# to every name and match nothing.
ISQ_HEADER_STACK = 3
# ISQ withholds the terminal year of the (t)→(t+1) flow table as '...'; a non-numeric cell is
# EXCLUDED from the closure check and the exclusion is reported, never silently coerced.
ISQ_WITHHELD = "..."
# The flow workbook's parts close on its province row only to a rounding step. Measured, and
# the tolerance is stated at the size the workbook's own arithmetic needs rather than at zero.
ISQ_FLOW_TOLERANCE = 10

REF_YEAR = "2021"
_DATA_CHUNK = 60

# The margin above the maximum innocent residual, in the shape P8's territory gate uses: the
# calibration is the measured maximum, this is only the cushion on top of it.
MARGIN_FRACTION = 0.25

# Everything the run measured, for tests/test_probe_p10.py to read directly rather than
# re-parsing prose. Rebound (never mutated in place) at the top of `_sections`.
LAST_RUN: dict = {}


class ProbeRefusal(RuntimeError):
    """A floor guard fired: this run may NOT publish a measurement.

    Deliberately not a network-class exception name — `tests/_probe_asserts.py`'s
    NETWORK_EXCEPTIONS decides fail-vs-skip, and a refusal that could be mistaken for an
    outage would launder the cardinal cheap all-clear into a green-looking skip.
    """

    def __init__(self, boundary: str, message: str) -> None:
        super().__init__(message)
        self.boundary = boundary


# ===========================================================================================
# Arithmetic — the definitions, as code
# ===========================================================================================
@dataclass(frozen=True)
class Cell:
    """One (geography, population-characteristic) reading: three published counts."""

    persons: int
    maintainers: int
    owner_maintainers: int

    @property
    def headship(self) -> float:
        if not self.persons:
            raise ProbeRefusal("arithmetic", f"headship over {self.persons} persons is not a "
                                             "rate — a 0.0 here would be a number where there "
                                             "is no measurement")
        return self.maintainers / self.persons

    @property
    def owner_propensity(self) -> float:
        if not self.maintainers:
            raise ProbeRefusal("arithmetic", f"owner propensity over {self.maintainers} "
                                             "maintainers is not a rate")
        return self.owner_maintainers / self.maintainers


def cell_minus(whole: Cell, part: Cell) -> Cell:
    return Cell(whole.persons - part.persons,
                whole.maintainers - part.maintainers,
                whole.owner_maintainers - part.owner_maintainers)


def cell_sum(cells) -> Cell:
    total = Cell(0, 0, 0)
    for cell in cells:
        total = Cell(total.persons + cell.persons,
                     total.maintainers + cell.maintainers,
                     total.owner_maintainers + cell.owner_maintainers)
    return total


def ownership_ratio(immigrant: Cell, non_immigrant: Cell) -> float:
    """§6's `ratio`: the owner-MAINTAINER propensity of one member over the other's."""
    return immigrant.owner_propensity / non_immigrant.owner_propensity


def complement_bound(total: Cell, non_immigrant: Cell) -> tuple:
    """An upper bound on ANY immigrant member's counts: everything that is not non-immigrant.

    Clamped at zero per field, and the clamp is reported rather than silent: both inputs are
    rounded to 5 independently, so a geography whose immigrant population is genuinely zero can
    publish a total one rounding step BELOW its non-immigrant count. A negative "bound" would
    make the upper end of an interval sit under its lower end.
    """
    fields = (total.persons - non_immigrant.persons,
              total.maintainers - non_immigrant.maintainers,
              total.owner_maintainers - non_immigrant.owner_maintainers)
    return Cell(*(max(f, 0) for f in fields)), sum(1 for f in fields if f < 0)


def relative_pct(new: float, old: float) -> float:
    """`new` against `old`, in percent OF `old` — a relative delta, never a difference.

    The quantity #12(B) is written about is a relative SCALING of ρ, because that is the one
    that cancels in ED. Reporting a percentage-point difference beside a sentence about
    scaling would be a different quantity wearing the same symbol.
    """
    if not old:
        raise ProbeRefusal("arithmetic", "a relative delta against zero is not a percentage")
    return (new / old - 1) * 100


def bounded_sum(rows: list) -> tuple:
    """Sum a set of geographies FIELD BY FIELD, carrying the withheld fields as an interval.

    `rows` is [(values, bounds)] per geography, both the same width, where a `None` in `values`
    marks a count StatCan withholds at that geography and the parallel entry in `bounds` is an
    upper bound on it drawn from a quantity the same cube DOES publish.

    FIELD-WISE is the whole point, and it is not a detail: a subdivision that publishes settled
    PERSONS while withholding settled MAINTAINERS is a real shape here, and dropping the whole
    geography would discard a published count — which does not merely widen the interval, it
    biases the point estimate, because the persons and the maintainers of one territory would
    then be netted out of different denominators.

    Returns (low, high, withheld_field_count): `low` counts only what is published, `high` adds
    every bound. The truth lies in the box between them, per field.
    """
    if not rows:
        raise ProbeRefusal("suppression", "a bounded sum over no geographies is not a sum")
    width = len(rows[0][0])
    low, high, withheld = [0] * width, [0] * width, 0
    for values, bounds in rows:
        for index, (value, bound) in enumerate(zip(values, bounds, strict=True)):
            if value is None:
                if bound is None:
                    raise ProbeRefusal(
                        "suppression",
                        f"field {index} is withheld and carries NO bound — an interval with an "
                        "unmeasured end is not an interval, and this run may not publish one.")
                high[index] += bound
                withheld += 1
            else:
                low[index] += value
                high[index] += value
    return tuple(low), tuple(high), withheld


def bounded_pair(shipped: Cell, low: Cell, high: Cell, non_immigrant: Cell) -> dict:
    """The aligned headship and ratio, with the ENVELOPE the withheld cells leave.

    The headline subtracts only what is published (`low`) — the one construction assembled
    entirely from counts the source states. The envelope is taken at the box's CORNERS, not by
    pairing the two sums: headship is largest when the fewest maintainers and the most persons
    are netted out, and those are opposite corners. Pairing low-with-low would report an
    interval narrower than the uncertainty actually is.
    """
    headline = cell_minus(shipped, low)
    h_high = (shipped.maintainers - low.maintainers) / (shipped.persons - high.persons)
    h_low = (shipped.maintainers - high.maintainers) / (shipped.persons - low.persons)
    p_high = ((shipped.owner_maintainers - low.owner_maintainers)
              / (shipped.maintainers - high.maintainers))
    p_low = ((shipped.owner_maintainers - high.owner_maintainers)
             / (shipped.maintainers - low.maintainers))
    base = non_immigrant.owner_propensity
    return {"cells": headline, "headship": headline.headship,
            "ratio": ownership_ratio(headline, non_immigrant),
            "headship_band": (h_low, h_high),
            "ratio_band": (p_low / base, p_high / base)}


def derive_threshold(innocent: dict) -> dict:
    """The membership gate's threshold, DERIVED from the innocent controls handed in."""
    if not innocent:
        raise ProbeRefusal("threshold", "no innocent controls were measured — a threshold over "
                                        "an empty calibration set has nothing behind it")
    name, value = max(innocent.items(), key=lambda kv: abs(kv[1]))
    largest = abs(value)
    if largest <= 0:
        raise ProbeRefusal("threshold", f"every innocent control measured exactly zero "
                                        f"({sorted(innocent)}) — a calibration set with no "
                                        "spread cannot bound anything")
    margin = largest * MARGIN_FRACTION
    return {"max_name": name, "max_abs": largest, "max_signed": value,
            "margin_pp": margin, "threshold": largest + margin}


def coord(*ids: int) -> str:
    """A WDS coordinate: the cube's dimension ids, zero-padded to ten positions."""
    return ".".join(str(i) for i in ids) + ".0" * (10 - len(ids))


# ===========================================================================================
# Boundaries — the injectable seams (so the whole green path runs OFFLINE in tests)
# ===========================================================================================
def _meta(pids) -> list:
    """POST `getCubeMetadata`. Boundary `wds-meta` (www150)."""
    return post(WDS_META, [{"productId": int(p)} for p in pids])


def _data(requests: list) -> list:
    """POST `getDataFromCubePidCoordAndLatestNPeriods`, chunked. Boundary `wds-data`.

    No try/except, no retry, no sentinel: a swallowed HTTP error would reach `_guard_response`
    as a missing cell and be reported as StatCan suppressing a count.
    """
    out: list = []
    for start in range(0, len(requests), _DATA_CHUNK):
        out += post(WDS_DATA, requests[start:start + _DATA_CHUNK])
    return out


def _isq_rows(name: str) -> list:
    """A pinned ISQ workbook's row grid. Boundary `isq-workbook` (a committed file)."""
    from openpyxl import load_workbook

    path = DATA_DIR / name
    verify_pin(path, name)
    book = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = book[book.sheetnames[0]]
        return list(sheet.iter_rows(values_only=True))
    finally:
        book.close()


def _spec_s6() -> str:
    """Spec §6 — the ruling text this note reproduces figures from. Boundary `spec`."""
    text = SPEC.read_text(encoding="utf-8")
    head, tail = "## 6. Demand side", "## 7. Outputs"
    if head not in text or tail not in text:
        raise ProbeRefusal("spec", f"{SPEC.name} no longer carries both section markers "
                                   f"{head!r} and {tail!r} — the citation gate cannot scope "
                                   "itself to the ruling text and must not guess")
    return text[text.index(head):text.index(tail)]


def _qfe_record() -> str:
    """The Task-26 QFE record — where the whole-CD bracket's other two variants are stated."""
    if not QFE.exists():
        raise ProbeRefusal("qfe", f"{QFE.name} is absent; this note may not reproduce bracket "
                                  "figures it cannot read at their source")
    return QFE.read_text(encoding="utf-8")


def _constants_source() -> str:
    """`loaders/constants.py` — where this tree records the private-household universe gap."""
    return CONSTANTS.read_text(encoding="utf-8")


def _p9_note() -> str:
    """P9's committed note — the catalogue closure this one carries BY REFERENCE."""
    if not P9_NOTE.exists():
        raise ProbeRefusal("p9", f"{P9_NOTE.name} is absent; this note may not restate a "
                                 "catalogue closure it cannot read at the level P9 earned it")
    return P9_NOTE.read_text(encoding="utf-8")


# ===========================================================================================
# Guards
# ===========================================================================================
def _guard_meta(pids, response) -> dict:
    if not isinstance(response, list) or len(response) != len(pids):
        raise ProbeRefusal("wds-meta", f"getCubeMetadata returned "
                                       f"{len(response) if isinstance(response, list) else type(response).__name__}"
                                       f" for {len(pids)} requested cube(s)")
    metas = {}
    for item in response:
        if item.get("status") != "SUCCESS":
            raise ProbeRefusal("wds-meta", f"getCubeMetadata status {item.get('status')!r}")
        obj = item["object"]
        if not (obj.get("dimension") or []):
            raise ProbeRefusal("wds-meta", f"{obj.get('productId')} answered with an EMPTY "
                                           "dimension list — every member this run resolves "
                                           "would be looked up in nothing")
        metas[int(obj["productId"])] = obj
    missing = [p for p in pids if int(p) not in metas]
    if missing:
        raise ProbeRefusal("wds-meta", f"getCubeMetadata answered for {sorted(metas)}, "
                                       f"missing {missing}")
    return metas


def _dimension(metas: dict, pid: int, position: int) -> dict:
    for dim in metas[pid]["dimension"]:
        if dim.get("dimensionPositionId") == position:
            return dim
    raise ProbeRefusal("wds-meta", f"{table_number(pid)} has no dimension at position "
                                   f"{position}")


def _member(metas: dict, pid: int, position: int, name: str, *,
            geo_level: int | None = None, parent: int | None = None) -> dict:
    """Resolve a member BY NAME, refusing on zero OR MORE THAN ONE match.

    Ambiguity is a refusal rather than a first-hit: `Gatineau` names both the census division
    (geoLevel 3) and the city inside it (geoLevel 5), and picking whichever comes first would
    publish a city's numbers under a region's label.
    """
    dim = _dimension(metas, pid, position)
    hits = [m for m in dim.get("member") or []
            if m.get("memberNameEn") == name
            and (geo_level is None or m.get("geoLevel") == geo_level)
            and (parent is None or m.get("parentMemberId") == parent)]
    if len(hits) != 1:
        raise ProbeRefusal(
            "wds-meta",
            f"{table_number(pid)} dimension {position}: {len(hits)} member(s) named {name!r} "
            f"at geoLevel {geo_level} under parent {parent} (of "
            f"{len(dim.get('member') or [])} members). Exactly one must match — zero means the "
            f"cube moved, more than one means the name alone does not identify a territory.")
    return hits[0]


def _member_by_code(pid: int, by_code: dict, code: str, level: int, label: str) -> dict:
    """Resolve a geography member by its SGC classification code AND geoLevel.

    The code is what joins the membership cube's child to the cube the counts are read from,
    and the geoLevel is what keeps a census subdivision from answering for the division that
    shares its name and its code prefix. Ambiguity refuses, exactly as `_member` does.
    """
    hits = by_code.get((code, level)) or []
    if len(hits) != 1:
        raise ProbeRefusal(
            "wds-meta",
            f"{table_number(pid)}: {len(hits)} member(s) with SGC {code} at geoLevel {level} "
            f"({label}). Exactly one must match — the code is the join between the membership "
            "and the counts, and a name would not survive an accented rename.")
    return hits[0]


def _guard_pinned_id(pid: int, name: str, member: dict, pinned: int) -> dict:
    """The resolved member must be the one the mandate pinned. A move is a FINDING."""
    if member["memberId"] != pinned:
        raise ProbeRefusal(
            "wds-meta",
            f"{table_number(pid)}: {name!r} resolves to memberId {member['memberId']}, but the "
            f"mandate pinned {pinned}. That is a finding about the cube, not a typo to paper "
            "over — every value keyed on the pinned id would be a different territory's.")
    return member


def _guard_response(requests: list, response, suppressible: set) -> tuple:
    """Key every returned cell by (productId, coordinate); refuse on any UNEXPECTED absence.

    Keying by POSITION is the defect this refuses: WDS returns its objects sorted by coordinate
    STRING, not in request order, so a positional read pairs every value with the wrong cell
    while every count still looks plausible.

    A `status: FAILED` cell carries an EMPTY `vectorDataPoint`. At the tiny census subdivisions
    inside the Ottawa-Gatineau Québec part that is a PUBLICATION RULE — StatCan withholds small
    immigrant and age counts — so those coordinates are declared suppressible IN ADVANCE by
    their geography and handled by the bound. Everywhere else an absent cell is refused: an
    outage that reached the arithmetic as "the source publishes nothing here" would be a
    fabricated publication rule.
    """
    if not isinstance(response, list):
        raise ProbeRefusal("wds-data", f"getData returned {type(response).__name__}, not a list")
    series: dict = {}
    for item in response:
        obj = item.get("object") or {}
        points = obj.get("vectorDataPoint") or []
        key = (int(obj["productId"]), obj["coordinate"]) if obj.get("productId") else None
        if item.get("status") != "SUCCESS" or not points:
            if key in suppressible:
                continue
            raise ProbeRefusal(
                "wds-data",
                f"cell {obj.get('productId')}/{obj.get('coordinate')} came back "
                f"status={item.get('status')!r} with {len(points)} data point(s), and it is "
                "NOT in the suppressible scope (the QC-part CSDs' immigrant and age cells). A "
                "withheld count is a property of a tiny geography; anywhere else it is an "
                "outage, and treating it as a publication rule would fabricate one.")
        period = str(points[0].get("refPer", ""))[:4]
        if period != REF_YEAR:
            raise ProbeRefusal(
                "wds-data",
                f"cell {obj.get('productId')}/{obj.get('coordinate')} answered for reference "
                f"period {period!r}, not the {REF_YEAR} census — mixing vintages inside one "
                "ratio is silent.")
        series[key] = points[0]["value"]
    wanted = {(int(r["productId"]), r["coordinate"]) for r in requests}
    missing = sorted(wanted - set(series) - suppressible)
    if missing:
        raise ProbeRefusal("wds-data", f"{len(missing)} of {len(wanted)} requested cells did "
                                       f"not come back and are not suppressible (first: "
                                       f"{missing[0]})")
    return series, sorted(wanted - set(series))


def _guard_required_complete(published: dict) -> None:
    """The members the suppression BOUND is built from must themselves be published.

    The bound is `Total - Age` minus `Non-immigrants`: an upper bound on any immigrant member
    at that geography. If either leg were itself withheld the bound would be built out of a
    hole, and an interval whose upper end is unmeasured bounds nothing.
    """
    holes = sorted(name for name, complete in published.items() if not complete)
    if holes:
        raise ProbeRefusal(
            "wds-data",
            f"{len(holes)} QC-part CSD(s) do not publish both members the suppression bound is "
            f"built from ({holes[:4]}). An interval whose upper end is itself unmeasured is not "
            "a bound, and this run may not publish one as if it were.")


def _guard_membership_closes(children: dict, whole: float) -> None:
    """The CMA's children must sum EXACTLY to the CMA — a PARTITION, not a selection.

    This is what makes "the Québec side of this CMA" a subtractable territory: if the published
    children did not close on their own parent, the Québec subset of them would be an arbitrary
    slice of an unknown whole.
    """
    if not children:
        raise ProbeRefusal("membership", "the CMA published no constituent members — a "
                                         "membership over an empty child list is vacuous")
    total = sum(children.values())
    if total != whole:
        raise ProbeRefusal(
            "membership",
            f"the CMA's {len(children)} published children sum to {total:,.0f} against the "
            f"CMA's own {whole:,.0f} (gap {total - whole:+,.0f}). The Québec side is a SUBSET "
            "of that child list, so a list that does not close on its parent cannot define a "
            "territory to subtract.")


def _guard_qc_split(children: list) -> None:
    """SGC prefix and structural province ancestry must agree on EVERY child.

    Two independent readings of one claim. The prefix is a string property of a code; the
    ancestry is a property of the geography tree in the cube the counts are read from. A
    disagreement is a finding about the axis, never a tie broken by whichever ran first.
    """
    if not children:
        raise ProbeRefusal("membership", "no children were classified — a split over an empty "
                                         "list is vacuous")
    broken = [(name, by_code, by_tree) for name, by_code, by_tree in children
              if by_code != by_tree]
    if broken:
        raise ProbeRefusal(
            "membership",
            "the SGC prefix and the census-tree ancestry disagree on which side of the "
            "provincial boundary a member sits: "
            + "; ".join(f"{n} (prefix says {'QC' if c else 'not QC'}, tree says "
                        f"{'QC' if t else 'not QC'})" for n, c, t in broken))


def _guard_membership_pop(residual: float, threshold: float, innocent: dict) -> None:
    """The QC part's census population must match the ISQ row it aligns to.

    Like-for-like: a census TOTAL-population count against an ISQ TOTAL-population estimate, so
    the private-household universe gap is not inside this comparison at all. The threshold is
    derived from the six wholly-QC CMAs measured the same way — territories whose census↔ISQ
    identity is settled — so what it bounds is the census-vs-July-1 gap, and what survives is
    a membership difference.
    """
    if abs(residual) > threshold:
        raise ProbeRefusal(
            "membership",
            f"the resolved Québec part measures {residual:+.3f}% against the ISQ row it aligns "
            f"to, outside the {threshold:.3f}% derived from innocent controls "
            f"({', '.join(f'{n} {v:+.3f}%' for n, v in sorted(innocent.items()))}). Either the "
            "membership is not the territory ISQ publishes, or the two delineations moved "
            "apart — both are findings, neither is a footnote.")


_ROUNDING_TOLERANCE = 5      # StatCan rounds counts to 5, INDEPENDENTLY per cube


def _guard_universe(label: str, left: dict, right: dict, bound: tuple, fields: tuple) -> list:
    """Two cubes' province rows, asserted at exactly their measured strength.

    Bit-identity is required on the quantities this run SUBTRACTS ACROSS the pair — that is
    what makes them one universe at two grains rather than two sources. It is NOT required
    elsewhere, because independent rounding-to-5 makes it false elsewhere; those cells are
    RETURNED as measured. Anything beyond the rounding step refuses.
    """
    for key in bound:
        if key not in left or key not in right:
            raise ProbeRefusal("universe", f"{label}: no province row for {key!r} — an identity "
                                           "check over an absent row is vacuous")
        if tuple(left[key]) != tuple(right[key]):
            raise ProbeRefusal(
                "universe",
                f"{label}: the province rows DISAGREE on {key!r} ({left[key]} vs {right[key]}). "
                "Bit-identity there is the whole basis for subtracting one cube's geography "
                "from the other's.")
    drifted = []
    for key in sorted(set(left) & set(right)):
        # `strict`: both sides are the same statistic tuple for the same key, so they are
        # aligned by construction — but a silent truncation would drop a field from the
        # comparison without changing a single count.
        for index, (a, b) in enumerate(zip(left[key], right[key], strict=True)):
            gap = abs(a - b)
            if gap > _ROUNDING_TOLERANCE:
                raise ProbeRefusal(
                    "universe",
                    f"{label}: province {key!r} {fields[index]} differs by {gap} across the two "
                    f"cubes ({a} vs {b}) — beyond the ±{_ROUNDING_TOLERANCE} rounding step, so "
                    "these are not one universe")
            if gap:
                drifted.append((key, fields[index], a, b))
    return drifted


def _guard_no_direct_source(scanned: dict) -> None:
    """No maintainer-cross cube may publish the Québec part directly.

    If one did, this run's CSD sum would be an approximation standing beside a published exact,
    which is the shape amendment #12 exists to close. The scan is over the four cubes P9's
    closure names, by geography GRAIN (a CMA-part member) and by NAME, and it is published
    beside its result so a reader can falsify it.
    """
    found = {table_number(pid): hits for pid, hits in scanned.items() if hits}
    if found:
        raise ProbeRefusal(
            "direct-source",
            f"a maintainer-cross cube DOES publish a Québec-part member: {found}. The exact "
            "construction must then read that member rather than summing census subdivisions, "
            "and this run refuses to approximate beside a published exact.")


def _guard_ratio_band(low: float, high: float) -> None:
    """The suppression bound may not straddle the finding it is published under.

    The qualitative claim is that the ALIGNED ratio crosses 1.0 — settled immigrants out-own in
    the aligned hors-RMR territory where they under-own in the contaminated one. If the bound
    contains 1.0 the crossing is not earned at this run's own resolution.
    """
    lo, hi = sorted((low, high))
    if lo <= 1.0 <= hi:
        raise ProbeRefusal(
            "suppression",
            f"the suppression bound on the aligned ratio is [{lo:.4f}, {hi:.4f}] and STRADDLES "
            "1.0, so the crossing this note would report is not earned at the resolution the "
            "withheld cells leave. A verdict inside its own uncertainty is not a verdict.")


def _guard_footnote(label: str, marker: str, footnote: str) -> None:
    """The quoted separability evidence must be READ, never typed.

    The note quotes the workbook's Gatineau row label and the footnote line that scopes it. If
    either is missing the quotation would be an assertion wearing quotation marks.
    """
    if not marker or not label.endswith(marker):
        raise ProbeRefusal(
            "isq-workbook",
            f"the Gatineau row label {label!r} does not carry the footnote marker {marker!r} "
            "this note quotes it by — the scoping footnote is the evidence that the ISQ row is "
            "the QUÉBEC part, and a quotation this run cannot read may not be printed.")
    if ISQ_PART_FOOTNOTE_MARKER not in footnote:
        raise ProbeRefusal(
            "isq-workbook",
            f"footnote {marker!r} reads {footnote!r}, which does not scope the row to the "
            f"Québec part ({ISQ_PART_FOOTNOTE_MARKER!r} absent). The separability claim rests "
            "on that sentence.")


def _isq_label(text: object) -> str:
    """A workbook label, normalised: trailing space and FOOTNOTE MARKER off.

    ISQ appends the footnote's NUMBER to the cell text (`RMR d'Ottawa-Gatineau2`) and the RMR
    rows carry a trailing space. Neither is part of a territory's identity, and a guard keyed
    on them would REFUSE when a footnote is renumbered.
    """
    return str(text).strip().rstrip("0123456789").strip()


def _isq_marker(text: object) -> str:
    """The footnote NUMBER a label carries, as published — the other half of `_isq_label`."""
    stripped = str(text).strip()
    return stripped[len(stripped.rstrip("0123456789")):]


def _isq_code(book: str, labels: dict, label: str, pinned: int) -> int:
    """Resolve an ISQ row BY LABEL, then assert the code this probe pinned for it.

    The ISQ mirror of `_member` + `_guard_pinned_id`: a code alone is a hand-typed integer no
    live surface confirms, so a re-pinned vintage that renumbered the axis would put another
    territory's numbers through the arithmetic with the note still rendering a PASS.
    """
    hits = sorted(code for code, text in labels.items() if _isq_label(text) == label)
    if len(hits) != 1:
        raise ProbeRefusal(
            "isq-workbook",
            f"{book}: {len(hits)} row(s) labelled {label!r} (of {len(labels)}). Exactly one "
            "must match — zero means the workbook renamed or dropped the row this run keys on, "
            "more than one means the label alone does not identify a territory.")
    if hits[0] != pinned:
        raise ProbeRefusal(
            "isq-workbook",
            f"{book}: {label!r} sits at code {hits[0]}, not the pinned {pinned}.")
    return hits[0]


def _guard_isq_parts(book: str, province: float, parts: dict, tolerance: int) -> float:
    """The workbook's own parts must close on its province row, within a stated tolerance.

    This is the ISQ-side separability, MEASURED rather than quoted: the published parts add up
    to the province ONLY with the Gatineau row counted as its own part, so the hors-RMR row
    cannot also contain it. The population workbook closes exactly; the flow workbook closes to
    a rounding step, and the tolerance is stated at the size its own arithmetic needs.
    """
    total = sum(parts.values())
    gap = total - province
    if abs(gap) > tolerance:
        raise ProbeRefusal(
            "isq-workbook",
            f"{book}: its {len(parts)} published parts sum to {total:,.0f} against a province "
            f"row of {province:,.0f} (gap {gap:+,.0f}, tolerance ±{tolerance}). If the parts do "
            "not partition the province, 'the hors-RMR row excludes Gatineau' is not something "
            "this arithmetic can show.")
    return gap


def _guard_citation(source_name: str, text: str, coupled: list) -> None:
    """Every figure this run REPRODUCES from a document must match that document's digits.

    Deliberately scoped to the REPRODUCED set. The ALIGNED values are new measurements: they
    are not in §6 and must not be, until the seat rules them in — so a gate demanding their
    presence would either fail today or force the wiring this run refuses to do.
    """
    absent = [(label, tok) for label, tok in coupled if tok not in text]
    if absent:
        raise ProbeRefusal(
            "citation",
            f"figures this run recomputes that {source_name} does NOT state: "
            + "; ".join(f"{label} -> {tok}" for label, tok in absent)
            + ". Either this run measured something the record contradicts (fork-class — raise "
              "it, do not paper over it) or the record moved and the coupling must move with it.")


def _guard_p9(text: str) -> dict:
    """The closure P9 earned, read from P9's own DECISION tokens."""
    wanted = ("DECISION-VERDICT", "DECISION-CLOSURE-LEVEL", "DECISION-MAINTAINER-CROSS",
              "DECISION-RESIDUAL")
    tokens = {}
    for name in wanted:
        found = re.search(rf"`{name}:\s*(.*?)`", text)
        if not found:
            raise ProbeRefusal("p9", f"P9's note carries no `{name}` token — this note may not "
                                     "restate a closure whose level it cannot read")
        tokens[name] = found.group(1).strip()
    return tokens


# ===========================================================================================
# Reading the sources
# ===========================================================================================
def _census_requests(pid: int, geo_ids: list, ids: dict) -> tuple:
    """One request per (geography, popchar, statistic) cell, plus the key map."""
    requests, keys = [], {}
    for geo in geo_ids:
        for name in POPCHARS:
            for field, (tenure, maintainer) in enumerate(
                    ((ids["tenure_total"], ids["maint_all"]),
                     (ids["tenure_total"], ids["maint_primary"]),
                     (ids["tenure_owner"], ids["maint_primary"]))):
                coordinate = coord(geo, tenure, ids["gender"], maintainer, ids["statistic"],
                                   ids["popchar"][name], ids["suitability"])
                requests.append({"productId": pid, "coordinate": coordinate, "latestN": 1})
                keys[(geo, name, field)] = (pid, coordinate)
    return requests, keys


def _census_fields(series: dict, keys: dict, geo: int, name: str) -> tuple:
    """The three counts as they came back — `None` in the position of a withheld one.

    The FIELD-WISE reading. `_census_cell` collapses a partially published geography to
    nothing, which is right where a whole `Cell` is needed and wrong where a sum is: these
    subdivisions publish persons while withholding maintainers, and that published count is
    part of the territory being netted out.
    """
    return tuple(series.get(keys[(geo, name, field)]) for field in range(3))


def _census_cell(series: dict, keys: dict, geo: int, name: str) -> Cell | None:
    """A `Cell`, or None when ANY of its three counts is withheld at this geography."""
    values = [series.get(keys[(geo, name, field)]) for field in range(3)]
    if any(v is None for v in values):
        return None
    return Cell(int(values[0]), int(values[1]), int(values[2]))


def _rho_requests(pid: int, pos: dict, geo_ids: list, ids: dict, ages: dict) -> tuple:
    """Ownership-by-maintainer-age cells: (total households, owner households) per age member.

    The two cubes' dimension LAYOUTS differ (98100231 carries `Condominium status`), so the
    coordinate is assembled from the position map rather than from a shared literal order.
    """
    requests, keys = [], {}
    for geo in geo_ids:
        for age_name, age_id in ages.items():
            for field, tenure in enumerate((ids["tenure_total"], ids["tenure_owner"])):
                parts = {pos["geo"]: geo, pos["struct"]: ids["struct"], pos["hh"]: ids["hh"],
                         pos["stat"]: ids["statistic"], pos["age"]: age_id,
                         pos["tenure"]: tenure}
                if "condo" in pos:
                    parts[pos["condo"]] = ids["condo"]
                coordinate = coord(*[parts[p] for p in sorted(parts)])
                requests.append({"productId": pid, "coordinate": coordinate, "latestN": 1})
                keys[(geo, age_name, field)] = (pid, coordinate)
    return requests, keys


def _band_pair(series: dict, keys: dict, geo: int, members) -> tuple | None:
    """(total households, owner households) summed over a band's published age members."""
    out = []
    for field in range(2):
        values = [series.get(keys[(geo, name, field)]) for name in members]
        if any(v is None for v in values):
            return None
        out.append(sum(values))
    return tuple(out)


def _isq_header_row(book: str, rows: list) -> int:
    """The workbook's own `Scénario` header row. Located by its cell text, never by index."""
    header = next((i for i, row in enumerate(rows)
                   if row and str(row[0]).strip() == ISQ_HEADER_CELL), None)
    if header is None:
        raise ProbeRefusal("isq-workbook", f"{book} carries no {ISQ_HEADER_CELL!r} header row — "
                                           "its layout is not the one this probe reads")
    return header


def _isq_table(book: str, rows: list, *, value_column: int | None = None,
               scenario: str | None = None) -> dict:
    """The workbook's own body, keyed on its `Code` column.

    Two workbooks, two shapes, one reader: the population book is read at the census year with
    its three scenario fans cross-checked (2021 is a `réel` observation, so a disagreement
    means the row is not the observation it is taken for), while the flow book is read as a
    per-year series on ONE named scenario. Both resolve the header row by its own `Scénario`
    cell and select body rows by scenario membership rather than by a magic offset.
    """
    header = _isq_header_row(book, rows)
    note = next((str(row[0]).strip() for row in rows[:header]
                 if row and row[0] and ISQ_NOTE_MARKER in str(row[0])), "")
    if not note:
        raise ProbeRefusal("isq-workbook", f"{book} states no {ISQ_NOTE_MARKER!r} header note — "
                                           "the delineation it uses is exactly what this run "
                                           "must quote rather than assume")
    footnote = next((str(row[0]).strip() for row in rows[:header]
                     if row and row[0] and ISQ_PART_FOOTNOTE_MARKER in str(row[0])), "")
    labels, values, withheld = {}, {}, []
    for row in rows[header + 1:]:
        if not row or row[0] not in ISQ_SCENARIOS or row[1] is None:
            continue
        code = int(row[1])
        labels[code] = str(row[2])
        if value_column is None:                       # the population book
            if row[3] != ISQ_YEAR or row[5] != ISQ_SEX_TOTAL:
                continue
            values.setdefault(code, {})[row[0]] = int(row[6])
        else:                                          # the flow book
            if row[0] != scenario:
                continue
            cell = row[value_column]
            if isinstance(cell, str):
                withheld.append((code, row[3], cell))
                continue
            values.setdefault(code, {})[int(row[3])] = cell
    if not values:
        raise ProbeRefusal("isq-workbook", f"{book} yielded no rows this run can read — a "
                                           "total over nothing")
    if value_column is None:
        totals = {}
        for code, fan in values.items():
            if set(fan) != set(ISQ_SCENARIOS) or len(set(fan.values())) != 1:
                raise ProbeRefusal(
                    "isq-workbook",
                    f"{book} code {code}: the {ISQ_YEAR} scenario fans disagree ({fan}). "
                    f"{ISQ_YEAR} is a `réel` observation in this workbook family, so three "
                    "different values mean the row being read is not the observation it is "
                    "taken for.")
            totals[code] = next(iter(fan.values()))
        values = totals
    return {"note": note, "footnote": footnote, "labels": labels, "values": values,
            "withheld": withheld, "header_row": header}


def _flow_column(book: str, rows: list, header_row: int) -> int:
    """Locate the arrival-flow column by its OWN stacked header text, never by index.

    The compo sheet stacks a column's name across four header rows (`Immigrants` /
    `permanents`), and its columns are neither contiguous nor labelled once. Resolving by the
    joined text and refusing on anything but a single hit is the same discipline the census
    side's `_member` applies — a hand-typed column index is an integer no live surface confirms.
    """
    width = max(len(row) for row in rows)
    hits = []
    for column in range(width):
        stacked = " ".join(
            str(rows[i][column]).strip()
            for i in range(header_row, min(header_row + ISQ_HEADER_STACK, len(rows)))
            if column < len(rows[i]) and rows[i][column] is not None)
        if stacked == ISQ_FLOW_COLUMN:
            hits.append(column)
    if len(hits) != 1:
        raise ProbeRefusal(
            "isq-workbook",
            f"{book}: {len(hits)} column(s) carry the stacked header {ISQ_FLOW_COLUMN!r}. "
            "Exactly one must — zero means the operand's own column was renamed or moved, more "
            "than one means the header text does not identify it.")
    return hits[0]


# ===========================================================================================
# Formatting
# ===========================================================================================
def _r4(x: float) -> str:
    return f"{x:.4f}"


def _r3(x: float) -> str:
    return f"{x:.3f}"


def _pct(x: float) -> str:
    return f"{x:+.3f}%"


def _r2(x: float) -> str:
    """Two decimals — the precision the spec and the audit record state these figures at.

    A coupling token must be rendered at the SOURCE's precision: `40.420` is the same
    measurement as `40.42` and would still fail a substring gate against a document that
    published two decimals.
    """
    return f"{x:.2f}"


def _n(x: float) -> str:
    return f"{int(x):,}"


def _plural(count: int, noun: str) -> str:
    """`1 cell` / `2 cells`. A generated note that writes '1 cells' is a hand-written sentence
    meeting a computed number, which is the seam this family keeps closing."""
    return f"{count} {noun}" if count == 1 else f"{count} {noun}s"


# ===========================================================================================
# The note
# ===========================================================================================
_SCOPE = (
    "SCOPE OF THIS HEADER (it claims only what it can enforce): every count, rate, ratio, "
    "residual, delta and threshold below is COMPUTED by this run from the live StatCan WDS "
    "responses and the two pinned ISQ workbooks it names — none is transcribed from the "
    "rulings or the audit record it is checked against. The check runs one way: those "
    "documents state the SHIPPED and BRACKET figures, this run recomputes them, and a "
    "disagreement REFUSES the run rather than being published. The ALIGNED figures are NEW "
    "measurements and are deliberately absent from the spec — they are proposed here, ruled "
    "elsewhere. Quoted strings are verbatim from a live response, a workbook cell or a named "
    "in-repo file, and every absence claim is scoped to the search that produced it."
)
_CITED_LABEL = "Quoted or cited verbatim (not computed here):"


def _summary(*, total: int, derived: int, cited: int) -> str:
    return (f"This run registered {total} provenance-tagged figures: {derived} DERIVED "
            f"(computed here from the live responses and pinned workbooks this run read) and "
            f"{cited} CITED (verbatim from a live response, a workbook cell, the spec, the "
            f"Task-26 audit record or a named in-repo file). The tagged set is the NARRATIVE "
            f"figures — the ones a sentence rests on. Untagged numerals fall in two other "
            f"classes, both stated rather than left to be assumed: TABLE CELLS, which are "
            f"counts and rates this run computed from the coordinate-keyed live responses "
            f"printed in the same row, and AUDIT METADATA — member ids, dimension positions, "
            f"member counts, geoLevels, SGC codes and row indices. Every one of the three is "
            f"traceable to a response or a file this run read.")


def _sections(where: list) -> list:            # noqa: C901 — one derivation, read top to bottom
    """The whole derivation. Appends nothing global; raises ProbeRefusal on any guard."""
    global LAST_RUN
    LAST_RUN = {}

    # ---------------------------------------------------------------- boundary: wds-meta
    where[0] = "wds-meta"
    pids = [CMA_PID, CD_PID, MEMBER_PID, RHO_CMA_PID, RHO_CD_PID, SHELTER_CMA_PID,
            SHELTER_CD_PID, PARTS_PID]
    metas = _guard_meta(pids, _meta(pids))

    def geo_members(pid: int) -> list:
        return _dimension(metas, pid, POS_GEO).get("member") or []

    # --- the ruled cube and its CSD sibling
    census_ids = {}
    for pid in (CMA_PID, CD_PID):
        census_ids[pid] = {
            "tenure_total": _member(metas, pid, POS_TENURE, TENURE_TOTAL)["memberId"],
            "tenure_owner": _member(metas, pid, POS_TENURE, TENURE_OWNER)["memberId"],
            "gender": _member(metas, pid, POS_GENDER, GENDER_TOTAL)["memberId"],
            "maint_all": _member(metas, pid, POS_MAINT, MAINT_ALL_PERSONS)["memberId"],
            "maint_primary": _member(metas, pid, POS_MAINT, MAINT_PRIMARY)["memberId"],
            "statistic": _member(metas, pid, POS_STAT, STAT_PEOPLE)["memberId"],
            "suitability": _member(metas, pid, POS_SUIT, SUIT_TOTAL)["memberId"],
            "popchar": {name: _guard_pinned_id(
                pid, name, _member(metas, pid, POS_POPCHAR, name),
                PINNED_POPCHAR_IDS[name])["memberId"] for name in POPCHARS},
        }
    province = {pid: _guard_pinned_id(
        pid, QUEBEC, _member(metas, pid, POS_GEO, QUEBEC, geo_level=PROVINCE_GEO_LEVEL),
        PINNED_GEO_IDS[(pid, QUEBEC)]) for pid in (CMA_PID, CD_PID)}

    cma_members = sorted(
        (m for m in geo_members(CMA_PID)
         if m.get("parentMemberId") == province[CMA_PID]["memberId"]
         and m.get("geoLevel") == CMA_GEO_LEVEL),
        key=lambda m: m["memberId"])
    if tuple(m["memberId"] for m in cma_members) != PINNED_CMA_IDS:
        raise ProbeRefusal(
            "wds-meta",
            f"the geoLevel-{CMA_GEO_LEVEL} children of Quebec in {table_number(CMA_PID)} are "
            f"{[m['memberId'] for m in cma_members]}, not the pinned {list(PINNED_CMA_IDS)} — "
            "the shipped residual is the province net of exactly this set")

    # The census-side INSEPARABILITY, read rather than asserted.
    ontario = _member(metas, CMA_PID, POS_GEO, ONTARIO, geo_level=PROVINCE_GEO_LEVEL)
    og_cma = _guard_pinned_id(CMA_PID, OG_CMA_NAME,
                              _member(metas, CMA_PID, POS_GEO, OG_CMA_NAME,
                                      geo_level=CMA_GEO_LEVEL,
                                      parent=ontario["memberId"]), PINNED_OG_CMA_ID)
    og_named = [m for m in geo_members(CMA_PID) if "Gatineau" in m["memberNameEn"]]

    # --- the membership cube
    og_member = _guard_pinned_id(
        MEMBER_PID, OG_MEMBER_NAME,
        _member(metas, MEMBER_PID, POS_GEO, OG_MEMBER_NAME, geo_level=CMA_GEO_LEVEL),
        PINNED_OG_MEMBER_ID)
    member_count = _member(metas, MEMBER_PID, 2, MEMBER_COUNT_2021)["memberId"]
    children = sorted((m for m in geo_members(MEMBER_PID)
                       if m.get("parentMemberId") == og_member["memberId"]),
                      key=lambda m: str(m["classificationCode"]))
    qc_children = [m for m in children
                   if str(m["classificationCode"]).startswith(QC_SGC_PREFIX)]
    if tuple(str(m["classificationCode"]) for m in qc_children) != PINNED_QC_PART_CODES:
        raise ProbeRefusal(
            "membership",
            f"the Québec-side children of {OG_MEMBER_NAME} are "
            f"{[str(m['classificationCode']) for m in qc_children]}, not the pinned "
            f"{list(PINNED_QC_PART_CODES)}. The delineation this run subtracts is exactly that "
            "set, so a change here silently redefines the aligned territory.")

    # --- the QC-part CSDs, and the bracket CDs, in the ruled cube's sibling
    cd_by_code = {}
    for m in geo_members(CD_PID):
        cd_by_code.setdefault((str(m.get("classificationCode")), m.get("geoLevel")), []).append(m)

    csd_members = {str(m["classificationCode"]): _member_by_code(
        CD_PID, cd_by_code, str(m["classificationCode"]), CSD_GEO_LEVEL, m["memberNameEn"])
        for m in qc_children}
    cd_members = {code: _member_by_code(CD_PID, cd_by_code, code, CD_GEO_LEVEL, name)
                  for code, name in BRACKET_CDS}
    _guard_pinned_id(CD_PID, "CD Gatineau", cd_members["2481"], PINNED_CD_GATINEAU_ID)
    _guard_pinned_id(CD_PID, "CSD Gatineau", csd_members["2481017"], PINNED_CSD_GATINEAU_ID)

    # Structural ancestry: is this CSD's census division a child of the Quebec member?
    cd_parent = {m["memberId"]: m for m in geo_members(CD_PID)}
    quebec_cds = {m["memberId"] for m in geo_members(CD_PID)
                  if m.get("parentMemberId") == province[CD_PID]["memberId"]
                  and m.get("geoLevel") == CD_GEO_LEVEL}
    split = []
    for m in children:
        code = str(m["classificationCode"])
        by_code = code.startswith(QC_SGC_PREFIX)
        sibling = (cd_by_code.get((code, CSD_GEO_LEVEL)) or [None])[0]
        by_tree = bool(sibling and sibling.get("parentMemberId") in quebec_cds)
        split.append((m["memberNameEn"], by_code, by_tree))
    _guard_qc_split(split)

    # --- the ρ cubes
    rho_ids = {}
    for pid, pos, bands, hh_total in ((RHO_CMA_PID, RHO_CMA_POS, RHO_BANDS_CMA,
                                       RHO_HH_TOTAL_CMA),
                                      (RHO_CD_PID, RHO_CD_POS, RHO_BANDS_CD, RHO_HH_TOTAL_CD)):
        entry = {
            "struct": _member(metas, pid, pos["struct"], RHO_STRUCT_TOTAL)["memberId"],
            "hh": _member(metas, pid, pos["hh"], hh_total)["memberId"],
            "statistic": _member(metas, pid, pos["stat"], RHO_STAT)["memberId"],
            "tenure_total": _member(metas, pid, pos["tenure"], RHO_TENURE_TOTAL)["memberId"],
            "tenure_owner": _member(metas, pid, pos["tenure"], RHO_TENURE_OWNER)["memberId"],
            "ages": {name: _member(metas, pid, pos["age"], name)["memberId"]
                     for name in [RHO_AGE_TOTAL] + [n for b in bands.values() for n in b]},
        }
        if "condo" in pos:
            entry["condo"] = _member(metas, pid, pos["condo"], RHO_CONDO_TOTAL)["memberId"]
        rho_ids[pid] = entry
    rho_province = {pid: _member(metas, pid, POS_GEO, QUEBEC, geo_level=PROVINCE_GEO_LEVEL)
                    for pid in (RHO_CMA_PID, RHO_CD_PID)}
    rho_cma_members = [_member(metas, RHO_CMA_PID, POS_GEO, m["memberNameEn"],
                               geo_level=CMA_GEO_LEVEL) for m in cma_members]
    rho_cd_by_code = {}
    for m in geo_members(RHO_CD_PID):
        rho_cd_by_code.setdefault((str(m.get("classificationCode")), m.get("geoLevel")),
                                  []).append(m)
    rho_csd = {code: _member_by_code(RHO_CD_PID, rho_cd_by_code, code, CSD_GEO_LEVEL, code)
               for code in csd_members}

    # --- the absence: does ANY maintainer-cross cube publish the Québec part directly?
    scanned = {}
    for pid in MAINTAINER_CROSS:
        scanned[pid] = [m["memberNameEn"] for m in geo_members(pid)
                        if m.get("geoLevel") == CMA_PART_GEO_LEVEL
                        or ("Gatineau" in m["memberNameEn"]
                            and "part" in m["memberNameEn"].casefold())]
    _guard_no_direct_source(scanned)
    parts_by_code = {}
    for m in geo_members(PARTS_PID):
        parts_by_code.setdefault((str(m.get("classificationCode")), m.get("geoLevel")),
                                 []).append(m)
    parts_member = _member_by_code(PARTS_PID, parts_by_code, PARTS_QC_PART_CODE,
                                   CMA_PART_GEO_LEVEL, "Ottawa-Gatineau, Québec part")
    parts_dims = [d["dimensionNameEn"] for d in sorted(metas[PARTS_PID]["dimension"],
                                                       key=lambda d: d["dimensionPositionId"])]

    # ---------------------------------------------------------------- boundary: wds-data
    where[0] = "wds-data"
    requests, keys = [], {}
    cma_geo_ids = [province[CMA_PID]["memberId"]] + [m["memberId"] for m in cma_members]
    cd_geo_ids = ([province[CD_PID]["memberId"]]
                  + [m["memberId"] for m in csd_members.values()]
                  + [m["memberId"] for m in cd_members.values()])
    for pid, geo_ids in ((CMA_PID, cma_geo_ids), (CD_PID, cd_geo_ids)):
        part, part_keys = _census_requests(pid, geo_ids, census_ids[pid])
        requests += part
        keys.update({(pid,) + k: v for k, v in part_keys.items()})

    pop_keys = {}
    for member in [og_member] + children + [
            m for m in geo_members(MEMBER_PID)
            if m.get("geoLevel") == CMA_GEO_LEVEL
            and str(m["classificationCode"]) in {str(c["classificationCode"])
                                                 for c in cma_members}]:
        coordinate = coord(member["memberId"], member_count)
        requests.append({"productId": MEMBER_PID, "coordinate": coordinate, "latestN": 1})
        pop_keys[member["memberId"]] = (MEMBER_PID, coordinate)

    rho_keys = {}
    for pid, pos, geo_ids in (
            (RHO_CMA_PID, RHO_CMA_POS,
             [rho_province[RHO_CMA_PID]["memberId"]] + [m["memberId"] for m in rho_cma_members]),
            (RHO_CD_PID, RHO_CD_POS,
             [rho_province[RHO_CD_PID]["memberId"]] + [m["memberId"] for m in rho_csd.values()])):
        part, part_keys = _rho_requests(pid, pos, geo_ids, rho_ids[pid], rho_ids[pid]["ages"])
        requests += part
        rho_keys.update({(pid,) + k: v for k, v in part_keys.items()})

    # The suppressible scope, declared BEFORE the response is read: the QC-part CSDs only.
    suppressible = {v for k, v in keys.items()
                    if k[0] == CD_PID and k[1] in {m["memberId"] for m in csd_members.values()}}
    suppressible |= {v for k, v in rho_keys.items()
                     if k[0] == RHO_CD_PID and k[1] in {m["memberId"] for m in rho_csd.values()}}
    series, withheld = _guard_response(requests, _data(requests), suppressible)

    # ---------------------------------------------------------------- boundary: isq-workbook
    where[0] = "isq-workbook"
    pop_rows, flow_rows = _isq_rows(ISQ_POP_BOOK), _isq_rows(ISQ_FLOW_BOOK)
    pop_book = _isq_table(ISQ_POP_BOOK, pop_rows)
    flow_column = _flow_column(ISQ_FLOW_BOOK, flow_rows,
                               _isq_header_row(ISQ_FLOW_BOOK, flow_rows))
    flow_book = _isq_table(ISQ_FLOW_BOOK, flow_rows, value_column=flow_column,
                           scenario=ISQ_REFERENCE)

    isq_codes = {}
    for book, table, hors_label in ((ISQ_POP_BOOK, pop_book, ISQ_HORS_POP_LABEL),
                                    (ISQ_FLOW_BOOK, flow_book, ISQ_HORS_FLOW_LABEL)):
        for label, pinned in ((ISQ_PROVINCE_LABEL, ISQ_PROVINCE_CODE),
                              (hors_label, ISQ_HORS_CODE),
                              (ISQ_GATINEAU_LABEL, ISQ_GATINEAU_CODE)):
            isq_codes[(book, label)] = _isq_code(book, table["labels"], label, pinned)
    _guard_footnote(str(pop_book["labels"][ISQ_GATINEAU_CODE]).strip(),
                    _isq_marker(pop_book["labels"][ISQ_GATINEAU_CODE]), pop_book["footnote"])
    _guard_footnote(str(flow_book["labels"][ISQ_GATINEAU_CODE]).strip(),
                    _isq_marker(flow_book["labels"][ISQ_GATINEAU_CODE]), flow_book["footnote"])

    isq_province_pop = pop_book["values"][ISQ_PROVINCE_CODE]
    _guard_isq_parts(ISQ_POP_BOOK, isq_province_pop,
                     {c: v for c, v in pop_book["values"].items() if c != ISQ_PROVINCE_CODE}, 0)
    isq_gatineau_pop = pop_book["values"][ISQ_GATINEAU_CODE]

    flow_years = sorted(flow_book["values"][ISQ_PROVINCE_CODE])
    flow_gaps = {}
    for year in flow_years:
        parts = {code: fan[year] for code, fan in flow_book["values"].items()
                 if code != ISQ_PROVINCE_CODE and year in fan}
        flow_gaps[year] = _guard_isq_parts(
            f"{ISQ_FLOW_BOOK} ({year})", flow_book["values"][ISQ_PROVINCE_CODE][year],
            parts, ISQ_FLOW_TOLERANCE)
    hors_flow = [flow_book["values"][ISQ_HORS_CODE][y] for y in flow_years]
    gatineau_flow = [flow_book["values"][ISQ_GATINEAU_CODE][y] for y in flow_years]
    hors_mean = sum(hors_flow) / len(hors_flow)
    gatineau_mean = sum(gatineau_flow) / len(gatineau_flow)

    # ---------------------------------------------------------------- the derivation
    where[0] = "derivation"
    cma_keys = {k[1:]: v for k, v in keys.items() if k[0] == CMA_PID}
    cd_keys = {k[1:]: v for k, v in keys.items() if k[0] == CD_PID}

    def cma_cell(geo: int, name: str) -> Cell:
        cell = _census_cell(series, cma_keys, geo, name)
        if cell is None:
            raise ProbeRefusal("derivation", f"{table_number(CMA_PID)} geography {geo} "
                                             f"publishes no complete {name!r} triple")
        return cell

    prov_cma = {name: cma_cell(province[CMA_PID]["memberId"], name) for name in POPCHARS}
    prov_cd = {name: _census_cell(series, cd_keys, province[CD_PID]["memberId"], name)
               for name in POPCHARS}
    universe_drift = _guard_universe(
        f"{table_number(CMA_PID)} vs {table_number(CD_PID)}",
        {n: (c.persons, c.maintainers, c.owner_maintainers) for n, c in prov_cma.items()},
        {n: (c.persons, c.maintainers, c.owner_maintainers) for n, c in prov_cd.items()
         if c is not None},
        bound=(PC_SETTLED,), fields=("persons", "maintainers", "owner-maintainers"))
    shipped = {name: cell_minus(prov_cma[name],
                                cell_sum([cma_cell(m["memberId"], name) for m in cma_members]))
               for name in POPCHARS}

    # The QC part, FIELD BY FIELD, with every withheld count bounded by what IS published.
    rows, csd_settled, complete, clamped_fields = [], {}, {}, 0
    for code, member in csd_members.items():
        geo = member["memberId"]
        total = _census_cell(series, cd_keys, geo, PC_TOTAL_AGE)
        nonimm = _census_cell(series, cd_keys, geo, PC_NONIMM)
        complete[f"{member['memberNameEn']} ({code})"] = bool(total and nonimm)
        values = _census_fields(series, cd_keys, geo, PC_SETTLED)
        # The bound: all immigrants and non-permanent residents together, which CONTAINS the
        # ruled `Before 2016` member by construction. Published at every one of these
        # geographies — `_guard_required_complete` is what makes that a checked precondition.
        bound, clamped = (complement_bound(total, nonimm) if (total and nonimm)
                          else (None, 0))
        clamped_fields += clamped
        csd_settled[code] = (values, bound)
        rows.append((values, (bound.persons, bound.maintainers, bound.owner_maintainers)
                     if bound else (None, None, None)))
    _guard_required_complete(complete)
    incomplete = {code for code, (values, _bound) in csd_settled.items()
                  if any(v is None for v in values)}
    # `incomplete` is the SETTLED TRIPLE's set of subdivisions; `withheld` spans BOTH cube
    # pairs. Composing one with the other reads as a single measurement of a single set and is
    # not one — a subdivision withholding only ρ cells is missed, one withholding only settled
    # cells is counted. So the union carries its own geographies, inverted PER CUBE — not
    # because the two cubes disagree on these ids (measured, they agree on all 16 at this
    # vintage) but because a geography id is a PER-CUBE namespace: that agreement is an
    # unasserted coincidence, and reading one cube's ids through another's map would be a
    # silent mis-key the first time it stops holding.
    csd_code_by_id = {CD_PID: {m["memberId"]: code for code, m in csd_members.items()},
                      RHO_CD_PID: {m["memberId"]: code for code, m in rho_csd.items()}}
    withheld_csds = {csd_code_by_id[pid][int(coordinate.split(".")[0])]
                     for pid, coordinate in withheld}
    withheld_by_cube = {pid: sum(1 for p, _ in withheld if p == pid)
                        for pid in (CD_PID, RHO_CD_PID)}
    low_counts, high_counts, withheld_fields = bounded_sum(rows)
    qc_low, qc_high = Cell(*low_counts), Cell(*high_counts)
    qc_totals = {name: cell_sum([_census_cell(series, cd_keys, m["memberId"], name)
                                 for m in csd_members.values()])
                 for name in (PC_TOTAL_AGE, PC_NONIMM)}

    aligned_low = {PC_SETTLED: cell_minus(shipped[PC_SETTLED], qc_low),
                   PC_NONIMM: cell_minus(shipped[PC_NONIMM], qc_totals[PC_NONIMM]),
                   PC_TOTAL_AGE: cell_minus(shipped[PC_TOTAL_AGE], qc_totals[PC_TOTAL_AGE])}
    aligned_high = dict(aligned_low, **{PC_SETTLED: cell_minus(shipped[PC_SETTLED], qc_high)})

    def pair(cells: dict) -> tuple:
        return cells[PC_SETTLED].headship, ownership_ratio(cells[PC_SETTLED], cells[PC_NONIMM])

    shipped_pair = pair(shipped)
    envelope = bounded_pair(shipped[PC_SETTLED], qc_low, qc_high, aligned_low[PC_NONIMM])
    aligned_pair = (envelope["headship"], envelope["ratio"])
    _guard_ratio_band(*envelope["ratio_band"])

    # The whole-CD bracket — the construction this run did NOT use, measured anyway.
    bracket, running = {}, dict(shipped)
    for code, name in BRACKET_CDS:
        member = cd_members[code]
        running = {n: cell_minus(running[n],
                                 _census_cell(series, cd_keys, member["memberId"], n))
                   for n in POPCHARS}
        bracket[code] = {"label": name, "cells": running, "pair": pair(running)}
    bracket_headship = [v["pair"][0] for v in bracket.values()]
    bracket_ratio = [v["pair"][1] for v in bracket.values()]
    encloses = {
        "headship": min(bracket_headship) <= aligned_pair[0] <= max(bracket_headship),
        "ratio": min(bracket_ratio) <= aligned_pair[1] <= max(bracket_ratio),
    }

    # CD Gatineau's weight in the shipped residual — the mandate's 40.42% against its own
    # person weight, both recomputed, because the two are different denominators.
    gatineau_cd = {n: _census_cell(series, cd_keys, cd_members["2481"]["memberId"], n)
                   for n in POPCHARS}
    weights = {
        "settled_persons": 100 * gatineau_cd[PC_SETTLED].persons / shipped[PC_SETTLED].persons,
        "persons": 100 * gatineau_cd[PC_TOTAL_AGE].persons / shipped[PC_TOTAL_AGE].persons,
        "maintainers": (100 * gatineau_cd[PC_TOTAL_AGE].maintainers
                        / shipped[PC_TOTAL_AGE].maintainers),
    }
    qc_weights = {
        "settled_persons": 100 * qc_low.persons / shipped[PC_SETTLED].persons,
        "persons": 100 * qc_totals[PC_TOTAL_AGE].persons / shipped[PC_TOTAL_AGE].persons,
    }

    # ---------------------------------------------------------------- the membership gate
    where[0] = "membership"
    child_pop = {str(m["classificationCode"]): series[pop_keys[m["memberId"]]]
                 for m in children}
    _guard_membership_closes(child_pop, series[pop_keys[og_member["memberId"]]])
    qc_part_pop = sum(child_pop[str(m["classificationCode"])] for m in qc_children)
    census_cma_pop = {str(m["classificationCode"]): series[pop_keys[m["memberId"]]]
                      for m in geo_members(MEMBER_PID)
                      if m["memberId"] in pop_keys and m.get("geoLevel") == CMA_GEO_LEVEL
                      and m["memberId"] != og_member["memberId"]}
    innocent = {}
    for member in cma_members:
        code = str(member["classificationCode"])
        if code not in census_cma_pop or int(code) not in pop_book["values"]:
            raise ProbeRefusal(
                "membership",
                f"innocent control {member['memberNameEn']} (SGC {code}) has no matching row on "
                "both sides — the calibration set's identity is what makes its spread a "
                "measure of the census-vs-estimate gap rather than of territory mismatch")
        innocent[member["memberNameEn"]] = relative_pct(census_cma_pop[code],
                                                        pop_book["values"][int(code)])
    bounds = derive_threshold(innocent)
    membership_residual = relative_pct(qc_part_pop, isq_gatineau_pop)
    _guard_membership_pop(membership_residual, bounds["threshold"], innocent)

    # The universe conversion, stated as the approximation it is. The province-wide ratio is
    # DERIVED from the two counts `loaders/constants.py` records — this run reads that line
    # rather than retyping either number, and rather than substituting the ISQ estimate for the
    # census total (which would fold the census-vs-July-1 gap into a universe ratio).
    recorded = re.search(r"\(([\d,]+) private-household persons vs ([\d,]+) published\)",
                         _constants_source())
    if not recorded:
        raise ProbeRefusal(
            "constants",
            "loaders/constants.py no longer records the private-household universe gap this "
            "note converts at — the province-wide ratio would become a hand-typed number.")
    recorded_ph, recorded_total = (int(g.replace(",", "")) for g in recorded.groups())
    province_universe = recorded_ph / recorded_total
    local_universe = qc_totals[PC_TOTAL_AGE].persons / qc_part_pop
    converted = isq_gatineau_pop * province_universe

    # ---------------------------------------------------------------- the ownership leg
    where[0] = "rho"
    rho_cma_keys = {k[1:]: v for k, v in rho_keys.items() if k[0] == RHO_CMA_PID}
    rho_cd_keys = {k[1:]: v for k, v in rho_keys.items() if k[0] == RHO_CD_PID}
    rho_prov = {}
    for pid, keymap, member in ((RHO_CMA_PID, rho_cma_keys, rho_province[RHO_CMA_PID]),
                                (RHO_CD_PID, rho_cd_keys, rho_province[RHO_CD_PID])):
        pair_ = _band_pair(series, keymap, member["memberId"], (RHO_AGE_TOTAL,))
        if pair_ is None:
            raise ProbeRefusal("rho", f"{table_number(pid)} publishes no province total for "
                                      f"{RHO_AGE_TOTAL!r}")
        rho_prov[pid] = pair_
    rho_drift = _guard_universe(
        f"{table_number(RHO_CMA_PID)} vs {table_number(RHO_CD_PID)}",
        {"all ages": rho_prov[RHO_CMA_PID]}, {"all ages": rho_prov[RHO_CD_PID]},
        bound=("all ages",), fields=("households", "owner households"))

    def rho_band(members) -> dict:
        shipped_pair_ = _band_pair(series, rho_cma_keys,
                                   rho_province[RHO_CMA_PID]["memberId"], members)
        if shipped_pair_ is None:
            raise ProbeRefusal("rho", f"{table_number(RHO_CMA_PID)} withholds a province cell "
                                      f"for {members} — the shipped curve is not measurable")
        totals = list(shipped_pair_)
        for member in rho_cma_members:
            part = _band_pair(series, rho_cma_keys, member["memberId"], members)
            if part is None:
                raise ProbeRefusal("rho", f"{table_number(RHO_CMA_PID)} withholds "
                                          f"{member['memberNameEn']} for {members}")
            totals = [totals[i] - part[i] for i in range(2)]
        return {"shipped": totals}

    def rho_subtract(band: str) -> tuple:
        """The QC part's ρ counts for one band, FIELD-WISE, with the withheld ones bounded.

        Same discipline as the immigrant side and for the same reason: a subdivision publishes
        its household count while withholding the owner count of the same band, and dropping
        the published one would net a territory's households out of one denominator and its
        owners out of another. The upper bound on a band's withheld mass is the CSD's own
        all-ages total MINUS everything it publishes across all four bands — the unpublished
        remainder, which contains the missing cell by construction.
        """
        low, high, withheld_here = [0, 0], [0, 0], 0
        for member in rho_csd.values():
            geo = member["memberId"]
            total = [series.get(rho_cd_keys[(geo, RHO_AGE_TOTAL, f)]) for f in range(2)]
            if any(t is None for t in total):
                raise ProbeRefusal(
                    "rho", f"{table_number(RHO_CD_PID)} geography {geo} withholds its all-ages "
                           "total, so the unpublished remainder that bounds its withheld bands "
                           "cannot be computed — an interval with no upper end is not one")
            published_all, band_low, band_missing = [0, 0], [0, 0], [False, False]
            for other in RHO_BAND_ORDER:
                for name in RHO_BANDS_CD[other]:
                    for field in range(2):
                        value = series.get(rho_cd_keys[(geo, name, field)])
                        if value is None:
                            if other == band:
                                band_missing[field] = True
                                withheld_here += 1
                            continue
                        published_all[field] += value
                        if other == band:
                            band_low[field] += value
            for field in range(2):
                remainder = max(total[field] - published_all[field], 0)
                low[field] += band_low[field]
                high[field] += band_low[field] + (remainder if band_missing[field] else 0)
        return tuple(low), tuple(high), withheld_here

    def rho_total_subtract() -> tuple:
        """The all-ages row, which every one of these subdivisions publishes in full."""
        low = [0, 0]
        for member in rho_csd.values():
            part = _band_pair(series, rho_cd_keys, member["memberId"], (RHO_AGE_TOTAL,))
            if part is None:
                raise ProbeRefusal("rho", f"{table_number(RHO_CD_PID)} geography "
                                          f"{member['memberId']} withholds its all-ages total")
            low = [low[i] + part[i] for i in range(2)]
        return tuple(low), tuple(low), 0

    rho = {}
    for band in RHO_BAND_ORDER + ("TOTAL",):
        members_cma = (RHO_BANDS_CMA[band] if band != "TOTAL" else (RHO_AGE_TOTAL,))
        shipped_band = rho_band(members_cma)["shipped"]
        low, high, withheld_here = (rho_subtract(band) if band != "TOTAL"
                                    else rho_total_subtract())
        p_shipped = shipped_band[1] / shipped_band[0]
        p_low = (shipped_band[1] - low[1]) / (shipped_band[0] - low[0])
        p_high = (shipped_band[1] - high[1]) / (shipped_band[0] - high[0])
        rho[band] = {"shipped_counts": shipped_band, "shipped": p_shipped, "aligned": p_low,
                     "aligned_bound": p_high, "delta": relative_pct(p_low, p_shipped),
                     "delta_bound": relative_pct(p_high, p_shipped), "withheld": withheld_here}
    band_deltas = [rho[b]["delta"] for b in RHO_BAND_ORDER]
    spread = max(band_deltas) - min(band_deltas)

    # The all-households propensity on the RULED cube, both constructions — the figure #12(B)
    # states (+1.49%) is the whole-CD one, and both are printed so the pair is comparable.
    allhh_shipped = shipped[PC_TOTAL_AGE].owner_propensity
    allhh_aligned = relative_pct(aligned_low[PC_TOTAL_AGE].owner_propensity, allhh_shipped)
    allhh_bracket = relative_pct(bracket["2481"]["cells"][PC_TOTAL_AGE].owner_propensity,
                                 allhh_shipped)

    # ---------------------------------------------------------------- citation coupling
    where[0] = "citation"
    s6 = _spec_s6()
    qfe = _qfe_record()
    # The two figures §6 states that this run does NOT recompute — a person weight and a
    # converted population. They are quoted, so they are READ: a retyped quotation is the
    # hand-written claim this probe family forbids, and a regex that finds nothing refuses.
    stated = {}
    for name, pattern in (("person_weight", r"([\d.]+)% person weight"),
                          ("multiple", r"% person weight — ([\d.]+)×"),
                          ("converted", r"≈([\d,]+) private-household persons")):
        found = re.search(pattern, s6)
        if not found:
            raise ProbeRefusal(
                "citation",
                f"spec §6 no longer states the {name.replace('_', ' ')} this note quotes it "
                f"for (pattern {pattern!r}). A quotation this run cannot read may not be "
                "printed, and the sentence comparing this run's measurement against it would "
                "be comparing against nothing.")
        stated[name] = found.group(1)
    spec_coupled = [
        ("shipped headship", _r4(shipped_pair[0])),
        ("shipped ratio", _r4(shipped_pair[1])),
        ("CD-Gatineau bracket headship", _r4(bracket["2481"]["pair"][0])),
        ("CD-Gatineau bracket ratio", _r4(bracket["2481"]["pair"][1])),
        ("CD Gatineau settled-persons weight", _r2(weights["settled_persons"])),
        ("all-households contamination, whole-CD", _r2(allhh_bracket)),
    ]
    _guard_citation(f"spec §6 ({SPEC.name})", s6, spec_coupled)
    qfe_coupled = [
        ("bracket variant 2 headship", _r4(bracket["2482"]["pair"][0])),
        ("bracket variant 2 ratio", _r4(bracket["2482"]["pair"][1])),
        ("bracket variant 3 headship", _r4(bracket["2480"]["pair"][0])),
        ("bracket variant 3 ratio", _r4(bracket["2480"]["pair"][1])),
        ("shipped settled persons", _n(shipped[PC_SETTLED].persons)),
        ("bracket variant 1 settled persons", _n(bracket["2481"]["cells"][PC_SETTLED].persons)),
        ("bracket variant 2 settled persons", _n(bracket["2482"]["cells"][PC_SETTLED].persons)),
        ("bracket variant 3 settled persons", _n(bracket["2480"]["cells"][PC_SETTLED].persons)),
    ]
    _guard_citation(f"the Task-26 QFE record ({QFE.name})", qfe, qfe_coupled)

    where[0] = "p9"
    p9 = _guard_p9(_p9_note())

    LAST_RUN = {
        "shipped": {"pair": shipped_pair, "cells": shipped},
        "aligned": {"pair": aligned_pair, "headship_band": envelope["headship_band"],
                    "ratio_band": envelope["ratio_band"], "cells": aligned_low,
                    "qc_low": qc_low, "qc_high": qc_high},
        "bracket": {code: v["pair"] for code, v in bracket.items()},
        "encloses": encloses,
        "membership": {"children": len(children), "qc": len(qc_children),
                       "closes": True, "residual": membership_residual,
                       "threshold": bounds["threshold"], "innocent": innocent,
                       "qc_part_pop": qc_part_pop, "isq_pop": isq_gatineau_pop},
        "suppression": {"withheld_cells": len(withheld), "fields": withheld_fields,
                        "csds": len(incomplete), "low": qc_low, "high": qc_high},
        "rho": {band: dict(rho[band]) for band in rho},
        "rho_spread": spread,
        "weights": weights,
        "qc_weights": qc_weights,
        "flow": {"hors_mean": hors_mean, "gatineau_mean": gatineau_mean,
                 "gaps": flow_gaps, "years": flow_years},
        "universe_drift": universe_drift + rho_drift,
        "scanned": scanned,
        "coupled": spec_coupled + qfe_coupled,
    }

    # ======================================================================= the note
    where[0] = "note"
    f_flow_footnote = Fact.cited(flow_book["footnote"],
                                 f"footnote line of the pinned {ISQ_FLOW_BOOK}, verbatim")
    f_gat_label = Fact.cited(str(flow_book["labels"][ISQ_GATINEAU_CODE]).strip(),
                             f"the Gatineau row label in {ISQ_FLOW_BOOK}, verbatim")
    f_hors_label = Fact.cited(str(flow_book["labels"][ISQ_HORS_CODE]).strip(),
                              f"the hors-RMR row label in {ISQ_FLOW_BOOK}, verbatim")
    f_flow_note = Fact.cited(flow_book["note"], f"header note of the pinned {ISQ_FLOW_BOOK}")
    f_p9_verdict = Fact.cited(p9["DECISION-VERDICT"],
                              f"probes/{P9_NOTE.name} DECISION-VERDICT, read this run")
    f_p9_cross = Fact.cited(p9["DECISION-MAINTAINER-CROSS"],
                            f"probes/{P9_NOTE.name} DECISION-MAINTAINER-CROSS, read this run")
    f_p9_residual = Fact.cited(p9["DECISION-RESIDUAL"],
                               f"probes/{P9_NOTE.name} DECISION-RESIDUAL, read this run")
    f_hors_mean = Fact.derived(
        f"{hors_mean:,.0f}", f"{ISQ_FLOW_BOOK} `{ISQ_FLOW_COLUMN}` column {flow_column}, "
                             f"{ISQ_REFERENCE}, mean over {len(flow_years)} published years")
    f_gat_mean = Fact.derived(
        f"{gatineau_mean:,.0f}", f"{ISQ_FLOW_BOOK} `{ISQ_FLOW_COLUMN}`, {ISQ_REFERENCE}, the "
                                 f"Gatineau row's mean over the same years")
    f_qc_pop = Fact.derived(_n(qc_part_pop),
                            f"{table_number(MEMBER_PID)}: the 2021 populations of the "
                            f"{len(qc_children)} Québec-side children, summed")
    f_isq_pop = Fact.derived(_n(isq_gatineau_pop),
                             f"{ISQ_POP_BOOK} code {ISQ_GATINEAU_CODE}, July-1 {ISQ_YEAR}")
    f_membership = Fact.derived(_pct(membership_residual),
                                "the resolved Québec part against the ISQ row it aligns to, "
                                "census total population vs ISQ estimate")
    f_threshold = Fact.derived(_r3(bounds["threshold"]) + "%",
                               f"max innocent |residual| {_r3(bounds['max_abs'])}% "
                               f"({bounds['max_name']}) + {MARGIN_FRACTION:.0%} of it")
    f_headship = Fact.derived(_r4(aligned_pair[0]),
                              "aligned residual: maintainers / persons on the "
                              f"{PC_SETTLED} member")
    f_ratio = Fact.derived(_r4(aligned_pair[1]),
                           "aligned residual: settled owner-maintainer propensity / "
                           "non-immigrant")
    f_headship_move = Fact.derived(_pct(relative_pct(aligned_pair[0], shipped_pair[0])),
                                   "aligned headship against the shipped one")
    f_ratio_move = Fact.derived(_pct(relative_pct(aligned_pair[1], shipped_pair[1])),
                                "aligned ratio against the shipped one")
    f_leg = Fact.derived(
        _pct(relative_pct(aligned_pair[0] * aligned_pair[1],
                          shipped_pair[0] * shipped_pair[1])),
        "the PRODUCT headship × ratio, aligned against shipped — the immigrant demand leg")
    f_allhh = Fact.derived(_pct(allhh_aligned),
                           "all-households owner propensity, aligned against shipped, on "
                           f"{table_number(CMA_PID)}")
    f_allhh_bracket = Fact.derived(_pct(allhh_bracket),
                                   "all-households owner propensity under the whole-CD "
                                   "construction — the figure amendment #12(B) states")
    # ONE measurement, printed in two unit words below: as a spread it is percentage POINTS of
    # relative contamination, and because ED is linear in ρ the same number bounds |ΔED/ED| as
    # a PERCENT of ED. Registering it twice would inflate the count with one measurement.
    f_spread = Fact.derived(f"{spread:.3f}",
                            "max minus min of the four model bands' relative contamination — "
                            "in ED's own relative terms, since ED is linear in ρ")
    f_gat_weight = Fact.derived(_r3(weights["settled_persons"]) + "%",
                                "CD Gatineau's share of the shipped residual's settled persons")
    f_flow_gap = Fact.derived(
        f"{max(abs(g) for g in flow_gaps.values()):.0f}",
        "largest |parts − province| gap across the published years")
    f_children = Fact.derived(
        len(children), "children of the Ottawa-Gatineau CMA member, read from the live "
                       "geography hierarchy")
    f_further_out = Fact.derived(
        sum(1 for v in innocent.values() if abs(v) > abs(membership_residual)),
        "innocent controls whose |residual| exceeds the Québec part |residual|")
    f_empty_complement = Fact.derived(
        sum(1 for _code, (values, bound) in csd_settled.items()
            if any(v is None for v in values) and bound == Cell(0, 0, 0)),
        "QC-part CSDs whose withheld settled counts are bounded at zero — the cube publishes "
        "the same total and non-immigrant count, so the withheld member is empty")
    f_encloses = Fact.derived(
        sum(1 for v in encloses.values() if v),
        "of the two aligned quantities, how many the whole-CD bracket encloses")
    f_local_universe = Fact.derived(
        _r4(local_universe), "the resolved QC part: its census private-household persons / its "
                             "own census total population")
    # The drifting province cells, NAMED with their two values rather than counted. P8 §3a
    # settled the shape: an unscoped bit-identity claim asserts what the run's own data
    # contradicts, so the note states the identity the guards bind and prints what drifts.
    f_drift = Fact.derived(
        "; ".join(f"`{key}` {field} {_n(left)} vs {_n(right)}"
                  for key, field, left, right in universe_drift + rho_drift) or "none measured",
        "every province cell measured as differing across a cube pair, with both values")
    f_drift_size = Fact.derived(
        f"{max((abs(a - b) for _k, _f, a, b in universe_drift), default=0):.0f} in "
        f"{_n(aligned_low[PC_NONIMM].maintainers)}",
        "the largest census-pair province drift against the aligned non-immigrant maintainer "
        "base it enters")

    lines = [
        "## 1. The defect, and the principle that settles it",
        "",
        f"HORS_RMR's immigrant inputs are measured over the Québec province NET of the six "
        f"wholly-QC CMAs. That residual INCLUDES the Québec side of Ottawa-Gatineau; the "
        f"arrival flows it multiplies EXCLUDE it. Amendment #12(A) rules the principle — **the "
        f"rate's territory must match the flow's territory** — and this run builds the residual "
        f"that satisfies it.",
        "",
        "| cube | title (live) | released | archive |",
        "|---|---|---|---|",
    ]
    for pid in pids:
        lines.append(f"| [{table_number(pid)}]({table_url(pid)}) | {metas[pid]['cubeTitleEn']} "
                     f"| {metas[pid]['releaseTime']} | "
                     f"{metas[pid]['archiveStatusEn'].split(' - ')[0]} |")
    lines += [
        "",
        f"`{table_number(CMA_PID)}` is the ruled source (unchanged), `{table_number(CD_PID)}` "
        f"its census-division/subdivision sibling, `{table_number(MEMBER_PID)}` the membership, "
        f"`{table_number(RHO_CMA_PID)}`/`{table_number(RHO_CD_PID)}` the ownership curve's own "
        f"pair, `{table_number(SHELTER_CMA_PID)}`/`{table_number(SHELTER_CD_PID)}` scanned only "
        f"for a Québec-part member, and `{table_number(PARTS_PID)}` the one cube that publishes "
        "that part — with the wrong metric, §3.",
        "",
        "## 2. The ISQ side is SEPARABLE — measured on the OPERAND's own workbook",
        "",
        f"The flows come from `{ISQ_FLOW_BOOK}`, and that workbook publishes "
        f"**{f_gat_label}** as a row of its own, beside **{f_hors_label}**. Its footnote "
        f"{_isq_marker(flow_book['labels'][ISQ_GATINEAU_CODE])!r} reads, verbatim: "
        f"*{f_flow_footnote}* — so the Gatineau row IS the Québec part, by the publisher's own "
        f"statement. Its header note fixes the delineation: *{f_flow_note}*.",
        "",
        f"Quoting is not the measurement. The measurement is that the workbook's own parts "
        f"CLOSE on its province row with Gatineau counted as one of them: over "
        f"{len(flow_years)} published years of `{ISQ_FLOW_COLUMN}` under {ISQ_REFERENCE}, the "
        f"{len(flow_book['values']) - 1} regional rows sum to the province row within "
        f"{f_flow_gap}"
        f" persons/yr (tolerance ±{ISQ_FLOW_TOLERANCE}, the workbook's own rounding step). A "
        f"hors-RMR row that also contained the Gatineau flow would double-count it, and the "
        f"same sum would then overshoot by that row's whole size — a mean of {f_gat_mean} "
        f"permanent immigrants a year against hors-RMR's own {f_hors_mean}. The operand this "
        "rate multiplies is the second number; the rate was measured over a territory carrying "
        "both.",
        "",
        f"({len(flow_book['withheld'])} cells are published as `{ISQ_WITHHELD}` — the terminal "
        "year of a (t)→(t+1) flow table — and are excluded from the closure rather than "
        "coerced to a number.)",
        "",
        "## 3. The census side is INSEPARABLE at CMA grain, and no cube with this cross "
        "publishes the part",
        "",
        f"`{table_number(CMA_PID)}` carries exactly "
        f"{Fact.derived(len(og_named), f'members of {table_number(CMA_PID)} geography whose name contains Gatineau')} "
        f"member naming Gatineau: `{og_cma['memberNameEn']}` (id {og_cma['memberId']}, "
        f"geoLevel {og_cma['geoLevel']}, SGC {og_cma['classificationCode']}), parented to "
        f"`{ontario['memberNameEn']}` (id {ontario['memberId']}). It is therefore NOT one of the "
        f"six geoLevel-{CMA_GEO_LEVEL} children of Quebec the residual is taken net of, and its "
        "Québec half stays inside the residual. There is no Québec-part member to subtract at "
        "this grain.",
        "",
        f"That absence is scoped to a search, not assumed. P9 closed the catalogue at "
        f"**{f_p9_verdict}** and named the cubes carrying the maintainer × population-"
        f"characteristics cross: **{f_p9_cross}**. This run scanned all four for a geography "
        f"member at CMA-PART grain (geoLevel {CMA_PART_GEO_LEVEL}) or a name carrying both "
        "`Gatineau` and `part`:",
        "",
        "| cube | geography members | at geoLevel 505 or named as a part |",
        "|---|---:|---:|",
    ]
    for pid in MAINTAINER_CROSS:
        lines.append(f"| {table_number(pid)} | {len(geo_members(pid)):,} | "
                     f"{len(scanned[pid])} |")
    lines += [
        "",
        f"`{table_number(PARTS_PID)}` DOES publish it — `{parts_member['memberNameEn']}` (id "
        f"{parts_member['memberId']}, geoLevel {parts_member['geoLevel']}, SGC "
        f"{parts_member['classificationCode']}) — and cannot serve, for a reason that is "
        f"structural rather than a preference: its dimensions are "
        + ", ".join(f"`{n}`" for n in parts_dims)
        + ". There is no household-maintainer axis and no count statistic; its indicators are "
        "person-weighted PERCENTAGES. Reading the ruled maintainer-denominated quantities off "
        "it is exactly the metric transport ruling S eliminated. The residual carried from "
        f"P9 stands: **{f_p9_residual}**.",
        "",
        "## 4. The EXACT construction — the CMA's Québec-part membership, resolved live",
        "",
        f"`{table_number(MEMBER_PID)}` publishes `{og_member['memberNameEn']}` (id "
        f"{og_member['memberId']}, geoLevel {og_member['geoLevel']}, SGC "
        f"{og_member['classificationCode']}) with its constituent census subdivisions as "
        f"geography-dimension CHILDREN. This run reads that hierarchy: "
        f"{f_children} "
        f"children, of which "
        f"{Fact.derived(len(qc_children), 'children whose SGC classification code begins 24 (Québec)')} "
        "carry a Québec SGC code.",
        "",
        "### 4a. The children CLOSE on the CMA — a partition, not a selection",
        "",
        f"Their {ISQ_YEAR} populations sum to "
        f"{Fact.derived(_n(sum(child_pop.values())), 'the CMA children populations, summed')} "
        f"against the CMA's own {_n(series[pop_keys[og_member['memberId']]])} — EXACTLY, with "
        f"the Québec side at {f_qc_pop} and the Ontario side at "
        f"{_n(sum(child_pop.values()) - qc_part_pop)}. A child list that did not close on its "
        "own parent would make 'the Québec side of this CMA' a slice of an unknown whole, so "
        "the closure is a refusal condition rather than a remark.",
        "",
        "### 4b. The Québec side, selected TWO ways",
        "",
        f"Every one of the {len(children)} children is classified twice: by SGC prefix "
        f"(`{QC_SGC_PREFIX}`) and STRUCTURALLY, by asking whether the same SGC code resolves in "
        f"`{table_number(CD_PID)}` to a geoLevel-{CSD_GEO_LEVEL} member whose census division is "
        f"a child of that cube's Quebec member. The two agree on all {len(children)}; a "
        "disagreement refuses the run. The Québec-side members, with the census division each "
        "sits in — four different CDs, which is why no whole-CD union is this territory:",
        "",
        "| SGC | census subdivision | census division | settled persons | maintainers | "
        "owner-maintainers |",
        "|---|---|---|---:|---:|---:|",
    ]
    for code, member in csd_members.items():
        parent = cd_parent[member["parentMemberId"]]
        values, bound = csd_settled[code]
        # NOT `bounds`: that name holds the derived membership threshold in this scope, and
        # shadowing it here made the note read a tuple as a dict — caught by the first run.
        bound_fields = (bound.persons, bound.maintainers, bound.owner_maintainers)
        shown = " | ".join(_n(v) if v is not None else f"withheld (≤ {_n(b)})"
                           for v, b in zip(values, bound_fields, strict=True))
        lines.append(f"| {code} | {member['memberNameEn']} | {parent['memberNameEn']} "
                     f"(SGC {parent['classificationCode']}) | {shown} |")
    lines += [
        "",
        f"`Gatineau` names both the census division (id {cd_members['2481']['memberId']}, "
        f"geoLevel {CD_GEO_LEVEL}) and the city inside it (id "
        f"{csd_members['2481017']['memberId']}, geoLevel {CSD_GEO_LEVEL}); every member above "
        "is resolved by SGC code AND geoLevel, so the two cannot be confused.",
        "",
        "### 4c. The membership, VALIDATED against the ISQ row it aligns to",
        "",
        f"Like-for-like, with no universe transport inside the comparison: census TOTAL "
        f"population against ISQ's TOTAL-population estimate. The resolved Québec part measures "
        f"{f_qc_pop} against ISQ's {f_isq_pop} — **{f_membership}**. The threshold is DERIVED "
        f"from innocent controls measured the same way, the six wholly-QC CMAs, whose "
        f"census↔ISQ identity is not in question:",
        "",
        "| geography | census 2021 population | ISQ July-1 2021 | residual |",
        "|---|---:|---:|---:|",
    ]
    for member in cma_members:
        code = str(member["classificationCode"])
        lines.append(f"| {member['memberNameEn']} | {_n(census_cma_pop[code])} | "
                     f"{_n(pop_book['values'][int(code)])} | "
                     f"{_pct(innocent[member['memberNameEn']])} |")
    lines += [
        f"| **Ottawa-Gatineau, Québec part (resolved here)** | **{_n(qc_part_pop)}** | "
        f"**{_n(isq_gatineau_pop)}** | **{_pct(membership_residual)}** |",
        "",
        f"The largest innocent |residual| is {_r3(bounds['max_abs'])}% ({bounds['max_name']}); "
        f"the threshold is that maximum plus {MARGIN_FRACTION:.0%} of it — {f_threshold}. The "
        f"resolved part sits INSIDE the innocent spread, not merely under the bound: "
        f"{f_further_out} "
        f"of the {len(innocent)} controls are further out. What this gate CANNOT do: it bounds "
        "the census-vs-July-1 estimate gap, it does not separate that gap from a membership "
        f"difference — a difference smaller than "
        f"{_n(bounds['threshold'] / 100 * isq_gatineau_pop)} persons would not be seen.",
        "",
        "### 4d. Suppression: BOUNDED by the published complement, never dropped",
        "",
        f"StatCan withholds small counts. Of the cells this run requested at the "
        f"{len(csd_members)} Québec-part CSDs, "
        f"{Fact.derived(len(withheld), 'requested cells withheld across the QC-part CSDs, both cubes')} "
        f"came back with no data point, at "
        f"{Fact.derived(len(withheld_csds), 'QC-part CSDs carrying at least one withheld cell, EITHER cube — the union those cells actually fall at')} "
        f"of the {len(csd_members)} subdivisions — all of them rural municipalities of at most "
        f"{_n(max((child_pop[c] for c in withheld_csds), default=0))} people. Two cubes are "
        f"involved and each bounds its own, so the split is stated rather than averaged: "
        f"{Fact.derived(withheld_by_cube[RHO_CD_PID], f'withheld maintainer-age cells at the QC-part CSDs of {table_number(RHO_CD_PID)}')} "
        f"of them are maintainer-age cells of `{table_number(RHO_CD_PID)}`, bounded against "
        f"that cube's own unpublished remainder and carried in §6. The other "
        f"{Fact.derived(withheld_by_cube[CD_PID], f'withheld settled-immigrant cells at the QC-part CSDs of {table_number(CD_PID)}')} "
        f"are settled-immigrant counts of `{table_number(CD_PID)}`, and they are neither "
        "dropped nor guessed: each is bounded above by a quantity the same cube DOES publish at "
        f"the same geography — `{PC_TOTAL_AGE}` minus `{PC_NONIMM}`, i.e. all immigrants and "
        "non-permanent residents together, which contains `Before 2016` by construction. Both "
        "ends are carried to the published figures:",
        "",
        f"- FIELD-WISE, because these subdivisions publish some counts of a triple and "
        f"withhold others: {Fact.derived(withheld_fields, 'withheld FIELDS across the settled triples of the QC-part CSDs')} "
        f"of the {3 * len(csd_members)} settled-member counts are withheld, at "
        f"{Fact.derived(len(incomplete), 'QC-part CSDs whose settled-immigrant triple is incomplete')} "
        f"subdivisions, and a published count is never discarded because a "
        "neighbouring one in the same triple is missing.",
        f"- subtracting only what is published: persons {_n(qc_low.persons)}, maintainers "
        f"{_n(qc_low.maintainers)}, owner-maintainers {_n(qc_low.owner_maintainers)};",
        f"- suppressed at their bound (subtract most): persons {_n(qc_high.persons)}, "
        f"maintainers {_n(qc_high.maintainers)}, owner-maintainers "
        f"{_n(qc_high.owner_maintainers)};",
        f"- the complement is EMPTY at "
        f"{f_empty_complement} "
        f"of those {len(incomplete)}, where the cube publishes the same total and "
        f"non-immigrant count: the "
        f"withheld cell is bounded at zero rather than estimated, and "
        f"{_plural(clamped_fields, 'field')} needed the bound clamped for a rounding-step "
        "negative.",
        f"- the whole withheld mass is at most "
        f"{Fact.derived(_n(qc_high.persons - qc_low.persons), 'the bounded withheld persons: subtract-most minus subtract-least')} "
        f"persons against a subtraction of {_n(qc_low.persons)}.",
        "",
        "## 5. The ALIGNED immigrant inputs",
        "",
        "| construction | settled persons | maintainers | owner-maintainers | HEADSHIP | "
        "non-imm propensity | settled propensity | RATIO |",
        "|---|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for label, cells in (("as shipped (Gatineau IN)", shipped),
                         ("**ALIGNED (published counts only)**", aligned_low),
                         ("aligned, every withheld field at its bound", aligned_high)):
        settled = cells[PC_SETTLED]
        lines.append(
            f"| {label} | {_n(settled.persons)} | {_n(settled.maintainers)} | "
            f"{_n(settled.owner_maintainers)} | {_r4(settled.headship)} | "
            f"{_r4(cells[PC_NONIMM].owner_propensity)} | {_r4(settled.owner_propensity)} | "
            f"{_r4(ownership_ratio(settled, cells[PC_NONIMM]))} |")
    lines += [
        "",
        f"**Aligned: headship {f_headship} ({f_headship_move}), ratio {f_ratio} "
        f"({f_ratio_move}) — and the ratio CROSSES 1.0.** Shipped, settled immigrants under-own "
        f"in hors-RMR; aligned, they out-own. The suppression bound is "
        f"[{_r4(envelope['ratio_band'][0])}, {_r4(envelope['ratio_band'][1])}] on the ratio "
        f"and [{_r4(envelope['headship_band'][0])}, {_r4(envelope['headship_band'][1])}] on the "
        f"headship — taken at the box's OPPOSITE corners (headship is largest when the fewest "
        f"maintainers and the most persons are netted out), so it is the envelope rather than "
        f"the two sums paired. Neither end straddles 1.0, which is a refusal condition here, "
        "not a remark.",
        "",
        f"The two factors MULTIPLY inside D_imm and their errors are same-signed, so the "
        f"immigrant demand leg moves by their product: **{f_leg}**. The contaminant's size "
        f"explains it — CD Gatineau alone holds {f_gat_weight} of the shipped residual's "
        f"settled persons against a {_r3(weights['persons'])}% share of its persons and "
        f"{_r3(weights['maintainers'])}% of its maintainers, and the full Québec part holds "
        f"{_r3(qc_weights['settled_persons'])}% against {_r3(qc_weights['persons'])}%. "
        f"(Amendment #12(A) states that weight's counterpart as "
        f"{Fact.cited(stated['person_weight'] + '%', 'spec §6 amendment #12(A), verbatim')} — "
        f"a PERSON weight — and draws "
        f"{Fact.cited(stated['multiple'] + '×', 'spec §6 amendment #12(A), verbatim')} from it. "
        f"Measured, {stated['person_weight']}% is CD Gatineau's MAINTAINER "
        f"share ({_r3(weights['maintainers'])}%); its person share is "
        f"{_r3(weights['persons'])}%. The concentration multiple is "
        f"{Fact.derived(_r2(weights['settled_persons'] / weights['maintainers']), 'settled-persons weight over the MAINTAINER weight')}× "
        f"at the maintainer denominator and "
        f"{Fact.derived(_r2(weights['settled_persons'] / weights['persons']), 'settled-persons weight over the PERSON weight')}× "
        f"at the person one, so the amendment's figure is the maintainer reading; the label is "
        "what moves, and the multiple with it.)",
        "",
        "### 5a. The whole-CD bracket, as sensitivity — and what it would have missed",
        "",
        "The mandate's fallback construction, measured anyway, because a bracket is only "
        "judgeable beside the exact value it was standing in for:",
        "",
        "| construction | settled persons | HEADSHIP | RATIO | residual persons |",
        "|---|---:|---:|---:|---:|",
    ]
    running_label = []
    for code, entry in bracket.items():
        running_label.append(f"{entry['label']} ({code})")
        lines.append(
            f"| − {' + '.join(running_label)} | {_n(entry['cells'][PC_SETTLED].persons)} | "
            f"{_r4(entry['pair'][0])} | {_r4(entry['pair'][1])} | "
            f"{_n(entry['cells'][PC_TOTAL_AGE].persons)} |")
    lines += [
        f"| **exact (16 QC-part CSDs)** | **{_n(aligned_low[PC_SETTLED].persons)}** | "
        f"**{_r4(aligned_pair[0])}** | **{_r4(aligned_pair[1])}** | "
        f"**{_n(aligned_low[PC_TOTAL_AGE].persons)}** |",
        "",
        f"The whole-CD range is {_r4(min(bracket_headship))}-{_r4(max(bracket_headship))} on "
        f"the headship and {_r4(min(bracket_ratio))}-{_r4(max(bracket_ratio))} on the ratio, "
        f"and it contains "
        f"{f_encloses} "
        f"of the two exact values: the ratio "
        f"{'INSIDE' if encloses['ratio'] else 'OUTSIDE'}, the headship "
        f"{'INSIDE' if encloses['headship'] else 'OUTSIDE'}. That is the measured cost of the "
        "fallback construction, and it is structural rather than bad luck: FOUR census "
        "divisions contribute to this CMA and no union of whole ones equals it — CD Gatineau "
        "entire, all seven Les Collines municipalities, seven of Papineau's and one of La "
        "Vallée-de-la-Gatineau's — so every whole-CD variant is simultaneously short of one "
        "territory and long on another.",
        "",
        "## 6. The OWNERSHIP propensity — SIZED here, and not corrected",
        "",
        f"Amendment #12(B) rules that HORS_RMR's ownership propensity ρ is NOT to be corrected, "
        f"because a BAND-UNIFORM relative scaling of ρ cancels exactly in ED. The premise is "
        f"the claim, so it is measured — in the model's OWN lattice: `{table_number(RHO_CMA_PID)}` "
        f"is the cube `loaders/census.py` reads, `{table_number(RHO_CD_PID)}` is its CSD "
        f"sibling (province rows bit-identical on the all-ages pair of counts the guard binds — "
        f"the four bands are subtracted across the pair too, at granularities the two cubes do "
        f"not share, and are NOT asserted identical), and the bands are "
        f"`census._AGE_BAND_SPEC`'s own.",
        "",
        "| band | shipped households | shipped owners | ρ shipped | ρ aligned | relative Δ | "
        "withheld CSD cells |",
        "|---|---:|---:|---:|---:|---:|---:|",
    ]
    for band in ("TOTAL",) + RHO_BAND_ORDER:
        entry = rho[band]
        label = "all ages" if band == "TOTAL" else band
        delta = Fact.derived(_pct(entry["delta"]),
                             f"{label}: aligned ρ against shipped ρ, "
                             f"{table_number(RHO_CMA_PID)} residual net of the QC-part CSDs")
        lines.append(
            f"| {label} | {_n(entry['shipped_counts'][0])} | {_n(entry['shipped_counts'][1])} | "
            f"{_r4(entry['shipped'])} | {_r4(entry['aligned'])} | **{delta}** | "
            f"{entry['withheld']} |")
    lines += [
        "",
        f"**The band-uniform premise is FALSE in its strict form.** The four model bands' "
        f"relative contamination runs "
        f"{_pct(min(band_deltas))} to {_pct(max(band_deltas))} — a spread of {f_spread} "
        f"percentage points, wider than the aggregate move itself. The shape is two-valued rather than noisy: 25-54 "
        f"carries {_pct(rho['25-54']['delta'])} and the three older bands sit within "
        f"{max(rho[b]['delta'] for b in ('55-64', '65-74', '75+')) - min(rho[b]['delta'] for b in ('55-64', '65-74', '75+')):.3f} "
        "pp of each other.",
        "",
        f"**What that costs in ED is NOT bounded by the spread, and the arithmetic says so.** "
        f"ED = (D − S) / OwnerStock is linear in ρ through every term, so under "
        f"ρ(a) → ρ(a)(1 + δ(a)) each of D, S and OwnerStock picks up its own ρ-weighted mean of "
        f"δ and, to first order,",
        "",
        "```",
        "ΔED / ED  =  (δ_S − δ_OS)  +  (δ_D − δ_S) × D / (D − S)",
        "```",
        "",
        f"The first term lies inside the spread. The SECOND does not: its multiplier is GROSS "
        f"demand over NET excess demand, which is ≈1 only when |ED| is of the order of D / "
        f"OwnerStock and grows without bound as the two flows approach balance — the regime this "
        f"module exists to measure. The numerator of ED is a DIFFERENCE of flows, so a "
        f"band-varying δ is amplified there rather than averaged. What the spread does bound is "
        f"the ABSOLUTE move, |ΔED| ≤ {f_spread}% × (D + |D − S|) / OwnerStock: second-order in "
        f"ED's own units, and of unmeasured size RELATIVE to ED. Both D/(D−S) and D/OwnerStock "
        "are outputs of the model, not quantities this probe reads, so it prices what it can "
        "and names what it cannot.",
        "",
        f"**And the structure is adverse rather than neutral.** The band carrying the largest "
        f"contamination is 25-54 at {_pct(rho['25-54']['delta'])} — the band D_native is built "
        f"from — while S rides ρ(75+) through `initialize_households`, the band carrying the "
        f"SMALLEST at {_pct(rho['75+']['delta'])}. So δ_D − δ_S sits at or near the FULL spread "
        f"rather than near zero, which is the worst arrangement of these four numbers for a "
        f"cancellation argument. **This run therefore does not certify #12(B)'s cost/signal "
        f"conclusion: it measures the premise FALSE and leaves the consequence sized by a "
        f"quantity outside the probe.** (The withheld CSD cells reach only the 75+ band: "
        f"netting them out at their bound instead of leaving them unsubtracted moves that "
        f"band's Δ to {_pct(rho['75+']['delta_bound'])}, widening the spread to "
        f"{max(band_deltas) - min(min(band_deltas), rho['75+']['delta_bound']):.3f} — the "
        "adverse structure holds at either end.)",
        "",
        f"Two things this measurement does NOT say. It does not say the correction is "
        f"unavailable: `{table_number(RHO_CD_PID)}` publishes the identical cross at CSD grain, "
        f"so an aligned ρ curve is extractable — what #12(B) rules is a cost/signal judgment, "
        f"not a data absence. And it does not say the SIGN is safe: for D < S — the decline "
        f"regime the module exists to measure — the multiplier D/(D − S) is NEGATIVE and "
        f"unbounded near balance, so a large enough amplification carries ED across zero rather "
        f"than merely rescaling it, and rank 1 is the most negative mean ED. How close to "
        f"balance this geography runs is the model's output, not this probe's. Two cubes, one "
        f"answer: the RULED cube "
        f"({table_number(CMA_PID)}, maintainer counts crossed with population characteristics "
        f"rather than with maintainer age) puts the same all-households contamination at "
        f"{f_allhh}, against the {_pct(rho['TOTAL']['delta'])} the ownership pair measures — "
        f"different cubes, different crosses, the same figure. Under the whole-CD construction "
        f"it reads {f_allhh_bracket}, which is the figure amendment #12(B) states.",
        "",
        "## 7. Rounding, and the universe conversion, stated as approximations",
        "",
        f"- Every count is rounded to 5 INDEPENDENTLY per cube. The aligned settled triple is a "
        f"province cell minus {len(cma_members)} CMA cells minus {len(csd_members)} CSD cells — "
        f"{Fact.derived(1 + len(cma_members) + len(csd_members), 'independently rounded cells entering one aligned figure')} "
        f"rounded cells — so its worst-case rounding envelope is ±"
        f"{Fact.derived((1 + len(cma_members) + len(csd_members)) * 2, 'worst-case stacked rounding, ±2 per rounded cell')} "
        f"persons against a base of {_n(aligned_low[PC_SETTLED].persons)}: it moves the fourth "
        "decimal of the headship at most.",
        f"- The two cube pairs' province rows are bit-identical on the RULED `{PC_SETTLED}` "
        f"triple and on the ρ pair's all-ages row — the two identities the guards bind, not a "
        f"cube-wide one — and differ by no more than {_ROUNDING_TOLERANCE} elsewhere: "
        f"{_plural(len(universe_drift) + len(rho_drift), 'province cell')} measured as "
        f"drifting ({f_drift}), and anything beyond the rounding step refuses the run. "
        f"That cell is not idle: the aligned non-immigrant propensity subtracts a "
        f"`{PC_NONIMM}` triple across the pair too, so the drift enters its denominator — at "
        f"{f_drift_size} it moves neither published figure at the resolution this note "
        "prints them.",
        f"- The private-household universe is NOT the total-population universe, and the "
        f"conversion between them is an approximation that carries a LOCAL error. This tree "
        f"records the province-wide gap in `loaders/constants.py` — "
        f"{Fact.cited(f'{recorded_ph:,} private-household persons vs {recorded_total:,} published', 'loaders/constants.py, verbatim')} "
        f"— i.e. a ratio of "
        f"{Fact.derived(_r4(province_universe), 'the two counts loaders/constants.py records, divided')}, "
        f"at which ISQ's Gatineau row converts to {_n(converted)}; §6 states that conversion as "
        f"{Fact.cited('≈' + stated['converted'], 'spec §6 amendment #12(A), verbatim')}. The "
        f"resolved membership actually carries "
        f"{_n(qc_totals[PC_TOTAL_AGE].persons)} private-household persons — a LOCAL ratio of "
        f"{f_local_universe}, "
        f"above the province-wide one, which is why the measured figure lands above both "
        f"conversions. Nothing in this note's arithmetic uses the conversion: the membership is "
        "validated total-against-total in §4c and the immigrant inputs are private-household on "
        "both sides of every ratio.",
        "",
        "## 8. Scope",
        "",
        f"MEASURE ONLY. This run wires nothing: `demand/immigrant_inputs.py`, "
        f"`probes/run_p8.py` and `probes/P8-immigrant-inputs.md` are untouched, and the ruled "
        f"§6 table still carries {_r4(shipped_pair[0])} / {_r4(shipped_pair[1])} — which this "
        "run RECOMPUTES and agrees with, as the contaminated measurement it is. The aligned "
        "values above are proposed to the spec, not applied: P8's note is citation-coupled to "
        "§6's stated figures, so wiring before the ruling would couple that note to numbers the "
        "ruling no longer carries. Order: spec, then P8, then the join table.",
        "",
        "## DECISION",
        "",
        "- `DECISION-VERDICT: MEASURED`",
        f"- `DECISION-CONSTRUCTION: EXACT — province {table_number(CMA_PID)} NET of the six "
        f"geoLevel-{CMA_GEO_LEVEL} children of Quebec, NET of the {len(qc_children)} "
        f"Québec-side census subdivisions of {table_number(MEMBER_PID)}'s Ottawa-Gatineau CMA "
        f"member {og_member['memberId']}, read from {table_number(CD_PID)}`",
        f"- `DECISION-ALIGNED-HEADSHIP: {_r4(aligned_pair[0])} "
        f"(envelope {_r4(envelope['headship_band'][0])}-"
        f"{_r4(envelope['headship_band'][1])}) — shipped {_r4(shipped_pair[0])}, "
        f"{_pct(relative_pct(aligned_pair[0], shipped_pair[0]))}`",
        f"- `DECISION-ALIGNED-RATIO: {_r4(aligned_pair[1])} "
        f"(envelope {_r4(envelope['ratio_band'][0])}-{_r4(envelope['ratio_band'][1])}) — "
        f"shipped {_r4(shipped_pair[1])}, "
        f"{_pct(relative_pct(aligned_pair[1], shipped_pair[1]))}, CROSSES 1.0`",
        f"- `DECISION-IMMIGRANT-LEG: headship × ratio moves {_pct(relative_pct(aligned_pair[0] * aligned_pair[1], shipped_pair[0] * shipped_pair[1]))} "
        f"— D_imm understated by that much at HORS_RMR, ED understated, rank-1-is-most-negative "
        "so the geography is ranked MORE RISKY than truth`",
        "- `DECISION-BRACKET: "
        + "; ".join(f"−{entry['label']} {_r4(entry['pair'][0])}/{_r4(entry['pair'][1])}"
                    for entry in bracket.values())
        + f" — cumulative whole-CD variants, published as sensitivity`",
        f"- `DECISION-BRACKET-ENCLOSURE: ratio "
        f"{'INSIDE' if encloses['ratio'] else 'OUTSIDE'}, headship "
        f"{'INSIDE' if encloses['headship'] else 'OUTSIDE'} the whole-CD range — the fallback "
        "construction could not have produced the exact headship`",
        f"- `DECISION-MEMBERSHIP: {len(children)} children close EXACTLY on the CMA "
        f"({_n(sum(child_pop.values()))}); {len(qc_children)} Québec-side by SGC prefix AND by "
        f"census-tree ancestry, agreeing on all {len(children)}`",
        f"- `DECISION-MEMBERSHIP-GATE: PASS — resolved Québec part {_n(qc_part_pop)} vs ISQ "
        f"{_n(isq_gatineau_pop)} = {_pct(membership_residual)}, against a threshold of "
        f"{_r3(bounds['threshold'])}% derived from the six wholly-QC CMAs "
        f"({_r3(bounds['max_abs'])}% max innocent + {MARGIN_FRACTION:.0%})`",
        f"- `DECISION-SUPPRESSION: {len(withheld)} withheld cells across both cubes; "
        f"{withheld_fields} of the {3 * len(csd_members)} settled-member counts, at "
        f"{len(incomplete)}/{len(csd_members)} QC-part CSDs; bounded FIELD-WISE above by "
        f"{PC_TOTAL_AGE} − {PC_NONIMM} at the same geography; envelope width "
        f"{envelope['ratio_band'][1] - envelope['ratio_band'][0]:.4f} on the ratio and "
        f"{envelope['headship_band'][1] - envelope['headship_band'][0]:.4f} on the headship, "
        "straddling nothing`",
        f"- `DECISION-ISQ-SEPARABILITY: MEASURED — {ISQ_FLOW_BOOK}'s "
        f"{len(flow_book['values']) - 1} regional rows close on its province row within "
        f"{max(abs(g) for g in flow_gaps.values()):.0f}/yr with Gatineau as its own row "
        f"(footnote: {flow_book['footnote']}); hors-RMR mean {hors_mean:,.0f} vs the Gatineau "
        f"row's {gatineau_mean:,.0f} permanents/yr`",
        f"- `DECISION-CENSUS-INSEPARABILITY: {table_number(CMA_PID)} member "
        f"{og_cma['memberId']} `{og_cma['memberNameEn']}` at geoLevel {og_cma['geoLevel']} "
        f"parented to {ontario['memberNameEn']} ({ontario['memberId']}) — one member, no "
        "Québec-part member at this grain`",
        f"- `DECISION-DIRECT-SOURCE: ABSENT across the {len(MAINTAINER_CROSS)} maintainer-cross "
        f"cubes P9 names, scanned live at geoLevel {CMA_PART_GEO_LEVEL} and by name; "
        f"{table_number(PARTS_PID)} publishes the part (member {parts_member['memberId']}) but "
        "carries no maintainer axis and person-weighted percentages — the metric transport "
        "ruling S eliminated`",
        f"- `DECISION-RHO-CONTAMINATION: all ages {_pct(rho['TOTAL']['delta'])}; "
        + "; ".join(f"{b} {_pct(rho[b]['delta'])}" for b in RHO_BAND_ORDER)
        + f" — {table_number(RHO_CMA_PID)}/{table_number(RHO_CD_PID)} in "
          f"census._AGE_BAND_SPEC's own bands`",
        f"- `DECISION-RHO-VERDICT: the band-uniform premise of #12(B) is FALSE as measured — "
        f"spread {spread:.3f} pp across the four model bands, all same-signed and "
        f"STRUCTURALLY ADVERSE (D_native's 25-54 band carries {_pct(rho['25-54']['delta'])}, "
        f"the 75+ band S rides carries {_pct(rho['75+']['delta'])}, so δ_D − δ_S is at or near "
        f"the full spread). ΔED/ED = (δ_S − δ_OS) + (δ_D − δ_S)·D/(D−S): second-order in ED's "
        f"OWN UNITS (|ΔED| ≤ {spread:.3f}% × (D + |D−S|)/OwnerStock) but amplified RELATIVE to "
        f"ED by D/(D−S), a model output this probe does not read and which grows without bound "
        f"near flow balance. This run PRICES the contamination and does NOT certify #12(B)'s "
        f"cost/signal conclusion. NOT a data absence: {table_number(RHO_CD_PID)} publishes the "
        "aligned curve's other half`",
        f"- `DECISION-CATALOGUE-CLOSURE: {p9['DECISION-VERDICT']} at "
        f"{p9['DECISION-CLOSURE-LEVEL']} — residual {p9['DECISION-RESIDUAL']} (read from "
        f"probes/{P9_NOTE.name} this run, not restated)`",
        "- `DECISION-SCOPE: MEASURE ONLY — nothing wired by this run; demand/immigrant_inputs.py"
        ", probes/run_p8.py and probes/P8-immigrant-inputs.md are untouched and the aligned "
        "values are NOT in the spec. Spec ruling first, then P8's regeneration, then the join "
        "table`",
        "",
    ]
    return lines


def _failure_sections(exc: BaseException, where: list) -> list:
    boundary = getattr(exc, "boundary", where[0])
    return [
        "## DECISION",
        "",
        "- `DECISION-VERDICT: UNKNOWN-PROBE-FAILED`",
        f"- `LIVE PROBE FAILED: {type(exc).__name__}: {exc}`",
        f"- `LIVE PROBE FAILED-AT: {boundary}`",
        "",
        "## What this run could not establish",
        "",
        "This probe resolves the Ottawa-Gatineau Québec part and measures the aligned immigrant "
        "inputs over a residual that excludes it. This run did not complete that derivation, so "
        "it decides NOTHING: no aligned headship, no aligned ratio, no membership verdict, no "
        "ownership size. A partial derivation is not a weaker measurement — it is no "
        "measurement, and recording it as one would be the cheap all-clear this probe family "
        "exists to refuse.",
        "",
    ]


def main() -> None:
    log = new_run()
    where = ["unattributed"]
    try:
        sections = _sections(where)
    except Exception as exc:              # noqa: BLE001 — every failure is RECORDED, never lost
        sections = _failure_sections(exc, where)
    header = provenance_header(log, written_by=_WRITTEN_BY, scope=_SCOPE, summary=_summary,
                               cited_label=_CITED_LABEL)
    text = "\n".join([_TITLE, ""] + header + sections) + "\n"
    if "[FILL" in text:
        raise ProbeRefusal("note", "an unresolved [FILL placeholder survived into the note — "
                                   "every value must be COMPUTED, never invited")
    OUT.write_text(text, encoding="utf-8")
    print("\n".join(line for line in sections
                    if line.startswith(("- `DECISION", "- `LIVE PROBE"))))


if __name__ == "__main__":
    main()
