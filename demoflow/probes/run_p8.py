"""P8 — the IMMIGRANT INPUTS (headship AND ownership ratio), measured under rulings S and T.

Writes `probes/P8-immigrant-inputs.md`. This module is that file's only writer. The name
carries BOTH quantities on purpose: ruling S binds a PAIR drawn from ONE member of ONE cube,
and a note titled for headship alone invites the ratio to be sourced elsewhere — which is the
exact transport ruling S eliminated.

WHAT IS MEASURED, exactly. From StatCan 98-10-0621-01 (CMA grain) and its census-division
sibling 98-10-0622-01, on the `Before 2016` member of `Population characteristics (46)`:

    headship(g)   = maintainers(g) / persons(g)
    propensity(g) = owner-maintainers(g) / maintainers(g)
    ratio(g)      = propensity(g, Before 2016) / propensity(g, Non-immigrants)

`maintainers` is the `Person is primary household maintainer` member; `persons` is the
`Total - Household maintainer` member, which is ALL persons rather than a maintainer subset —
the two names are adjacent and reading them the other way inverts headship. HORS_RMR is the
province NET of the six wholly-QC CMAs, computed here from published counts rather than
borrowed, and the six are resolved STRUCTURALLY as the geoLevel-503 children of the Quebec
member — never from a typed id list (the pinned ids are asserted, not trusted).

HORS_RMR'S TERRITORY IS THE OPERAND-ALIGNED ONE (spec §6 amendment #13). That residual alone
still carries the Québec side of Ottawa-Gatineau — parented to Ontario in this cube, so not
one of the six — while the ISQ arrival flows it multiplies exclude it, and the rate's
territory must match the flow's. So the residual is netted a second time, against the
Ottawa-Gatineau CMA's Québec-side census subdivisions, FIELD-WISE, with every withheld count
carried at a bound the same cube publishes (`Total - Age` − `Non-immigrants`) rather than
dropped. The MEMBERSHIP — which subdivisions — is carried BY REFERENCE from P10, the
committed probe that derived it, and every member of it is then RE-RESOLVED live here by SGC
code and geoLevel with its Québec-side property re-established two ways in this cube. The
VALUES are recomputed from live cells and checked digit for digit against BOTH documents that
state them: the rates against §6 (ruling S's table row for the superseded construction,
amendment #13 for the ruled one) and the aligned counts against P10's note. The superseded
pair stays in the note under its own name, because §6 still states it and §2a's per-member
readings are properties of that same territory.

RULING T'S TERRITORY GATE, AND WHY THIS PROBE DOES NOT IMPLEMENT ITS ORIGINAL WORDING. As
first written, the gate compared 98-10-0622-01's population against the ISQ RA total. That is
a PRIVATE-HOUSEHOLD count against a TOTAL-POPULATION estimate, and measured it trips at both
RAs AND at the Québec province CONTROL, where territory identity is not in question. A gate
that fires where the answer is known is measuring the universe gap, not the territory.
Amendment #11 REFUSED that construction and replaced it with the one implemented here:

    residual(g) = ( part_census(g)/total_census  ÷  part_ISQ(g)/total_ISQ ) − 1

Each geography's share of its OWN source's provincial total, census against ISQ. A uniform
universe offset cancels by construction (`share_residual_pct` is tested on exactly that
property), so what remains is the geography-VARYING part — which is what a territory
mismatch would look like, and also what a region-varying universe gap looks like. The two are
not separated by arithmetic, so the THRESHOLD is calibrated on INNOCENT CONTROLS measured in
the same construction: the six wholly-QC CMAs, whose ISQ rows join the census rows on an
EXACT classification code and whose identity is therefore not in question. Their spread IS
the size of the geography-varying component at this grain, and the threshold is their maximum
plus a stated margin. The 1% of ruling T's original wording is deliberately NOT used: it was
calibrated against the refuted construction's semantics and does not transfer.

WHAT THE GATE CANNOT DO, stated beside what it can. Passing does not establish that a census
division IS an ISQ région administrative. It fails to REFUTE that at a resolution set by the
innocent spread — roughly one and a half percent of a population share. A mismatch of a few
thousand people sits well inside that and would not be seen. The note publishes the
resolution beside the verdict so a reader can price the claim rather than take it.

THE FLOOR GUARDS (each raises `ProbeRefusal`, which routes the run to UNKNOWN-PROBE-FAILED —
never to a friendlier verdict; the name is deliberately NOT network-class so a refusal cannot
launder itself into `pytest.skip`):

  * `_guard_meta`      — every cube answered SUCCESS with a populated dimension list.
  * `_member`          — every member resolves by NAME to EXACTLY one candidate; ambiguity
                         refuses. `Laval` names both the census division (1730, geoLevel 3)
                         and a census subdivision inside it (1731, geoLevel 5), so a
                         name-only resolution picks whichever comes first.
  * `_guard_pinned_id` — the resolved id equals the id the ruling pinned. A move is a
                         FINDING about the cube, not a typo to paper over.
  * `_guard_response`  — every requested series came back, non-empty, at refPer 2021. A
                         `status: FAILED` cell carries an EMPTY `vectorDataPoint`, so a blind
                         `[0]` would raise an IndexError that reads like a code bug.
  * `_guard_universe`  — the province rows of the two cubes are BIT-IDENTICAL on the RULED
                         TRIPLE and within ±5 elsewhere. Not cube-wide identity: independent
                         rounding-to-5 per cube makes four other cells differ by exactly 5,
                         and a gate pinning cube-wide identity would pin a falsehood.
  * `_guard_isq`       — each ISQ workbook's own parts sum EXACTLY to its province row, and
                         the two workbooks agree on that province row. Both share
                         denominators are then literally the same number.
  * `_isq_code`        — every ISQ row this run keys on is resolved BY LABEL in the workbook
                         and its pinned code then ASSERTED: the ISQ mirror of the census
                         side's resolve-by-name / `_guard_pinned_id` pair. Without it the
                         gate's ISQ side fetched on a hand-typed integer no live surface ever
                         confirmed, so a re-pinned vintage that renumbered the
                         région-administrative axis would have moved the gate onto another
                         région's population with the note still rendering a PASS.
  * `_guard_code_join` — every innocent control's ISQ code equals the census member's
                         classification code. This is what makes the calibration set's
                         identity measured rather than asserted.
  * `_guard_citation`  — every figure the record states is RECOMPUTED here and matches, keyed
                         on the DIGITS (a stable token) and never on a prose prefix. Two
                         records: §6 for the rates, P10's note for the aligned counts.
  * `_guard_amendment13` — §6 states HORS_RMR's pair TWICE, and the prose statement is the
                         operative one. A table-keyed parser cannot see it, which is exactly
                         how a superseded pair rode through a green suite; this reads #13 and
                         refuses if the ruling is missing or unparseable.
  * `_guard_p10`       — P10's note yields all 16 Québec-part subdivisions and a MEASURED
                         verdict. A membership read short would net out a smaller territory
                         and publish the result as the aligned one.
  * `_guard_qc_split`  — SGC prefix AND census-tree ancestry both place every one of them
                         in Québec. Agreement alone would pass a member the two readings
                         agree is outside it, and a prefix-filtered parser would leave the
                         first reading unable to fail at all.
  * `_guard_required_complete` — the two members each suppression bound is cut from are
                         themselves published; an interval with an unmeasured end is not one.
  * `_guard_withheld_accounted` — every cell the boundary reported absent is a field the
                         bound carried. Dropping one biases the point estimate rather than
                         widening the interval.
  * `_guard_ratio_band` — the aligned ratio's suppression envelope does not straddle 1.0.
                         #13 rules the crossing with BOTH ends above it; a verdict inside its
                         own uncertainty is not a verdict.
  * `_guard_floor`     — ruled headship EXCEEDS the settled living-alone share at the same
                         geography. Each person living alone maintains exactly one household.
  * `_guard_territory` — |residual| ≤ the derived threshold, two-sided.

SHAPE, stated because `_sections` is long and that is a CHOICE. Every pure computation and
every guard is a module-level function, unit-tested and mutation-tested on its own. What
stays inline is the orchestration and the prose — deliberately, and for the defect class this
arc grades: each narrative sentence sits lexically beside the expression that computes it, so
a gloss cannot drift from its number without both moving in one diff. Hoisting the note into
a separate renderer would put sixty measured values through a handoff and reintroduce exactly
the distance the inline form removes. The byte-equality gate in tests/test_probe_p8.py is
what makes either shape safe to change later.

SCOPE: this is a PROBE. No value here is wired into `demand/immigrant_inputs.py`; plan task
25b is that consumer and a separate run.

Run:  cd demoflow && uv run python probes/run_p8.py
"""
import re
from dataclasses import dataclass
from pathlib import Path

# Flat, NOT `probes._wds`: probes/ is deliberately not a package, so in script mode
# sys.path[0] IS probes/ and this resolves natively. See probes/_wds.py.
from _wds import (WDS_DATA, WDS_META, Fact, new_run, post, provenance_header, table_number,
                  table_url)

from demoflow.loaders.pins import DATA_DIR, WORKBOOK_SHA256, verify_pin

_WRITTEN_BY = Path(__file__).name
OUT = Path(__file__).resolve().parent / "P8-immigrant-inputs.md"
_TITLE = "# P8 — the immigrant inputs: HEADSHIP and OWNERSHIP RATIO (rulings S and T)"

_REPO_ROOT = Path(__file__).resolve().parents[2]
SPEC = (_REPO_ROOT / "docs" / "specs"
        / "2026-07-21-demoflow-demographic-scenario-module-design.md")
P9_NOTE = Path(__file__).resolve().parent / "P9-catalogue-closure.md"
P10_NOTE = Path(__file__).resolve().parent / "P10-hors-operand-alignment.md"
CONSTANTS = _REPO_ROOT / "demoflow" / "src" / "demoflow" / "loaders" / "constants.py"

# --- the cubes ----------------------------------------------------------------------------
CMA_PID = 98100621          # ruling S: both inputs, CMA grain
CD_PID = 98100622           # ruling T: RA06/RA13, census-division grain
SIBLING_PID = 43100060      # the COARSE cross-check + the living-alone floor
TOTALPOP_PID = 98100007     # the SECOND DIAGNOSTIC — never the gate

# 98100621/98100622 share this dimension order, name for name.
POS_GEO, POS_TENURE, POS_GENDER, POS_MAINT, POS_STAT, POS_POPCHAR, POS_SUIT = range(1, 8)
TENURE_TOTAL = ("Total - Tenure including presence of mortgage payments and subsidized "
                "housing (totals include farm operators)")
TENURE_OWNER = "Owner"
GENDER_TOTAL = "Total - Gender"
MAINT_ALL_PERSONS = "Total - Household maintainer"
MAINT_PRIMARY = "Person is primary household maintainer"
STAT_PEOPLE = "Number of people"
SUIT_TOTAL = "Total - Housing suitability"

# Population characteristics. `TOTAL_AGE` is the universe row (all persons in the cube's
# private-household population) and is what the territory gate's share is taken over.
PC_TOTAL_AGE = "Total - Age"
PC_TOTAL_IMM = "Total - Immigrant status and period of immigration"
PC_NONIMM = "Non-immigrants"
PC_ALL_IMM = "Total immigrants"
PC_SETTLED = "Before 2016"                        # the RULED member
PC_RECENT = "Recent immigrants: 2016 to 2021"
PC_NPR = "Non-permanent residents"
POPCHARS = (PC_TOTAL_AGE, PC_TOTAL_IMM, PC_NONIMM, PC_ALL_IMM, PC_SETTLED, PC_RECENT, PC_NPR)
PINNED_POPCHAR_IDS = {PC_TOTAL_AGE: 1, PC_TOTAL_IMM: 9, PC_NONIMM: 10, PC_ALL_IMM: 11,
                      PC_SETTLED: 12, PC_RECENT: 13, PC_NPR: 14}

QUEBEC = "Quebec"
PINNED_GEO_IDS = {(CMA_PID, QUEBEC): 24, (CD_PID, QUEBEC): 884,
                  (CD_PID, "Montréal"): 1732, (CD_PID, "Laval"): 1730,
                  (SIBLING_PID, QUEBEC): 13, (TOTALPOP_PID, QUEBEC): 53}
# Recorded so a move REDS; never used to fetch. The six are resolved as the geoLevel-503
# children of the Quebec member, which is a property of the cube rather than of this list.
PINNED_CMA_IDS = (30, 35, 36, 40, 48, 51)
CMA_GEO_LEVEL = 503
CD_GEO_LEVEL = 3
CSD_GEO_LEVEL = 5
PROVINCE_GEO_LEVEL = 2

# --- the OPERAND-ALIGNED HORS_RMR territory (spec §6 amendment #13) -------------------------
# The residual above includes the Québec side of Ottawa-Gatineau; the ISQ arrival flows it
# multiplies exclude it. #13 rules the corrected territory: the same province-net-of-six
# residual, NET of the Ottawa-Gatineau CMA's Québec-side census subdivisions.
#
# The MEMBERSHIP — which subdivisions those are — is carried BY REFERENCE from P10, the
# committed probe that derived it (98-10-0003-01's 25 CMA children closing exactly on the CMA,
# 16 Québec-side by SGC prefix AND census-tree ancestry, validated against the ISQ row at
# −0.752%). This run does not re-derive that set: it READS P10's own §4b table and then
# re-resolves every member of it live, by SGC code and geoLevel, re-checking the Québec-side
# property in the cube the counts come from. The same treatment P9's catalogue closure gets —
# a committed measurement carried at the level its own note earned, never restated as a
# hand-typed list, and never taken on trust where a live surface can re-confirm it.
QC_SGC_PREFIX = "24"
PINNED_QC_PART_COUNT = 16        # asserted against what P10's note yields; never used to fetch
# The three members the aligned arithmetic needs at CSD grain: the ruled member, the ratio's
# base, and the universe row the suppression bound is built out of.
ALIGNED_POPCHARS = (PC_TOTAL_AGE, PC_NONIMM, PC_SETTLED)
# The CSD that shares the census division's name one level down. Named so the avoidance is a
# recorded property of this run rather than a coincidence of resolution order.
CSD_LAVAL_TRAP_ID = 1731

# 43-10-0060-01. Values are PERCENTAGES; every read is divided by 100 at the boundary.
SIB_POS_GEO, SIB_POS_GENDER, SIB_POS_AGE, SIB_POS_IMM = 1, 2, 3, 4
SIB_POS_VISMIN, SIB_POS_EDU, SIB_POS_IND = 5, 6, 7
SIB_GENDER_TOTAL = "Total – Gender"          # en dash, unlike 98100621's hyphen
SIB_AGE_TOTAL = "Total – Age group"
SIB_VISMIN_TOTAL = "Total – Visible minority"
SIB_EDU_TOTAL = "Total – Highest certificate, diploma or degree"
SIB_IND_OWNERSHIP = "Population living in a dwelling owned by some members of the household"
SIB_IND_ALONE = "Population living alone"
SIB_NONIMM = "Non-immigrants"
SIB_ALL_IMM = "Immigrants"
SIB_RECENT = "Admitted to Canada in the last 10 years"
SIB_SETTLED = "Admitted to Canada more than 10 years ago"     # the member the floor uses
SIB_NPR = "Non-permanent residents"
SIB_MEMBERS = (SIB_NONIMM, SIB_ALL_IMM, SIB_RECENT, SIB_SETTLED, SIB_NPR)
SIB_GEO = {"MTL_RMR": ("Montréal (CMA), Que.", 15), "QC_RMR": ("Québec (CMA), Que.", 17)}

TOTALPOP_POS_GEO, TOTALPOP_POS_COUNT = 1, 2
TOTALPOP_COUNT_2021 = "Population, 2021"

# --- the ISQ side -------------------------------------------------------------------------
ISQ_RA_BOOK = "pop-as-ra-base.xlsx"
ISQ_RMR_BOOK = "pop-as-rmr-base.xlsx"
ISQ_PROVINCE_CODE = 0
ISQ_PROVINCE_LABEL = "Le Québec"
ISQ_YEAR = 2021
ISQ_SEX_TOTAL = 3
ISQ_SCENARIOS = ("Référence (A2026)", "Faible (D2026)", "Fort (E2026)")
ISQ_HEADER_CELL = "Scénario"
ISQ_BODY_OFFSET = 3          # the header row is followed by a units row and a blank row
ISQ_NOTE_MARKER = "découpage"
# The ISQ rows this run keys on, each as the workbook's own LABEL beside the code pinned for
# it. Every one is resolved BY LABEL and its code then ASSERTED (`_isq_code`) — a code alone
# is a hand-typed integer no live surface confirms, and the census side does not accept one.
ISQ_RA_LABELS = {"MTL_ISLAND_RA06": "Montréal", "LAVAL_RA13": "Laval"}
ISQ_RA_CODES = {"MTL_ISLAND_RA06": 6, "LAVAL_RA13": 13}
# The RMR workbook's OWN published outside-the-CMAs row, and the CMA-part row that makes it a
# DIFFERENT geography from this model's HORS_RMR. Both read live; the difference is measured.
ISQ_HORS_CODE = 999
ISQ_HORS_LABEL = "Territoire hors des RMR"
ISQ_GATINEAU_CODE = 505
ISQ_GATINEAU_LABEL = "RMR d'Ottawa-Gatineau"
# Name markers for an outside-the-CMAs member, used to SEARCH the sibling cube rather than to
# assert its absence. A predicate published beside its result is falsifiable; a bare "no such
# member" is not.
HORS_NAME_MARKERS = ("outside", "non-cma", "rest of", "remainder", "hors")

REF_YEAR = "2021"
_DATA_CHUNK = 60

# §6 states HORS_RMR's immigrant pair TWICE and the two statements DISAGREE, deliberately:
# ruling S's table row carries the contaminated residual's 0.5169 / 0.9600, and amendment #13
# carries the operand-aligned 0.5234 / 1.0248 that supersedes it. Both are recomputed here and
# each is checked against the statement that rules IT — the table row against the shipped
# construction, #13 against the aligned one. Binding both is what keeps the note honest about
# a spec that (correctly) still carries the superseded row as the record of what was shipped.
AMD13_HEAD = "**AMENDMENT #13 (2026-08-15)"
AMD13_TAIL = "**Costs, recorded rather than argued away:**"

# The margin above the maximum innocent residual. STATED, and small: the innocent set has six
# members, so its maximum is itself a noisy estimate of the innocent spread's upper edge, and
# a quarter of it is enough headroom for a seventh innocent territory without turning the
# gate into a formality. Changing this changes the threshold and nothing else — the
# calibration is the max, this is only the cushion on top of it.
MARGIN_FRACTION = 0.25

# Everything the run measured, for the gates in tests/test_probe_p8.py to read directly
# rather than re-parsing prose. Rebound (never mutated in place) at the top of `_sections`.
LAST_RUN: dict = {}


class ProbeRefusal(RuntimeError):
    """A floor guard fired: this run may NOT publish a measurement.

    Deliberately not a network-class exception name — `tests/_probe_asserts.py`'s
    NETWORK_EXCEPTIONS decides fail-vs-skip, and a refusal that could be mistaken for an
    outage would launder the cardinal cheap all-clear into a green-looking skip.
    """

    def __init__(self, boundary: str, message: str) -> None:
        super().__init__(message)
        self.boundary = boundary


# ===========================================================================================
# Arithmetic — the definitions, as code
# ===========================================================================================
@dataclass(frozen=True)
class Cell:
    """One (geography, population-characteristic) reading: three published counts."""

    persons: int
    maintainers: int
    owner_maintainers: int

    @property
    def headship(self) -> float:
        if not self.persons:
            raise ProbeRefusal("arithmetic", f"headship over {self.persons} persons is not a "
                                             "rate — a cell with no persons has no headship, "
                                             "and a 0.0 here would be a number where there is "
                                             "no measurement")
        return self.maintainers / self.persons

    @property
    def owner_propensity(self) -> float:
        if not self.maintainers:
            raise ProbeRefusal("arithmetic", f"owner propensity over {self.maintainers} "
                                             "maintainers is not a rate")
        return self.owner_maintainers / self.maintainers


def cell_minus(whole: Cell, part: Cell) -> Cell:
    """The residual cell. Used ONCE, for HORS_RMR = province net of the six wholly-QC CMAs."""
    return Cell(whole.persons - part.persons,
                whole.maintainers - part.maintainers,
                whole.owner_maintainers - part.owner_maintainers)


def cell_sum(cells) -> Cell:
    total = Cell(0, 0, 0)
    for cell in cells:
        total = Cell(total.persons + cell.persons,
                     total.maintainers + cell.maintainers,
                     total.owner_maintainers + cell.owner_maintainers)
    return total


def complement_bound(total: Cell, non_immigrant: Cell) -> tuple:
    """An upper bound on ANY immigrant member's counts: everything that is not non-immigrant.

    The bound the aligned construction's withheld cells are carried at. Clamped at zero per
    field, and the clamp is REPORTED rather than silent: both inputs are rounded to 5
    independently, so a geography whose immigrant population is genuinely zero can publish a
    total one rounding step BELOW its non-immigrant count, and a negative "bound" would put
    the upper end of an interval under its lower end.

    The same definition P10 measured this territory with (`run_p10.complement_bound`), and
    deliberately the same arithmetic rather than a paraphrase of it: the two probes publish
    figures that are compared digit for digit, and a fourth-decimal divergence between two
    spellings of one definition would surface as a citation refusal with no defect behind it.
    """
    fields = (total.persons - non_immigrant.persons,
              total.maintainers - non_immigrant.maintainers,
              total.owner_maintainers - non_immigrant.owner_maintainers)
    return Cell(*(max(f, 0) for f in fields)), sum(1 for f in fields if f < 0)


def bounded_sum(rows: list) -> tuple:
    """Sum a set of geographies FIELD BY FIELD, carrying the withheld fields as an interval.

    `rows` is [(values, bounds)] per geography, both the same width, where a `None` in `values`
    marks a count StatCan withholds at that geography and the parallel entry in `bounds` is an
    upper bound on it drawn from a quantity the same cube DOES publish.

    FIELD-WISE is the whole point: a subdivision that publishes settled PERSONS while
    withholding settled MAINTAINERS is a real shape in this territory, and dropping the whole
    geography would discard a published count — which does not merely widen the interval, it
    BIASES the point estimate, because the persons and the maintainers of one territory would
    then be netted out of different denominators.

    Returns (low, high, withheld_field_count): `low` counts only what is published, `high` adds
    every bound. The truth lies in the box between them, per field.
    """
    if not rows:
        raise ProbeRefusal("suppression", "a bounded sum over no geographies is not a sum")
    width = len(rows[0][0])
    low, high, withheld = [0] * width, [0] * width, 0
    for values, bounds in rows:
        for index, (value, bound) in enumerate(zip(values, bounds, strict=True)):
            if value is None:
                if bound is None:
                    raise ProbeRefusal(
                        "suppression",
                        f"field {index} is withheld and carries NO bound — an interval with an "
                        "unmeasured end is not an interval, and this run may not publish one.")
                high[index] += bound
                withheld += 1
            else:
                low[index] += value
                high[index] += value
    return tuple(low), tuple(high), withheld


def bounded_pair(shipped: Cell, low: Cell, high: Cell, non_immigrant: Cell) -> dict:
    """The aligned headship and ratio, with the ENVELOPE the withheld cells leave.

    The headline subtracts only what is published (`low`) — the one construction assembled
    entirely from counts the source states. The envelope is taken at the box's CORNERS, not by
    pairing the two sums: headship is largest when the fewest maintainers and the most persons
    are netted out, and those are opposite corners. Pairing low-with-low would report an
    interval narrower than the uncertainty actually is.
    """
    headline = cell_minus(shipped, low)
    h_high = (shipped.maintainers - low.maintainers) / (shipped.persons - high.persons)
    h_low = (shipped.maintainers - high.maintainers) / (shipped.persons - low.persons)
    p_high = ((shipped.owner_maintainers - low.owner_maintainers)
              / (shipped.maintainers - high.maintainers))
    p_low = ((shipped.owner_maintainers - high.owner_maintainers)
             / (shipped.maintainers - low.maintainers))
    base = non_immigrant.owner_propensity
    return {"cells": headline, "headship": headline.headship,
            "ratio": ownership_ratio(headline, non_immigrant),
            "headship_band": (h_low, h_high),
            "ratio_band": (p_low / base, p_high / base)}


def relative_pct(new: float, old: float) -> float:
    """`new` against `old`, in percent OF `old` — a relative delta, never a difference."""
    if not old:
        raise ProbeRefusal("arithmetic", "a relative delta against zero is not a percentage")
    return (new / old - 1) * 100


def ownership_ratio(immigrant: Cell, non_immigrant: Cell) -> float:
    """§6's `ratio`: the owner-MAINTAINER propensity of one member over the other's.

    Not a person-weighted ownership share — that is 43-10-0060-01's metric, and the whole
    point of ruling S is that this cube publishes the quantity §6 defines directly.
    """
    return immigrant.owner_propensity / non_immigrant.owner_propensity


def share_residual_pct(*, part_census: float, total_census: float,
                       part_isq: float, total_isq: float) -> float:
    """The PROVINCE-CONTROLLED share residual, in percent.

    Each side's share is taken against its OWN source's provincial total, so a universe offset
    that is uniform across the province cancels by construction rather than being argued away.
    """
    if not total_census or not total_isq or not part_isq:
        raise ProbeRefusal("arithmetic", "a share residual needs non-zero totals on both "
                                         "sides and a non-zero ISQ part")
    return ((part_census / total_census) / (part_isq / total_isq) - 1) * 100


def derive_threshold(innocent: dict) -> dict:
    """The territory gate's threshold, DERIVED from the innocent controls handed in.

    Never inherited: ruling T's original 1% was calibrated against the construction amendment
    #11 refused, and a threshold that does not move with its calibration set is a hand-typed
    number one level up.
    """
    if not innocent:
        raise ProbeRefusal("threshold", "no innocent controls were measured — a threshold over "
                                        "an empty calibration set has nothing behind it")
    name, value = max(innocent.items(), key=lambda kv: abs(kv[1]))
    largest = abs(value)
    if largest <= 0:
        raise ProbeRefusal("threshold", f"every innocent control measured exactly zero "
                                        f"({sorted(innocent)}) — a calibration set with no "
                                        "spread cannot bound anything")
    margin = largest * MARGIN_FRACTION
    return {"max_name": name, "max_abs": largest, "max_signed": value,
            "margin_pp": margin, "threshold": largest + margin}


def coord(*ids: int) -> str:
    """A WDS coordinate: the cube's dimension ids, zero-padded to ten positions."""
    return ".".join(str(i) for i in ids) + ".0" * (10 - len(ids))


# ===========================================================================================
# Boundaries — the four injectable seams (so the whole green path runs OFFLINE in tests)
# ===========================================================================================
def _meta(pids) -> list:
    """POST `getCubeMetadata`. Boundary `wds-meta` (www150)."""
    return post(WDS_META, [{"productId": int(p)} for p in pids])


def _data(requests: list) -> list:
    """POST `getDataFromCubePidCoordAndLatestNPeriods`, chunked. Boundary `wds-data`.

    No try/except, no retry, no sentinel: a swallowed HTTP error would reach `_guard_response`
    as a missing series and be reported as StatCan not publishing the cell.
    """
    out: list = []
    for start in range(0, len(requests), _DATA_CHUNK):
        out += post(WDS_DATA, requests[start:start + _DATA_CHUNK])
    return out


def _isq_rows(name: str) -> list:
    """The pinned ISQ workbook's row grid. Boundary `isq-workbook` (a committed file).

    The pin is verified BEFORE a cell is read — the same PIT discipline the loaders use — and
    `openpyxl` directly rather than pandas: nothing here needs a DataFrame, and the raw cell
    grid is what the header search reads.
    """
    from openpyxl import load_workbook

    path = DATA_DIR / name
    verify_pin(path, name)
    book = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = book[book.sheetnames[0]]
        return list(sheet.iter_rows(values_only=True))
    finally:
        book.close()


def _spec_s6() -> str:
    """Spec §6, the ruling text this note is citation-coupled to. Boundary `spec`."""
    text = SPEC.read_text(encoding="utf-8")
    head, tail = "## 6. Demand side", "## 7. Outputs"
    if head not in text or tail not in text:
        raise ProbeRefusal("spec", f"{SPEC.name} no longer carries both section markers "
                                   f"{head!r} and {tail!r} — the citation gate cannot scope "
                                   "itself to the ruling text and must not guess")
    return text[text.index(head):text.index(tail)]


def _p9_note() -> str:
    """P9's committed note — the source of the closure this one carries BY REFERENCE."""
    if not P9_NOTE.exists():
        raise ProbeRefusal("p9", f"{P9_NOTE.name} is absent; this note may not restate a "
                                 "catalogue closure it cannot read at the level P9 earned it")
    return P9_NOTE.read_text(encoding="utf-8")


def _p10_note() -> str:
    """P10's committed note — the source of the aligned MEMBERSHIP this run carries."""
    if not P10_NOTE.exists():
        raise ProbeRefusal("p10", f"{P10_NOTE.name} is absent; this run may not net out a "
                                  "territory whose membership it cannot read at the level P10 "
                                  "earned it")
    return P10_NOTE.read_text(encoding="utf-8")


def _constants_source() -> str:
    """`loaders/constants.py` — where the POOLED-ratio anti-pattern is recorded."""
    return CONSTANTS.read_text(encoding="utf-8")


# ===========================================================================================
# Guards
# ===========================================================================================
def _guard_meta(pids, response) -> dict:
    if not isinstance(response, list) or len(response) != len(pids):
        raise ProbeRefusal("wds-meta", f"getCubeMetadata returned "
                                       f"{len(response) if isinstance(response, list) else type(response).__name__}"
                                       f" for {len(pids)} requested cube(s)")
    metas = {}
    for item in response:
        if item.get("status") != "SUCCESS":
            raise ProbeRefusal("wds-meta", f"getCubeMetadata status {item.get('status')!r}")
        obj = item["object"]
        if not (obj.get("dimension") or []):
            raise ProbeRefusal("wds-meta", f"{obj.get('productId')} answered with an EMPTY "
                                           "dimension list — every member this run resolves "
                                           "would be looked up in nothing")
        metas[int(obj["productId"])] = obj
    missing = [p for p in pids if int(p) not in metas]
    if missing:
        raise ProbeRefusal("wds-meta", f"getCubeMetadata answered for {sorted(metas)}, "
                                       f"missing {missing}")
    return metas


def _dimension(metas: dict, pid: int, position: int) -> dict:
    for dim in metas[pid]["dimension"]:
        if dim.get("dimensionPositionId") == position:
            return dim
    raise ProbeRefusal("wds-meta", f"{table_number(pid)} has no dimension at position "
                                   f"{position}")


def _member(metas: dict, pid: int, position: int, name: str, *,
            geo_level: int | None = None, parent: int | None = None) -> dict:
    """Resolve a member BY NAME, refusing on zero OR MORE THAN ONE match.

    Ambiguity is a refusal rather than a first-hit: `Laval` names both the census division
    (geoLevel 3) and a census subdivision inside it (geoLevel 5), and picking whichever comes
    first would publish a city's numbers under a region's label.
    """
    dim = _dimension(metas, pid, position)
    hits = [m for m in dim.get("member") or []
            if m.get("memberNameEn") == name
            and (geo_level is None or m.get("geoLevel") == geo_level)
            and (parent is None or m.get("parentMemberId") == parent)]
    if len(hits) != 1:
        raise ProbeRefusal(
            "wds-meta",
            f"{table_number(pid)} dimension {position}: {len(hits)} member(s) named {name!r} "
            f"at geoLevel {geo_level} under parent {parent} (of "
            f"{len(dim.get('member') or [])} members). Exactly one must match — zero means the "
            f"cube moved, more than one means the name alone does not identify a territory.")
    return hits[0]


def _guard_pinned_id(pid: int, name: str, member: dict, pinned: int) -> dict:
    """The resolved member must be the one the ruling pinned. A move is a FINDING."""
    if member["memberId"] != pinned:
        raise ProbeRefusal(
            "wds-meta",
            f"{table_number(pid)}: {name!r} resolves to memberId {member['memberId']}, but the "
            f"ruling pinned {pinned}. That is a finding about the cube, not a typo to paper "
            "over — every value keyed on the pinned id would be a different territory's.")
    return member


def _member_by_code(pid: int, by_code: dict, code: str, level: int, label: str) -> dict:
    """Resolve a geography member by its SGC classification code AND geoLevel.

    The code is what joins P10's membership to the cube these counts are read from, and the
    geoLevel is what keeps a census subdivision from answering for the division that shares its
    name and its code prefix. Ambiguity refuses, exactly as `_member` does.
    """
    hits = by_code.get((code, level)) or []
    if len(hits) != 1:
        raise ProbeRefusal(
            "wds-meta",
            f"{table_number(pid)}: {len(hits)} member(s) with SGC {code} at geoLevel {level} "
            f"({label}). Exactly one must match — the code is the join between the membership "
            "and the counts, and a name would not survive an accented rename.")
    return hits[0]


def _guard_qc_split(children: list) -> None:
    """BOTH readings must place EVERY member P10 hands over on the Québec side.

    Two independent readings of one claim, re-run HERE rather than inherited: the prefix is a
    string property of the SGC code this cube gives the member, the ancestry is a property of
    the geography tree in the same cube these counts are read from. P10 established the SET;
    this run re-establishes that each of its members really sits on the Québec side of the
    cube it is subtracted from. Neither reading is taken on the other's word, and AGREEMENT is
    not the test: a member both readings place OUTSIDE Québec is refused — its counts would be
    netted out of a residual they were never inside — and so is a member they disagree about,
    which is a finding about the axis, never a tie broken by whichever ran first.
    """
    if not children:
        raise ProbeRefusal("membership", "no members were classified — a split over an empty "
                                         "list is vacuous")
    broken = [(name, by_code, by_tree) for name, by_code, by_tree in children
              if by_code != by_tree]
    if broken:
        raise ProbeRefusal(
            "membership",
            "the SGC prefix and the census-tree ancestry disagree on which side of the "
            "provincial boundary a member sits: "
            + "; ".join(f"{n} (prefix says {'QC' if c else 'not QC'}, tree says "
                        f"{'QC' if t else 'not QC'})" for n, c, t in broken))
    outside = [name for name, by_code, by_tree in children if not (by_code and by_tree)]
    if outside:
        raise ProbeRefusal(
            "membership",
            f"{len(outside)} member(s) of the membership sit OUTSIDE Québec by the SGC prefix "
            f"and by the census tree alike ({outside[:4]}). Two readings agreeing does not put "
            "a subdivision inside this territory, and its counts may not be netted out of a "
            "residual they were never inside.")


def _guard_required_complete(published: dict) -> None:
    """The members the suppression BOUND is built from must themselves be published.

    The bound is `Total - Age` minus `Non-immigrants`: an upper bound on any immigrant member
    at that geography. If either leg were itself withheld the bound would be built out of a
    hole, and an interval whose upper end is unmeasured bounds nothing.
    """
    holes = sorted(name for name, complete in published.items() if not complete)
    if holes:
        raise ProbeRefusal(
            "wds-data",
            f"{len(holes)} QC-part CSD(s) do not publish both members the suppression bound is "
            f"built from ({holes[:4]}). An interval whose upper end is itself unmeasured is not "
            "a bound, and this run may not publish one as if it were.")


def _guard_withheld_accounted(*, reported: int, carried: int) -> None:
    """Every cell the BOUNDARY reported absent must be a field the BOUND actually carried.

    The two counts are taken on opposite sides of the derivation — one from the response, one
    from inside the arithmetic — so a divergence means a withheld count went missing between
    them. That does not widen the interval, it BIASES the point estimate, which is the failure
    the whole field-wise construction exists to avoid.
    """
    if reported != carried:
        raise ProbeRefusal(
            "suppression",
            f"{reported} cell(s) came back withheld but the bounded sum carried {carried} — a "
            "withheld count was dropped between the boundary and the arithmetic.")


def _guard_ratio_band(low: float, high: float) -> None:
    """The suppression bound may not straddle the finding it is published under.

    Amendment #13 rules the aligned ratio as CROSSING 1.0 with BOTH envelope ends above it —
    settled immigrants OUT-own in the aligned hors-RMR territory where they under-own in the
    contaminated one. If the bound contained 1.0 the crossing would not be earned at this
    run's own resolution, and a verdict inside its own uncertainty is not a verdict.
    """
    lo, hi = sorted((low, high))
    if lo <= 1.0 <= hi:
        raise ProbeRefusal(
            "suppression",
            f"the suppression bound on the aligned ratio is [{lo:.4f}, {hi:.4f}] and STRADDLES "
            "1.0, so the crossing amendment #13 rules is not earned at the resolution the "
            "withheld cells leave.")


def _guard_response(requests: list, response, suppressible: set | None = None) -> tuple:
    """Key every returned cell by (productId, coordinate) and prove none is UNEXPECTEDLY absent.

    Keying by POSITION is the defect this refuses: WDS returns its objects sorted by coordinate
    STRING, not in request order, so a positional read pairs every value with the wrong cell
    while every count still looks plausible.

    A `status: FAILED` cell carries an EMPTY `vectorDataPoint`. At the tiny census subdivisions
    of the Ottawa-Gatineau Québec part that is a PUBLICATION RULE — StatCan withholds small
    immigrant counts — so those coordinates are declared suppressible IN ADVANCE, by geography,
    and are carried by the bound rather than dropped. Everywhere else an absent cell is refused:
    an outage reaching the arithmetic as "the source publishes nothing here" would fabricate a
    publication rule. Returns (series, withheld coordinates).
    """
    suppressible = set(suppressible or ())
    if not isinstance(response, list):
        raise ProbeRefusal("wds-data", f"getData returned {type(response).__name__}, not a list")
    series: dict = {}
    for item in response:
        obj = item.get("object") or {}
        points = obj.get("vectorDataPoint") or []
        key = (int(obj["productId"]), obj["coordinate"]) if obj.get("productId") else None
        if item.get("status") != "SUCCESS" or not points:
            if key in suppressible:
                continue
            raise ProbeRefusal(
                "wds-data",
                f"cell {obj.get('productId')}/{obj.get('coordinate')} came back "
                f"status={item.get('status')!r} with {len(points)} data point(s), and it is NOT "
                "in the suppressible scope (the QC-part CSDs' cells). A FAILED cell carries an "
                "EMPTY vectorDataPoint, so a blind [0] would raise an IndexError that reads "
                "like a code bug — and outside a tiny geography an absence is an outage, never "
                "a publication rule.")
        period = str(points[0].get("refPer", ""))[:4]
        if period != REF_YEAR:
            raise ProbeRefusal(
                "wds-data",
                f"cell {obj.get('productId')}/{obj.get('coordinate')} answered for reference "
                f"period {period!r}, not the {REF_YEAR} census — a census cube answering a "
                "different period is drift, and mixing vintages inside one ratio is silent.")
        series[key] = points[0]["value"]
    wanted = {(int(r["productId"]), r["coordinate"]) for r in requests}
    missing = sorted(wanted - set(series) - suppressible)
    if missing:
        raise ProbeRefusal("wds-data", f"{len(missing)} of {len(wanted)} requested cells did "
                                       f"not come back and are not suppressible "
                                       f"(first: {missing[0]})")
    return series, sorted(wanted - set(series))


_RULED_TRIPLE_FIELDS = ("persons", "maintainers", "owner-maintainers")
_ROUNDING_TOLERANCE = 5      # StatCan rounds counts to 5, INDEPENDENTLY per cube


def _guard_universe(cma_rows: dict, cd_rows: dict) -> list:
    """The one-universe claim, asserted at exactly its measured strength.

    Bit-identity is required on the RULED TRIPLE (`Before 2016`: persons, maintainers,
    owner-maintainers) — that is what makes the two cubes one universe at two grains rather
    than two sources. It is NOT required cube-wide, because it is not TRUE cube-wide:
    independent rounding-to-5 gives four other province cells that differ by exactly 5. Those
    are RETURNED as measured. Anything beyond the rounding step refuses.
    """
    ruled = PINNED_POPCHAR_IDS[PC_SETTLED]
    for source, rows in (("98100621", cma_rows), ("98100622", cd_rows)):
        if ruled not in rows:
            raise ProbeRefusal("universe", f"{source} carries no row for the ruled member "
                                           f"(popchar {ruled}) — an identity check over an "
                                           "absent row is vacuous")
    if cma_rows[ruled] != cd_rows[ruled]:
        raise ProbeRefusal(
            "universe",
            f"the province rows DISAGREE on the ruled member (popchar {ruled}): "
            f"{cma_rows[ruled]} vs {cd_rows[ruled]}. Bit-identity there is the whole basis for "
            "reading these two cubes as one universe at two geography grains.")
    drifted = []
    for popchar in sorted(set(cma_rows) & set(cd_rows)):
        # `strict`: these are (persons, maintainers, owner-maintainers) triples from the same
        # popchar key, so they are aligned by construction — but a silent truncation here
        # would drop a field from the comparison without changing a single count.
        for index, (left, right) in enumerate(
                zip(cma_rows[popchar], cd_rows[popchar], strict=True)):
            gap = abs(left - right)
            if gap > _ROUNDING_TOLERANCE:
                raise ProbeRefusal(
                    "universe",
                    f"province popchar {popchar} {_RULED_TRIPLE_FIELDS[index]} differs by "
                    f"{gap} across the two cubes ({left} vs {right}) — beyond the ±"
                    f"{_ROUNDING_TOLERANCE} rounding step, so these are not one universe")
            if gap:
                drifted.append((popchar, index, left, right))
    return drifted


def _guard_isq(book: str, totals: dict, province_code: int) -> int:
    """Each workbook's own parts must sum EXACTLY to its province row.

    This is what earns the province total as a share denominator: if the published parts did
    not close on it, the "share of its own source's provincial total" would be a share of
    something the source does not actually decompose.
    """
    if province_code not in totals:
        raise ProbeRefusal("isq-workbook", f"{book} carries no province row (code "
                                           f"{province_code})")
    province = totals[province_code]
    parts = sum(v for k, v in totals.items() if k != province_code)
    if parts != province:
        raise ProbeRefusal(
            "isq-workbook",
            f"{book}: its {len(totals) - 1} published parts sum to {parts:,} against a province "
            f"row of {province:,} (gap {parts - province:+,}). The province total is this "
            "gate's share denominator; a total its own parts do not close on cannot serve.")
    return province


def _isq_label(text: object) -> str:
    """A workbook label, normalised for comparison: trailing space and FOOTNOTE MARKER off.

    ISQ appends the footnote's NUMBER to the cell text (`RMR d'Ottawa-Gatineau2`) and the RMR
    rows carry a trailing space. Neither is part of a territory's identity, and a guard keyed
    on them would REFUSE when a footnote is renumbered — a spurious trip is a real cost, and
    this is the same reason `_digits` strips the sign instead of matching §6's U+2212 glyph.
    No région or RMR name in either workbook ends in a digit, so the strip cannot eat a name.
    """
    return str(text).strip().rstrip("0123456789").strip()


def _isq_code(book: str, labels: dict, label: str, pinned: int) -> int:
    """Resolve an ISQ row BY LABEL, then assert the code this probe pinned for it.

    The ISQ mirror of `_member` + `_guard_pinned_id`. The census side refuses a member whose
    id moved; before this the ISQ side simply fetched on the typed code and PRINTED whatever
    label sat there, so a re-pinned vintage that renumbered the région-administrative axis
    would have taken another région's population into the territory gate and still rendered a
    PASS. Ambiguity refuses for the same reason a name-only `Laval` lookup does.
    """
    hits = sorted(code for code, text in labels.items() if _isq_label(text) == label)
    if len(hits) != 1:
        raise ProbeRefusal(
            "isq-workbook",
            f"{book}: {len(hits)} row(s) labelled {label!r} (of {len(labels)}). Exactly one "
            "must match — zero means the workbook renamed or dropped the row this gate keys "
            "on, more than one means the label alone does not identify a territory.")
    if hits[0] != pinned:
        raise ProbeRefusal(
            "isq-workbook",
            f"{book}: {label!r} sits at code {hits[0]}, not the pinned {pinned}. The code is "
            "what this gate FETCHES on, so a renumbered axis would put another territory's "
            "population through the residual with the note still rendering a PASS.")
    return hits[0]


def _guard_isq_labels(isq: dict) -> dict:
    """Every ISQ row this run keys on, resolved by label with its pinned code asserted.

    Returns {(book, label): code}. The run then fetches on THOSE codes, so the resolution is
    load-bearing rather than decorative — exactly as the census side fetches on the memberId
    its by-name lookup returned.
    """
    pins = {
        ISQ_RA_BOOK: {ISQ_PROVINCE_LABEL: ISQ_PROVINCE_CODE,
                      **{ISQ_RA_LABELS[tag]: ISQ_RA_CODES[tag] for tag in ISQ_RA_LABELS}},
        ISQ_RMR_BOOK: {ISQ_PROVINCE_LABEL: ISQ_PROVINCE_CODE,
                       ISQ_HORS_LABEL: ISQ_HORS_CODE,
                       ISQ_GATINEAU_LABEL: ISQ_GATINEAU_CODE},
    }
    return {(book, label): _isq_code(book, isq[book]["labels"], label, code)
            for book, wanted in pins.items() for label, code in wanted.items()}


def _guard_code_join(joined: list) -> None:
    """Every innocent control joins census↔ISQ on an EXACT code.

    The calibration set's identity is what makes its spread a measure of the universe's
    geography-varying component rather than of territory mismatch. Measured, not asserted.
    """
    broken = [(name, census, isq) for name, census, isq in joined if census != isq]
    if broken:
        raise ProbeRefusal(
            "code-join",
            f"{len(broken)} innocent control(s) do not join on a code: {broken}. Their "
            "identity is the only reason their spread can calibrate the threshold.")


def _guard_floor(headship: dict, alone: dict) -> None:
    """Ruled headship must EXCEED the settled living-alone share at the same geography.

    Each person living alone maintains exactly one household, so headship at or below the
    living-alone share is arithmetically impossible: a defect, not a datum.
    """
    if not headship:
        raise ProbeRefusal("floor", "the floor gate was handed no headship to check")
    for geography, value in sorted(headship.items()):
        if geography not in alone:
            raise ProbeRefusal(
                "floor",
                f"no living-alone share was measured for {geography}, so its headship "
                f"{value:.4f} has no floor to clear. A gate that cannot verify REFUSES.")
        if value <= alone[geography]:
            raise ProbeRefusal(
                "floor",
                f"{geography}: headship {value:.4f} does NOT exceed the settled living-alone "
                f"floor {alone[geography]:.3f}. Every person living alone maintains exactly "
                "one household, so this is impossible — a defect, not a datum.")


def _guard_territory(residuals: dict, threshold: float) -> None:
    """|residual| ≤ threshold, two-sided. A territory mismatch has no preferred sign."""
    if not residuals:
        raise ProbeRefusal("territory", "the territory gate was handed no residual to check")
    over = {name: value for name, value in residuals.items() if abs(value) > threshold}
    if over:
        raise ProbeRefusal(
            "territory",
            f"the territory gate FAILED for {sorted(over)}: "
            + "; ".join(f"{n} {v:+.3f}%" for n, v in sorted(over.items()))
            + f" against a derived threshold of {threshold:.3f}%. Ruling T flags these "
            "geographies `cited` on the strength of a 1:1 territory match; above threshold "
            "that is a seat question, never a footnote.")


_GEO_ROW_TOKENS = ("MTL_RMR", "QC_RMR", "HORS_RMR", "MTL_ISLAND_RA06", "LAVAL_RA13")
_DIGITS = re.compile(r"\d+\.\d+")


def _guard_s6_rows(s6: str) -> dict:
    """Parse §6's two ruled tables, keyed on the GEOGRAPHY token and the DIGITS.

    Deliberately not on a prose prefix: a reworded sentence would silently detach the coupling
    and leave a gate that cannot fail. The geography name and the numerals are the two stable
    tokens in that table, so those are what this reads.
    """
    rows: dict = {}
    for line in s6.splitlines():
        if not line.strip().startswith("|"):
            continue
        cells = [c.strip().strip("*` ") for c in line.strip().strip("|").split("|")]
        if cells and cells[0] in _GEO_ROW_TOKENS:
            found = _DIGITS.findall(" ".join(cells[1:]))
            if len(found) >= 2:
                rows[cells[0]] = found
    missing = [g for g in _GEO_ROW_TOKENS if g not in rows]
    if missing:
        raise ProbeRefusal(
            "spec",
            f"spec §6 no longer yields a ruled table row for {missing} (parsed {sorted(rows)}). "
            "A citation gate over a section it could not parse passes vacuously, which is the "
            "can't-fail shape this whole family refuses.")
    return rows


def _guard_citation(source_name: str, text: str, coupled: list) -> None:
    """Every recomputed figure must match the digits the document it is checked against states.

    A measurement contradicting the record is fork-class: the run refuses rather than
    publishing a note whose numbers quietly disagree with what it claims to implement.

    Parameterised by SOURCE because two documents rule different halves here: §6 states the
    ruled rates (both HORS_RMR constructions, since #13 supersedes ruling S's row in prose),
    while the aligned territory's COUNTS are stated only by P10, the committed probe that
    derived that membership. Coupling the counts to §6 would demand digits no ruling carries;
    leaving them uncoupled would let the one construction this run does not derive end to end
    drift silently. So each figure is bound to the document that actually states it.
    """
    absent = [(label, tok) for label, tok in coupled if tok not in text]
    if absent:
        raise ProbeRefusal(
            "citation",
            f"recomputed figures {source_name} does NOT state: "
            + "; ".join(f"{label} -> {tok}" for label, tok in absent)
            + ". Either this run measured something the record contradicts (fork-class — raise "
              "it, do not paper over it) or the record moved and the coupling must move with "
              "it.")


def _guard_rows_match_spec(spec_rows: dict, computed: dict) -> None:
    """Each modeled geography's recomputed PAIR must be the pair its own §6 row carries.

    HORS_RMR enters here as the SHIPPED residual, which is what its ruling-S row states. Its
    ruled pair is the ALIGNED one and is bound to amendment #13 instead (`_guard_amendment13`)
    — a table row cannot carry a value a later amendment moved into prose. If a future seat
    rewrites that row to the aligned pair, this gate REDS and the coupling must be re-pointed
    deliberately, which is the intended cost of a spec that keeps both statements.
    """
    for geography, (headship, ratio) in sorted(computed.items()):
        published = spec_rows[geography]
        for label, token in (("headship", headship), ("ratio", ratio)):
            if token not in published:
                raise ProbeRefusal(
                    "citation",
                    f"{geography}: recomputed {label} {token} is not among the figures §6's own "
                    f"row publishes ({published}) — the row is keyed on the geography token, so "
                    "this is a disagreement about the value, not about the wording.")


# Every membership row P10's §4b table publishes, keyed on the 7-digit SGC code its first
# column carries — NOT on the `24` prefix. Selecting on the prefix HERE would make the
# province a property of the parser: an off-province row would be dropped without a word,
# and `_guard_qc_split`'s code leg could never fail. The prefix is enforced downstream,
# where failing it is a refusal rather than a silent filter.
_CSD_CODE_ROW = re.compile(r"^\|\s*(\d{7})\s*\|\s*([^|]+?)\s*\|")


def _guard_p10(text: str) -> dict:
    """The aligned MEMBERSHIP and its gate, read from P10's own note and DECISION tokens.

    Two things are carried, and only two: WHICH census subdivisions make up the Ottawa-Gatineau
    Québec part (P10's §4b table, one row per member, keyed on the SGC code the row starts
    with), and the verdicts P10 earned about that set — the 25-child closure and the −0.752%
    membership gate. The values this note publishes are NOT carried: they are recomputed here
    from the live cells at exactly those geographies.

    Parsed on the CODE COLUMN rather than a prose prefix, EVERY row of it rather than the
    Québec-coded ones, and refused unless the set is complete: a membership gate over a table
    this run could not read would net out a smaller territory than the ruling names and
    publish the result as the aligned one, and a parser that kept only the rows it already
    believed were Québec would answer the province question before `_guard_qc_split` asked it.
    """
    wanted = ("DECISION-VERDICT", "DECISION-CONSTRUCTION", "DECISION-MEMBERSHIP",
              "DECISION-MEMBERSHIP-GATE", "DECISION-ALIGNED-HEADSHIP", "DECISION-ALIGNED-RATIO")
    tokens = {}
    for name in wanted:
        found = re.search(rf"`{name}:\s*(.*?)`", text)
        if not found:
            raise ProbeRefusal("p10", f"P10's note carries no `{name}` token — this run may not "
                                      "net out a territory whose derivation it cannot read")
        tokens[name] = found.group(1).strip()
    if tokens["DECISION-VERDICT"] != "MEASURED":
        raise ProbeRefusal("p10", f"P10's verdict is {tokens['DECISION-VERDICT']!r}, not "
                                  "MEASURED — an unmeasured membership is not a territory")
    members = [(m.group(1), m.group(2)) for m in
               (_CSD_CODE_ROW.match(line) for line in text.splitlines()) if m]
    if len(members) != PINNED_QC_PART_COUNT or len({c for c, _n in members}) != len(members):
        raise ProbeRefusal(
            "p10",
            f"P10's note yields {len(members)} membership row(s) "
            f"({sorted(c for c, _n in members)[:4]}...), not the {PINNED_QC_PART_COUNT} the "
            "ruled construction names. A residual netted against a partial membership is a "
            "third territory, neither the shipped one nor the aligned one.")
    tokens["members"] = dict(members)
    return tokens


def _guard_amendment13(s6: str) -> dict:
    """Amendment #13's ruled pair and envelope, parsed out of §6's own text.

    §6 states HORS_RMR's pair twice — ruling S's table row (the shipped residual) and #13 (the
    aligned territory that supersedes it). `_guard_s6_rows` reads the first; this reads the
    second, because a table-keyed parser cannot see a ruling stated in prose. That blindness is
    exactly why the note could carry a superseded pair through a green suite.
    """
    if AMD13_HEAD not in s6 or AMD13_TAIL not in s6:
        raise ProbeRefusal(
            "citation",
            f"spec §6 no longer carries amendment #13 between {AMD13_HEAD!r} and "
            f"{AMD13_TAIL!r} — the aligned HORS_RMR pair would be coupled to nothing, which is "
            "the can't-fail shape this family refuses.")
    block = s6[s6.index(AMD13_HEAD):s6.index(AMD13_TAIL)]
    pair = re.search(r"headship \*\*([0-9.]+)\*\*, ratio\s*\n?\*\*([0-9.]+)\*\*", block)
    envelope = re.search(r"suppression envelope ([0-9.]+)-([0-9.]+) and ([0-9.]+)-([0-9.]+)",
                         block)
    if not pair or not envelope:
        raise ProbeRefusal(
            "citation",
            "amendment #13 yields no parseable ruled pair and envelope — this run recomputes "
            "both and may not publish them against a ruling it could not read.")
    return {"block": block, "headship": pair.group(1), "ratio": pair.group(2),
            "headship_band": (envelope.group(1), envelope.group(2)),
            "ratio_band": (envelope.group(3), envelope.group(4))}


def _guard_amendment13_match(ruled: dict, headship: str, ratio: str, bands: dict) -> None:
    """The recomputed aligned pair and envelope must be the ones amendment #13 states."""
    for label, computed, stated in (("headship", headship, ruled["headship"]),
                                    ("ratio", ratio, ruled["ratio"]),
                                    ("headship envelope", bands["headship"],
                                     "-".join(ruled["headship_band"])),
                                    ("ratio envelope", bands["ratio"],
                                     "-".join(ruled["ratio_band"]))):
        if computed != stated:
            raise ProbeRefusal(
                "citation",
                f"HORS_RMR aligned {label}: this run recomputes {computed}, amendment #13 rules "
                f"{stated}. Either the measurement contradicts the ruling (fork-class — raise "
                "it, do not paper over it) or §6 moved and the coupling must move with it.")


def _guard_p9(text: str) -> dict:
    """The closure P9 earned, read from P9's own DECISION tokens."""
    wanted = ("DECISION-VERDICT", "DECISION-CLOSURE-LEVEL", "DECISION-RESIDUAL")
    tokens = {}
    for name in wanted:
        found = re.search(rf"`{name}:\s*(.*?)`", text)
        if not found:
            raise ProbeRefusal("p9", f"P9's note carries no `{name}` token — this note may not "
                                     "restate a closure whose level it cannot read")
        tokens[name] = found.group(1).strip()
    return tokens


# ===========================================================================================
# Reading the sources
# ===========================================================================================
def _census_requests(pid: int, geo_ids: list, member_ids: dict) -> tuple:
    """One request per (geography, popchar, statistic) cell, plus the key map."""
    tenure_total = member_ids["tenure_total"]
    tenure_owner = member_ids["tenure_owner"]
    gender = member_ids["gender"]
    all_persons = member_ids["maint_all"]
    primary = member_ids["maint_primary"]
    statistic = member_ids["statistic"]
    suitability = member_ids["suitability"]

    requests, keys = [], {}
    for geo in geo_ids:
        for name in POPCHARS:
            popchar = member_ids["popchar"][name]
            for field, (tenure, maintainer) in enumerate(
                    ((tenure_total, all_persons), (tenure_total, primary),
                     (tenure_owner, primary))):
                coordinate = coord(geo, tenure, gender, maintainer, statistic, popchar,
                                   suitability)
                requests.append({"productId": pid, "coordinate": coordinate, "latestN": 1})
                keys[(geo, name, field)] = (pid, coordinate)
    return requests, keys


def _census_cells(series: dict, keys: dict, geo_ids: list) -> dict:
    """{geography id: {popchar name: Cell}} — every value read back BY COORDINATE."""
    out: dict = {}
    for geo in geo_ids:
        out[geo] = {}
        for name in POPCHARS:
            values = [series[keys[(geo, name, field)]] for field in range(3)]
            out[geo][name] = Cell(int(values[0]), int(values[1]), int(values[2]))
    return out


def _csd_requests(pid: int, geo_ids: list, member_ids: dict) -> tuple:
    """The QC-part subdivisions' cells: the three members the aligned arithmetic needs.

    A narrower cross than `_census_requests` on purpose. The aligned residual is a pair, and
    the pair needs the ruled member, the base it is divided by, and the universe row the
    suppression bound is cut from — 3 members rather than 7, at 16 geographies whose withheld
    cells each cost a guard.
    """
    requests, keys = [], {}
    for geo in geo_ids:
        for name in ALIGNED_POPCHARS:
            for field, (tenure, maintainer) in enumerate(
                    ((member_ids["tenure_total"], member_ids["maint_all"]),
                     (member_ids["tenure_total"], member_ids["maint_primary"]),
                     (member_ids["tenure_owner"], member_ids["maint_primary"]))):
                coordinate = coord(geo, tenure, member_ids["gender"], maintainer,
                                   member_ids["statistic"], member_ids["popchar"][name],
                                   member_ids["suitability"])
                requests.append({"productId": pid, "coordinate": coordinate, "latestN": 1})
                keys[(geo, name, field)] = (pid, coordinate)
    return requests, keys


def _census_fields(series: dict, keys: dict, geo: int, name: str) -> tuple:
    """The three counts as they came back — `None` in the position of a withheld one.

    The FIELD-WISE reading. `_census_cell` collapses a partially published geography to
    nothing, which is right where a whole `Cell` is needed and wrong where a SUM is: these
    subdivisions publish persons while withholding maintainers, and that published count is
    part of the territory being netted out.
    """
    return tuple(series.get(keys[(geo, name, field)]) for field in range(3))


def _census_cell(series: dict, keys: dict, geo: int, name: str) -> Cell | None:
    """A `Cell`, or None when ANY of its three counts is withheld at this geography."""
    values = [series.get(keys[(geo, name, field)]) for field in range(3)]
    if any(v is None for v in values):
        return None
    return Cell(int(values[0]), int(values[1]), int(values[2]))


def _isq_totals(book: str, rows: list) -> dict:
    """{code: (label, July-1 2021 total)} plus the workbook's own header note.

    The three scenario fans must AGREE at the census year: 2021 is a `réel` observation, so a
    disagreement would mean the row being read is not the observation it is taken for.
    """
    header = next((i for i, row in enumerate(rows)
                   if row and str(row[0]).strip() == ISQ_HEADER_CELL), None)
    if header is None:
        raise ProbeRefusal("isq-workbook", f"{book} carries no {ISQ_HEADER_CELL!r} header row — "
                                           "its layout is not the one this probe reads")
    note = next((str(row[0]).strip() for row in rows[:header]
                 if row and row[0] and ISQ_NOTE_MARKER in str(row[0])), "")
    if not note:
        raise ProbeRefusal("isq-workbook", f"{book} states no {ISQ_NOTE_MARKER!r} header note — "
                                           "the delineation it uses is exactly what this gate "
                                           "must quote rather than assume")
    fans: dict = {}
    labels: dict = {}
    for row in rows[header + ISQ_BODY_OFFSET:]:
        if not row or row[0] is None:
            continue
        scenario, code, label, year, _statut, sex = row[0], row[1], row[2], row[3], row[4], row[5]
        if year != ISQ_YEAR or sex != ISQ_SEX_TOTAL or scenario not in ISQ_SCENARIOS:
            continue
        fans.setdefault(int(code), {})[scenario] = int(row[6])
        labels[int(code)] = str(label)
    if not fans:
        raise ProbeRefusal("isq-workbook", f"{book} yielded no {ISQ_YEAR} rows at sex code "
                                           f"{ISQ_SEX_TOTAL} — a total over nothing")
    totals = {}
    for code, fan in fans.items():
        if set(fan) != set(ISQ_SCENARIOS) or len(set(fan.values())) != 1:
            raise ProbeRefusal(
                "isq-workbook",
                f"{book} code {code}: the {ISQ_YEAR} scenario fans disagree ({fan}). {ISQ_YEAR} "
                "is a `réel` observation in this workbook family, so three different values "
                "mean the row being read is not the observation it is taken for.")
        totals[code] = next(iter(fan.values()))
    return {"note": note, "totals": totals, "labels": labels, "header_row": header}


# ===========================================================================================
# Formatting
# ===========================================================================================
def _r4(x: float) -> str:
    return f"{x:.4f}"


def _r3(x: float) -> str:
    return f"{x:.3f}"


def _pct(x: float) -> str:
    return f"{x:+.3f}%"


def _digits(x: float) -> str:
    """The DIGIT token of a percentage — sign stripped, because §6 writes its minus as U+2212
    and a gate keyed on the glyph would break on a typographic choice."""
    return f"{abs(x):.3f}"


def _n(x: float) -> str:
    return f"{int(x):,}"


# ===========================================================================================
# The note
# ===========================================================================================
_SCOPE = (
    "SCOPE OF THIS HEADER (it claims only what it can enforce): every rate, ratio, count, "
    "residual and threshold below is COMPUTED by this run from the live StatCan WDS responses "
    "and the two pinned ISQ workbooks it names — none is transcribed from the ruling it is "
    "checked against. The direction of that check is one-way: §6 states the ruled figures — "
    "ruling S's table row for HORS_RMR's superseded territory and amendment #13 for its ruled "
    "one, both recomputed here — and P10's committed note states the aligned territory's "
    "counts; this run recomputes each and a disagreement REFUSES the run rather than being "
    "published. What is CARRIED rather than recomputed is named where it is used and listed "
    "below: P10's membership derivation and P9's catalogue closure, each at the level its own "
    "note earned. "
    "The threshold in §4 is derived from innocent controls measured here in the same "
    "construction, never inherited. Quoted strings are verbatim from a live response, a "
    "workbook cell or a named in-repo file, and every absence claim is scoped to the search "
    "that produced it."
)
_CITED_LABEL = "Quoted or cited verbatim (not computed here):"


def _summary(*, total: int, derived: int, cited: int) -> str:
    return (f"This run registered {total} provenance-tagged figures: {derived} DERIVED "
            f"(computed here from the live responses and pinned workbooks this run read) and "
            f"{cited} CITED (verbatim from a live response, a workbook header, the spec or a "
            f"named in-repo file). The tagged set is the NARRATIVE figures — the ones a "
            f"sentence rests on. Untagged numerals fall in two other classes, both stated "
            f"rather than left to be assumed: TABLE CELLS, which are counts and rates this run "
            f"computed from the coordinate-keyed live responses printed in the same row, and "
            f"AUDIT METADATA — member ids, dimension positions, member counts, geoLevels and "
            f"row indices. Every one of the three is traceable to a response or a file this "
            f"run read; none is transcribed from the ruling this note is checked against.")


def _sections(where: list) -> list:
    """The whole derivation. Appends nothing global; raises ProbeRefusal on any guard."""
    global LAST_RUN
    LAST_RUN = {}

    # ---------------------------------------------------------------- boundary: wds-meta
    where[0] = "wds-meta"
    pids = [CMA_PID, CD_PID, SIBLING_PID, TOTALPOP_PID]
    metas = _guard_meta(pids, _meta(pids))

    dim_names = {pid: [(d["dimensionPositionId"], d["dimensionNameEn"],
                        len(d.get("member") or []))
                       for d in sorted(metas[pid]["dimension"],
                                       key=lambda d: d["dimensionPositionId"])]
                 for pid in pids}
    shared_shape = ([n for _p, n, _c in dim_names[CMA_PID]]
                    == [n for _p, n, _c in dim_names[CD_PID]])
    if not shared_shape:
        raise ProbeRefusal("wds-meta", "98100621 and 98100622 no longer publish the same "
                                       "dimension list name for name — ruling T's 'one universe "
                                       "at two grains' reading rests on that")

    def geo_members(pid: int) -> list:
        return _dimension(metas, pid, POS_GEO).get("member") or []

    nbsp = {pid: sum(1 for m in geo_members(pid)
                     if m["memberNameEn"] != m["memberNameEn"].rstrip("\xa0")) for pid in pids}

    census_ids = {}
    for pid in (CMA_PID, CD_PID):
        census_ids[pid] = {
            "tenure_total": _member(metas, pid, POS_TENURE, TENURE_TOTAL)["memberId"],
            "tenure_owner": _member(metas, pid, POS_TENURE, TENURE_OWNER)["memberId"],
            "gender": _member(metas, pid, POS_GENDER, GENDER_TOTAL)["memberId"],
            "maint_all": _member(metas, pid, POS_MAINT, MAINT_ALL_PERSONS)["memberId"],
            "maint_primary": _member(metas, pid, POS_MAINT, MAINT_PRIMARY)["memberId"],
            "statistic": _member(metas, pid, POS_STAT, STAT_PEOPLE)["memberId"],
            "suitability": _member(metas, pid, POS_SUIT, SUIT_TOTAL)["memberId"],
            "popchar": {name: _guard_pinned_id(
                pid, name, _member(metas, pid, POS_POPCHAR, name),
                PINNED_POPCHAR_IDS[name])["memberId"] for name in POPCHARS},
        }

    province = {pid: _guard_pinned_id(
        pid, QUEBEC, _member(metas, pid, POS_GEO, QUEBEC, geo_level=PROVINCE_GEO_LEVEL),
        PINNED_GEO_IDS[(pid, QUEBEC)]) for pid in (CMA_PID, CD_PID)}

    # The six wholly-QC CMAs, resolved STRUCTURALLY as the geoLevel-503 children of the Quebec
    # member. The pinned ids are then ASSERTED against that computation, never used to fetch.
    cma_members = sorted(
        (m for m in geo_members(CMA_PID)
         if m.get("parentMemberId") == province[CMA_PID]["memberId"]
         and m.get("geoLevel") == CMA_GEO_LEVEL),
        key=lambda m: m["memberId"])
    if tuple(m["memberId"] for m in cma_members) != PINNED_CMA_IDS:
        raise ProbeRefusal(
            "wds-meta",
            f"the geoLevel-{CMA_GEO_LEVEL} children of Quebec in {table_number(CMA_PID)} are "
            f"{[m['memberId'] for m in cma_members]}, not the pinned {list(PINNED_CMA_IDS)}. "
            "HORS_RMR is the province NET of exactly this set, so a change here silently "
            "redefines the residual geography.")

    ra_members = {
        "MTL_ISLAND_RA06": _guard_pinned_id(
            CD_PID, "Montréal", _member(metas, CD_PID, POS_GEO, "Montréal",
                                        geo_level=CD_GEO_LEVEL,
                                        parent=province[CD_PID]["memberId"]),
            PINNED_GEO_IDS[(CD_PID, "Montréal")]),
        "LAVAL_RA13": _guard_pinned_id(
            CD_PID, "Laval", _member(metas, CD_PID, POS_GEO, "Laval", geo_level=CD_GEO_LEVEL,
                                     parent=province[CD_PID]["memberId"]),
            PINNED_GEO_IDS[(CD_PID, "Laval")]),
    }
    csd_trap = [m for m in geo_members(CD_PID)
                if m["memberNameEn"] == "Laval" and m["memberId"] != ra_members["LAVAL_RA13"]["memberId"]]

    # ------------------------------------------------------- the aligned territory (#13)
    # P10's membership, re-resolved live. The codes come from P10's committed note; every
    # member behind them is looked up in THIS cube by code AND geoLevel, and its Québec-side
    # property is re-established two ways — the SGC prefix and the census tree — in the cube
    # the counts are actually subtracted from.
    where[0] = "p10"
    p10_text = _p10_note()
    p10 = _guard_p10(p10_text)
    where[0] = "wds-meta"
    by_code: dict = {}
    for member in geo_members(CD_PID):
        by_code.setdefault((str(member.get("classificationCode")), member.get("geoLevel")),
                           []).append(member)
    cd_by_id = {m["memberId"]: m for m in geo_members(CD_PID)}
    qc_part, qc_part_split = {}, []
    for code, name in sorted(p10["members"].items()):
        member = _member_by_code(CD_PID, by_code, code, CSD_GEO_LEVEL,
                                 f"{name} — from P10's membership")
        if member["memberNameEn"] != name:
            raise ProbeRefusal(
                "p10",
                f"SGC {code} names {member['memberNameEn']!r} in {table_number(CD_PID)} and "
                f"{name!r} in P10's note. The membership is carried BY REFERENCE, so the two "
                "documents must be describing the same subdivision.")
        parent = cd_by_id.get(member.get("parentMemberId"))
        in_tree = bool(parent and parent.get("geoLevel") == CD_GEO_LEVEL
                       and parent.get("parentMemberId") == province[CD_PID]["memberId"])
        qc_part_split.append(
            (f"{name} ({code})",
             str(member.get("classificationCode")).startswith(QC_SGC_PREFIX), in_tree))
        qc_part[code] = member
    _guard_qc_split(qc_part_split)
    qc_part_parents = {code: cd_by_id[m["parentMemberId"]] for code, m in qc_part.items()}

    sibling_ids = {
        "gender": _member(metas, SIBLING_PID, SIB_POS_GENDER, SIB_GENDER_TOTAL)["memberId"],
        "age": _member(metas, SIBLING_PID, SIB_POS_AGE, SIB_AGE_TOTAL)["memberId"],
        "vismin": _member(metas, SIBLING_PID, SIB_POS_VISMIN, SIB_VISMIN_TOTAL)["memberId"],
        "education": _member(metas, SIBLING_PID, SIB_POS_EDU, SIB_EDU_TOTAL)["memberId"],
        "ownership": _member(metas, SIBLING_PID, SIB_POS_IND, SIB_IND_OWNERSHIP)["memberId"],
        "alone": _member(metas, SIBLING_PID, SIB_POS_IND, SIB_IND_ALONE)["memberId"],
        "immigrant": {name: _member(metas, SIBLING_PID, SIB_POS_IMM, name)["memberId"]
                      for name in SIB_MEMBERS},
    }
    sibling_geo = {"QC_PROVINCE": _guard_pinned_id(
        SIBLING_PID, QUEBEC,
        _member(metas, SIBLING_PID, SIB_POS_GEO, QUEBEC, geo_level=PROVINCE_GEO_LEVEL),
        PINNED_GEO_IDS[(SIBLING_PID, QUEBEC)])}
    for tag, (name, pinned) in SIB_GEO.items():
        hits = [m for m in geo_members(SIBLING_PID)
                if m["memberNameEn"].rstrip("\xa0") == name]
        if len(hits) != 1:
            raise ProbeRefusal("wds-meta", f"{table_number(SIBLING_PID)}: {len(hits)} member(s) "
                                           f"named {name!r} after stripping trailing NBSPs")
        sibling_geo[tag] = _guard_pinned_id(SIBLING_PID, name, hits[0], pinned)

    totalpop_ids = {
        "count": _member(metas, TOTALPOP_PID, TOTALPOP_POS_COUNT,
                         TOTALPOP_COUNT_2021)["memberId"],
        "province": _guard_pinned_id(
            TOTALPOP_PID, QUEBEC,
            _member(metas, TOTALPOP_PID, TOTALPOP_POS_GEO, QUEBEC,
                    geo_level=PROVINCE_GEO_LEVEL), PINNED_GEO_IDS[(TOTALPOP_PID, QUEBEC)]),
    }
    for tag, name in (("MTL_ISLAND_RA06", "Montréal"), ("LAVAL_RA13", "Laval")):
        totalpop_ids[tag] = _member(metas, TOTALPOP_PID, TOTALPOP_POS_GEO, name,
                                    geo_level=CD_GEO_LEVEL,
                                    parent=totalpop_ids["province"]["memberId"])

    # ---------------------------------------------------------------- boundary: wds-data
    where[0] = "wds-data"
    cma_geo_ids = [province[CMA_PID]["memberId"]] + [m["memberId"] for m in cma_members]
    cd_geo_ids = [province[CD_PID]["memberId"]] + [ra_members[t]["memberId"]
                                                   for t in ("MTL_ISLAND_RA06", "LAVAL_RA13")]
    requests, keys = [], {}
    for pid, geo_ids in ((CMA_PID, cma_geo_ids), (CD_PID, cd_geo_ids)):
        part, part_keys = _census_requests(pid, geo_ids, census_ids[pid])
        requests += part
        keys.update({(pid,) + k: v for k, v in part_keys.items()})

    sib_keys = {}
    for tag, member in sorted(sibling_geo.items()):
        for name in SIB_MEMBERS:
            for indicator in ("ownership", "alone"):
                coordinate = coord(member["memberId"], sibling_ids["gender"],
                                   sibling_ids["age"], sibling_ids["immigrant"][name],
                                   sibling_ids["vismin"], sibling_ids["education"],
                                   sibling_ids[indicator])
                requests.append({"productId": SIBLING_PID, "coordinate": coordinate,
                                 "latestN": 1})
                sib_keys[(tag, name, indicator)] = (SIBLING_PID, coordinate)

    pop_keys = {}
    for tag in ("province", "MTL_ISLAND_RA06", "LAVAL_RA13"):
        coordinate = coord(totalpop_ids[tag]["memberId"], totalpop_ids["count"])
        requests.append({"productId": TOTALPOP_PID, "coordinate": coordinate, "latestN": 1})
        pop_keys[tag] = (TOTALPOP_PID, coordinate)

    csd_part, csd_keys = _csd_requests(CD_PID, [m["memberId"] for m in qc_part.values()],
                                       census_ids[CD_PID])
    requests += csd_part
    # The suppressible scope, declared BEFORE the response is read and by GEOGRAPHY: these 16
    # subdivisions and nothing else. Declaring it after would let an outage anywhere name
    # itself a publication rule.
    suppressible = set(csd_keys.values())
    series, withheld = _guard_response(requests, _data(requests), suppressible)

    cma_cells = _census_cells(series, {k[1:]: v for k, v in keys.items() if k[0] == CMA_PID},
                              cma_geo_ids)
    cd_cells = _census_cells(series, {k[1:]: v for k, v in keys.items() if k[0] == CD_PID},
                             cd_geo_ids)
    sibling = {k: series[v] / 100 for k, v in sib_keys.items()}
    total_pop = {k: series[v] for k, v in pop_keys.items()}

    # ---------------------------------------------------------------- boundary: isq-workbook
    where[0] = "isq-workbook"
    isq = {book: _isq_totals(book, _isq_rows(book))
           for book in (ISQ_RA_BOOK, ISQ_RMR_BOOK)}
    # Resolve BY LABEL first, assert the pinned codes, and fetch on what came back — so the
    # ISQ side of the gate is keyed on the workbook's own identity for the row, never on a
    # typed integer. Everything below reads `ra_code`/`hors_code`/`gatineau_code`.
    isq_code = _guard_isq_labels(isq)
    ra_code = {tag: isq_code[(ISQ_RA_BOOK, ISQ_RA_LABELS[tag])] for tag in ISQ_RA_LABELS}
    hors_code = isq_code[(ISQ_RMR_BOOK, ISQ_HORS_LABEL)]
    gatineau_code = isq_code[(ISQ_RMR_BOOK, ISQ_GATINEAU_LABEL)]
    isq_province = {book: _guard_isq(book, isq[book]["totals"],
                                     isq_code[(book, ISQ_PROVINCE_LABEL)])
                    for book in isq}
    if isq_province[ISQ_RA_BOOK] != isq_province[ISQ_RMR_BOOK]:
        raise ProbeRefusal(
            "isq-workbook",
            f"the two ISQ workbooks disagree on the {ISQ_YEAR} province total "
            f"({isq_province[ISQ_RA_BOOK]:,} vs {isq_province[ISQ_RMR_BOOK]:,}). The RA and CMA "
            "controls are only commensurable because their share denominators are the SAME "
            "number on the ISQ side.")
    isq_total = isq_province[ISQ_RA_BOOK]

    # ---------------------------------------------------------------- the derivation
    where[0] = "derivation"
    prov_cma = cma_cells[province[CMA_PID]["memberId"]]
    prov_cd = cd_cells[province[CD_PID]["memberId"]]
    universe_drift = _guard_universe(
        {PINNED_POPCHAR_IDS[n]: (c.persons, c.maintainers, c.owner_maintainers)
         for n, c in prov_cma.items()},
        {PINNED_POPCHAR_IDS[n]: (c.persons, c.maintainers, c.owner_maintainers)
         for n, c in prov_cd.items()})
    if prov_cma[PC_TOTAL_AGE].persons != prov_cd[PC_TOTAL_AGE].persons:
        raise ProbeRefusal("universe", "the two cubes disagree on the province's total "
                                       "private-household persons, which is the share "
                                       "denominator BOTH sides of the territory gate use")
    census_total = prov_cma[PC_TOTAL_AGE].persons
    # The rounding explanation for the ±5 cells, MEASURED rather than asserted: if every
    # published province count in both cubes is a multiple of 5, a 5-person disagreement is
    # what independent rounding produces. If some value were NOT a multiple of 5, that
    # explanation would be unavailable and the note must not offer it.
    province_counts = [v for cells in (prov_cma, prov_cd) for cell in cells.values()
                       for v in (cell.persons, cell.maintainers, cell.owner_maintainers)]

    named_cma = {m["memberNameEn"]: m for m in cma_members}
    mtl = cma_cells[named_cma["Montréal (CMA), Que."]["memberId"]]
    qc = cma_cells[named_cma["Québec (CMA), Que."]["memberId"]]
    six_sum = {name: cell_sum([cma_cells[m["memberId"]][name] for m in cma_members])
               for name in POPCHARS}
    hors = {name: cell_minus(prov_cma[name], six_sum[name]) for name in POPCHARS}

    geographies = {"MTL_RMR": mtl, "QC_RMR": qc, "HORS_RMR": hors,
                   "MTL_ISLAND_RA06": cd_cells[ra_members["MTL_ISLAND_RA06"]["memberId"]],
                   "LAVAL_RA13": cd_cells[ra_members["LAVAL_RA13"]["memberId"]],
                   "QC_PROVINCE": prov_cma}
    ruled = {g: (cells[PC_SETTLED].headship,
                 ownership_ratio(cells[PC_SETTLED], cells[PC_NONIMM]))
             for g, cells in geographies.items()}
    recent = {g: (cells[PC_RECENT].headship,
                  ownership_ratio(cells[PC_RECENT], cells[PC_NONIMM]))
              for g, cells in geographies.items()}
    pooled = {g: (cells[PC_ALL_IMM].headship,
                  ownership_ratio(cells[PC_ALL_IMM], cells[PC_NONIMM]))
              for g, cells in geographies.items()}

    # ------------------------------------------------- HORS_RMR at ALIGNED territory (#13)
    # The rate's territory must match the flow's. `hors` above is the shipped residual and
    # still carries the Québec side of Ottawa-Gatineau; the ISQ flow row it multiplies does
    # not. Netting P10's 16 Québec-part subdivisions out of it makes the two the same
    # territory. FIELD-WISE, with every withheld count carried at a bound the same cube
    # publishes — never dropped, because dropping a geography discards published counts and
    # biases the point estimate rather than widening the interval.
    where[0] = "suppression"
    rows, csd_settled, complete = [], {}, {}
    for code, member in sorted(qc_part.items()):
        geo = member["memberId"]
        total = _census_cell(series, csd_keys, geo, PC_TOTAL_AGE)
        nonimm = _census_cell(series, csd_keys, geo, PC_NONIMM)
        complete[f"{member['memberNameEn']} ({code})"] = bool(total and nonimm)
        values = _census_fields(series, csd_keys, geo, PC_SETTLED)
        # The bound: all immigrants and non-permanent residents together, which CONTAINS the
        # ruled `Before 2016` member by construction and is published at every one of these
        # geographies — `_guard_required_complete` makes that a checked precondition.
        bound, clamped = (complement_bound(total, nonimm) if (total and nonimm) else (None, 0))
        csd_settled[code] = (values, bound, clamped)
        rows.append((values, (bound.persons, bound.maintainers, bound.owner_maintainers)
                     if bound else (None, None, None)))
    _guard_required_complete(complete)
    clamped_fields = sum(c for _v, _b, c in csd_settled.values())
    withheld_codes = sorted(code for code, (values, _b, _c) in csd_settled.items()
                            if any(v is None for v in values))
    low_counts, high_counts, withheld_fields = bounded_sum(rows)
    _guard_withheld_accounted(reported=len(withheld), carried=withheld_fields)
    qc_low, qc_high = Cell(*low_counts), Cell(*high_counts)
    qc_totals = {name: cell_sum([_census_cell(series, csd_keys, m["memberId"], name)
                                 for m in qc_part.values()])
                 for name in (PC_TOTAL_AGE, PC_NONIMM)}
    aligned_cells = {PC_SETTLED: cell_minus(hors[PC_SETTLED], qc_low),
                     PC_NONIMM: cell_minus(hors[PC_NONIMM], qc_totals[PC_NONIMM]),
                     PC_TOTAL_AGE: cell_minus(hors[PC_TOTAL_AGE], qc_totals[PC_TOTAL_AGE])}
    envelope = bounded_pair(hors[PC_SETTLED], qc_low, qc_high, aligned_cells[PC_NONIMM])
    _guard_ratio_band(*envelope["ratio_band"])
    aligned = (envelope["headship"], envelope["ratio"])
    shipped_pair = ruled["HORS_RMR"]
    aligned_move = {
        "headship": relative_pct(aligned[0], shipped_pair[0]),
        "ratio": relative_pct(aligned[1], shipped_pair[1]),
        "leg": relative_pct(aligned[0] * aligned[1], shipped_pair[0] * shipped_pair[1]),
    }
    # The RULED pair for HORS_RMR is the aligned one from here on: `ruled` is what the note's
    # headline table, its DECISION tokens and task 25b's join table all read. The shipped
    # residual stays measured and published under its own name — §6's ruling-S row still
    # states it, and §2a's recent-member readings are properties of that same territory.
    ruled = dict(ruled, HORS_RMR=aligned)

    # ---------------------------------------------------------------- the territory gate
    where[0] = "territory"
    joined, innocent = [], {}
    for member in cma_members:
        code = int(member["classificationCode"])
        label = isq[ISQ_RMR_BOOK]["labels"].get(code, "")
        joined.append((member["memberNameEn"], code,
                       code if code in isq[ISQ_RMR_BOOK]["totals"] else None))
        if code in isq[ISQ_RMR_BOOK]["totals"]:
            innocent[member["memberNameEn"]] = share_residual_pct(
                part_census=cma_cells[member["memberId"]][PC_TOTAL_AGE].persons,
                total_census=census_total,
                part_isq=isq[ISQ_RMR_BOOK]["totals"][code], total_isq=isq_total)
    _guard_code_join(joined)
    innocent_labels = {m["memberNameEn"]: isq[ISQ_RMR_BOOK]["labels"][
        int(m["classificationCode"])] for m in cma_members}
    bounds = derive_threshold(innocent)

    ra_residual = {tag: share_residual_pct(
        part_census=cd_cells[ra_members[tag]["memberId"]][PC_TOTAL_AGE].persons,
        total_census=census_total,
        part_isq=isq[ISQ_RA_BOOK]["totals"][ra_code[tag]], total_isq=isq_total)
        for tag in ISQ_RA_LABELS}
    _guard_territory(ra_residual, bounds["threshold"])

    # The SIGN, decomposed. The gate rules on |residual| (amendment #11 sets the threshold on
    # the maximum innocent |residual|), but the sign is the informative direction for a
    # territory difference and here it does NOT run the same way as the magnitude. Publishing
    # only the absolute comparison would leave the adverse half of this run's own measurement
    # unsaid; attributing the sign would claim what §4c says this construction cannot decide.
    signed_low_name, signed_low = min(innocent.items(), key=lambda kv: kv[1])
    signed_high_name, signed_high = max(innocent.items(), key=lambda kv: kv[1])
    negative_controls = sorted(n for n, v in innocent.items() if v < 0)
    positive_controls = sorted(n for n, v in innocent.items() if v > 0)
    ra_positive = [t for t, v in ra_residual.items() if v > 0]
    top_positive = max((innocent[n] for n in positive_controls), default=None)
    above_every_positive = sorted(
        t for t, v in ra_residual.items() if top_positive is not None and v > top_positive)

    diagnostic = {
        "MTL_ISLAND_RA06": (total_pop["MTL_ISLAND_RA06"]
                            / isq[ISQ_RA_BOOK]["totals"][ra_code["MTL_ISLAND_RA06"]] - 1) * 100,
        "LAVAL_RA13": (total_pop["LAVAL_RA13"]
                       / isq[ISQ_RA_BOOK]["totals"][ra_code["LAVAL_RA13"]] - 1) * 100,
        "QC_PROVINCE": (total_pop["province"] / isq_total - 1) * 100,
    }

    # ---------------------------------------------------------------- the floor gate
    where[0] = "floor"
    alone = {tag: sibling[(tag, SIB_SETTLED, "alone")] for tag in SIB_GEO}
    alone_pooled = {tag: sibling[(tag, SIB_ALL_IMM, "alone")] for tag in SIB_GEO}
    _guard_floor({tag: ruled[tag][0] for tag in SIB_GEO}, alone)
    covered = set(SIB_GEO)
    modeled = ("MTL_RMR", "QC_RMR", "HORS_RMR", "MTL_ISLAND_RA06", "LAVAL_RA13")
    not_covered = [g for g in modeled if g not in covered]
    # The NOT-COVERED verdict, EARNED. A prose "none of its members is a census division"
    # beside a member COUNT is the depth-3 shape this arc grades: the count is correct and it
    # entails nothing about census divisions. So the cube's members are actually searched —
    # by SGC code for the two census divisions, and by geoLevel for the class as a whole.
    sib_members = geo_members(SIBLING_PID)
    sib_levels: dict = {}
    for member in sib_members:
        sib_levels[member.get("geoLevel")] = sib_levels.get(member.get("geoLevel"), 0) + 1
    floor_search = {}
    for tag in not_covered:
        wanted_code = (ra_members[tag]["classificationCode"] if tag in ra_members else None)
        floor_search[tag] = {
            "sgc": wanted_code,
            "predicate": (f"classificationCode == {wanted_code!r}" if wanted_code
                          else f"memberNameEn containing any of {list(HORS_NAME_MARKERS)}"),
            "code_hits": [m["memberNameEn"] for m in sib_members
                          if (str(m.get("classificationCode")) == str(wanted_code)
                              if wanted_code is not None
                              else any(mark in m["memberNameEn"].casefold()
                                       for mark in HORS_NAME_MARKERS))],
        }
    if any(hit["code_hits"] for hit in floor_search.values()):
        raise ProbeRefusal(
            "floor",
            f"{table_number(SIBLING_PID)} DOES publish a member for a geography this run "
            f"recorded as uncovered: {floor_search}. The absence claim would be false, and a "
            "floor left unwired over an available member is a gate declining to run.")
    sibling_ratio = {tag: (sibling[(tag, SIB_SETTLED, "ownership")]
                           / sibling[(tag, SIB_NONIMM, "ownership")]) for tag in SIB_GEO}
    sibling_ratio["QC_PROVINCE"] = (sibling[("QC_PROVINCE", SIB_SETTLED, "ownership")]
                                    / sibling[("QC_PROVINCE", SIB_NONIMM, "ownership")])

    # ---------------------------------------------------------------- citation coupling
    where[0] = "citation"
    s6 = _spec_s6()
    spec_rows = _guard_s6_rows(s6)
    # §6 rules HORS_RMR twice, and each construction is bound to the statement that rules IT:
    # the SHIPPED residual against ruling S's table row (which still states it, correctly, as
    # the record of what shipped) and the ALIGNED pair against amendment #13. Binding only the
    # table row is how a superseded pair rode through a green suite: the table cannot state a
    # value #13 moved into prose, so a table-keyed gate is blind to exactly the move it exists
    # to catch.
    _guard_rows_match_spec(spec_rows, dict({g: (_r4(ruled[g][0]), _r4(ruled[g][1]))
                                            for g in modeled},
                                           HORS_RMR=(_r4(shipped_pair[0]),
                                                     _r4(shipped_pair[1]))))
    amd13 = _guard_amendment13(s6)
    _guard_amendment13_match(
        amd13, _r4(aligned[0]), _r4(aligned[1]),
        {"headship": "-".join(_r4(b) for b in envelope["headship_band"]),
         "ratio": "-".join(_r4(b) for b in envelope["ratio_band"])})
    coupled = [(f"{g} headship", _r4(ruled[g][0])) for g in modeled]
    coupled += [(f"{g} ratio", _r4(ruled[g][1])) for g in modeled]
    coupled += [
        ("HORS_RMR shipped headship", _r4(shipped_pair[0])),
        ("HORS_RMR shipped ratio", _r4(shipped_pair[1])),
        ("HORS_RMR aligned headship envelope low", _r4(envelope["headship_band"][0])),
        ("HORS_RMR aligned headship envelope high", _r4(envelope["headship_band"][1])),
        ("HORS_RMR aligned ratio envelope low", _r4(envelope["ratio_band"][0])),
        ("HORS_RMR aligned ratio envelope high", _r4(envelope["ratio_band"][1])),
        ("HORS_RMR immigrant demand leg move", _r3(aligned_move["leg"])),
        ("province headship", _r4(ruled["QC_PROVINCE"][0])),
        ("province ratio", _r4(ruled["QC_PROVINCE"][1])),
        ("island non-immigrant owner propensity",
         _r4(geographies["MTL_ISLAND_RA06"][PC_NONIMM].owner_propensity)),
        ("Laval non-immigrant owner propensity",
         _r4(geographies["LAVAL_RA13"][PC_NONIMM].owner_propensity)),
        ("Laval p_imm product",
         _r3(geographies["LAVAL_RA13"][PC_NONIMM].owner_propensity * ruled["LAVAL_RA13"][1])),
        ("general MTL headship", _r4(mtl[PC_TOTAL_AGE].headship)),
        ("pooled MTL headship", _r4(pooled["MTL_RMR"][0])),
        ("sibling MTL ratio", _r4(sibling_ratio["MTL_RMR"])),
        ("sibling QC_RMR ratio", _r4(sibling_ratio["QC_RMR"])),
        ("sibling province ratio", _r4(sibling_ratio["QC_PROVINCE"])),
        ("settled living-alone MTL", _r3(alone["MTL_RMR"])),
        ("settled living-alone QC_RMR", _r3(alone["QC_RMR"])),
        ("pooled living-alone MTL", _r3(alone_pooled["MTL_RMR"])),
        ("pooled living-alone QC_RMR", _r3(alone_pooled["QC_RMR"])),
        ("QC total maintainers", _n(prov_cma[PC_TOTAL_AGE].maintainers)),
        ("QC total persons", _n(prov_cma[PC_TOTAL_AGE].persons)),
        ("ruled triple persons", _n(prov_cma[PC_SETTLED].persons)),
        ("ruled triple maintainers", _n(prov_cma[PC_SETTLED].maintainers)),
        ("ruled triple owner-maintainers", _n(prov_cma[PC_SETTLED].owner_maintainers)),
        ("RA06 share residual", _digits(ra_residual["MTL_ISLAND_RA06"])),
        ("RA13 share residual", _digits(ra_residual["LAVAL_RA13"])),
        ("RA06 total-population diagnostic", _digits(diagnostic["MTL_ISLAND_RA06"])),
        ("RA13 total-population diagnostic", _digits(diagnostic["LAVAL_RA13"])),
        ("province total-population diagnostic", _digits(diagnostic["QC_PROVINCE"])),
    ]
    coupled += [(f"{g} recent headship", _r4(recent[g][0]))
                for g in ("MTL_RMR", "QC_RMR", "HORS_RMR")]
    coupled += [(f"{g} recent ratio", _r4(recent[g][1]))
                for g in ("MTL_RMR", "QC_RMR", "HORS_RMR")]
    _guard_citation("spec §6", s6, coupled)
    # The aligned territory's COUNTS are stated by P10 and by no ruling — the spec rules rates.
    # Bound to the note that measured them, so the one construction this run carries a
    # membership for cannot drift from the probe that derived it.
    coupled_p10 = [
        ("aligned settled persons", _n(aligned_cells[PC_SETTLED].persons)),
        ("aligned settled maintainers", _n(aligned_cells[PC_SETTLED].maintainers)),
        ("aligned settled owner-maintainers",
         _n(aligned_cells[PC_SETTLED].owner_maintainers)),
        ("aligned non-immigrant propensity",
         _r4(aligned_cells[PC_NONIMM].owner_propensity)),
        ("aligned settled propensity", _r4(aligned_cells[PC_SETTLED].owner_propensity)),
        ("aligned headship", _r4(aligned[0])),
        ("aligned ratio", _r4(aligned[1])),
        ("aligned headship move", _r3(aligned_move["headship"])),
        ("aligned ratio move", _r3(aligned_move["ratio"])),
        ("immigrant demand leg move", _r3(aligned_move["leg"])),
    ]
    _guard_citation(f"probes/{P10_NOTE.name}", p10_text, coupled_p10)

    where[0] = "p9"
    p9 = _guard_p9(_p9_note())

    where[0] = "constants"
    # The MODULE DOCSTRING only, then its paragraphs. Splitting the whole file on blank lines
    # ran the quote past the closing triple-quote and pulled three import statements into a
    # sentence presented as a verbatim citation — measured on the first generated note.
    source = _constants_source()
    doc = source.split('"""')[1] if source.count('"""') >= 2 else ""
    anti = next((" ".join(block.split()) for block in doc.split("\n\n")
                 if "ANTI-PATTERN" in block and "pooled" in block.lower()), "")
    if not anti:
        raise ProbeRefusal("constants", "loaders/constants.py no longer records the POOLED "
                                        "anti-pattern block this note distinguishes itself "
                                        "from — the distinction would be an unsourced claim")

    LAST_RUN = {
        "territory": {"innocent": innocent, "ra": ra_residual,
                      "threshold": bounds["threshold"], "bounds": bounds,
                      "source_pids": [CMA_PID, CD_PID, "ISQ"],
                      "isq_code": ra_code,
                      "signed": {"low": signed_low, "high": signed_high,
                                 "negative": negative_controls,
                                 "positive": positive_controls,
                                 "ra_positive": ra_positive,
                                 "above_every_positive": above_every_positive}},
        "floor": {"member": SIB_SETTLED, "share": alone, "pooled": alone_pooled,
                  "searched_members": len(sib_members), "search": floor_search,
                  "geo_levels": sib_levels, "not_covered": not_covered},
        "universe": {"drifted": universe_drift},
        "coupled": coupled,
        "coupled_p10": coupled_p10,
        "ruled": ruled,
        "aligned": {"pair": aligned, "shipped": shipped_pair, "move": aligned_move,
                    "cells": aligned_cells, "headship_band": envelope["headship_band"],
                    "ratio_band": envelope["ratio_band"],
                    "members": {code: m["memberId"] for code, m in qc_part.items()},
                    "withheld_fields": withheld_fields, "withheld_codes": withheld_codes,
                    "withheld_cells": len(withheld), "clamped_fields": clamped_fields,
                    "qc_low": qc_low, "qc_high": qc_high,
                    "p10": {k: v for k, v in p10.items() if k != "members"}},
    }

    # ======================================================================= the note
    where[0] = "note"
    f_prov_pers = Fact.derived(_n(prov_cma[PC_TOTAL_AGE].persons),
                               f"{table_number(CMA_PID)} Quebec, Total-Age, all persons")
    f_prov_maint = Fact.derived(_n(prov_cma[PC_TOTAL_AGE].maintainers),
                                f"{table_number(CMA_PID)} Quebec, Total-Age, primary "
                                f"household maintainers")
    f_p3 = Fact.cited("8,308,475", "probe P3, quoted in spec §6")
    f_t13b = Fact.cited("3,749,035", "loaders/census.py T13b docstring")
    f_anti = Fact.cited("1.144", f"loaders/constants.py, verbatim: \"{anti}\"")
    f_isq_ra_note = Fact.cited(isq[ISQ_RA_BOOK]["note"],
                               f"header note of the pinned {ISQ_RA_BOOK}")
    f_isq_rmr_note = Fact.cited(isq[ISQ_RMR_BOOK]["note"],
                                f"header note of the pinned {ISQ_RMR_BOOK}")
    # §2b. The aligned pair and its envelope are DERIVED — recomputed here from live cells at
    # P10's membership — while the membership's own derivation and gate are CITED from P10's
    # note. The split is the point: this run measures the values and carries the territory.
    f_aligned_headship = Fact.derived(
        _r4(aligned[0]), "HORS_RMR aligned: maintainers / persons on the Before 2016 member, "
                         "shipped residual net of the QC-part CSDs' published counts")
    f_aligned_ratio = Fact.derived(
        _r4(aligned[1]), "HORS_RMR aligned: settled owner-maintainer propensity / "
                         "non-immigrant, both netted over the same aligned territory")
    f_move_headship = Fact.derived(_pct(aligned_move["headship"]),
                                   "aligned headship against the shipped one")
    f_move_ratio = Fact.derived(_pct(aligned_move["ratio"]),
                                "aligned ratio against the shipped one")
    f_move_leg = Fact.derived(_pct(aligned_move["leg"]),
                              "the PRODUCT headship × ratio, aligned against shipped — the "
                              "immigrant demand leg at HORS_RMR")
    f_band_headship_low = Fact.derived(_r4(envelope["headship_band"][0]),
                                       "aligned headship at the suppression box's low corner")
    f_band_headship_high = Fact.derived(_r4(envelope["headship_band"][1]),
                                        "aligned headship at the suppression box's high corner")
    f_band_ratio_low = Fact.derived(_r4(envelope["ratio_band"][0]),
                                    "aligned ratio at the suppression box's low corner")
    f_band_ratio_high = Fact.derived(_r4(envelope["ratio_band"][1]),
                                     "aligned ratio at the suppression box's high corner")
    f_p10_membership = Fact.cited(p10["DECISION-MEMBERSHIP"],
                                  f"probes/{P10_NOTE.name} DECISION-MEMBERSHIP, read this run")
    f_p10_gate = Fact.cited(p10["DECISION-MEMBERSHIP-GATE"],
                            f"probes/{P10_NOTE.name} DECISION-MEMBERSHIP-GATE, read this run")
    f_p10_construction = Fact.cited(
        p10["DECISION-CONSTRUCTION"],
        f"probes/{P10_NOTE.name} DECISION-CONSTRUCTION, read this run")
    f_p9_verdict = Fact.cited(p9["DECISION-VERDICT"],
                              f"probes/{P9_NOTE.name} DECISION-VERDICT, read this run")
    f_p9_level = Fact.cited(p9["DECISION-CLOSURE-LEVEL"],
                            f"probes/{P9_NOTE.name} DECISION-CLOSURE-LEVEL, read this run")
    f_p9_residual = Fact.cited(p9["DECISION-RESIDUAL"],
                               f"probes/{P9_NOTE.name} DECISION-RESIDUAL, read this run")
    f_isq_total = Fact.derived(_n(isq_total), f"the {ISQ_YEAR} province row of both pinned ISQ "
                                              f"workbooks, each closing exactly on its own parts")
    f_threshold = Fact.derived(_r3(bounds["threshold"]) + "%",
                               f"max innocent |residual| {_r3(bounds['max_abs'])}% "
                               f"({bounds['max_name']}) + {MARGIN_FRACTION:.0%} of it")
    f_ra06 = Fact.derived(_pct(ra_residual["MTL_ISLAND_RA06"]),
                          "province-controlled share residual, census 98100622 vs ISQ RA")
    f_ra13 = Fact.derived(_pct(ra_residual["LAVAL_RA13"]),
                          "province-controlled share residual, census 98100622 vs ISQ RA")
    # §4d's two code axes, both registered rather than printed bare. The ISQ codes are
    # DERIVED — this run resolves them by label in the workbook it read — while the SGC codes
    # are CITED: they are read verbatim off live metadata, not computed from anything.
    f_ra_code = {tag: Fact.derived(
        ra_code[tag], f"{ISQ_RA_BOOK}: the code of the row labelled "
                      f"{_isq_label(isq[ISQ_RA_BOOK]['labels'][ra_code[tag]])!r}, resolved by "
                      f"label this run and asserted against the pinned code")
        for tag in ISQ_RA_LABELS}
    f_sgc_ra06 = Fact.cited(
        ra_members["MTL_ISLAND_RA06"]["classificationCode"],
        f"{table_number(CD_PID)} `{ra_members['MTL_ISLAND_RA06']['memberNameEn']}` "
        f"classificationCode, verbatim from live metadata")
    f_sgc_ra13 = Fact.cited(
        ra_members["LAVAL_RA13"]["classificationCode"],
        f"{table_number(CD_PID)} `{ra_members['LAVAL_RA13']['memberNameEn']}` "
        f"classificationCode, verbatim from live metadata")

    def rate_rows(rows) -> list:
        """One row per (label, cells) pair — the counts and the rates from the SAME cells.

        Keyed on the cells rather than on a geography name, because HORS_RMR now has two
        constructions in this table and a row that took its counts from one and its rates from
        the other would be arithmetically impossible while looking entirely plausible.
        """
        out = []
        for label, cells in rows:
            settled = cells[PC_SETTLED]
            headship = settled.headship
            ratio = ownership_ratio(settled, cells[PC_NONIMM])
            out.append(
                f"| {label} | {_n(settled.persons)} | {_n(settled.maintainers)} | "
                f"{_n(settled.owner_maintainers)} | "
                f"**{Fact.derived(_r4(headship), f'{label}: maintainers / persons on the {PC_SETTLED} member')}** | "
                f"{_r4(cells[PC_NONIMM].owner_propensity)} | "
                f"{_r4(settled.owner_propensity)} | "
                f"**{Fact.derived(_r4(ratio), f'{label}: settled owner-maintainer propensity / non-immigrant')}** |")
        return out

    def member_rows(geos) -> list:
        """Every population-characteristic member × geography, both quantities.

        The FULL cross rather than the ruled row plus a summary sentence: the ruled cut is
        only judgeable against the members it was chosen over, and a reader who can see
        `Before 2016` but not `Recent immigrants` beside it cannot price the flow-vs-stock
        choice §6 names.
        """
        out = []
        for geo in geos:
            cells = geographies[geo]
            for name in (PC_NONIMM, PC_ALL_IMM, PC_SETTLED, PC_RECENT, PC_NPR):
                cell = cells[name]
                out.append(
                    f"| {geo} | {name} | {_n(cell.persons)} | {_n(cell.maintainers)} | "
                    f"{_n(cell.owner_maintainers)} | {_r4(cell.headship)} | "
                    f"{_r4(cell.owner_propensity)} | "
                    f"{_r4(cell.owner_propensity / cells[PC_NONIMM].owner_propensity)} |")
        return out

    _MEMBER_HEADER = [
        "| geography | population-characteristic member | persons | maintainers | "
        "owner-maintainers | headship | owner propensity | ratio vs non-imm |",
        "|---|---|---:|---:|---:|---:|---:|---:|",
    ]

    lines = [
        "## 1. What this measures, and out of which universe",
        "",
        f"Both immigrant inputs come from ONE cube on ONE member: **{table_number(CMA_PID)}** "
        f"(`{PC_SETTLED}` of `Population characteristics (46)`), with its census-division "
        f"sibling **{table_number(CD_PID)}** supplying the two census divisions ruling T "
        "measures. Zero geography transport and zero metric transport: the cube publishes the "
        "owner-MAINTAINER propensity §6 defines, as counts, in one universe.",
        "",
        "| cube | title (live) | released | archive |",
        "|---|---|---|---|",
    ]
    for pid in pids:
        lines.append(f"| [{table_number(pid)}]({table_url(pid)}) | {metas[pid]['cubeTitleEn']} "
                     f"| {metas[pid]['releaseTime']} | "
                     f"{metas[pid]['archiveStatusEn'].split(' - ')[0]} |")
    lines += [
        "",
        f"- `{table_number(CMA_PID)}` and `{table_number(CD_PID)}` publish the SAME dimension "
        f"list, name for name, in this order: "
        + ", ".join(f"`{n}`" for _p, n, _c in dim_names[CMA_PID]) + ".",
        f"- headship = `{MAINT_PRIMARY}` ÷ `{MAINT_ALL_PERSONS}`. The second member is ALL "
        "persons, not a maintainer subset — the two names sit next to each other and reading "
        "them the other way inverts the rate.",
        f"- ratio = the `{PC_SETTLED}` owner-maintainer propensity ÷ the `{PC_NONIMM}` one, "
        f"where propensity = `{TENURE_OWNER}` maintainers ÷ all maintainers.",
        f"- Every member id below was resolved BY NAME from the live metadata and then checked "
        f"against the id the ruling pinned; a move refuses the run rather than being corrected "
        f"silently.",
        "",
        "### 1a. Three universe corroborations",
        "",
        f"1. Québec total primary household maintainers read **{f_prov_maint}** from this "
        f"cube's maintainer axis. The tree already cites that same published private-household "
        f"count in `loaders/census.py`'s T13b docstring: {f_t13b}. The agreement is the "
        "universe check — this cube's maintainer universe IS the private-household universe, "
        "which is what lets a headship computed here sit beside rates derived from that count.",
        f"2. Québec total persons read **{f_prov_pers}** against probe P3's independently "
        f"measured {f_p3} private-household persons — a gap of "
        f"{prov_cma[PC_TOTAL_AGE].persons - 8308475} persons, one rounding step, from a "
        "different cube and a different extraction path.",
        f"3. Geography labels: `{table_number(CMA_PID)}` carries "
        f"{Fact.derived(nbsp[CMA_PID], 'geography members ending in a non-breaking space')} of "
        f"{len(geo_members(CMA_PID))} members with a trailing non-breaking space and "
        f"`{table_number(CD_PID)}` carries {nbsp[CD_PID]} of {len(geo_members(CD_PID))}, while "
        f"`{table_number(SIBLING_PID)}` carries "
        f"{Fact.derived(nbsp[SIBLING_PID], 'sibling geography members ending in a NBSP')} of "
        f"{len(geo_members(SIBLING_PID))}. Scanned over EVERY geography member of each cube, "
        "not over the handful this run resolves.",
        "",
        "## 2. The ruled inputs, RECOMPUTED (98-10-0621-01, CMA grain)",
        "",
        "| geography | persons | maintainers | owner-maintainers | HEADSHIP | non-imm "
        "propensity | settled propensity | RATIO |",
        "|---|---:|---:|---:|---:|---:|---:|---:|",
    ]
    lines += rate_rows([("MTL_RMR", geographies["MTL_RMR"]),
                        ("QC_RMR", geographies["QC_RMR"]),
                        ("HORS_RMR", aligned_cells),
                        ("HORS_RMR (SUPERSEDED — Gatineau IN)", hors),
                        ("QC_PROVINCE", prov_cma)])
    lines += [
        "",
        f"`HORS_RMR` is the province NET of the six wholly-QC CMAs AND net of the "
        f"Ottawa-Gatineau CMA's Québec-side census subdivisions — the operand-aligned territory "
        f"amendment #13 rules, §2b. The six are resolved STRUCTURALLY as the "
        f"geoLevel-{CMA_GEO_LEVEL} children of the Quebec member "
        f"(memberId {province[CMA_PID]['memberId']}): "
        + ", ".join(f"{m['memberNameEn']} (id {m['memberId']}, code {m['classificationCode']})"
                    for m in cma_members)
        + ". The Québec side of Ottawa-Gatineau is parented to Ontario in this cube, so it is "
        "NOT one of those six and stays inside the residual until it is netted out by "
        "subdivision — which is the whole of the correction below. The SUPERSEDED row is the "
        "construction that stopped there: it is published because §6's ruling-S table still "
        f"states its pair ({_r4(shipped_pair[0])} / {_r4(shipped_pair[1])}) as the record of "
        "what shipped, and because §2a's per-member readings are properties of that same "
        "territory. It is not a candidate; it is the measurement being corrected.",
        "",
        f"**The ISQ RMR workbook's own `{isq[ISQ_RMR_BOOK]['labels'][hors_code].strip()}` "
        f"row is what the ALIGNMENT closes on, and the gap is measured rather than argued.** "
        f"ISQ publishes {_n(isq[ISQ_RMR_BOOK]['totals'][hors_code])} for {ISQ_YEAR}; province "
        f"net of the same six CMAs is "
        f"{_n(isq_total - sum(isq[ISQ_RMR_BOOK]['totals'][int(m['classificationCode'])] for m in cma_members))}"
        f" — a difference of "
        f"{Fact.derived(_n(isq_total - sum(isq[ISQ_RMR_BOOK]['totals'][int(m['classificationCode'])] for m in cma_members) - isq[ISQ_RMR_BOOK]['totals'][hors_code]), 'ISQ province net of the six wholly-QC CMAs, minus ISQ published outside-the-CMAs row')}"
        f", which is exactly the workbook's `{isq[ISQ_RMR_BOOK]['labels'][gatineau_code].strip()}` "
        f"row ({_n(isq[ISQ_RMR_BOOK]['totals'][gatineau_code])}). ISQ nets Gatineau out on its "
        "side; the SUPERSEDED census construction kept it in, and that mismatch IS the defect "
        "#13 corrects. Two rows with the same name for two different territories is precisely "
        "the substitution this note refuses to make — and the aligned row above is the one "
        "that no longer makes it.",
        "",
        f"The province row differs from the residual precisely because the province CONTAINS "
        f"the CMAs: {_r4(ruled['QC_PROVINCE'][0])} / {_r4(ruled['QC_PROVINCE'][1])} at "
        f"province level against {_r4(shipped_pair[0])} / {_r4(shipped_pair[1])} net "
        "of them.",
        "",
        "### 2a. EVERY population-characteristic member at EVERY CMA-grain geography",
        "",
        "Both quantities, all five immigrant-status members, at all four geographies this cube "
        "serves — so the ruled cut can be judged against the members it was chosen over rather "
        "than presented alone.",
        "",
    ] + _MEMBER_HEADER
    lines += member_rows(("MTL_RMR", "QC_RMR", "HORS_RMR", "QC_PROVINCE"))
    lines += [
        "",
        f"The RECENT rows are what make the flow-vs-stock modeling choice §6 names visible "
        f"rather than hidden — the operand is an arrival FLOW and the ruled member is a settled "
        f"STOCK: headship "
        + " / ".join(_r4(recent[g][0]) for g in ("MTL_RMR", "QC_RMR", "HORS_RMR"))
        + " and ratio "
        + " / ".join(_r4(recent[g][1]) for g in ("MTL_RMR", "QC_RMR", "HORS_RMR"))
        + " at MTL_RMR / QC_RMR / HORS_RMR. Crediting an arrival cohort at the settled rate is "
        "a choice taken deliberately, and this table is its size.",
        "",
        f"Immigrant headship measures HIGHER than the general population at the settled reading "
        f"({_r4(ruled['MTL_RMR'][0])} vs {_r4(mtl[PC_TOTAL_AGE].headship)} at MTL_RMR), and the "
        f"pooled stock also clears the general population ({_r4(pooled['MTL_RMR'][0])}) though "
        f"it sits BELOW the settled reading — the recent member "
        f"({_r4(recent['MTL_RMR'][0])}) is what pulls the pool down. So the immigrant channel "
        "contributes MORE household formation per settled person than a general-rate model "
        "would have credited. Measured, not assumed either way.",
        "",
        f"**These rows are the SUPERSEDED territory's**, and they are the right ones to print "
        f"here: they come from `{table_number(CMA_PID)}` alone, which publishes no Québec-part "
        f"member to net out at this grain (§2b), so every per-member reading at CMA grain is a "
        f"property of the residual that keeps Gatineau in. The aligned correction is measured "
        f"at the three members the ruled pair needs — {', '.join(f'`{n}`' for n in ALIGNED_POPCHARS)} "
        "— and is deliberately NOT extended to the rest: an aligned `Recent immigrants` row "
        "would be a figure this run did not measure.",
        "",
        "### 2b. THE OPERAND ALIGNMENT — HORS_RMR's ruled territory (amendment #13)",
        "",
        f"The rate's territory must match the flow's territory. The residual above is measured "
        f"over a census territory that INCLUDES the Québec side of Ottawa-Gatineau, while the "
        f"arrival flows it multiplies come from ISQ's `{isq[ISQ_RMR_BOOK]['labels'][hors_code].strip()}` "
        f"row, which EXCLUDES it. Amendment #13 rules the corrected construction: the same "
        f"province-net-of-six residual, NET of that CMA's Québec-side census subdivisions.",
        "",
        f"**The MEMBERSHIP is carried BY REFERENCE from P10 and RE-RESOLVED here; the VALUES "
        f"are recomputed.** P10 derived which subdivisions those are — "
        f"{f_p10_membership} — and validated the resolved part against the ISQ row it aligns "
        f"to: {f_p10_gate}. This run does not repeat that derivation. It reads P10's own §4b "
        f"table, and then looks EVERY one of its {len(qc_part)} members up live in "
        f"`{table_number(CD_PID)}` by SGC code AND geoLevel {CSD_GEO_LEVEL}, checking the name "
        f"the two documents give it and re-establishing the Québec-side property two "
        f"independent ways in the cube the counts are actually subtracted from — the `"
        f"{QC_SGC_PREFIX}` province prefix on the SGC code that cube gives the member, and the "
        f"census tree (the member's census division is a geoLevel-{CD_GEO_LEVEL} child of that "
        f"cube's Quebec member, id {province[CD_PID]['memberId']}). BOTH readings place all "
        f"{len(qc_part)} inside Québec: a member either one puts outside it refuses the run, "
        f"and so does a disagreement between them. P10's rows are read by their SGC code and "
        f"never SELECTED on that prefix — a parser that filtered on it would drop an "
        f"off-province row without a word and leave the first reading unable to fail. Four "
        f"census divisions contribute — "
        + ", ".join(f"`{name}` (SGC {code})" for code, name in sorted(
            {m["classificationCode"]: m["memberNameEn"]
             for m in qc_part_parents.values()}.items()))
        + " — which is why no whole-CD union is this territory.",
        "",
        f"**Suppression: BOUNDED by the published complement, never dropped.** StatCan withholds "
        f"small counts at these subdivisions. Of the "
        f"{len(qc_part) * len(ALIGNED_POPCHARS) * 3} cells this run requested across the "
        f"{len(qc_part)} of them, "
        f"{Fact.derived(len(withheld), 'cells returned with no data point at the QC-part CSDs')} "
        f"came back with no data point, every one a `{PC_SETTLED}` field, at "
        f"{Fact.derived(len(withheld_codes), 'QC-part subdivisions withholding a settled count')} "
        f"of the {len(qc_part)} subdivisions — the other two members are published at every one "
        f"of them, which is what makes the bound available at all, and the count the bounded "
        f"sum carried is asserted equal to the count the boundary reported. Each is bounded "
        f"above by a quantity the SAME cube "
        f"publishes at the SAME geography — `{PC_TOTAL_AGE}` minus `{PC_NONIMM}`, i.e. all "
        f"immigrants and non-permanent residents together, which contains `{PC_SETTLED}` by "
        f"construction — and the bound's own two legs are required to be published "
        f"(`_guard_required_complete`), because an interval whose upper end is itself unmeasured "
        f"is not a bound. FIELD-WISE: a subdivision publishing settled persons while withholding "
        f"settled maintainers keeps the published count, since dropping the geography would net "
        f"its persons and its maintainers out of different denominators — a bias, not a wider "
        f"interval. {Fact.derived(clamped_fields, 'bound fields clamped at zero for a rounding-step negative')}"
        f" field(s) needed the clamp for a rounding-step negative.",
        "",
        "| construction | settled persons | maintainers | owner-maintainers | HEADSHIP | RATIO |",
        "|---|---:|---:|---:|---:|---:|",
        f"| as shipped (Gatineau IN) — SUPERSEDED | {_n(hors[PC_SETTLED].persons)} | "
        f"{_n(hors[PC_SETTLED].maintainers)} | {_n(hors[PC_SETTLED].owner_maintainers)} | "
        f"{_r4(shipped_pair[0])} | {_r4(shipped_pair[1])} |",
        f"| **ALIGNED (published counts only) — RULED** | "
        f"**{Fact.derived(_n(aligned_cells[PC_SETTLED].persons), 'aligned settled persons: shipped residual net of the published QC-part counts')}** | "
        f"**{Fact.derived(_n(aligned_cells[PC_SETTLED].maintainers), 'aligned settled maintainers')}** | "
        f"**{Fact.derived(_n(aligned_cells[PC_SETTLED].owner_maintainers), 'aligned settled owner-maintainers')}** | "
        f"**{f_aligned_headship}** | **{f_aligned_ratio}** |",
        f"| aligned, every withheld field at its bound | {_n(hors[PC_SETTLED].persons - qc_high.persons)} | "
        f"{_n(hors[PC_SETTLED].maintainers - qc_high.maintainers)} | "
        f"{_n(hors[PC_SETTLED].owner_maintainers - qc_high.owner_maintainers)} | "
        f"{_r4(cell_minus(hors[PC_SETTLED], qc_high).headship)} | "
        f"{_r4(ownership_ratio(cell_minus(hors[PC_SETTLED], qc_high), aligned_cells[PC_NONIMM]))} |",
        "",
        f"**Aligned: headship {f_aligned_headship} ({f_move_headship}), ratio "
        f"{f_aligned_ratio} ({f_move_ratio}) — and the ratio CROSSES 1.0.** Shipped, settled "
        f"immigrants under-own in hors-RMR; aligned, they OUT-own. The non-immigrant base moves "
        f"too and is netted across the same territory: {_r4(hors[PC_NONIMM].owner_propensity)} "
        f"shipped against {_r4(aligned_cells[PC_NONIMM].owner_propensity)} aligned. The "
        f"suppression envelope is "
        f"[{f_band_ratio_low}, {f_band_ratio_high}] on the ratio and "
        f"[{f_band_headship_low}, {f_band_headship_high}] on the headship — taken at the box's "
        f"OPPOSITE corners (headship is largest when the fewest maintainers and the most "
        f"persons are netted out), so it is the envelope rather than the two sums paired. "
        f"Neither end straddles 1.0, which is a refusal condition here rather than a remark: "
        f"amendment #13 rules the crossing with BOTH ends above 1.0, and a verdict inside its "
        f"own uncertainty is not a verdict.",
        "",
        f"The two factors MULTIPLY inside D_imm and their errors are same-signed, so the "
        f"immigrant demand leg moves by their product: **{f_move_leg}**. Rank 1 is the most "
        f"negative ED, so the shipped construction ranked HORS_RMR more risky than truth — "
        f"which is why #13 rules the aligned pair rather than recording the gap as a caveat.",
        "",
        f"**What this run does NOT re-derive, stated as the reference it is.** The 25-child "
        f"closure on the Ottawa-Gatineau CMA, the two-way selection of the Québec side at the "
        f"membership cube, and the population gate against the ISQ row are P10's measurements, "
        f"read from its committed note this run: {f_p10_construction}. What this run adds is "
        f"independent of them only where it can be: every member is re-resolved live, its "
        f"Québec-side property re-checked in this cube, and every count above is read from the "
        f"live responses rather than transcribed from P10 — the figures are then checked "
        f"against P10's own digits, so a divergence refuses instead of publishing two "
        f"measurements of one territory that disagree.",
        "",
        "## 3. The census-division sibling (98-10-0622-01) and the one-universe check",
        "",
        "| geography | census division | persons | maintainers | owner-maintainers | HEADSHIP "
        "| non-imm propensity | settled propensity | RATIO |",
        "|---|---|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for tag in ("MTL_ISLAND_RA06", "LAVAL_RA13"):
        member = ra_members[tag]
        cells = geographies[tag]
        settled = cells[PC_SETTLED]
        lines.append(
            f"| {tag} | `{member['memberNameEn']}` (id {member['memberId']}, geoLevel "
            f"{member['geoLevel']}, SGC {member['classificationCode']}) | "
            f"{_n(settled.persons)} | {_n(settled.maintainers)} | "
            f"{_n(settled.owner_maintainers)} | "
            f"**{Fact.derived(_r4(ruled[tag][0]), f'{tag}: maintainers / persons, {PC_SETTLED}')}** | "
            f"{_r4(cells[PC_NONIMM].owner_propensity)} | {_r4(settled.owner_propensity)} | "
            f"**{Fact.derived(_r4(ruled[tag][1]), f'{tag}: settled propensity / non-immigrant')}** |")
    lines += [
        "",
        f"`Laval` names BOTH the census division (id {ra_members['LAVAL_RA13']['memberId']}, "
        f"geoLevel {CD_GEO_LEVEL}) and a census SUBDIVISION inside it "
        + (f"(id {csd_trap[0]['memberId']}, geoLevel {csd_trap[0]['geoLevel']}, SGC "
           f"{csd_trap[0]['classificationCode']})" if csd_trap else "(absent this run)")
        + ". Resolution therefore requires name AND geoLevel AND parent, and refuses on more "
        "than one match — a name-only lookup would publish a city's numbers under a region's "
        "label.",
        "",
        "Every member at both census divisions, on the same cross as §2a:",
        "",
    ] + _MEMBER_HEADER
    lines += member_rows(("MTL_ISLAND_RA06", "LAVAL_RA13"))
    lines += [
        "",
        "### 3a. BIT-IDENTITY, asserted on the RULED TRIPLE and nowhere else",
        "",
        f"The two cubes' Québec province rows agree EXACTLY on the ruled `{PC_SETTLED}` triple "
        f"— persons {_n(prov_cma[PC_SETTLED].persons)}, maintainers "
        f"{_n(prov_cma[PC_SETTLED].maintainers)}, owner-maintainers "
        f"{_n(prov_cma[PC_SETTLED].owner_maintainers)} — which is what makes them ONE universe "
        "at two geography grains rather than two sources. They do NOT agree cube-wide, and "
        "this note does not claim they do: "
        f"{Fact.derived(len(universe_drift), 'province cells differing across the two cubes')} "
        f"province cells differ by exactly 5. Of the {len(province_counts)} province counts "
        f"read from the two cubes, "
        f"{Fact.derived(sum(1 for v in province_counts if v % 5), 'province counts read this run that are NOT a multiple of 5')} "
        "are not a multiple of 5 — so a 5-person disagreement is exactly what rounding applied "
        "independently per cube produces, an explanation this run can offer only because it "
        "measured that property rather than assuming it.",
        "",
        "| popchar member | field | 98-10-0621-01 | 98-10-0622-01 | gap |",
        "|---|---|---:|---:|---:|",
    ]
    id_to_name = {v: k for k, v in PINNED_POPCHAR_IDS.items()}
    for popchar, field, left, right in universe_drift:
        lines.append(f"| {id_to_name[popchar]} | {_RULED_TRIPLE_FIELDS[field]} | {_n(left)} | "
                     f"{_n(right)} | {left - right:+d} |")
    lines += [
        "",
        f"A note claiming cube-wide identity would assert what its own data contradicts, and a "
        f"gate pinning that claim would pin a falsehood. So the gate binds the ruled triple "
        f"EXACTLY, tolerates ±{_ROUNDING_TOLERANCE} elsewhere, and refuses beyond it.",
        "",
        "### 3b. The ratio EXCEEDS 1 at both, and that is COMPOSITION — measured, not asserted",
        "",
        f"On the island the NON-immigrant base is renter-heavy: its owner-maintainer propensity "
        f"is only **{_r4(geographies['MTL_ISLAND_RA06'][PC_NONIMM].owner_propensity)}**, against "
        f"the settled-immigrant "
        f"**{_r4(geographies['MTL_ISLAND_RA06'][PC_SETTLED].owner_propensity)}** — so the ratio "
        f"is {_r4(ruled['MTL_ISLAND_RA06'][1])} LOCALLY while the same measurement CMA-wide is "
        f"{_r4(ruled['MTL_RMR'][1])}, where the non-immigrant base is "
        f"{_r4(mtl[PC_NONIMM].owner_propensity)}. Both are true at their own scale; printing "
        "only the ratio would leave the reader to take the explanation on trust.",
        f"At Laval the same shape: non-immigrant "
        f"{_r4(geographies['LAVAL_RA13'][PC_NONIMM].owner_propensity)} against settled "
        f"{_r4(geographies['LAVAL_RA13'][PC_SETTLED].owner_propensity)}, and the product the "
        f"[0,1] assertion binds is "
        f"{_r4(geographies['LAVAL_RA13'][PC_NONIMM].owner_propensity)} × "
        f"{_r4(ruled['LAVAL_RA13'][1])} = "
        f"{_r3(geographies['LAVAL_RA13'][PC_NONIMM].owner_propensity * ruled['LAVAL_RA13'][1])}.",
        "",
        f"**This is NOT the pooled-ratio anti-pattern** recorded in `loaders/constants.py` "
        f"({f_anti}). That one pools ACROSS RECENCY, and this run can show what pooling costs "
        f"from its own numbers: pooling `{PC_SETTLED}` with `{PC_RECENT}` into "
        f"`{PC_ALL_IMM}` moves the island ratio from "
        f"{_r4(ruled['MTL_ISLAND_RA06'][1])} to {_r4(pooled['MTL_ISLAND_RA06'][1])} — i.e. it "
        f"erases the >1 finding entirely — and at MTL_RMR from {_r4(ruled['MTL_RMR'][1])} to "
        f"{_r4(pooled['MTL_RMR'][1])}, with the recent member sitting at "
        f"{_r4(recent['MTL_ISLAND_RA06'][1])} and {_r4(recent['MTL_RMR'][1])} respectively. A "
        "single pooled number stands in for a spread that wide, which is what defeats the "
        "netting by construction. This reading is decomposed: one member, named.",
        "",
        f"## 4. Ruling T's TERRITORY GATE — the province-controlled share residual",
        "",
        "**The construction ruling T first named is REFUSED and is not implemented here.** It "
        f"compared `{table_number(CD_PID)}`'s population against the ISQ RA total — a "
        "PRIVATE-HOUSEHOLD count against a TOTAL-POPULATION estimate — and measured, it trips "
        "at the Québec province CONTROL, where territory identity is not in question. A gate "
        "that fires where the answer is known is measuring the universe gap, not the territory. "
        "Amendment #11 replaced it with:",
        "",
        "```",
        "residual(g) = ( part_census(g) / total_census  ÷  part_ISQ(g) / total_ISQ ) − 1",
        "```",
        "",
        f"Each geography's share of its OWN source's provincial total. Census side: "
        f"`{table_number(CD_PID)}` persons against a province total of {_n(census_total)} — "
        f"the province person count `{table_number(CMA_PID)}` publishes, which this run "
        f"compared cube against cube and refuses on a disagreement — so the RA residuals and "
        f"the innocent controls share one denominator rather than two that happen to be "
        f"close. ISQ side: the pinned workbooks against {f_isq_total}, "
        f"likewise checked equal across the two workbooks. A universe "
        "offset that is uniform across the province cancels by construction; what survives is "
        "the geography-VARYING part, and that is exactly what the innocent controls measure.",
        "",
        "### 4a. The innocent controls — territories whose identity is NOT in question",
        "",
        "The six wholly-QC CMAs. Two independent facts, both measured this run, put their "
        "identity beyond question:",
        "",
        f"1. Every one of the six joins on an EXACT code. The `code` column below IS the "
        f"census member's `classificationCode`, and it is the key this run looked the ISQ row "
        f"up under — the workbook's own `Code` axis — so the `ISQ row` beside it is the row "
        f"that code retrieved. A census code the workbook does not publish refuses the run "
        f"rather than printing.",
        f"2. The RMR workbook DECLARES the delineation it uses, in its own header note: "
        f"*{f_isq_rmr_note}*. So its rows are census territories by the publisher's own "
        "statement, not by inference from the code agreement.",
        "",
        f"The RA workbook declares a DIFFERENT one: *{f_isq_ra_note}* — an administrative "
        "delineation dated four years after the census, with no correspondence to the SGC in "
        "this tree. That is why RA06 and RA13 need a gate and why these six can calibrate it: "
        "the six are census territories by declaration, the two under test are not.",
        "",
        "| CMA (census member) | code | ISQ row | census persons | ISQ population | share "
        "residual |",
        "|---|---:|---|---:|---:|---:|",
    ]
    # Registered ONCE here and re-interpolated in §4b's signed paragraph: the same measured
    # residual reprinted under a second `Fact` would inflate the provenance count with a
    # figure this run measured only once.
    f_innocent = {}
    for member in cma_members:
        code = int(member["classificationCode"])
        name = member["memberNameEn"]
        f_innocent[name] = Fact.derived(
            _pct(innocent[name]), f"{name}: province-controlled share residual")
        lines.append(
            f"| {name} | {code} | {innocent_labels[name].strip()} | "
            f"{_n(cma_cells[member['memberId']][PC_TOTAL_AGE].persons)} | "
            f"{_n(isq[ISQ_RMR_BOOK]['totals'][code])} | "
            f"**{f_innocent[name]}** |")
    lines += [
        "",
        f"The largest innocent |residual| is **{_r3(bounds['max_abs'])}%** "
        f"({bounds['max_name']}). The threshold is that maximum PLUS a stated margin of "
        f"{MARGIN_FRACTION:.0%} of it — {_r3(bounds['margin_pp'])} percentage points — giving "
        f"**{f_threshold}**. The margin is a cushion, not the calibration: the calibration is "
        "the measured maximum, and the cushion exists because six controls make that maximum a "
        "noisy estimate of the innocent spread's upper edge. Ruling T's original 1% is "
        "deliberately NOT used — it was calibrated against the refuted construction's "
        "semantics and does not transfer.",
        "",
        "### 4b. The two geographies under test",
        "",
        "| geography | census persons | ISQ population | share residual | vs threshold |",
        "|---|---:|---:|---:|---|",
    ]
    for tag in ("MTL_ISLAND_RA06", "LAVAL_RA13"):
        residual = ra_residual[tag]
        lines.append(
            f"| {tag} | "
            f"{_n(cd_cells[ra_members[tag]['memberId']][PC_TOTAL_AGE].persons)} | "
            f"{_n(isq[ISQ_RA_BOOK]['totals'][ra_code[tag]])} | "
            f"**{f_ra06 if tag == 'MTL_ISLAND_RA06' else f_ra13}** | "
            f"PASS ({abs(residual):.3f}% ≤ {bounds['threshold']:.3f}%) |")
    inside = [n for n, v in innocent.items() if abs(v) > max(abs(x) for x in ra_residual.values())]
    # Branched on the measured set, never assumed: with no positive innocent control there is
    # no "largest positive" to compare against, and `signed_high_name` would then be the
    # LEAST-NEGATIVE one. A shape, not a defect — so the prose changes rather than refusing.
    if not positive_controls:
        positive_clause = ("No innocent control is positive at all, so the positive side "
                           "carries no calibration.")
    else:
        above = (f"{', '.join(above_every_positive)} measures ABOVE it"
                 if above_every_positive else
                 "no geography under test measures above it")
        positive_clause = (f"The largest positive innocent control is {signed_high_name} at "
                           f"{f_innocent[signed_high_name]}, and {above}.")
    lines += [
        "",
        f"Both residuals sit INSIDE the innocent |residual| range: {len(inside)} of the "
        f"{len(innocent)} controls — {', '.join(sorted(inside))} — carry a LARGER |share "
        "residual| than either geography under test, and every one of those is a census "
        "territory by the ISQ workbook's own declaration. That is stronger than the threshold "
        "comparison alone — the two geographies under test are not merely under a bound, they "
        "are less discrepant IN ABSOLUTE VALUE than territories whose identity is settled — "
        "and absolute value is what the gate rules on, since amendment #11 sets the threshold "
        "on the maximum innocent |residual|.",
        "",
        f"**THE SIGNED DECOMPOSITION DOES NOT RUN THE SAME WAY, and is published rather than "
        f"left inside the absolute value.** The innocent controls span "
        f"{f_innocent[signed_low_name]} to {f_innocent[signed_high_name]} "
        f"({signed_low_name} to {signed_high_name}), {len(negative_controls)} of the "
        f"{len(innocent)} of them NEGATIVE ({', '.join(negative_controls)}); "
        f"{len(ra_positive)} of {len(ra_residual)} geographies under test measure POSITIVE ("
        + ", ".join(f"{t} {_pct(ra_residual[t])}" for t in ra_positive)
        + f"). {positive_clause} The gate PASSES on |residual| and this note does not "
        "restate that as agreement in SIGN. What the sign would MEAN is not decidable from "
        "this construction: §4c's second limit — the residual cannot separate a territory "
        "difference from a region-varying universe component — binds the sign exactly as it "
        "binds the magnitude, and six controls do not fix a direction. Measured and recorded "
        "here, deliberately un-attributed.",
        "",
        "### 4c. What this gate CANNOT do",
        "",
        "Passing does not establish that a census division IS an ISQ région administrative. It "
        "fails to REFUTE that, at the resolution this construction has — the gate trips only "
        f"above {_r3(bounds['threshold'])}%, and the innocent spread it is calibrated on "
        f"already runs to {_r3(bounds['max_abs'])}%. In persons, at each geography's own ISQ "
        "population, that trip point is "
        + ", ".join(
            f"**{_n(bounds['threshold'] / 100 * isq[ISQ_RA_BOOK]['totals'][ra_code[t]])}** "
            f"at {t}" for t in ("MTL_ISLAND_RA06", "LAVAL_RA13"))
        + " — a territory difference smaller than that would not be seen at all. The residual "
        "also cannot SEPARATE a territory difference from a region-varying universe "
        "difference; the innocent controls bound the second, they do not remove it. Both "
        "limits are published beside the verdict so a reader can price the claim rather than "
        "take it.",
        "",
        "### 4d. Code axes — RECORDED, and deliberately NOT the gate",
        "",
        f"- ISQ keys the RA workbook on région-administrative codes 0-17. This run resolved "
        f"`{isq[ISQ_RA_BOOK]['labels'][ra_code['MTL_ISLAND_RA06']]}` and "
        f"`{isq[ISQ_RA_BOOK]['labels'][ra_code['LAVAL_RA13']]}` BY LABEL in the workbook and "
        f"read back codes {f_ra_code['MTL_ISLAND_RA06']} and {f_ra_code['LAVAL_RA13']}, each "
        f"then asserted against the code pinned for it — the ISQ mirror of the by-name "
        f"resolution the census side uses, so the gate fetches on the workbook's own identity "
        f"for the row rather than on a typed integer. The delineation is the one its own "
        f"header note states.",
        f"- The census publishes SGC classification codes: CD Montréal {f_sgc_ra06} and CD "
        f"Laval {f_sgc_ra13} in `{table_number(CD_PID)}`, the same two codes in "
        f"`{table_number(TOTALPOP_PID)}` "
        f"({totalpop_ids['MTL_ISLAND_RA06']['classificationCode']} and "
        f"{totalpop_ids['LAVAL_RA13']['classificationCode']}).",
        "- **SGC agreement across two CENSUS cubes establishes that two CENSUS cubes mean the "
        "same census division. It does NOT establish that the census division equals the ISQ "
        "région administrative** — no correspondence between the two code systems exists in "
        "this tree. So the population residual carries the gate and the code agreement is "
        "corroboration. (The six CMAs above are a different case: there the ISQ workbook and "
        "the census share ONE code system, by the RMR workbook's own declaration.)",
        "",
        f"### 4e. Second diagnostic — census TOTAL population ({table_number(TOTALPOP_PID)}), "
        "never the gate",
        "",
        "A like-for-like universe check: total population against total population, no "
        "private-household restriction on either side.",
        "",
        "| geography | census 2021 population | ISQ July-1 2021 | delta |",
        "|---|---:|---:|---:|",
    ]
    for tag, isq_value in (("MTL_ISLAND_RA06",
                            isq[ISQ_RA_BOOK]["totals"][ra_code["MTL_ISLAND_RA06"]]),
                           ("LAVAL_RA13",
                            isq[ISQ_RA_BOOK]["totals"][ra_code["LAVAL_RA13"]]),
                           ("QC_PROVINCE", isq_total)):
        census_value = total_pop["province" if tag == "QC_PROVINCE" else tag]
        lines.append(
            f"| {tag} | {_n(census_value)} | {_n(isq_value)} | "
            f"**{Fact.derived(_pct(diagnostic[tag]), f'{tag}: census total population vs ISQ estimate')}** |")
    lines += [
        "",
        "All three are the same small negative — a census count taken in May against a July-1 "
        "estimate, plus net undercoverage — and the two geographies under test are CLOSER to "
        "their ISQ counterpart than the province is. It is a corroboration and nothing more: "
        f"admitting `{table_number(TOTALPOP_PID)}` into the GATE would mean repairing a gate by "
        "adding cubes until it passed, which is what amendment #11 refuses. The gate above "
        "uses only the two sources ruling T names.",
        "",
        "## 5. The FLOOR GATE — cross-cube, and wired knowingly",
        "",
        f"`{table_number(CMA_PID)}` publishes no living-alone indicator, so the floor comes from "
        f"`{table_number(SIBLING_PID)}` indicator `{SIB_IND_ALONE}` — a DIFFERENT cube, with a "
        f"different immigrant axis. The member must be the ruled-adjacent one: "
        f"`{SIB_SETTLED}`, not the pooled `{SIB_ALL_IMM}`.",
        "",
        "| geography | ruled headship | settled living-alone | pooled living-alone (NOT used) "
        "| clears by |",
        "|---|---:|---:|---:|---:|",
    ]
    for tag in sorted(SIB_GEO):
        lines.append(
            f"| {tag} | {_r4(ruled[tag][0])} | "
            f"**{Fact.derived(_r3(alone[tag]), f'{table_number(SIBLING_PID)} {tag}: {SIB_IND_ALONE}, {SIB_SETTLED} member, percent / 100')}** "
            f"| {_r3(alone_pooled[tag])} | {_r3(ruled[tag][0] - alone[tag])} |")
    lines += [
        "",
        f"Each person living alone maintains exactly one household, so headship must EXCEED "
        f"the living-alone share; a value at or below it is a defect, not a datum. §6 names "
        f"the settled member on two grounds — nearer in definition to the ruled `{PC_SETTLED}` "
        f"cut, and stricter as a bound. This run measures the SECOND: "
        f"{_r3(alone['MTL_RMR'])} > {_r3(alone_pooled['MTL_RMR'])} and "
        f"{_r3(alone['QC_RMR'])} > {_r3(alone_pooled['QC_RMR'])}. It does not measure the "
        "first, which is a claim about definitions rather than a quantity.",
        "",
        f"**COVERAGE, and the absence is EARNED rather than assumed.** The floor binds at "
        f"{', '.join(sorted(SIB_GEO))} only. For the other three modeled geographies — "
        f"{', '.join(not_covered)} — all {len(sib_members)} of this cube's geography members "
        "were searched this run, and the search is reported rather than summarised:",
        "",
        "  | uncovered geography | searched for | members matching |",
        "  |---|---|---:|",
    ]
    for tag in not_covered:
        hit = floor_search[tag]
        lines.append(f"  | {tag} | `{hit['predicate']}` | {len(hit['code_hits'])} |")
    lines += [
        "",
        f"The cube's geography members carry these geoLevels: "
        + ", ".join(f"`{level}` × {count}" for level, count in sorted(
            sib_levels.items(), key=lambda kv: (kv[0] is None, kv[0])))
        + f" — **{sib_levels.get(CD_GEO_LEVEL, 0)}** of them at geoLevel {CD_GEO_LEVEL}, the "
        "census-division level RA06 and RA13 are published at. Substituting the province "
        f"figure ({_r3(sibling[('QC_PROVINCE', SIB_SETTLED, 'alone')])}, measured here) would "
        "be a geography transport the ruling does not make, so these three are recorded NOT "
        "COVERED rather than floored against a stand-in. Scoped to this cube and these "
        "predicates — not a claim that no source anywhere carries them.",
        "",
        "## 6. The SIBLING CROSS-CHECK, at exactly its coarse strength",
        "",
        f"All five members of this cube's `Immigrant and generation status` axis, both "
        f"indicators, at all three geographies it serves. Values are published as PERCENTAGES "
        "and divided by 100 at the boundary; nothing here is a maintainer propensity.",
        "",
        "| geography | immigrant member | ownership share | living-alone share | ownership vs "
        "non-immigrant |",
        "|---|---|---:|---:|---:|",
    ]
    for tag in ("MTL_RMR", "QC_RMR", "QC_PROVINCE"):
        for name in SIB_MEMBERS:
            own = sibling[(tag, name, "ownership")]
            base = sibling[(tag, SIB_NONIMM, "ownership")]
            marker = "**" if name in (SIB_NONIMM, SIB_SETTLED) else ""
            lines.append(
                f"| {tag} | {name} | {marker}{_r3(own)}{marker} | "
                f"{_r3(sibling[(tag, name, 'alone')])} | {_r4(own / base)} |")
    lines += [
        "",
        "The two bolded members are the only ones this note uses: `" + SIB_NONIMM
        + "` as the base and `" + SIB_SETTLED + "` as the comparator, giving "
        + ", ".join(
            f"**{Fact.derived(_r4(sibling_ratio[t]), f'{table_number(SIBLING_PID)} {t}: settled ownership share / non-immigrant, percent-denominated')}** at {t}"
            for t in ("MTL_RMR", "QC_RMR", "QC_PROVINCE"))
        + f" — against ruling S's {_r4(ruled['MTL_RMR'][1])} / {_r4(ruled['QC_RMR'][1])} / "
        f"{_r4(ruled['QC_PROVINCE'][1])} from {table_number(CMA_PID)}.",
        "",
        f"**This is a COARSE consistency check across two named axes, never a like-for-like "
        f"agreement.** The sibling's {_r4(sibling_ratio['MTL_RMR'])} and ruling S's "
        f"{_r4(ruled['MTL_RMR'][1])} differ in BOTH the member cut (`{SIB_SETTLED}`, i.e. more "
        f"than 10 years, against `{PC_SETTLED}`, i.e. at least five) AND the metric (a "
        "person-weighted ownership share against a maintainer propensity). Their closeness at "
        "Montréal therefore bounds the COMBINED size of those two differences and asserts "
        "nothing stronger — in particular it does not validate either quantity, and neither "
        "figure may be substituted for the other.",
        "",
        "## 7. The catalogue search, carried BY REFERENCE",
        "",
        f"This note makes no absence claim of its own. P9 closed the search and this run reads "
        f"P9's own tokens rather than restating them: verdict **{f_p9_verdict}**, at "
        f"**{f_p9_level}**. The residual P9 records and this note therefore inherits: "
        f"**{f_p9_residual}**. A re-run of P9 that narrows its closure narrows this sentence "
        "with it, which a restated absolute would not.",
        "",
        "## 8. Scope",
        "",
        f"PROBE ONLY. Nothing here is wired into `demand/immigrant_inputs.py`; the plan's task "
        f"25b is that consumer and a separate run. The values above are the measurement task "
        f"25b consumes, and the per-field provenance §6 requires is visible in the DECISION "
        f"block: MTL_RMR and QC_RMR direct and `cited`; RA06 and RA13 measured at census-"
        f"division grain and `cited`, on the strength of the gate in §4; HORS_RMR computed as "
        "the operand-aligned residual of §2b — the province net of the six wholly-QC CMAs and "
        "of the Ottawa-Gatineau Québec part, with the membership carried from P10.",
        "",
        "## DECISION",
        "",
        "- `DECISION-VERDICT: MEASURED`",
        "- `DECISION-HEADSHIP: "
        + "; ".join(f"{g} {_r4(ruled[g][0])}" for g in modeled)
        + f" — {table_number(CMA_PID)}/{table_number(CD_PID)} `{PC_SETTLED}`, maintainers ÷ "
          "persons, recomputed this run`",
        "- `DECISION-RATIO: "
        + "; ".join(f"{g} {_r4(ruled[g][1])}" for g in modeled)
        + " — settled owner-maintainer propensity ÷ non-immigrant, same member, same universe`",
        f"- `DECISION-SOURCE-MTL_RMR: {table_number(CMA_PID)} member "
        f"{named_cma['Montréal (CMA), Que.']['memberId']} — DIRECT, cited, both fields`",
        f"- `DECISION-SOURCE-QC_RMR: {table_number(CMA_PID)} member "
        f"{named_cma['Québec (CMA), Que.']['memberId']} — DIRECT, cited, both fields`",
        f"- `DECISION-SOURCE-HORS_RMR: {table_number(CMA_PID)} province member "
        f"{province[CMA_PID]['memberId']} NET of the six geoLevel-{CMA_GEO_LEVEL} children "
        f"{list(PINNED_CMA_IDS)} NET of the {len(qc_part)} Ottawa-Gatineau Québec-part census "
        f"subdivisions read from {table_number(CD_PID)} (membership carried from "
        f"probes/{P10_NOTE.name}, every member re-resolved live by SGC code and geoLevel "
        f"{CSD_GEO_LEVEL}) — OPERAND-ALIGNED computed residual per amendment #13, both fields`",
        f"- `DECISION-HORS-ALIGNMENT: RULED {_r4(aligned[0])} / {_r4(aligned[1])} "
        f"(suppression envelope {_r4(envelope['headship_band'][0])}-"
        f"{_r4(envelope['headship_band'][1])} and {_r4(envelope['ratio_band'][0])}-"
        f"{_r4(envelope['ratio_band'][1])}, neither ratio end straddling 1.0); SUPERSEDED "
        f"{_r4(shipped_pair[0])} / {_r4(shipped_pair[1])} — the same residual with the Québec "
        f"side of Ottawa-Gatineau still IN, which §6's ruling-S row states and amendment #13 "
        f"supersedes. Immigrant demand leg {_pct(aligned_move['leg'])}; headship "
        f"{_pct(aligned_move['headship'])}, ratio {_pct(aligned_move['ratio'])}. "
        f"{withheld_fields} withheld settled fields at {len(withheld_codes)} of "
        f"{len(qc_part)} subdivisions, bounded FIELD-WISE by `{PC_TOTAL_AGE}` − "
        f"`{PC_NONIMM}` at the same geography`",
        f"- `DECISION-SOURCE-MTL_ISLAND_RA06: {table_number(CD_PID)} CD Montréal member "
        f"{ra_members['MTL_ISLAND_RA06']['memberId']} (SGC "
        f"{ra_members['MTL_ISLAND_RA06']['classificationCode']}) — MEASURED, cited, both "
        "fields`",
        f"- `DECISION-SOURCE-LAVAL_RA13: {table_number(CD_PID)} CD Laval member "
        f"{ra_members['LAVAL_RA13']['memberId']} (SGC "
        f"{ra_members['LAVAL_RA13']['classificationCode']}) — MEASURED, cited, both fields`",
        f"- `DECISION-TERRITORY-GATE: PASS — province-controlled share residual RA06 "
        f"{_pct(ra_residual['MTL_ISLAND_RA06'])}, RA13 {_pct(ra_residual['LAVAL_RA13'])}; both "
        f"inside the innocent |residual| range ({len(inside)} of {len(innocent)} innocent "
        f"controls are further out in |.|). SIGNED, recorded and NOT attributed: innocent "
        f"span {_pct(signed_low)} to {_pct(signed_high)}, {len(negative_controls)} of "
        f"{len(innocent)} negative; {len(ra_positive)} of {len(ra_residual)} under test "
        f"positive; above every positive innocent control: "
        + (", ".join(above_every_positive) if above_every_positive else "none") + "`",
        f"- `DECISION-TERRITORY-THRESHOLD: {_r3(bounds['threshold'])}% = max innocent "
        f"|residual| {_r3(bounds['max_abs'])}% ({bounds['max_name']}) + "
        f"{_r3(bounds['margin_pp'])} pp margin ({MARGIN_FRACTION:.0%} of the max); DERIVED "
        f"from the six wholly-QC CMAs, NOT ruling T's inherited 1%`",
        f"- `DECISION-TERRITORY-DIAGNOSTIC: {table_number(TOTALPOP_PID)} total population RA06 "
        f"{_pct(diagnostic['MTL_ISLAND_RA06'])}, RA13 {_pct(diagnostic['LAVAL_RA13'])}, "
        f"province {_pct(diagnostic['QC_PROVINCE'])} — corroborating, NEVER the gate`",
        f"- `DECISION-FLOOR-GATE: PASS — headship exceeds the settled living-alone share at "
        + "; ".join(f"{t} {_r4(ruled[t][0])} > {_r3(alone[t])}" for t in sorted(SIB_GEO))
        + f" ({table_number(SIBLING_PID)} indicator `{SIB_IND_ALONE}`, `{SIB_SETTLED}` member)`",
        f"- `DECISION-FLOOR-COVERAGE: binds at {', '.join(sorted(SIB_GEO))}; "
        f"{', '.join(not_covered)} NOT COVERED — searched all {len(sib_members)} geography "
        f"members of {table_number(SIBLING_PID)} this run: "
        f"{sib_levels.get(CD_GEO_LEVEL, 0)} at geoLevel {CD_GEO_LEVEL} (census division) and "
        f"{sum(len(h['code_hits']) for h in floor_search.values())} carrying SGC "
        + "/".join(str(h["sgc"]) for h in floor_search.values() if h["sgc"])
        + "; no stand-in substituted`",
        f"- `DECISION-UNIVERSE-IDENTITY: the two cubes' province rows are BIT-IDENTICAL on the "
        f"ruled {PC_SETTLED} triple ({_n(prov_cma[PC_SETTLED].persons)} / "
        f"{_n(prov_cma[PC_SETTLED].maintainers)} / "
        f"{_n(prov_cma[PC_SETTLED].owner_maintainers)}) and differ by exactly 5 on "
        f"{len(universe_drift)} other province cells — NOT cube-wide identity`",
        f"- `DECISION-SIBLING-CROSS-CHECK: COARSE — {table_number(SIBLING_PID)} "
        f"{_r4(sibling_ratio['MTL_RMR'])} vs ruled {_r4(ruled['MTL_RMR'][1])} at MTL_RMR "
        "bounds the COMBINED size of a member-cut difference and a metric difference; asserts "
        "nothing stronger`",
        f"- `DECISION-CATALOGUE-CLOSURE: {p9['DECISION-VERDICT']} at {p9['DECISION-CLOSURE-LEVEL']}"
        f" — residual {p9['DECISION-RESIDUAL']} (read from probes/{P9_NOTE.name} this run, not "
        "restated)`",
        "- `DECISION-SCOPE: PROBE ONLY — NOT wired into demand/immigrant_inputs.py by this "
        "run and NOT to be read as wired; plan task 25b is that consumer and a separate run`",
        "",
    ]
    return lines


def _failure_sections(exc: BaseException, where: list) -> list:
    boundary = getattr(exc, "boundary", where[0])
    return [
        "## DECISION",
        "",
        "- `DECISION-VERDICT: UNKNOWN-PROBE-FAILED`",
        f"- `LIVE PROBE FAILED: {type(exc).__name__}: {exc}`",
        f"- `LIVE PROBE FAILED-AT: {boundary}`",
        "",
        "## What this run could not establish",
        "",
        "This probe recomputes every immigrant input ruling S and T bind and decides ruling T's "
        "territory gate. This run did not complete that derivation, so it decides NOTHING: no "
        "headship, no ratio, no territory verdict, no floor result. A partial derivation is not "
        "a weaker measurement — it is no measurement, and recording it as one would be the "
        "cheap all-clear this probe family exists to refuse.",
        "",
    ]


def main() -> None:
    log = new_run()
    where = ["unattributed"]
    try:
        sections = _sections(where)
    except Exception as exc:              # noqa: BLE001 — every failure is RECORDED, never lost
        sections = _failure_sections(exc, where)
    header = provenance_header(log, written_by=_WRITTEN_BY, scope=_SCOPE, summary=_summary,
                               cited_label=_CITED_LABEL)
    text = "\n".join([_TITLE, ""] + header + sections) + "\n"
    if "[FILL" in text:
        raise ProbeRefusal("note", "an unresolved [FILL placeholder survived into the note — "
                                   "every value must be COMPUTED, never invited")
    OUT.write_text(text, encoding="utf-8")
    print("\n".join(line for line in sections
                    if line.startswith(("- `DECISION", "- `LIVE PROBE"))))


if __name__ == "__main__":
    main()
