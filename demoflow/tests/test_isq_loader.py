"""Task 11 contract: the ISQ population loader (spec §8 junctions + §4 loader contracts).

The first five tests are the plan's numbered contract verbatim. The rest close gates the
plan body left with no executable check — the RUN-1 addendum items (required
WORKBOOK_GEOGRAPHIES lookup, `is_file`) plus the plan's own traceability row "Loader
schema-drift raises (age-block group, header tokens, ADDITIVITY) | 10,11,12 |
test_isq_loader.py".

Workbook mutations run on COPIES in tmp_path, never against demoflow/data/. Each one
RE-PINS the copy's sha256 (load_population verifies the pin, so a mutated copy would
otherwise be rejected AT THE PIN and never reach the gate under test) and asserts every
cell's PRE-IMAGE, so a geometry drift fails the TEST instead of silently mutating the
wrong cell. Sheet geometry is MEASURED per workbook, never hardcoded (offsets differ:
pop-rmr 6, pop-ra 5, pop-qc 4). The ONE exception is
test_unpinned_workbook_bytes_raise_end_to_end, which deliberately does NOT re-pin: the
re-pinning that makes every other RED possible is exactly what leaves the pin CALL SITE
untested, so one RED has to exercise the drift branch through load_population.
"""
import hashlib
import shutil

import pandas as pd
import pytest

from openpyxl import load_workbook

from demoflow.errors import LoaderError
from demoflow.geography import (
    IGNORED, WORKBOOK_GEOGRAPHIES, Geography, Scenario, classify_geography,
)
from demoflow.loaders import isq, pins
from demoflow.loaders.isq import load_population
from demoflow.loaders.isq_ages import SHEET, build_single_year_long
from demoflow.loaders.pins import DATA_DIR

_RMR = "pop-as-rmr-base.xlsx"
_ID_TOKENS = ("Scénario", "Année", "Statut", "Sexe", "Région1", "Région")
_AGE_GROUP = "Âge"


# --- the plan's numbered contract ----------------------------------------------------

def test_load_population_returns_tidy_long_frame():
    df = load_population("pop-as-rmr-base.xlsx")
    assert set(df.columns) >= {"geography", "scenario", "year", "sex", "age", "population", "status"}
    assert set(df["scenario"].unique()) <= {Scenario.REFERENCE, Scenario.LOW, Scenario.HIGH}
    assert set(df["sex"].unique()) <= {"M", "F"}    # code 3 excluded after additivity
    assert {Geography.MTL_RMR, Geography.QC_RMR, Geography.HORS_RMR} <= set(df["geography"].unique())
    assert set(df["geography"].unique()) <= set(Geography)   # IGNORED rows dropped
    # year lattice pinned to 2021-2051 (loader would have raised otherwise)
    assert df["year"].min() == 2021 and df["year"].max() == 2051


def test_all_three_scenarios_present_for_mtl_rmr():
    df = load_population("pop-as-rmr-base.xlsx")
    mtl = df[df["geography"] == Geography.MTL_RMR]
    assert {Scenario.REFERENCE, Scenario.LOW, Scenario.HIGH} <= set(mtl["scenario"].unique())


def test_sex_additivity_violation_raises():
    with pytest.raises(LoaderError, match="additivity"):
        isq._check_sex_additivity(code3=100.0, code1=40.0, code2=40.0, ctx="MTL/2035/ref/age75")


def test_sex_orientation_guard_raises_on_swapped_map():   # RED (codex r2-F5)
    swapped = pd.DataFrame({
        "geography": [Geography.MTL_RMR, Geography.MTL_RMR],
        "scenario": [Scenario.REFERENCE, Scenario.REFERENCE],
        "year": [2035, 2035], "sex": ["M", "F"], "age": [85, 85],
        "population": [1000.0, 500.0],   # M > F at 85+ => swapped 1<->2 map
    })
    with pytest.raises(LoaderError, match="orientation"):
        isq._check_sex_orientation(swapped, "test")


def _orientation_frame(sexes, ages, years, populations):
    return pd.DataFrame({
        "geography": [Geography.MTL_RMR] * len(sexes),
        "scenario": [Scenario.REFERENCE] * len(sexes),
        "year": years, "sex": sexes, "age": ages, "population": populations,
    })


@pytest.mark.parametrize("frame", [
    # (a) NO 85+ rows at all: the guard's own comparison window is empty.
    _orientation_frame(["M", "F"], [40, 40], [2035, 2035], [100.0, 150.0]),
    # (b) the 85+ block carries only ONE sex: `piv` has no 'F' column to compare.
    _orientation_frame(["M"], [85], [2035], [100.0]),
    # (c) FAIL-OPEN leg: 2035 is a complete, non-violating key, 2036 is short its 'F' rows,
    # so `piv` keeps an 'F' column with a NaN cell — and `NaN <= M` is False, i.e. the
    # comparison reports "no violation" for a key it never saw. `_sex_lattice` refuses this
    # frame earlier in load_population today, but this guard must not depend on another
    # gate's POSITION for its own correctness (a reordering would silently open the hole).
    _orientation_frame(["M", "F", "M"], [85, 85, 85], [2035, 2035, 2036],
                       [100.0, 150.0, 100.0]),
], ids=["no-85+-rows", "one-sex-at-85+", "a-key-short-its-female-rows"])
def test_orientation_guard_refuses_when_it_cannot_verify(frame):
    """A guard that cannot verify REFUSES; it never passes by default. Each case removes a
    different part of the guard's own input, and none of them is a F<=M VIOLATION — the two
    violation REDs cannot reach any of these branches."""
    with pytest.raises(LoaderError, match="cannot run"):
        isq._check_sex_orientation(frame, "test")


def test_missing_workbook_raises():
    with pytest.raises(LoaderError, match="not found"):
        load_population("does-not-exist.xlsx")


def test_enum_columns_support_equality_filtering():
    """The typed junction output must be FILTERABLE by its enums. pandas 3.x infers its
    `str` dtype for a column of str-SUBCLASS objects and str()-ifies the scalar operand on
    comparison, so `df["scenario"] == Scenario.LOW` matches NOTHING while `.unique()` still
    returns enum members — an empty frame with no error anywhere. Every downstream consumer
    (rankings, joins) filters these two columns by equality, so the dtype is part of the
    contract. MEASURED failure on pandas 3.0.3 before the loader forced dtype=object."""
    df = load_population(_RMR)
    assert df["geography"].dtype == object, "geography column lost its enum identity"
    assert df["scenario"].dtype == object, "scenario column lost its enum identity"
    for geography in (Geography.MTL_RMR, Geography.QC_RMR, Geography.HORS_RMR):
        assert int((df["geography"] == geography).sum()) > 0, f"no rows filter to {geography}"
    for scenario in (Scenario.REFERENCE, Scenario.LOW, Scenario.HIGH):
        assert int((df["scenario"] == scenario).sum()) > 0, f"no rows filter to {scenario}"


# --- RUN-1 addendum gates ------------------------------------------------------------

def test_directory_is_not_a_workbook(tmp_path):
    """`path.exists()` passes for a DIRECTORY and leaks IsADirectoryError out of
    verify_pin's read_bytes, past the named guard. The PINNED name is deliberate: an
    unpinned name would make the exists() form raise LoaderError('no sha256 pin') — a
    pass for the wrong reason."""
    (tmp_path / _RMR).mkdir()
    with pytest.raises(LoaderError, match="not found"):
        load_population(_RMR, data_dir=tmp_path)


def test_unregistered_workbook_geography_set_raises(monkeypatch):
    """`WORKBOOK_GEOGRAPHIES.get(name, frozenset())` makes require_all_geographies pass
    trivially for an unregistered workbook (run-1 mutation survivor): the completeness
    gate is then unfalsifiable. The lookup must be REQUIRED."""
    monkeypatch.delitem(WORKBOOK_GEOGRAPHIES, _RMR)
    with pytest.raises(LoaderError, match="no expected-geography set registered"):
        load_population(_RMR)


def test_missing_modeled_geography_raises(monkeypatch):
    """Positive completeness: a modeled geography that never appears must raise — the
    guard against a renamed label silently vanishing from the rankings."""
    monkeypatch.setitem(WORKBOOK_GEOGRAPHIES, _RMR,
                        WORKBOOK_GEOGRAPHIES[_RMR] | {Geography.LAVAL_RA13})
    with pytest.raises(LoaderError, match="LAVAL_RA13"):
        load_population(_RMR)


def test_workbook_with_no_modeled_geography_raises():
    """pop-as-qc-base ships ONE row label ('Le Québec' → IGNORED), so every row is
    dropped: the QC-total workbook is NOT a T1 population source (WORKBOOK_GEOGRAPHIES
    declares frozenset() for it, and the pipeline's _POP_WORKBOOKS excludes it). An empty
    frame must raise, never return."""
    with pytest.raises(LoaderError, match="no modeled geography"):
        load_population("pop-as-qc-base.xlsx")


# --- sex + scenario junction gates (spec §8, absent from the plan body) --------------

def test_unknown_sex_code_raises():
    """Spec §8: 'Any other code → raise'. The plan body dropped unknown codes silently
    via `.isin(SEX_CODE_TO_GENDER)`."""
    long = pd.DataFrame({
        "geography": [Geography.MTL_RMR] * 3, "scenario": [Scenario.REFERENCE] * 3,
        "year": [2035] * 3, "age": [75] * 3, "status": ["proj"] * 3,
        "sex_code": [1, 2, 4], "population": [10.0, 12.0, 22.0],
    })
    with pytest.raises(LoaderError, match="sex code"):
        isq._sex_lattice(long, "test")


@pytest.mark.parametrize("sex_codes, ages, why", [
    ([1, 2, 3, 1, 2], [75, 75, 75, 76, 76], "age 76 has no code-3 row"),
    ([1, 2, 3, 3], [75, 75, 75, 75], "age 75 carries a DUPLICATE code-3 row"),
])
def test_incomplete_sex_lattice_raises(sex_codes, ages, why):
    """pivot_table silently omits NaN-keyed rows and silently SUMS duplicates — either
    would let the additivity check pass over data it never saw."""
    long = pd.DataFrame({
        "geography": [Geography.MTL_RMR] * len(ages), "scenario": [Scenario.REFERENCE] * len(ages),
        "year": [2035] * len(ages), "age": ages, "status": ["proj"] * len(ages),
        "sex_code": sex_codes, "population": [10.0] * len(ages),
    })
    with pytest.raises(LoaderError, match="sex lattice"):
        isq._sex_lattice(long, "test")


def test_missing_scenario_for_a_geography_year_raises():
    """Spec §8 Scenario: 'missing any of the three for a geography×year → raise'. The plan
    body mapped the labels but never checked completeness."""
    long = pd.DataFrame({
        "geography": [Geography.MTL_RMR, Geography.MTL_RMR],
        "year": [2035, 2035],
        "scenario": [Scenario.REFERENCE, Scenario.LOW],   # 'high' fan absent
    })
    with pytest.raises(LoaderError, match="missing scenario"):
        isq._check_scenario_completeness(long, "test")


# --- frame-level gate WIRING, mutated at the Task-10 seam ----------------------------
# The four §4 frame gates (PK / pinned year lattice / uniform year domain / Statut
# sub-lattice) have their PREDICATES unit-tested in test_validate.py (Task 8b). What is
# untested is that THIS loader calls them with the right frame and the right group keys —
# deleting a call leaves the suite green. These REDs substitute the output of the Task-10
# builder (the loader's own input contract) instead of re-writing a 2.5 MB workbook per
# case: ~0.6s vs ~5s each. The real-file path stays covered by the two workbook-mutation
# REDs below.

@pytest.fixture(scope="module")
def rmr_long():
    return build_single_year_long(DATA_DIR / _RMR)


def _mtl_terminal_slice(long):
    """Mask of ONE geography's whole terminal-year slice (every scenario × every sex).

    That granularity is not cosmetic: the sex lattice gate runs FIRST and refuses any
    partial slice (a single sex row moved leaves its key short a code), so a per-row
    mutation can never reach the frame-level gates. It is also the shape a real edition
    drift has — spec §4: "a missing terminal year for one geography must raise"."""
    labels = {label for label in long["label"].unique()
              if classify_geography(label) is Geography.MTL_RMR}
    return long["label"].isin(labels) & (long["year"] == 2051)


@pytest.mark.parametrize("mutate, expect", [
    (lambda df, mask: df.assign(year=df["year"].mask(mask, 2052)), "year span"),
    (lambda df, mask: df.assign(status=df["status"].mask(mask, "p")), "reversal|PROJECTED"),
    (lambda df, mask: df[~mask], "non-uniform year domain"),
    # 'Fort (E2026)' is the measured ISQ label of the high fan (geography.py's junction map).
    (lambda df, mask: df[~(mask & (df["scenario_label"] == "Fort (E2026)"))], "missing scenario"),
    # The PRIMARY KEY excludes `status`, so a transition year shipping the same block under
    # two Statut values duplicates the key while every sex-lattice invariant holds: both
    # Statut blocks pivot to complete {1,2,3} keys, the row-count identity is intact, and
    # additivity/orientation duplicate proportionally. Only the PK gate can see this.
    (lambda df, mask: pd.concat([df, df[mask].assign(status="p")], ignore_index=True),
     "duplicate primary key"),
], ids=["terminal-year-past-the-pinned-span", "proj-relabelled-provisional",
        "one-geography-loses-its-terminal-year", "one-geography-year-loses-the-high-fan",
        "same-block-under-two-statut-values"])
def test_frame_level_gates_are_wired(monkeypatch, rmr_long, mutate, expect):
    mutated = mutate(rmr_long, _mtl_terminal_slice(rmr_long))
    monkeypatch.setattr(isq, "build_single_year_long", lambda path: mutated)
    with pytest.raises(LoaderError, match=expect):
        load_population(_RMR)


# --- §4 degenerate-population gate (codex r4-F3), mutated at the same seam -----------
# "finite non-negative populations" had no executable check: the gate could be deleted
# wholesale with the suite green. Each construction below leaves EVERY other gate satisfied
# — the target cell sits at age 50 (outside the 85+ orientation window), the sex lattice
# stays complete, and additivity is preserved wherever the arithmetic allows it — so the
# finiteness gate is the only thing that can raise. The `match` strings carry the CAUSE
# COUNTS ("1 non-finite and 0 negative"), which pins which half of the gate fired.

_CELL = {"scenario_label": "Référence (A2026)", "year": 2035, "age": 50}


def _cell_mask(long, sex_code):
    """Mask of the ONE (MTL_RMR × reference × 2035 × age 50 × sex_code) row. Cardinality is
    ASSERTED: a geometry drift must fail the test, not mutate nothing (a vacuous pass)."""
    labels = {label for label in long["label"].unique()
              if classify_geography(label) is Geography.MTL_RMR}
    mask = (long["label"].isin(labels)
            & (long["scenario_label"] == _CELL["scenario_label"])
            & (long["year"] == _CELL["year"])
            & (long["age"] == _CELL["age"])
            & (long["sex_code"] == sex_code))
    assert int(mask.sum()) == 1, (
        f"fixture drift: {int(mask.sum())} row(s) for sex code {sex_code} at {_CELL}")
    return mask


def _nan_population(long):
    """NaN in one code-1 cell. Additivity canNOT be preserved through NaN (pivot_table's
    `sum` SKIPS it, so the code-1 total silently becomes 0.0) — with the gate neutered the
    additivity check raises with a different message, so the `match` string discriminates."""
    return long.assign(population=long["population"].mask(_cell_mask(long, 1), float("nan")))


def _inf_population(long):
    """+Inf in code 1 AND in its code-3 total: additivity SURVIVES intact
    (`math.isclose(inf, inf)` is True), so nothing but the finiteness gate can see this."""
    pop = long["population"].mask(_cell_mask(long, 1), float("inf"))
    return long.assign(population=pop.mask(_cell_mask(long, 3), float("inf")))


def _negative_population(long):
    """code1 → −code1 with the code-3 total re-derived to code2 − code1, so additivity stays
    byte-exact (measured residual 0.0 in the pinned edition). With the gate neutered this
    frame loads CLEAN and a negative population reaches the model."""
    m1, m3 = _cell_mask(long, 1), _cell_mask(long, 3)
    code1 = float(long.loc[m1, "population"].iloc[0])
    code2 = float(long.loc[_cell_mask(long, 2), "population"].iloc[0])
    return long.assign(population=long["population"].mask(m1, -code1).mask(m3, code2 - code1))


@pytest.mark.parametrize("mutate, expect", [
    (_nan_population, "1 non-finite and 0 negative"),
    (_inf_population, "2 non-finite and 0 negative"),
    (_negative_population, "0 non-finite and 1 negative"),
], ids=["nan-population-cell", "inf-population-cell", "negative-population-cell"])
def test_degenerate_population_raises(monkeypatch, rmr_long, mutate, expect):
    """Spec §4 degenerate policy / codex r4-F3: a non-finite or negative population cell
    must RAISE, never flow into the cohort roll-forward (a negative cohort silently reverses
    the excess-demand ranking it feeds)."""
    mutated = mutate(rmr_long)
    monkeypatch.setattr(isq, "build_single_year_long", lambda path: mutated)
    with pytest.raises(LoaderError, match=expect):
        load_population(_RMR)


def test_unknown_geography_label_raises_end_to_end(monkeypatch, rmr_long):
    """The geography junction's THIRD leg. modeled → enum and known-unmodeled → IGNORED are
    both asserted by test_load_population_returns_tidy_long_frame; "a label outside the
    verified set → raise" had no check at the LOADER (test_geography.py pins the PREDICATE
    only, so swapping this call site for a fail-open `_LABEL_TO_GEOGRAPHY.get(., IGNORED)`
    left the suite green).

    The victim is an ALREADY-IGNORED label, which is what makes the RED discriminating: under
    a fail-open lookup the retyped rows classify IGNORED, get dropped, and the workbook loads
    CLEAN. Retyping a MODELED label instead would raise for the wrong reason (that geography
    goes missing → require_all_geographies). The replacement carries no trailing digit run —
    `normalize_label` strips one, which would map a '…2' spelling straight back."""
    victim = "RMR de Saguenay "          # measured byte-exact spelling, trailing space
    assert classify_geography(victim) is IGNORED, "fixture drift: victim is no longer IGNORED"
    mask = rmr_long["label"] == victim
    assert int(mask.sum()) > 0, "fixture drift: the IGNORED victim label is gone"
    mutated = rmr_long.assign(label=rmr_long["label"].mask(mask, "RMR de Saguenay–Nouvelle"))
    monkeypatch.setattr(isq, "build_single_year_long", lambda path: mutated)
    with pytest.raises(LoaderError, match="outside verified set"):
        load_population(_RMR)


# --- end-to-end schema-drift REDs on mutated COPIES ----------------------------------

def _positions(name: str):
    """MEASURE the sheet geometry: (raw frame, body frame, id col positions, age→col).
    All 0-indexed; the body frame keeps the raw 0-indexed sheet rows as its index."""
    raw = pd.read_excel(DATA_DIR / name, sheet_name=SHEET, header=None, engine="openpyxl")
    group_row = next(
        row for row in range(min(30, len(raw)))
        if {"Scénario", "Année", "Statut", "Sexe"} <= {str(v).strip() for v in raw.iloc[row]}
    )
    header, labels = raw.iloc[group_row], raw.iloc[group_row + 1]
    groups = header.ffill()
    pos: dict[str, int] = {}
    age_col: dict[int, int] = {}
    for col in range(raw.shape[1]):
        token = str(header.iloc[col]).strip()
        if token in _ID_TOKENS:
            pos.setdefault(token, col)
        if str(groups.iloc[col]).strip() == _AGE_GROUP:
            try:
                age = int(float(str(labels.iloc[col]).strip()))
            except ValueError:
                continue                      # spacer / '100+' / 'Âge moyen' / 'Âge médian'
            age_col.setdefault(age, col)
    return raw, raw.iloc[group_row + 3:], pos, age_col


def _mtl_rows(body, pos):
    """Body rows of a MODELED geography. A mutation on an IGNORED row (the workbook's first
    'Sexe' rows belong to 'Le Québec') is DROPPED before every gate and cannot raise —
    measured: the additivity RED below passed vacuously until it was scoped this way."""
    geo_col = pos.get("Région1", pos.get("Région"))
    return body[[classify_geography(v) is Geography.MTL_RMR for v in body.iloc[:, geo_col]]]


def _mutated_dir(tmp_path, monkeypatch, name: str, edits, repin: bool = True) -> object:
    """Copy `name` into tmp_path, apply `edits` = [(row0, col0, pre_image, new)], re-pin
    the copy, and return the data_dir to load from. `repin=False` leaves the pin table
    pointing at the pristine edition — only test_unpinned_workbook_bytes_raise_end_to_end
    wants that, and it wants it as the property under test."""
    dst = tmp_path / name
    shutil.copyfile(DATA_DIR / name, dst)
    book = load_workbook(dst)
    sheet = book[SHEET]
    for row0, col0, pre, new in edits:
        cell = sheet.cell(row=row0 + 1, column=col0 + 1)      # openpyxl is 1-indexed
        assert cell.value == pre, (
            f"fixture geometry drift at 0-indexed ({row0}, {col0}): expected {pre!r}, got {cell.value!r}"
        )
        cell.value = new
    book.save(dst)
    if repin:
        monkeypatch.setitem(pins.WORKBOOK_SHA256, name,
                            hashlib.sha256(dst.read_bytes()).hexdigest())
    return tmp_path


def test_unpinned_workbook_bytes_raise_end_to_end(tmp_path, monkeypatch):
    """load_population must VERIFY the pin, not merely own a pin table. Every other workbook
    RED re-pins its copy, so `verify_pin(path, name)` could be deleted from `_resolve_path`
    with the suite green while a workbook that is not the pinned edition loads clean — the
    spec §4/§7c edition tripwire gone silently. test_pins.py covers the PREDICATE only.

    The edit touches the sheet's TITLE cell, which no loader code path reads (the header row
    is LOCATED by its id tokens, and this row has none), so the copy is a faithful model of
    the real hazard: an edition whose data reads back identically but whose bytes are not the
    ones that were reviewed. Matched on 'sha256 drift', never bare 'sha256' — the
    unregistered-name branch says 'no sha256 pin registered' (test_pins.py:25)."""
    raw, _body, _pos, _age_col = _positions(_RMR)
    title = raw.iat[0, 0]
    assert isinstance(title, str) and "Population selon" in title, "fixture drift: title cell"
    bad_dir = _mutated_dir(tmp_path, monkeypatch, _RMR,
                           [(0, 0, title, title + " [edition B]")], repin=False)
    with pytest.raises(LoaderError, match="sha256 drift"):
        load_population(_RMR, data_dir=bad_dir)


def test_sex_additivity_drift_raises_end_to_end(tmp_path, monkeypatch):
    """The scalar RED above proves the predicate; only a real workbook proves the LOADER
    calls it. Perturb ONE code-3 (both-sexes) cell: additivity is byte-exact in the pinned
    edition (measured max |code3 − (code1+code2)| = 0.0), so +1000 must raise."""
    raw, body, pos, age_col = _positions(_RMR)
    mtl = _mtl_rows(body, pos)
    row0 = int(mtl.index[mtl.iloc[:, pos["Sexe"]] == 3][0])
    col0 = age_col[50]
    pre = raw.iat[row0, col0]
    bad_dir = _mutated_dir(tmp_path, monkeypatch, _RMR, [(row0, col0, pre, float(pre) + 1000.0)])
    with pytest.raises(LoaderError, match="additivity"):
        load_population(_RMR, data_dir=bad_dir)


def test_swapped_sex_codes_raise_orientation_end_to_end(tmp_path, monkeypatch):
    """Swap the `Sexe` CODES of one MTL_RMR (scenario, year) pair — 2 cells. Additivity
    (code3 = code1+code2 is label-symmetric), the sex lattice, the primary key, scenario
    completeness and the year domain ALL survive this mutation; the orientation guard is
    the only thing standing between a swapped 1↔2 map and sex-specific mortality applied
    to the wrong cohort (codex r2-F5)."""
    _raw, body, pos, _age_col = _positions(_RMR)
    mtl = _mtl_rows(body, pos)
    male = mtl[mtl.iloc[:, pos["Sexe"]] == 1].iloc[0]
    female = mtl[(mtl.iloc[:, pos["Sexe"]] == 2)
                 & (mtl.iloc[:, pos["Scénario"]] == male.iloc[pos["Scénario"]])
                 & (mtl.iloc[:, pos["Année"]] == male.iloc[pos["Année"]])].iloc[0]
    bad_dir = _mutated_dir(tmp_path, monkeypatch, _RMR, [
        (int(male.name), pos["Sexe"], 1, 2),
        (int(female.name), pos["Sexe"], 2, 1),
    ])
    with pytest.raises(LoaderError, match="orientation"):
        load_population(_RMR, data_dir=bad_dir)
