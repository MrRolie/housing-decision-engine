"""Task 10 contract: the ISQ single-year age-block builder (spec §8 Age junction).

The first two tests are the plan's numbered contract verbatim. The two schema-drift
REDs below are required by the plan's own traceability table ("Loader schema-drift
raises (age-block group, header tokens) | 10 | test_isq_ages.py") and were absent
from the plan's Task 10 body — the mutations run on COPIES written to tmp_path, never
against demoflow/data/.
"""
import shutil

import pandas as pd
import pytest
from openpyxl import load_workbook

from demoflow.errors import LoaderError
from demoflow.loaders.isq_ages import SHEET, build_single_year_long
from demoflow.loaders.pins import DATA_DIR

# openpyxl is 1-indexed; pandas 0-indexed. pop-as-rmr-base "Années d'âge" header GROUP
# row is 0-indexed 6 -> openpyxl row 7; 'Sexe' at 0-indexed col 5 -> openpyxl col 6;
# the single-year 'Âge' group anchor at 0-indexed col 13 -> openpyxl col 14.
_GROUP_ROW_XL = 7
_SEXE_COL_XL = 6
_AGE_GROUP_COL_XL = 14

# --- pairing reconciliation (test_population_stays_paired_with_its_age) --------------
# The sheet's grouped-age reference block: level-0 group "Groupe d'âge", level-1 bands.
# 0-19 / 20-64 / 65+ partition 0..100 exactly; 85+ is a SUB-band of 65+ (the spec's
# orientation bucket) and so also discriminates permutations inside the terminal tail.
_BANDS = {"0-19": (0, 19), "20-64": (20, 64), "65+": (65, 100), "85+": (85, 100)}
_GROUPED_GROUP = "Groupe d'âge"
_ID_TOKENS = ("Scénario", "Année", "Statut", "Sexe", "Région1", "Région")
_KEYS = ["label", "scenario_label", "year", "status", "sex_code"]
# 0-indexed level-0 GROUP header row, MEASURED per workbook (pre-image asserted below).
# Both geometry families are covered: 'Région1'/offset 6 and 'Région'/offset 4.
_GROUP_ROW_0IDX = {"pop-as-rmr-base.xlsx": 6, "pop-as-qc-base.xlsx": 4}


def _mutated_copy(tmp_path, row_xl: int, col_xl: int, expect: str, new: str):
    """Copy pop-as-rmr-base.xlsx to tmp_path and overwrite one header cell.
    Asserts the pre-image so a geometry drift fails the TEST, never passes silently."""
    src = DATA_DIR / "pop-as-rmr-base.xlsx"
    dst = tmp_path / "mutated.xlsx"
    shutil.copyfile(src, dst)
    wb = load_workbook(dst)
    cell = wb[SHEET].cell(row=row_xl, column=col_xl)
    assert cell.value == expect, f"fixture geometry drift: expected {expect!r}, got {cell.value!r}"
    cell.value = new
    wb.save(dst)
    return dst


def test_builder_yields_pre_junction_long_rows():
    df = build_single_year_long(DATA_DIR / "pop-as-rmr-base.xlsx")
    assert set(df.columns) == {
        "label", "scenario_label", "year", "status", "sex_code", "age", "population",
    }
    # single-year ages only, 0..100 (100+ capped at 100), no grouped-age labels leaked
    assert df["age"].min() >= 0 and df["age"].max() == 100
    # numeric sex codes {1,2,3} preserved (additivity checked later, Task 11)
    assert set(df["sex_code"].unique()) <= {1, 2, 3}
    # scenario labels are the raw ISQ strings (mapped to enum in Task 11)
    assert "Référence (A2026)" in set(df["scenario_label"])


def test_builder_selects_single_year_block_not_grouped():
    df = build_single_year_long(DATA_DIR / "pop-as-rmr-base.xlsx")
    # single-year block => a contiguous 0..100 age set present for a given key
    ages = sorted(df["age"].unique())
    assert ages == list(range(0, 101))


@pytest.mark.parametrize("name", ["pop-as-ra-base.xlsx", "pop-as-qc-base.xlsx"])
def test_builder_handles_the_other_pop_workbooks(name):
    """Header offsets differ per workbook (pop-rmr=6, pop-ra=5, pop-qc=4) and pop-as-qc
    ships 'Région' (not 'Région1') and NO 'Code' column — Task 11 calls this builder for
    all three, so a hardcoded geometry would break it."""
    df = build_single_year_long(DATA_DIR / name)
    assert sorted(df["age"].unique()) == list(range(0, 101))
    assert set(df["sex_code"].unique()) <= {1, 2, 3}
    assert not df["label"].isna().any()


def test_age_group_label_drift_raises(tmp_path):
    """Header-GROUP selection is the spec-pinned method: rename the 'Âge' group anchor
    and the builder must refuse, never fall back to bare column names."""
    bad = _mutated_copy(tmp_path, _GROUP_ROW_XL, _AGE_GROUP_COL_XL, "Âge", "Ages")
    with pytest.raises(LoaderError, match="single-year"):
        build_single_year_long(bad)


def test_missing_id_header_token_raises(tmp_path):
    bad = _mutated_copy(tmp_path, _GROUP_ROW_XL, _SEXE_COL_XL, "Sexe", "Genre")
    with pytest.raises(LoaderError, match="header"):
        build_single_year_long(bad)


def _as_str(column: pd.Series):
    """Join key -> object array of str. BOTH sides are coerced through the same helper:
    a dtype mismatch on a merge key reads exactly like a real misalignment."""
    return column.astype(str).to_numpy(dtype=object)


def _as_i64(column: pd.Series):
    return column.astype("int64").to_numpy()


@pytest.mark.parametrize("name", sorted(_GROUP_ROW_0IDX))
def test_population_stays_paired_with_its_age(name):
    """Every population value must stay paired with ITS single-year age.

    The builder broadcasts the ids (`np.repeat`), the ages (`np.tile`) and the populations
    (`reshape`) INDEPENDENTLY, so a rotated age vector, a mistiled id column or a
    Fortran-order reshape would corrupt every value while leaving the age SET, the row
    count, the primary key and sex additivity intact — invisible to every other test here
    and to the Task 11 gate. Reconciled against the sheet's own grouped-age block, which
    the builder never reads: an independent, AGE-RESOLVED reference. The row 'TOTAL'
    column is NOT one — it is invariant under any within-row permutation, so a
    sum-vs-TOTAL check cannot fail for a misalignment. The single-year/grouped identity
    holds byte-exactly (measured 0.0 on all three workbooks) because the workbooks are
    sha256-pinned in pins.py; an edition that broke it would trip the pin first.
    """
    grp = _GROUP_ROW_0IDX[name]
    raw = pd.read_excel(DATA_DIR / name, sheet_name=SHEET, header=None, engine="openpyxl")
    header, labels = raw.iloc[grp], raw.iloc[grp + 1]
    assert {"Scénario", "Année", "Statut", "Sexe"} <= {str(v).strip() for v in header}, (
        f"fixture geometry drift: 0-indexed row {grp} of {name} is not the group header row"
    )

    id_pos: dict[str, int] = {}
    band_pos: dict[str, int] = {}
    groups = header.ffill()
    for col in range(raw.shape[1]):
        token = str(header.iloc[col]).strip()
        if token in _ID_TOKENS:
            id_pos.setdefault(token, col)
        if str(groups.iloc[col]).strip() == _GROUPED_GROUP:
            band_pos.setdefault(str(labels.iloc[col]).strip(), col)
    geo = id_pos.get("Région1", id_pos.get("Région"))
    assert set(_BANDS) <= set(band_pos), f"{name}: reference block drift {sorted(band_pos)}"

    body = raw.iloc[grp + 3:]
    expected = pd.DataFrame({
        "label": _as_str(body.iloc[:, geo]),
        "scenario_label": _as_str(body.iloc[:, id_pos["Scénario"]]),
        "year": _as_i64(body.iloc[:, id_pos["Année"]]),
        "status": _as_str(body.iloc[:, id_pos["Statut"]]),
        "sex_code": _as_i64(body.iloc[:, id_pos["Sexe"]]),
        **{band: body.iloc[:, band_pos[band]].astype(float).to_numpy() for band in _BANDS},
    })

    out = build_single_year_long(DATA_DIR / name)
    keyed = pd.DataFrame({
        "label": _as_str(out["label"]),
        "scenario_label": _as_str(out["scenario_label"]),
        "year": _as_i64(out["year"]),
        "status": _as_str(out["status"]),
        "sex_code": _as_i64(out["sex_code"]),
        "age": _as_i64(out["age"]),
        "population": out["population"].to_numpy(dtype=float),
    })

    for band, (low, high) in _BANDS.items():
        summed = (keyed[keyed["age"].between(low, high)]
                  .groupby(_KEYS, as_index=False)["population"].sum())
        # outer join + validate: an unmatched key means an id column is misbroadcast;
        # one_to_one pins the 5-tuple as the sheet's primary key.
        merged = pd.merge(summed, expected[_KEYS + [band]], on=_KEYS,
                          how="outer", validate="one_to_one")
        unmatched = int(merged.isna().any(axis=1).sum())
        assert unmatched == 0, (
            f"{name} band {band}: {unmatched} key(s) do not match the sheet's rows "
            f"— an id column is misbroadcast across the age block"
        )
        worst = (merged["population"] - merged[band]).abs().max()
        assert worst == 0.0, (
            f"{name} band {band}: single-year sum vs the grouped-age cell differs by "
            f"{worst} — population is not paired with its age"
        )
